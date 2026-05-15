"""
Test cases for Bulk Mentor Upload feature
Tests:
- Template download endpoint
- Bulk upload with valid Excel
- Bulk upload with invalid Excel (missing fields)
- Bulk upload with duplicate emails
- Verification that mentors are created as hidden
"""

import pytest
import requests
import os
import io

# Get BASE_URL from environment
BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

class TestBulkMentorUpload:
    """Test bulk mentor upload feature"""
    
    @pytest.fixture(autouse=True)
    def setup(self):
        """Setup - login as admin before each test"""
        self.session = requests.Session()
        # Login as admin
        response = self.session.post(f"{BASE_URL}/api/auth/mock-login?user_type=admin")
        assert response.status_code == 200, f"Admin login failed: {response.text}"
        self.admin_user = response.json()
        print(f"Logged in as admin: {self.admin_user.get('email')}")
    
    def test_template_download_returns_xlsx(self):
        """Test that template download returns a valid Excel file"""
        response = self.session.get(f"{BASE_URL}/api/admin/mentors/template")
        
        # Status code assertion
        assert response.status_code == 200, f"Template download failed: {response.status_code}"
        
        # Content type assertion
        content_type = response.headers.get('Content-Type', '')
        assert 'spreadsheet' in content_type or 'excel' in content_type or 'octet-stream' in content_type, \
            f"Unexpected content type: {content_type}"
        
        # Content disposition assertion
        content_disposition = response.headers.get('Content-Disposition', '')
        assert 'attachment' in content_disposition, "Missing attachment header"
        assert 'mentor_upload_template.xlsx' in content_disposition, "Missing filename in header"
        
        # File size assertion (should be non-empty)
        assert len(response.content) > 1000, f"Template file too small: {len(response.content)} bytes"
        
        print(f"Template downloaded successfully: {len(response.content)} bytes")
    
    def test_bulk_upload_valid_excel(self):
        """Test bulk upload with valid Excel file creates mentors"""
        # Create test Excel file
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mentors"
        
        # Headers
        headers = [
            "Full Name*", "Email*", "Phone*", "LinkedIn*", "Location*",
            "Consulting Position*", "Consulting Firm*", "Current Company*",
            "Consulting Is Current (Y/N)", "Previous Company 1", "Previous Company 2",
            "Years Experience*", "Hourly Rate*", "Session Price*", "Headline", "Top Coach (Y/N)"
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Test data with unique email
        import uuid
        unique_id = str(uuid.uuid4())[:8]
        test_email = f"bulktest_{unique_id}@example.com"
        
        test_data = [
            f"Bulk Test Mentor {unique_id}", test_email, "+91 98765 43210", "linkedin.com/in/bulktest",
            "Mumbai, India", "Senior Consultant", "McKinsey & Company", "Google",
            "N", "Amazon", "Microsoft", "8", "12000", "1500", "Test Mentor", "N"
        ]
        for col, value in enumerate(test_data, 1):
            ws.cell(row=2, column=col, value=value)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Upload
        files = {'file': ('test_mentors.xlsx', output, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = self.session.post(f"{BASE_URL}/api/admin/mentors/bulk-upload", files=files)
        
        # Status code assertion
        assert response.status_code == 200, f"Bulk upload failed: {response.text}"
        
        # Data assertions
        data = response.json()
        assert data.get('created') >= 1, f"Expected at least 1 mentor created, got {data.get('created')}"
        assert 'errors' in data, "Missing errors field in response"
        
        print(f"Bulk upload result: {data.get('created')} created, {len(data.get('errors', []))} errors")
        
        # Verify mentor was created
        mentors_response = self.session.get(f"{BASE_URL}/api/admin/mentors")
        assert mentors_response.status_code == 200
        mentors = mentors_response.json().get('mentors', [])
        
        # Find our test mentor
        test_mentor = next((m for m in mentors if m.get('email') == test_email), None)
        assert test_mentor is not None, f"Test mentor with email {test_email} not found"
        
        # Verify mentor is hidden by default
        assert test_mentor.get('is_hidden') == True, "Mentor should be hidden by default"
        
        print(f"Verified mentor created: {test_mentor.get('name')} (hidden={test_mentor.get('is_hidden')})")
    
    def test_bulk_upload_duplicate_email_shows_error(self):
        """Test that uploading duplicate email shows error"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        # First, get an existing mentor email
        mentors_response = self.session.get(f"{BASE_URL}/api/admin/mentors")
        assert mentors_response.status_code == 200
        mentors = mentors_response.json().get('mentors', [])
        
        if not mentors:
            pytest.skip("No existing mentors to test duplicate")
        
        existing_email = mentors[0].get('email')
        
        # Create Excel with duplicate email
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mentors"
        
        headers = [
            "Full Name*", "Email*", "Phone*", "LinkedIn*", "Location*",
            "Consulting Position*", "Consulting Firm*", "Current Company*",
            "Consulting Is Current (Y/N)", "Previous Company 1", "Previous Company 2",
            "Years Experience*", "Hourly Rate*", "Session Price*", "Headline", "Top Coach (Y/N)"
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        test_data = [
            "Duplicate Test", existing_email, "+91 98765 43210", "linkedin.com/in/duplicate",
            "Mumbai, India", "Consultant", "BCG", "BCG",
            "Y", "", "", "5", "10000", "1000", "", "N"
        ]
        for col, value in enumerate(test_data, 1):
            ws.cell(row=2, column=col, value=value)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Upload
        files = {'file': ('duplicate_test.xlsx', output, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = self.session.post(f"{BASE_URL}/api/admin/mentors/bulk-upload", files=files)
        
        # Status code assertion
        assert response.status_code == 200, f"Bulk upload failed: {response.text}"
        
        # Data assertions
        data = response.json()
        assert data.get('created') == 0, f"Expected 0 mentors created for duplicate, got {data.get('created')}"
        assert len(data.get('errors', [])) > 0, "Expected error for duplicate email"
        
        # Verify error message mentions duplicate
        errors = data.get('errors', [])
        has_duplicate_error = any('already exists' in err.lower() for err in errors)
        assert has_duplicate_error, f"Expected 'already exists' error, got: {errors}"
        
        print(f"Duplicate email error correctly shown: {errors}")
    
    def test_bulk_upload_invalid_file_type_rejected(self):
        """Test that non-Excel files are rejected"""
        # Create a text file
        files = {'file': ('test.txt', b'This is not an Excel file', 'text/plain')}
        response = self.session.post(f"{BASE_URL}/api/admin/mentors/bulk-upload", files=files)
        
        # Should return 400 error
        assert response.status_code == 400, f"Expected 400 for invalid file type, got {response.status_code}"
        
        data = response.json()
        assert 'excel' in data.get('detail', '').lower() or 'xlsx' in data.get('detail', '').lower(), \
            f"Expected error about Excel file, got: {data.get('detail')}"
        
        print(f"Invalid file type correctly rejected: {data.get('detail')}")
    
    def test_mentors_created_as_hidden_by_default(self):
        """Test that bulk uploaded mentors are hidden by default"""
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        
        # Create test Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mentors"
        
        headers = [
            "Full Name*", "Email*", "Phone*", "LinkedIn*", "Location*",
            "Consulting Position*", "Consulting Firm*", "Current Company*",
            "Consulting Is Current (Y/N)", "Previous Company 1", "Previous Company 2",
            "Years Experience*", "Hourly Rate*", "Session Price*", "Headline", "Top Coach (Y/N)"
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        import uuid
        unique_id = str(uuid.uuid4())[:8]
        test_email = f"hiddentest_{unique_id}@example.com"
        
        test_data = [
            f"Hidden Test {unique_id}", test_email, "+91 98765 43210", "linkedin.com/in/hiddentest",
            "Delhi, India", "Manager", "Bain", "Bain",
            "Y", "", "", "6", "15000", "2000", "Hidden Test", "N"
        ]
        for col, value in enumerate(test_data, 1):
            ws.cell(row=2, column=col, value=value)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Upload
        files = {'file': ('hidden_test.xlsx', output, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = self.session.post(f"{BASE_URL}/api/admin/mentors/bulk-upload", files=files)
        
        assert response.status_code == 200
        data = response.json()
        assert data.get('created') >= 1, "Expected mentor to be created"
        
        # Verify mentor is hidden
        mentors_response = self.session.get(f"{BASE_URL}/api/admin/mentors")
        mentors = mentors_response.json().get('mentors', [])
        
        test_mentor = next((m for m in mentors if m.get('email') == test_email), None)
        assert test_mentor is not None, "Test mentor not found"
        assert test_mentor.get('is_hidden') == True, f"Mentor should be hidden, got is_hidden={test_mentor.get('is_hidden')}"
        
        print(f"Verified mentor is hidden by default: {test_mentor.get('name')}")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
