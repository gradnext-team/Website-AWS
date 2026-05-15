"""
Case Competition API
Timed quiz competitions for consulting interview practice
"""
from fastapi import APIRouter, HTTPException, Request, Depends
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
from bson import ObjectId
import random
import uuid
import io

router = APIRouter()

# ============ Helper for datetime handling ============

def ensure_utc(dt):
    """Ensure datetime is timezone-aware UTC"""
    if dt is None:
        return None
    if dt.tzinfo is None:
        # Assume naive datetime is UTC
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)

def serialize_datetime(dt):
    """Serialize datetime to ISO format with Z suffix for UTC"""
    if dt is None:
        return None
    if isinstance(dt, str):
        return dt if dt.endswith('Z') else dt + 'Z'
    # Ensure it's UTC and format with Z suffix
    utc_dt = ensure_utc(dt)
    return utc_dt.strftime('%Y-%m-%dT%H:%M:%SZ')

def serialize_competition(comp):
    """Serialize competition with proper datetime formatting"""
    if comp is None:
        return None
    result = dict(comp)
    # Ensure datetime fields have Z suffix for proper JS parsing
    for field in ['quiz_start_time', 'quiz_end_time', 'window_start', 'window_end', 'created_at', 'updated_at']:
        if field in result and result[field]:
            result[field] = serialize_datetime(result[field])
    return result

# ============ Pydantic Models ============

class CompetitionCreate(BaseModel):
    name: str
    description: Optional[str] = ""
    rules: Optional[str] = ""  # Competition rules/instructions shown before starting
    show_in_nav: bool = True  # Whether to show in navigation sidebar
    window_start: datetime   # When candidates CAN START taking the quiz
    window_end: datetime     # Deadline to START the quiz (after this, no new attempts)
    quiz_duration_minutes: int = 10  # Time limit per user once they start
    questions_per_user: int = 10  # How many questions each user sees
    scoring: Dict[str, int] = {"correct": 3, "wrong": -1, "skip": 0}
    is_active: bool = True

class CompetitionUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    rules: Optional[str] = None
    show_in_nav: Optional[bool] = None  # Toggle navigation visibility
    window_start: Optional[datetime] = None
    window_end: Optional[datetime] = None
    quiz_duration_minutes: Optional[int] = None
    questions_per_user: Optional[int] = None
    scoring: Optional[Dict[str, int]] = None
    is_active: Optional[bool] = None

class QuestionCreate(BaseModel):
    competition_id: str
    question: str
    question_type: str  # "multiple_choice", "text_input", "numerical"
    options: Optional[List[str]] = None  # For MCQ
    correct_answer: str
    acceptable_answers: Optional[List[str]] = None  # For text/numerical
    explanation: Optional[str] = ""
    category: str  # "guesstimate", "case_math", "structuring", etc.
    difficulty: str = "medium"  # easy, medium, hard

class AnswerSubmit(BaseModel):
    question_id: str
    answer: Optional[str] = None  # None means skipped
    time_taken_seconds: int = 0

# ============ Helper Functions ============

async def get_current_user(request: Request):
    """Get current user from session - supports both cookie and Authorization header"""
    db = request.app.state.db
    
    # Try to get token from cookie first
    session_token = request.cookies.get('session_token')
    
    # If no cookie, try Authorization header (for mobile Safari and production)
    if not session_token:
        auth_header = request.headers.get('Authorization', '')
        if auth_header.startswith('Bearer '):
            session_token = auth_header[7:]  # Remove 'Bearer ' prefix
    
    if not session_token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Check user_sessions collection (used by auth routes)
    session = await db.user_sessions.find_one({"session_token": session_token})
    
    # Also try auth_token field (JWT tokens)
    if not session:
        session = await db.user_sessions.find_one({"auth_token": session_token})
    
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    
    user_id = session["user_id"]
    
    # Handle both ObjectId and string user IDs (mock users use string IDs)
    try:
        user = await db.users.find_one({"_id": ObjectId(user_id)})
    except:
        # Try with string ID (for mock users)
        user = await db.users.find_one({"id": user_id})
    
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    
    return user

async def get_optional_user(request: Request):
    """Get current user from session if logged in, otherwise return None"""
    db = request.app.state.db
    
    # Try to get token from cookie first
    session_token = request.cookies.get('session_token')
    
    # If no cookie, try Authorization header (for mobile Safari and production)
    if not session_token:
        auth_header = request.headers.get('Authorization', '')
        if auth_header.startswith('Bearer '):
            session_token = auth_header[7:]  # Remove 'Bearer ' prefix
    
    if not session_token:
        return None
    
    # Check user_sessions collection (used by auth routes)
    session = await db.user_sessions.find_one({"session_token": session_token})
    
    # Also try auth_token field (JWT tokens)
    if not session:
        session = await db.user_sessions.find_one({"auth_token": session_token})
    
    if not session:
        return None
    
    user_id = session.get("user_id")
    
    # Handle both ObjectId and string user IDs (mock users use string IDs)
    try:
        user = await db.users.find_one({"_id": ObjectId(user_id)})
        if user:
            return user
        # Try with string ID (for mock users)
        user = await db.users.find_one({"id": user_id})
        return user
    except:
        return None

async def get_admin_user(request: Request):
    """Get current admin user"""
    user = await get_current_user(request)
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")
    return user

# ============ Admin Endpoints ============

@router.post("/admin/competitions")
async def create_competition(request: Request, data: CompetitionCreate):
    """Create a new competition with flexible time window"""
    await get_admin_user(request)
    db = request.app.state.db
    
    # Validate that window_end is after window_start
    if data.window_end <= data.window_start:
        raise HTTPException(status_code=400, detail="Competition close time must be after open time")
    
    competition = {
        "id": str(uuid.uuid4()),
        "name": data.name,
        "description": data.description,
        "rules": data.rules,
        "show_in_nav": data.show_in_nav,
        "window_start": data.window_start,  # When candidates CAN START
        "window_end": data.window_end,      # Deadline to START
        "quiz_duration_minutes": data.quiz_duration_minutes,  # Time limit once started
        # Legacy fields for backward compatibility
        "quiz_start_time": data.window_start,
        "quiz_end_time": data.window_end,
        "duration_minutes": data.quiz_duration_minutes,
        "questions_per_user": data.questions_per_user,
        "scoring": data.scoring,
        "is_active": data.is_active,
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.competitions.insert_one(competition)
    
    # Serialize with proper datetime formatting
    return {"success": True, "competition": serialize_competition({k: v for k, v in competition.items() if k != "_id"})}

@router.get("/admin/competitions")
async def get_all_competitions(request: Request):
    """Get all competitions (admin)"""
    await get_admin_user(request)
    db = request.app.state.db
    
    competitions = await db.competitions.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)
    
    # Serialize with proper datetime formatting and get question counts
    result = []
    for comp in competitions:
        count = await db.competition_questions.count_documents({"competition_id": comp["id"]})
        serialized = serialize_competition(comp)
        serialized["question_count"] = count
        result.append(serialized)
    
    return {"competitions": result}

@router.put("/admin/competitions/{competition_id}")
async def update_competition(request: Request, competition_id: str, data: CompetitionUpdate):
    """Update a competition"""
    await get_admin_user(request)
    db = request.app.state.db
    
    update_data = {k: v for k, v in data.dict().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc)
    
    # Handle new window-based fields
    if "window_start" in update_data:
        update_data["quiz_start_time"] = update_data["window_start"]  # Legacy compatibility
    if "window_end" in update_data:
        update_data["quiz_end_time"] = update_data["window_end"]  # Legacy compatibility
    if "quiz_duration_minutes" in update_data:
        update_data["duration_minutes"] = update_data["quiz_duration_minutes"]  # Legacy compatibility
    
    # Validate window times if both are being updated
    comp = await db.competitions.find_one({"id": competition_id})
    if comp:
        window_start = update_data.get("window_start", comp.get("window_start", comp.get("quiz_start_time")))
        window_end = update_data.get("window_end", comp.get("window_end", comp.get("quiz_end_time")))
        if window_end and window_start and window_end <= window_start:
            raise HTTPException(status_code=400, detail="Competition close time must be after open time")
    
    result = await db.competitions.update_one(
        {"id": competition_id},
        {"$set": update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Competition not found")
    
    return {"success": True}

@router.delete("/admin/competitions/{competition_id}")
async def delete_competition(request: Request, competition_id: str):
    """Delete a competition and its questions"""
    await get_admin_user(request)
    db = request.app.state.db
    
    await db.competitions.delete_one({"id": competition_id})
    await db.competition_questions.delete_many({"competition_id": competition_id})
    await db.competition_attempts.delete_many({"competition_id": competition_id})
    
    return {"success": True}

# ============ Question Management ============

@router.post("/admin/competitions/{competition_id}/questions")
async def add_question(request: Request, competition_id: str, data: QuestionCreate):
    """Add a question to a competition"""
    await get_admin_user(request)
    db = request.app.state.db
    
    # Verify competition exists
    comp = await db.competitions.find_one({"id": competition_id})
    if not comp:
        raise HTTPException(status_code=404, detail="Competition not found")
    
    question = {
        "id": str(uuid.uuid4()),
        "competition_id": competition_id,
        "question": data.question,
        "question_type": data.question_type,
        "options": data.options,
        "correct_answer": data.correct_answer,
        "acceptable_answers": data.acceptable_answers or [data.correct_answer],
        "explanation": data.explanation,
        "category": data.category,
        "difficulty": data.difficulty,
        "created_at": datetime.now(timezone.utc)
    }
    
    await db.competition_questions.insert_one(question)
    
    return {"success": True, "question": {k: v for k, v in question.items() if k != "_id"}}

@router.post("/admin/competitions/{competition_id}/questions/bulk")
async def add_questions_bulk(request: Request, competition_id: str):
    """Add multiple questions to a competition"""
    await get_admin_user(request)
    db = request.app.state.db
    
    body = await request.json()
    questions_data = body.get("questions", [])
    
    if not questions_data:
        raise HTTPException(status_code=400, detail="No questions provided")
    
    # Verify competition exists
    comp = await db.competitions.find_one({"id": competition_id})
    if not comp:
        raise HTTPException(status_code=404, detail="Competition not found")
    
    questions = []
    for q in questions_data:
        question = {
            "id": str(uuid.uuid4()),
            "competition_id": competition_id,
            "question": q["question"],
            "question_type": q.get("question_type", "multiple_choice"),
            "options": q.get("options"),
            "correct_answer": q["correct_answer"],
            "acceptable_answers": q.get("acceptable_answers", [q["correct_answer"]]),
            "explanation": q.get("explanation", ""),
            "category": q.get("category", "case_math"),
            "difficulty": q.get("difficulty", "medium"),
            "created_at": datetime.now(timezone.utc)
        }
        questions.append(question)
    
    if questions:
        await db.competition_questions.insert_many(questions)
    
    return {"success": True, "count": len(questions)}

@router.get("/admin/competitions/{competition_id}/questions")
async def get_competition_questions(request: Request, competition_id: str):
    """Get all questions for a competition (admin)"""
    await get_admin_user(request)
    db = request.app.state.db
    
    questions = await db.competition_questions.find(
        {"competition_id": competition_id},
        {"_id": 0}
    ).to_list(500)
    
    return {"questions": questions}

@router.delete("/admin/questions/{question_id}")
async def delete_question(request: Request, question_id: str):
    """Delete a question"""
    await get_admin_user(request)
    db = request.app.state.db
    
    await db.competition_questions.delete_one({"id": question_id})
    return {"success": True}

# ============ Import from Existing Drills ============

@router.post("/admin/competitions/{competition_id}/import-drills")
async def import_from_drills(request: Request, competition_id: str):
    """Import hard questions from existing drill database"""
    await get_admin_user(request)
    db = request.app.state.db
    
    body = await request.json()
    count = body.get("count", 10)
    # Note: difficulty and drill_types can be used for filtering in future
    # difficulty = body.get("difficulty", "advanced")
    # drill_types = body.get("drill_types", ["case_math", "case_structuring"])
    
    # This will pull from the PRE_GENERATED_DRILLS in ai_drills.py
    # For now, we'll create some hard questions directly
    
    imported = 0
    
    # Import from existing drill sessions if available
    existing_drills = await db.drill_questions.find(
        {"difficulty": {"$in": ["intermediate", "advanced"]}}
    ).to_list(count * 2)
    
    for drill in existing_drills[:count]:
        question = {
            "id": str(uuid.uuid4()),
            "competition_id": competition_id,
            "question": drill.get("question", ""),
            "question_type": drill.get("type", "multiple_choice"),
            "options": drill.get("options"),
            "correct_answer": drill.get("correct_answer", ""),
            "acceptable_answers": drill.get("acceptable_answers", []),
            "explanation": drill.get("explanation", ""),
            "category": drill.get("drill_type", "case_math"),
            "difficulty": drill.get("difficulty", "medium"),
            "created_at": datetime.now(timezone.utc)
        }
        await db.competition_questions.insert_one(question)
        imported += 1
    
    return {"success": True, "imported": imported}

# ============ User-Facing Endpoints (Public - No Login Required) ============

@router.get("/competitions/active")
async def get_active_competitions(request: Request):
    """Get active competitions visible to everyone (no login required)"""
    user = await get_optional_user(request)
    db = request.app.state.db
    
    now = datetime.now(timezone.utc)
    
    # Find competitions where show_in_nav is true and is_active
    competitions = await db.competitions.find(
        {
            "is_active": True,
            "show_in_nav": True  # Only show if navigation visibility is enabled
        },
        {"_id": 0}
    ).to_list(20)
    
    # Add user's attempt status to each (if logged in)
    user_id = str(user["_id"]) if user else None
    result = []
    for comp in competitions:
        # Serialize with proper datetime formatting
        serialized = serialize_competition(comp)
        
        if user_id:
            attempt = await db.competition_attempts.find_one({
                "competition_id": comp["id"],
                "user_id": user_id
            }, {"_id": 0})
            
            serialized["user_attempt"] = attempt
            serialized["has_started"] = attempt is not None
            serialized["has_submitted"] = attempt.get("submitted", False) if attempt else False
        else:
            serialized["user_attempt"] = None
            serialized["has_started"] = False
            serialized["has_submitted"] = False
        
        # Calculate time status (ensure timezone-aware comparison)
        quiz_start = ensure_utc(comp["quiz_start_time"])
        quiz_end = ensure_utc(comp["quiz_end_time"])
        
        if now < quiz_start:
            serialized["status"] = "upcoming"
            serialized["time_until_start"] = (quiz_start - now).total_seconds()
        elif now > quiz_end:
            serialized["status"] = "ended"
        else:
            serialized["status"] = "live"
            serialized["time_remaining"] = (quiz_end - now).total_seconds()
        
        result.append(serialized)
    
    return {"competitions": result}

@router.get("/competitions/{competition_id}")
async def get_competition_details(request: Request, competition_id: str):
    """Get competition details (no login required)"""
    user = await get_optional_user(request)
    db = request.app.state.db
    
    comp = await db.competitions.find_one({"id": competition_id}, {"_id": 0})
    if not comp:
        raise HTTPException(status_code=404, detail="Competition not found")
    
    # Serialize with proper datetime formatting
    result = serialize_competition(comp)
    
    user_id = str(user["_id"]) if user else None
    now = datetime.now(timezone.utc)
    
    # Get user's attempt if logged in
    if user_id:
        attempt = await db.competition_attempts.find_one({
            "competition_id": competition_id,
            "user_id": user_id
        }, {"_id": 0})
        result["user_attempt"] = attempt
    else:
        result["user_attempt"] = None
    
    # Calculate status (ensure timezone-aware comparison)
    quiz_start = ensure_utc(comp["quiz_start_time"])
    quiz_end = ensure_utc(comp["quiz_end_time"])
    
    if now < quiz_start:
        result["status"] = "upcoming"
        result["time_until_start"] = (quiz_start - now).total_seconds()
    elif now > quiz_end:
        result["status"] = "ended"
    else:
        result["status"] = "live"
        result["time_remaining"] = (quiz_end - now).total_seconds()
    
    return result

@router.post("/competitions/{competition_id}/start")
async def start_competition(request: Request, competition_id: str):
    """Start a competition attempt - assigns random questions to user"""
    user = await get_current_user(request)
    db = request.app.state.db
    
    user_id = str(user["_id"])
    now = datetime.now(timezone.utc)
    
    # Get competition
    comp = await db.competitions.find_one({"id": competition_id})
    if not comp:
        raise HTTPException(status_code=404, detail="Competition not found")
    
    # Ensure datetime comparisons work (handle naive vs aware)
    quiz_start = ensure_utc(comp["quiz_start_time"])
    quiz_end = ensure_utc(comp["quiz_end_time"])
    
    # Check if quiz is live
    if now < quiz_start:
        raise HTTPException(status_code=400, detail="Quiz has not started yet")
    
    if now > quiz_end:
        raise HTTPException(status_code=400, detail="Quiz has ended")
    
    # Check for existing attempt
    existing = await db.competition_attempts.find_one({
        "competition_id": competition_id,
        "user_id": user_id
    })
    
    if existing:
        # Return existing attempt (resume)
        # Get the questions for this attempt
        questions = await db.competition_questions.find(
            {"id": {"$in": existing["question_ids"]}},
            {"_id": 0, "correct_answer": 0, "acceptable_answers": 0, "explanation": 0}
        ).to_list(100)
        
        # Sort questions in the order they were assigned
        question_order = {qid: idx for idx, qid in enumerate(existing["question_ids"])}
        questions.sort(key=lambda q: question_order.get(q["id"], 0))
        
        # Calculate remaining time based on user's start time + quiz duration (NOT window end)
        quiz_duration = comp.get("quiz_duration_minutes", comp.get("duration_minutes", 10))
        user_started_at = existing.get("started_at", now)
        if isinstance(user_started_at, str):
            user_started_at = datetime.fromisoformat(user_started_at.replace('Z', '+00:00'))
        user_started_at = ensure_utc(user_started_at)
        user_quiz_end = user_started_at + timedelta(minutes=quiz_duration)
        time_remaining = (user_quiz_end - now).total_seconds()
        
        # Also check if window has ended (can't continue after window closes)
        window_end = ensure_utc(comp.get("window_end", comp.get("quiz_end_time")))
        if now > window_end:
            time_remaining = 0
        
        return {
            "attempt": {k: v for k, v in existing.items() if k != "_id"},
            "questions": questions,
            "time_remaining": max(0, time_remaining),
            "resumed": True
        }
    
    # Create new attempt - assign random questions
    all_questions = await db.competition_questions.find(
        {"competition_id": competition_id}
    ).to_list(500)
    
    if len(all_questions) < comp["questions_per_user"]:
        raise HTTPException(
            status_code=400, 
            detail=f"Not enough questions. Need {comp['questions_per_user']}, have {len(all_questions)}"
        )
    
    # Randomly select questions
    selected = random.sample(all_questions, comp["questions_per_user"])
    question_ids = [q["id"] for q in selected]
    
    attempt = {
        "id": str(uuid.uuid4()),
        "competition_id": competition_id,
        "user_id": user_id,
        "user_name": user.get("name", ""),
        "user_email": user.get("email", ""),
        "question_ids": question_ids,
        "answers": {},  # {question_id: {"answer": str, "time_taken": int, "is_correct": bool}}
        "current_question_index": 0,
        "score": 0,
        "started_at": now,
        "submitted": False,
        "submitted_at": None
    }
    
    await db.competition_attempts.insert_one(attempt)
    
    # Return questions without answers
    questions = []
    for q in selected:
        questions.append({
            "id": q["id"],
            "question": q["question"],
            "question_type": q["question_type"],
            "options": q.get("options"),
            "category": q["category"],
            "difficulty": q["difficulty"]
        })
    
    # Calculate remaining time based on quiz duration (user just started, so full duration)
    quiz_duration = comp.get("quiz_duration_minutes", comp.get("duration_minutes", 10))
    time_remaining = quiz_duration * 60  # Convert to seconds
    
    return {
        "attempt": {k: v for k, v in attempt.items() if k != "_id"},
        "questions": questions,
        "time_remaining": max(0, time_remaining),
        "resumed": False
    }

@router.post("/competitions/{competition_id}/answer")
async def submit_answer(request: Request, competition_id: str, data: AnswerSubmit):
    """Submit an answer for a question (cannot go back)"""
    user = await get_current_user(request)
    db = request.app.state.db
    
    user_id = str(user["_id"])
    now = datetime.now(timezone.utc)
    
    # Get competition
    comp = await db.competitions.find_one({"id": competition_id})
    if not comp:
        raise HTTPException(status_code=404, detail="Competition not found")
    
    # Check if quiz is still live (ensure timezone-aware comparison)
    quiz_end = ensure_utc(comp["quiz_end_time"])
    if now > quiz_end:
        raise HTTPException(status_code=400, detail="Quiz has ended")
    
    # Get attempt
    attempt = await db.competition_attempts.find_one({
        "competition_id": competition_id,
        "user_id": user_id
    })
    
    if not attempt:
        raise HTTPException(status_code=400, detail="No active attempt found")
    
    if attempt.get("submitted"):
        raise HTTPException(status_code=400, detail="Quiz already submitted")
    
    # Check if question already answered
    if data.question_id in attempt.get("answers", {}):
        raise HTTPException(status_code=400, detail="Question already answered - cannot go back")
    
    # Get the question to check answer
    question = await db.competition_questions.find_one({"id": data.question_id})
    if not question:
        raise HTTPException(status_code=404, detail="Question not found")
    
    # Determine if answer is correct
    scoring = comp.get("scoring", {"correct": 3, "wrong": -1, "skip": 0})
    question_type = question.get("question_type", "multiple_choice")
    
    if data.answer is None:
        # Skipped
        is_correct = None
        points = scoring.get("skip", 0)
    else:
        # Check answer based on question type
        if question_type == "multi_select":
            # Multi-select: answer is pipe-separated, compare as sets
            user_answers = set(a.strip().lower() for a in data.answer.split('|') if a.strip())
            correct_answers = set(a.strip().lower() for a in question["correct_answer"].split('|') if a.strip())
            is_correct = user_answers == correct_answers
        elif question_type == "ordering":
            # Ordering: answer is pipe-separated, compare exact order
            user_order = [a.strip().lower() for a in data.answer.split('|') if a.strip()]
            correct_order = [a.strip().lower() for a in question["correct_answer"].split('|') if a.strip()]
            is_correct = user_order == correct_order
        else:
            # Standard answer check (multiple_choice, text_input, numerical)
            acceptable = question.get("acceptable_answers", [question["correct_answer"]])
            acceptable_lower = [a.lower().strip() for a in acceptable]
            is_correct = data.answer.lower().strip() in acceptable_lower
        
        points = scoring.get("correct", 3) if is_correct else scoring.get("wrong", -1)
    
    # Update attempt
    answer_data = {
        "answer": data.answer,
        "time_taken": data.time_taken_seconds,
        "is_correct": is_correct,
        "points": points
    }
    
    await db.competition_attempts.update_one(
        {"id": attempt["id"]},
        {
            "$set": {
                f"answers.{data.question_id}": answer_data,
                "current_question_index": attempt["current_question_index"] + 1
            },
            "$inc": {"score": points}
        }
    )
    
    return {
        "success": True,
        "is_correct": is_correct,
        "points": points,
        "correct_answer": question["correct_answer"],
        "explanation": question.get("explanation", "")
    }

@router.post("/competitions/{competition_id}/submit")
async def submit_competition(request: Request, competition_id: str):
    """Submit the competition (finish quiz)"""
    user = await get_current_user(request)
    db = request.app.state.db
    
    user_id = str(user["_id"])
    now = datetime.now(timezone.utc)
    
    # Get attempt
    attempt = await db.competition_attempts.find_one({
        "competition_id": competition_id,
        "user_id": user_id
    })
    
    if not attempt:
        raise HTTPException(status_code=400, detail="No active attempt found")
    
    if attempt.get("submitted"):
        raise HTTPException(status_code=400, detail="Already submitted")
    
    # Calculate final score
    answers = attempt.get("answers", {})
    total_score = sum(a.get("points", 0) for a in answers.values())
    correct_count = sum(1 for a in answers.values() if a.get("is_correct") is True)
    wrong_count = sum(1 for a in answers.values() if a.get("is_correct") is False)
    skipped_count = sum(1 for a in answers.values() if a.get("is_correct") is None)
    
    # Mark as submitted
    await db.competition_attempts.update_one(
        {"id": attempt["id"]},
        {
            "$set": {
                "submitted": True,
                "submitted_at": now,
                "score": total_score,
                "correct_count": correct_count,
                "wrong_count": wrong_count,
                "skipped_count": skipped_count
            }
        }
    )
    
    return {
        "success": True,
        "score": total_score,
        "correct_count": correct_count,
        "wrong_count": wrong_count,
        "skipped_count": skipped_count,
        "total_questions": len(attempt["question_ids"])
    }

@router.get("/competitions/{competition_id}/results")
async def get_results(request: Request, competition_id: str):
    """Get user's results after submission"""
    user = await get_current_user(request)
    db = request.app.state.db
    
    user_id = str(user["_id"])
    
    attempt = await db.competition_attempts.find_one({
        "competition_id": competition_id,
        "user_id": user_id
    }, {"_id": 0})
    
    if not attempt:
        raise HTTPException(status_code=404, detail="No attempt found")
    
    # Get questions with answers for review
    questions = await db.competition_questions.find(
        {"id": {"$in": attempt["question_ids"]}},
        {"_id": 0}
    ).to_list(100)
    
    # Add user's answers to questions
    for q in questions:
        user_answer = attempt.get("answers", {}).get(q["id"], {})
        q["user_answer"] = user_answer.get("answer")
        q["is_correct"] = user_answer.get("is_correct")
        q["points"] = user_answer.get("points", 0)
    
    return {
        "attempt": attempt,
        "questions": questions
    }

@router.get("/competitions/{competition_id}/leaderboard")
async def get_leaderboard(request: Request, competition_id: str):
    """Get competition leaderboard (public - no login required)"""
    # No authentication required - leaderboard is public
    db = request.app.state.db
    
    # Get all submitted attempts
    attempts = await db.competition_attempts.find(
        {
            "competition_id": competition_id,
            "submitted": True
        },
        {"_id": 0}
    ).sort("score", -1).to_list(100)
    
    # Format for leaderboard
    leaderboard = []
    for idx, attempt in enumerate(attempts):
        leaderboard.append({
            "rank": idx + 1,
            "user_name": attempt.get("user_name", "Anonymous"),
            "score": attempt.get("score", 0),
            "correct_count": attempt.get("correct_count", 0),
            "time_taken": (attempt.get("submitted_at") - attempt.get("started_at")).total_seconds() if attempt.get("submitted_at") and attempt.get("started_at") else 0
        })
    
    return {"leaderboard": leaderboard}

# ============ Admin Stats ============

@router.get("/admin/competitions/{competition_id}/stats")
async def get_competition_stats(request: Request, competition_id: str):
    """Get competition statistics (admin)"""
    await get_admin_user(request)
    db = request.app.state.db
    
    # Get all attempts
    attempts = await db.competition_attempts.find(
        {"competition_id": competition_id},
        {"_id": 0}
    ).to_list(1000)
    
    total_participants = len(attempts)
    submitted_count = sum(1 for a in attempts if a.get("submitted"))
    
    scores = [a.get("score", 0) for a in attempts if a.get("submitted")]
    avg_score = sum(scores) / len(scores) if scores else 0
    max_score = max(scores) if scores else 0
    min_score = min(scores) if scores else 0
    
    return {
        "total_participants": total_participants,
        "submitted_count": submitted_count,
        "in_progress": total_participants - submitted_count,
        "average_score": round(avg_score, 2),
        "highest_score": max_score,
        "lowest_score": min_score
    }


@router.get("/admin/competitions/{competition_id}/participants")
async def get_competition_participants(request: Request, competition_id: str):
    """Get all participants/attempts for a competition (admin)"""
    await get_admin_user(request)
    db = request.app.state.db
    
    # Get all attempts for this competition
    attempts = await db.competition_attempts.find(
        {"competition_id": competition_id},
        {"_id": 0}
    ).sort("started_at", -1).to_list(1000)
    
    # Enrich with additional calculated fields
    participants = []
    for attempt in attempts:
        answers = attempt.get("answers", {})
        total_questions = len(attempt.get("question_ids", []))
        answered_count = len(answers)
        skipped_count = sum(1 for a in answers.values() if a.get("is_correct") is None)
        correct_count = sum(1 for a in answers.values() if a.get("is_correct") is True)
        wrong_count = sum(1 for a in answers.values() if a.get("is_correct") is False)
        
        participant = {
            "id": attempt.get("id"),
            "user_id": attempt.get("user_id"),
            "user_name": attempt.get("user_name", "Unknown"),
            "user_email": attempt.get("user_email", ""),
            "started_at": attempt.get("started_at"),
            "submitted_at": attempt.get("submitted_at"),
            "submitted": attempt.get("submitted", False),
            "score": attempt.get("score", 0),
            "total_questions": total_questions,
            "answered_count": answered_count,
            "skipped_count": skipped_count,
            "correct_count": correct_count,
            "wrong_count": wrong_count,
            "unanswered_count": total_questions - answered_count
        }
        participants.append(participant)
    
    return {"participants": participants, "total": len(participants)}


@router.get("/admin/competitions/{competition_id}/participants/export")
async def export_competition_participants(request: Request, competition_id: str):
    """Export all participants data as Excel file (admin)"""
    await get_admin_user(request)
    db = request.app.state.db
    
    # Get competition info
    competition = await db.competitions.find_one({"id": competition_id})
    if not competition:
        raise HTTPException(status_code=404, detail="Competition not found")
    
    # Get all attempts
    attempts = await db.competition_attempts.find(
        {"competition_id": competition_id},
        {"_id": 0}
    ).sort("started_at", -1).to_list(1000)
    
    # Create Excel workbook
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Participants"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Rank", "Candidate Name", "Email", "Started At", "Submitted At",
        "Duration (mins)", "Score", "Correct", "Wrong", "Skipped", 
        "Unanswered", "Total Questions", "Accuracy %", "Status"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # Sort participants by score for ranking
    sorted_attempts = sorted(
        attempts,
        key=lambda x: (x.get("submitted", False), x.get("score", 0)),
        reverse=True
    )
    
    # Data rows
    for row_idx, attempt in enumerate(sorted_attempts, 2):
        answers = attempt.get("answers", {})
        total_questions = len(attempt.get("question_ids", []))
        answered_count = len(answers)
        skipped_count = sum(1 for a in answers.values() if a.get("is_correct") is None)
        correct_count = sum(1 for a in answers.values() if a.get("is_correct") is True)
        wrong_count = sum(1 for a in answers.values() if a.get("is_correct") is False)
        unanswered = total_questions - answered_count
        
        # Calculate duration
        started = attempt.get("started_at")
        submitted = attempt.get("submitted_at")
        duration_mins = ""
        if started and submitted:
            try:
                start_dt = started if isinstance(started, datetime) else datetime.fromisoformat(str(started).replace('Z', '+00:00'))
                end_dt = submitted if isinstance(submitted, datetime) else datetime.fromisoformat(str(submitted).replace('Z', '+00:00'))
                duration_mins = round((end_dt - start_dt).total_seconds() / 60, 1)
            except:
                duration_mins = ""
        
        # Calculate accuracy
        accuracy = round((correct_count / answered_count * 100), 1) if answered_count > 0 else 0
        
        # Format datetime for Excel
        def format_dt(dt):
            if not dt:
                return ""
            try:
                if isinstance(dt, str):
                    dt = datetime.fromisoformat(dt.replace('Z', '+00:00'))
                return dt.strftime("%Y-%m-%d %H:%M:%S")
            except:
                return str(dt)
        
        row_data = [
            row_idx - 1 if attempt.get("submitted") else "-",  # Rank (only for submitted)
            attempt.get("user_name", "Unknown"),
            attempt.get("user_email", ""),
            format_dt(started),
            format_dt(submitted),
            duration_mins,
            attempt.get("score", 0),
            correct_count,
            wrong_count,
            skipped_count,
            unanswered,
            total_questions,
            f"{accuracy}%",
            "Completed" if attempt.get("submitted") else "In Progress"
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if col != 2 else 'left')
            
            # Color code status
            if col == 14:  # Status column
                if value == "Completed":
                    cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    
    # Auto-adjust column widths
    column_widths = [8, 25, 30, 20, 20, 15, 10, 10, 10, 10, 12, 15, 12, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    comp_name = competition.get("name", "competition").replace(" ", "_")
    filename = f"{comp_name}_participants_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
