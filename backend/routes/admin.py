"""
Admin Panel API Routes
Handles all admin operations for managing users, mentors, content, and programs
"""

from fastapi import APIRouter, HTTPException, Request, UploadFile, File, Form
from pydantic import BaseModel, EmailStr
from typing import List, Optional, Dict, Any
from datetime import datetime, timedelta, timezone
import uuid
import os
import shutil
import logging

from routes.auth import get_current_user, get_db
from services.wati_service import wati_service

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/admin", tags=["admin"])


# ============ Email Helper Functions ============

async def send_mentor_invite_email(db, mentor_email: str, mentor_name: str):
    """Send an email invite to a new mentor"""
    from routes.auth import send_email_via_gmail
    
    # Get the frontend URL for login
    frontend_url = os.environ.get("REACT_APP_BACKEND_URL", "").replace("/api", "")
    if not frontend_url or frontend_url == "/api":
        frontend_url = os.environ.get("FRONTEND_URL", "")
    
    login_url = f"{frontend_url}?login=mentor"
    
    subject = "Welcome to gradnext - You've been added as a Mentor!"
    
    html_content = f"""
    <html>
    <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px; background-color: #f8fafc;">
        <div style="background: white; border-radius: 16px; padding: 40px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
            <div style="text-align: center; margin-bottom: 30px;">
                <h1 style="color: #0ea5e9; margin: 0; font-size: 28px;">gradnext</h1>
                <p style="color: #64748b; margin-top: 5px;">Interview Preparation Platform</p>
            </div>
            
            <h2 style="color: #1e293b; margin-bottom: 20px;">Welcome, {mentor_name}! 🎉</h2>
            
            <p style="color: #475569; line-height: 1.6;">
                You've been added as a <strong>Mentor</strong> on gradnext! As a mentor, you'll be able to:
            </p>
            
            <ul style="color: #475569; line-height: 1.8;">
                <li>Conduct coaching sessions with candidates</li>
                <li>Manage your availability and schedule</li>
                <li>Provide feedback and guidance</li>
                <li>Track your sessions and earnings</li>
            </ul>
            
            <div style="text-align: center; margin: 30px 0;">
                <a href="{login_url}" 
                   style="display: inline-block; background: linear-gradient(135deg, #0ea5e9, #0284c7); 
                          color: white; padding: 14px 32px; text-decoration: none; 
                          border-radius: 8px; font-weight: bold; font-size: 16px;">
                    Login to Your Mentor Dashboard
                </a>
            </div>
            
            <p style="color: #64748b; font-size: 14px;">
                If you don't have an account yet, you can sign up using this email address ({mentor_email}) 
                and you'll automatically have mentor access.
            </p>
            
            <hr style="border: none; border-top: 1px solid #e2e8f0; margin: 30px 0;">
            
            <p style="color: #94a3b8; font-size: 12px; text-align: center;">
                This email was sent from gradnext. If you believe this was sent in error, 
                please contact support.
            </p>
        </div>
    </body>
    </html>
    """
    
    # Try to send email via Gmail
    email_sent = await send_email_via_gmail(db, mentor_email, subject, html_content)
    
    if email_sent:
        print(f"[Mentor Invite] Email sent successfully to {mentor_email}")
    else:
        print(f"[Mentor Invite] Failed to send email to {mentor_email} (Gmail not configured)")
    
    return email_sent


# ============ Pydantic Models ============

class UserUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    plan: Optional[str] = None
    is_mentor: Optional[bool] = None
    is_admin: Optional[bool] = None
    coaching_sessions_total: Optional[int] = None
    coaching_sessions_used: Optional[int] = None
    peer_sessions_total: Optional[int] = None
    peer_sessions_used: Optional[int] = None
    strategy_calls_total: Optional[int] = None
    strategy_calls_used: Optional[int] = None
    custom_access: Optional[dict] = None  # For courses, drills, workshops, etc.
    plan_end_date: Optional[str] = None
    subscription_end: Optional[str] = None

class AccessUpdate(BaseModel):
    user_id: str
    access_type: str  # 'videos', 'workshops', 'drills', 'materials', 'peer_practice', 'coaching', 'cohort'
    granted: bool

class MentorCreate(BaseModel):
    name: str
    email: str
    title: str
    company: str
    bio: str
    expertise: List[str]
    picture: Optional[str] = None
    linkedin: Optional[str] = None
    price_per_session: int = 1500
    is_active: bool = True
    can_take_strategy_calls: bool = False  # Whether mentor can take strategy calls

class MentorUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    title: Optional[str] = None
    company: Optional[str] = None
    bio: Optional[str] = None
    expertise: Optional[List[str]] = None
    picture: Optional[str] = None
    linkedin: Optional[str] = None
    specialization: Optional[str] = None
    hourly_rate: Optional[int] = None
    price_per_session: Optional[int] = None
    is_active: Optional[bool] = None
    is_hidden: Optional[bool] = None  # Hide from candidate dashboard
    can_take_strategy_calls: Optional[bool] = None  # Whether mentor can take strategy calls

class AvailabilityOverride(BaseModel):
    mentor_id: Optional[str] = None  # Optional since it's in the URL
    availability: List[Dict[str, Any]]  # [{day: "Monday", slots: [{from: "09:00", to: "17:00"}]}]
    blocked_days: Optional[List[str]] = []  # ["2026-02-10", "2026-02-15"]

class VideoCreate(BaseModel):
    title: str
    description: str
    module: str
    duration: str
    video_url: str
    thumbnail: Optional[str] = None
    order: int = 0
    is_free: bool = False

class VideoUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    module: Optional[str] = None
    duration: Optional[str] = None
    video_url: Optional[str] = None
    thumbnail: Optional[str] = None
    order: Optional[int] = None
    is_free: Optional[bool] = None

# Course Management Models
class CourseCreate(BaseModel):
    title: str
    description: str = ""
    thumbnail: Optional[str] = None
    order: int = 0

class CourseUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    thumbnail: Optional[str] = None
    order: Optional[int] = None

class ModuleCreate(BaseModel):
    course_id: str
    title: str
    order: int = 0

class ModuleUpdate(BaseModel):
    title: Optional[str] = None
    order: Optional[int] = None

class SubmoduleCreate(BaseModel):
    module_id: str
    title: str
    order: int = 0

class SubmoduleUpdate(BaseModel):
    title: Optional[str] = None
    order: Optional[int] = None

class SessionCreate(BaseModel):
    module_id: str
    title: str
    description: str = ""
    duration: str = ""
    content_type: str = "video"  # 'video', 'quiz', 'pdf', 'mixed'
    content_url: Optional[str] = None  # Generic URL field from frontend (video or pdf)
    video_url: Optional[str] = None
    pdf_url: Optional[str] = None
    quiz_questions: Optional[List[Dict[str, Any]]] = None  # For quiz content type
    attachments: Optional[List[Dict[str, Any]]] = None  # [{type: 'pdf'/'video', url: '...', title: '...'}]
    thumbnail: Optional[str] = None
    order: int = 0
    is_free: bool = False

class SessionUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    duration: Optional[str] = None
    content_type: Optional[str] = None
    content_url: Optional[str] = None  # Generic URL field from frontend (video or pdf)
    video_url: Optional[str] = None
    pdf_url: Optional[str] = None
    quiz_questions: Optional[List[Dict[str, Any]]] = None
    attachments: Optional[List[Dict[str, Any]]] = None
    thumbnail: Optional[str] = None
    order: Optional[int] = None
    is_free: Optional[bool] = None

# Plan Management Models - Comprehensive System
class PlanFeatureAccess(BaseModel):
    course_recordings: bool = True
    course_recordings_limited: bool = False  # For free trial
    drills_exercises: bool = True
    drills_limited: bool = False  # For free trial
    case_materials: bool = True
    case_materials_limited: bool = False  # For free trial
    workshops: str = "none"  # "none", "only_recorded", "recorded_and_live"
    workshops_limited: bool = False  # For free trial
    peer_sessions_per_month: Optional[int] = None  # Sessions per month (None = use peer_to_peer, 0 = none, -1 = unlimited)
    peer_to_peer: str = "none"  # LEGACY: kept for backward compatibility
    coaching_sessions: int = 0  # Number of sessions (0 = none, -1 = unlimited)
    strategy_calls: int = 0  # Number of calls (0 = none, -1 = unlimited)
    dedicated_coach: bool = False

class PlanPricing(BaseModel):
    one_month: Optional[float] = None
    six_month: Optional[float] = None
    one_time: Optional[float] = None

class PlanCreate(BaseModel):
    name: str  # Display name
    plan_key: str  # Internal key (e.g., 'basic_plan', 'pro_plan', 'last_mile')
    category: str = "subscription"  # "subscription", "coaching", "cohort", "addon"
    description: str = ""
    pricing: PlanPricing = PlanPricing()
    currency: str = "INR"
    duration_months: Optional[int] = None  # Duration in months (None for single session/NA)
    is_auto_renew: bool = False  # True for subscription auto-renewal
    features: PlanFeatureAccess = PlanFeatureAccess()
    display_features: List[str] = []  # Features to show on plan cards (e.g., ["Full course access", "Live workshops"])
    is_active: bool = True
    is_hidden: bool = False  # Hidden from public but still accessible
    order: int = 0
    highlight: bool = False  # For highlighting popular plans
    badge: Optional[str] = None  # e.g., "Most Popular", "Best Value"
    application_only: bool = False  # Requires application (like Pinnacle)
    show_on_pages: List[str] = ["home"]  # Which landing pages to show this plan on
    auto_add_to_subscription: bool = False  # Auto-add this addon to subscriptions
    requires_base_plan: bool = False  # Addon requires a base subscription

class PlanUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    pricing: Optional[Dict[str, Any]] = None
    currency: Optional[str] = None
    duration_months: Optional[int] = None
    is_auto_renew: Optional[bool] = None
    features: Optional[Dict[str, Any]] = None
    display_features: Optional[List[str]] = None  # Features to show on plan cards
    is_active: Optional[bool] = None
    is_hidden: Optional[bool] = None
    is_visible: Optional[bool] = None  # Visibility on homepage
    order: Optional[int] = None
    highlight: Optional[bool] = None
    badge: Optional[str] = None
    application_only: Optional[bool] = None
    show_on_pages: Optional[List[str]] = None
    auto_add_to_subscription: Optional[bool] = None
    requires_base_plan: Optional[bool] = None

class UserPlanAssignment(BaseModel):
    user_id: str
    plan_key: str
    start_date: Optional[str] = None  # ISO date string
    end_date: Optional[str] = None  # ISO date string
    is_trial: bool = False
    coaching_sessions_granted: int = 0
    strategy_calls_granted: int = 0
    peer_sessions_granted: int = 0

class WorkshopCreate(BaseModel):
    title: str
    description: str
    instructor: str
    instructor_title: str
    date: str
    time: str
    duration: str
    topics: List[str] = []
    thumbnail: Optional[str] = None  # Legacy field - kept for backwards compatibility
    thumbnail_hero: Optional[str] = None  # 21:9 aspect ratio (2100x900) - for featured hero section
    thumbnail_card: Optional[str] = None  # 16:9 aspect ratio (1280x720) - for cards
    thumbnail_recording: Optional[str] = None  # 16:9 aspect ratio (1280x720) - for past recordings grid
    video_url: Optional[str] = None
    meeting_link: Optional[str] = None
    status: str = "upcoming"  # 'upcoming', 'live', 'completed'
    is_past: bool = False
    is_free: bool = False
    max_participants: int = 50

class WorkshopUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    instructor: Optional[str] = None
    instructor_title: Optional[str] = None
    date: Optional[str] = None
    time: Optional[str] = None
    duration: Optional[str] = None
    topics: Optional[List[str]] = None
    thumbnail: Optional[str] = None  # Legacy field
    thumbnail_hero: Optional[str] = None  # 21:9 aspect ratio (2100x900)
    thumbnail_card: Optional[str] = None  # 16:9 aspect ratio (1280x720)
    thumbnail_recording: Optional[str] = None  # 16:9 aspect ratio (1280x720)
    video_url: Optional[str] = None
    meeting_link: Optional[str] = None
    status: Optional[str] = None  # 'upcoming', 'live', 'completed'
    is_past: Optional[bool] = None
    is_free: Optional[bool] = None
    max_participants: Optional[int] = None

class DrillCreate(BaseModel):
    title: str
    category: str
    difficulty: str
    description: str
    time_limit: str
    questions: List[Dict[str, Any]]
    tags: List[str] = []
    is_free: bool = False

class DrillUpdate(BaseModel):
    title: Optional[str] = None
    category: Optional[str] = None
    difficulty: Optional[str] = None
    description: Optional[str] = None
    time_limit: Optional[str] = None
    questions: Optional[List[Dict[str, Any]]] = None
    tags: Optional[List[str]] = None
    is_free: Optional[bool] = None

class MaterialCreate(BaseModel):
    title: str
    category: str
    description: str
    file_type: str
    file_url: Optional[str] = None
    content: Optional[str] = None
    is_free: bool = False

class MaterialUpdate(BaseModel):
    title: Optional[str] = None
    category: Optional[str] = None
    description: Optional[str] = None
    file_type: Optional[str] = None
    file_url: Optional[str] = None
    content: Optional[str] = None
    is_free: Optional[bool] = None

class CohortSectionCreate(BaseModel):
    title: str
    description: str
    order: int = 0

class CohortResourceCreate(BaseModel):
    section_id: str
    title: str
    description: str
    resource_type: str  # 'video', 'document', 'link'
    resource_url: str
    order: int = 0


# ============ Admin Verification ============

async def verify_admin(request: Request):
    """Verify the current user is an admin"""
    user = await get_current_user(request)
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")
    return user


# ============ Dashboard Stats ============

@router.post("/abandoned-cart/sync")
async def sync_abandoned_cart_to_sheet(request: Request):
    """One-time backfill: writes every currently-pending subscription order
    (status='created' with plan_key and no explicit type) to the
    'Abandoned Cart' Google Sheet tab. Safe to re-run — uses upsert by email."""
    await verify_admin(request)
    db = get_db(request)

    from services.google_sheets_service import append_abandoned_cart_to_sheet

    pending = await db.payment_orders.find({
        "status": "created",
        "plan_key": {"$exists": True},
        "type": {"$exists": False},
    }, {"_id": 0}).to_list(5000)

    synced = 0
    for order in pending:
        email = order.get("user_email")
        if not email:
            continue
        user = await db.users.find_one({"email": email}, {"_id": 0}) or {}
        try:
            await append_abandoned_cart_to_sheet(
                user,
                {
                    "plan_attempted_key": order.get("plan_key"),
                    "plan_attempted_name": order.get("plan_name"),
                    "attempted_at": order.get("created_at"),
                },
            )
            synced += 1
        except Exception as e:
            logger.warning(f"Abandoned cart sync error for {email}: {e}")

    return {"synced": synced, "total_pending": len(pending)}


@router.get("/stats")
async def get_admin_stats(request: Request):
    """Get admin dashboard statistics"""
    await verify_admin(request)
    db = get_db(request)
    
    # Get counts
    users_count = await db.users.count_documents({})
    mentors_count = await db.mentors.count_documents({})
    videos_count = await db.videos.count_documents({})
    workshops_count = await db.workshops.count_documents({})
    drills_count = await db.drills.count_documents({})
    materials_count = await db.materials.count_documents({})
    bookings_count = await db.bookings.count_documents({})
    peer_sessions_count = await db.peer_sessions.count_documents({})
    
    return {
        "users": users_count,
        "mentors": mentors_count,
        "videos": videos_count,
        "workshops": workshops_count,
        "drills": drills_count,
        "materials": materials_count,
        "bookings": bookings_count,
        "peer_sessions": peer_sessions_count
    }


# ============ User Management ============

@router.get("/users")
async def get_all_users(request: Request, skip: int = 0, limit: int = 100, search: str = None, plan_category: str = None):
    """Get all users with pagination, search, and plan defaults merged"""
    await verify_admin(request)
    db = get_db(request)
    
    # Build filter
    user_filter = {"is_deleted": {"$ne": True}}
    
    # Search filter
    if search:
        user_filter["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"first_name": {"$regex": search, "$options": "i"}},
            {"last_name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"phone_number": {"$regex": search, "$options": "i"}}
        ]
    
    # Get total count first (before category filtering which happens in memory)
    base_total = await db.users.count_documents(user_filter)
    
    # Get all plans for default values and category lookup
    plans = await db.plans.find({}, {"_id": 0}).to_list(50)
    plans_map = {p.get("plan_key"): p for p in plans}
    
    # Push plan_category filter to the DB query using known plan keys
    if plan_category and plan_category != 'all':
        matching_keys = [pk for pk, p in plans_map.items() if p.get("category") == plan_category]
        if matching_keys:
            user_filter["plan"] = {"$in": matching_keys}
        else:
            return {"users": [], "total": 0}

        total = await db.users.count_documents(user_filter)
        users = await db.users.find(user_filter, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
        enriched_users = [enrich_user_with_plan_defaults(u, plans_map.get(u.get("plan"), {})) for u in users]
        return {"users": enriched_users, "total": total}
    
    # No category filter - use efficient DB pagination
    users = await db.users.find(user_filter, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    total = base_total
    
    # Enrich users with plan defaults
    enriched_users = []
    for user in users:
        plan_key = user.get("plan")
        plan_config = plans_map.get(plan_key, {})
        enriched_user = enrich_user_with_plan_defaults(user, plan_config)
        enriched_users.append(enriched_user)
    
    return {"users": enriched_users, "total": total}


def enrich_user_with_plan_defaults(user: dict, plan_config: dict) -> dict:
    """Helper to enrich user with plan default values"""
    features = plan_config.get("features", {})
    enriched_user = {**user}
    
    # Coaching sessions - use user value if set, else plan default
    if enriched_user.get("coaching_sessions_total") is None:
        enriched_user["coaching_sessions_total"] = features.get("coaching_sessions", 0)
    if enriched_user.get("coaching_sessions_used") is None:
        enriched_user["coaching_sessions_used"] = 0
        
    # Peer sessions - use user value if set, else plan default (-1 = unlimited)
    if enriched_user.get("peer_sessions_total") is None:
        peer_default = features.get("peer_sessions_per_month")
        enriched_user["peer_sessions_total"] = peer_default if peer_default is not None else -1
    if enriched_user.get("peer_sessions_used") is None:
        enriched_user["peer_sessions_used"] = 0
        
    # Strategy calls - use user value if set, else plan default
    if enriched_user.get("strategy_calls_total") is None:
        enriched_user["strategy_calls_total"] = features.get("strategy_calls", 0)
    if enriched_user.get("strategy_calls_used") is None:
        enriched_user["strategy_calls_used"] = 0
    
    # Add plan category for filtering
    enriched_user["plan_category"] = plan_config.get("category", "subscription")
    
    return enriched_user


@router.get("/users/counts")
async def get_user_counts(request: Request):
    """Get user counts by plan category for filter badges"""
    await verify_admin(request)
    db = get_db(request)
    
    # Get all plans for category lookup
    plans = await db.plans.find({}, {"_id": 0, "plan_key": 1, "category": 1}).to_list(50)
    plan_categories = {p.get("plan_key"): p.get("category", "subscription") for p in plans}
    
    # Get all users (excluding deleted)
    users = await db.users.find(
        {"is_deleted": {"$ne": True}, "is_admin": {"$ne": True}, "is_mentor": {"$ne": True}},
        {"_id": 0, "plan": 1}
    ).to_list(50000)
    
    # Count by category
    counts = {"subscription": 0, "coaching": 0, "cohort": 0, "addon": 0, "total": len(users)}
    
    for user in users:
        plan_key = user.get("plan", "free_trial")
        category = plan_categories.get(plan_key, "subscription")
        if category in counts:
            counts[category] += 1
        else:
            counts["subscription"] += 1  # Default to subscription
    
    return counts


# NOTE: Static routes must come BEFORE dynamic routes to avoid being captured
@router.post("/users/import-excel")
async def import_users_from_excel(
    request: Request,
    file: UploadFile = File(...),
    skip_existing: bool = Form(True)
):
    """
    Import candidates from an Excel file.
    Expected columns: name, email, plan, plan_start_date, plan_end_date,
    coaching_sessions_total, coaching_sessions_used, strategy_calls_total, etc.
    """
    import pandas as pd
    from datetime import timezone
    
    await verify_admin(request)
    db = get_db(request)
    
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    # Save uploaded file temporarily
    temp_path = f"/tmp/import_{uuid.uuid4().hex}.xlsx"
    try:
        with open(temp_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        # Read Excel file
        df = pd.read_excel(temp_path)
        
        results = {
            "total_rows": len(df),
            "imported": 0,
            "skipped_existing": 0,
            "skipped_invalid": 0,
            "errors": [],
            "imported_users": [],
            "skipped_users": []
        }
        
        # Helper functions
        def normalize_plan(plan):
            if not plan or pd.isna(plan):
                return "free_trial"
            plan_lower = str(plan).lower().strip().replace(" ", "_").replace("-", "_")
            mappings = {
                "full_prep": "full_prep", "fullprep": "full_prep",
                "mid_mile": "mid_mile", "midmile": "mid_mile",
                "last_mile": "last_mile", "lastmile": "last_mile",
                "pinnacle": "pinnacle", "basic": "basic_plan",
                "basic_plan": "basic_plan", "pro": "pro_plan",
                "pro_plan": "pro_plan", "pro_plus": "pro_plus",
                "free_trial": "free_trial", "free": "free_trial",
            }
            return mappings.get(plan_lower, plan_lower)
        
        def safe_int(val, default=0):
            if pd.isna(val) or val is None:
                return default
            try:
                return int(float(val))
            except Exception:
                return default
        
        def safe_bool(val, default=False):
            if pd.isna(val) or val is None:
                return default
            if isinstance(val, bool):
                return val
            if isinstance(val, (int, float)):
                return val > 0 or val == -1
            if isinstance(val, str):
                return val.lower() in ('true', 'yes', '1', '-1')
            return default
        
        def safe_str(val, default=""):
            if pd.isna(val) or val is None:
                return default
            return str(val).strip()
        
        def parse_date(val):
            if pd.isna(val) or val is None:
                return None
            if isinstance(val, datetime):
                return val.replace(tzinfo=timezone.utc)
            if isinstance(val, str):
                try:
                    return datetime.fromisoformat(val).replace(tzinfo=timezone.utc)
                except Exception:
                    return None
            return None
        
        now = datetime.now(timezone.utc)
        
        for idx, row in df.iterrows():
            email = safe_str(row.get('email')).lower()
            name = safe_str(row.get('name'))
            
            if not email or not name:
                results["skipped_invalid"] += 1
                results["errors"].append(f"Row {idx + 2}: Missing email or name")
                continue
            
            # Check existing
            if skip_existing:
                existing = await db.users.find_one({"email": email})
                if existing:
                    results["skipped_existing"] += 1
                    results["skipped_users"].append({"email": email, "name": name, "reason": "already exists"})
                    continue
            
            # Build user document
            total_sessions = safe_int(row.get('coaching_sessions_total'), 0)
            used_sessions = safe_int(row.get('coaching_sessions_used'), 0)
            remaining_sessions = safe_int(row.get('coaching_sessions_remaining'), 0)
            if remaining_sessions == 0 and total_sessions > 0:
                remaining_sessions = max(0, total_sessions - used_sessions)
            
            strategy_total = safe_int(row.get('strategy_calls_total'), 0)
            strategy_used = safe_int(row.get('strategy_calls_used'), 0)
            strategy_remaining = safe_int(row.get('strategy_calls_remaining'), 0)
            if strategy_remaining == 0 and strategy_total > 0:
                strategy_remaining = max(0, strategy_total - strategy_used)
            
            user_doc = {
                "id": str(uuid.uuid4()),
                "email": email,
                "name": name,
                "picture": safe_str(row.get('picture')) or None,
                "phone": safe_str(row.get('phone')) or None,
                "college": safe_str(row.get('college')) or None,
                "linkedin_url": safe_str(row.get('linkedin_url')) or None,
                "timezone": safe_str(row.get('timezone')) or None,
                "plan": normalize_plan(row.get('plan')),
                "plan_start_date": parse_date(row.get('plan_start_date')),
                "plan_end_date": parse_date(row.get('plan_end_date')),
                "coaching_sessions_total": total_sessions,
                "coaching_sessions_used": used_sessions,
                "coaching_sessions_remaining": remaining_sessions,
                "is_unlimited_coaching": safe_bool(row.get('is_unlimited_coaching'), False),
                "strategy_calls_total": strategy_total,
                "strategy_calls_used": strategy_used,
                "strategy_calls_remaining": strategy_remaining,
                "is_unlimited_strategy_calls": safe_bool(row.get('is_unlimited_strategy_calls'), False),
                "peer_sessions_per_month": safe_int(row.get('peer_sessions_per_month'), 0),
                "peer_sessions_used_this_month": safe_int(row.get('peer_sessions_used_this_month'), 0),
                "is_unlimited_peer_sessions": safe_bool(row.get('is_unlimited_peer_sessions'), False),
                "google_id": None,
                "is_mentor": False,
                "is_admin": False,
                "is_candidate": True,
                "mentor_id": None,
                "peer_rating": 5.0,
                "peer_sessions_done": 0,
                "peer_availability": [],
                "bio": None,
                "target_companies": [],
                "preparation_stage": None,
                "custom_access": {},
                "cohort_batch": None,
                "cohort_id": None,
                "cohort_enrolled_at": None,
                "created_at": now,
                "updated_at": now,
                "imported_at": now,
                "import_source": "excel_admin_upload"
            }
            
            try:
                await db.users.insert_one(user_doc)
                results["imported"] += 1
                results["imported_users"].append({
                    "email": email,
                    "name": name,
                    "plan": user_doc["plan"]
                })
            except Exception as e:
                results["errors"].append(f"Failed to insert {email}: {str(e)}")
        
        return results
        
    finally:
        # Cleanup temp file
        if os.path.exists(temp_path):
            os.remove(temp_path)


@router.get("/users/import-template")
async def get_import_template(request: Request):
    """Get the expected columns for Excel import"""
    await verify_admin(request)
    
    return {
        "required_columns": ["name", "email"],
        "optional_columns": [
            "plan", "plan_start_date", "plan_end_date",
            "coaching_sessions_total", "coaching_sessions_used", "coaching_sessions_remaining",
            "is_unlimited_coaching",
            "strategy_calls_total", "strategy_calls_used", "strategy_calls_remaining",
            "is_unlimited_strategy_calls",
            "peer_sessions_per_month", "peer_sessions_used_this_month", "is_unlimited_peer_sessions",
            "picture", "phone", "timezone", "college", "linkedin_url"
        ],
        "plan_values": [
            "free_trial", "basic_plan", "pro_plan", "pro_plus",
            "last_mile", "mid_mile", "full_prep", "pinnacle"
        ],
        "notes": [
            "Date columns should be in YYYY-MM-DD format",
            "Boolean columns accept: true/false, yes/no, 1/0, -1 (for unlimited)",
            "Plan names are case-insensitive and can use spaces or underscores"
        ]
    }


@router.post("/mentors/import-excel")
async def import_mentors_from_excel(
    request: Request,
    file: UploadFile = File(...)
):
    """
    Import/update mentor data from an Excel file.
    Expected columns: email (required), rating, sessions_conducted
    Matches mentors by email address.
    """
    import pandas as pd
    
    await verify_admin(request)
    db = get_db(request)
    
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    # Save uploaded file temporarily
    temp_path = f"/tmp/mentor_import_{uuid.uuid4().hex}.xlsx"
    try:
        with open(temp_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        # Read Excel file
        df = pd.read_excel(temp_path)
        
        # Normalize column names (lowercase, strip whitespace)
        df.columns = [str(col).lower().strip().replace(' ', '_') for col in df.columns]
        
        results = {
            "total_rows": len(df),
            "updated": 0,
            "not_found": 0,
            "errors": [],
            "updated_mentors": [],
            "not_found_emails": []
        }
        
        # Check for required email column
        if 'email' not in df.columns:
            raise HTTPException(
                status_code=400, 
                detail="Excel file must have an 'email' column to identify mentors"
            )
        
        for idx, row in df.iterrows():
            try:
                email = str(row.get('email', '')).strip().lower()
                if not email or email == 'nan':
                    results["errors"].append(f"Row {idx + 2}: Missing email")
                    continue
                
                # Find mentor by email
                mentor = await db.mentors.find_one({"email": {"$regex": f"^{email}$", "$options": "i"}})
                
                if not mentor:
                    results["not_found"] += 1
                    results["not_found_emails"].append(email)
                    continue
                
                # Build update data
                update_data = {"updated_at": datetime.utcnow()}
                
                # Rating (1-5 scale)
                if 'rating' in df.columns and pd.notna(row.get('rating')):
                    try:
                        rating = float(row.get('rating'))
                        if 0 <= rating <= 5:
                            update_data["rating"] = round(rating, 1)
                        else:
                            results["errors"].append(f"Row {idx + 2}: Rating must be between 0 and 5")
                    except (ValueError, TypeError):
                        results["errors"].append(f"Row {idx + 2}: Invalid rating value")
                
                # Sessions conducted
                if 'sessions_conducted' in df.columns and pd.notna(row.get('sessions_conducted')):
                    try:
                        sessions = int(float(row.get('sessions_conducted')))
                        if sessions >= 0:
                            update_data["sessions_conducted"] = sessions
                            update_data["sessions_done"] = sessions  # Keep both in sync
                        else:
                            results["errors"].append(f"Row {idx + 2}: Sessions must be >= 0")
                    except (ValueError, TypeError):
                        results["errors"].append(f"Row {idx + 2}: Invalid sessions value")
                
                # Optional: Update name if provided
                if 'name' in df.columns and pd.notna(row.get('name')):
                    name = str(row.get('name')).strip()
                    if name and name != 'nan':
                        update_data["name"] = name
                
                # Optional: Update title if provided
                if 'title' in df.columns and pd.notna(row.get('title')):
                    title = str(row.get('title')).strip()
                    if title and title != 'nan':
                        update_data["title"] = title
                
                # Optional: Update company if provided
                if 'company' in df.columns and pd.notna(row.get('company')):
                    company = str(row.get('company')).strip()
                    if company and company != 'nan':
                        update_data["company"] = company
                
                # Optional: Update years_experience if provided
                if 'years_experience' in df.columns and pd.notna(row.get('years_experience')):
                    try:
                        years = int(float(row.get('years_experience')))
                        if years >= 0:
                            update_data["years_experience"] = years
                    except (ValueError, TypeError):
                        pass
                
                # Optional: Update hourly_rate if provided
                if 'hourly_rate' in df.columns and pd.notna(row.get('hourly_rate')):
                    try:
                        rate = int(float(row.get('hourly_rate')))
                        if rate >= 0:
                            update_data["hourly_rate"] = rate
                    except (ValueError, TypeError):
                        pass
                
                # Update mentor if we have data to update
                if len(update_data) > 1:  # More than just updated_at
                    await db.mentors.update_one(
                        {"id": mentor["id"]},
                        {"$set": update_data}
                    )
                    results["updated"] += 1
                    results["updated_mentors"].append({
                        "email": email,
                        "name": mentor.get("name"),
                        "fields_updated": list(update_data.keys())
                    })
                    
            except Exception as e:
                results["errors"].append(f"Row {idx + 2}: {str(e)}")
        
        return results
        
    finally:
        # Clean up temp file
        import os
        if os.path.exists(temp_path):
            os.remove(temp_path)


@router.get("/mentors/import-template")
async def get_mentor_import_template(request: Request):
    """Get the expected format for mentor Excel import"""
    await verify_admin(request)
    
    return {
        "description": "Excel template for updating mentor ratings and session counts",
        "required_columns": ["email"],
        "optional_columns": [
            "rating", "sessions_conducted", "name", "title", 
            "company", "years_experience", "hourly_rate"
        ],
        "example_rows": [
            {"email": "mentor@example.com", "rating": 4.5, "sessions_conducted": 25},
            {"email": "coach@example.com", "rating": 4.8, "sessions_conducted": 42}
        ],
        "notes": [
            "Email is required to identify the mentor",
            "Rating should be between 0 and 5 (decimal allowed)",
            "Sessions_conducted should be a whole number >= 0",
            "Only mentors that exist in the system will be updated",
            "Mentors not found will be listed in the response"
        ]
    }


@router.post("/mentors/import-feedback")
async def import_mentor_feedback(
    request: Request,
    file: UploadFile = File(...)
):
    """
    Import historical feedback/testimonials for mentors from Excel.
    Expected columns: mentor_email (required), feedback (required), rating (optional), candidate_name (optional), date (optional)
    """
    import pandas as pd
    
    await verify_admin(request)
    db = get_db(request)
    
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
    
    temp_path = f"/tmp/feedback_import_{uuid.uuid4().hex}.xlsx"
    try:
        with open(temp_path, "wb") as buffer:
            content = await file.read()
            buffer.write(content)
        
        df = pd.read_excel(temp_path)
        df.columns = [str(col).lower().strip().replace(' ', '_') for col in df.columns]
        
        results = {
            "total_rows": len(df),
            "imported": 0,
            "mentor_not_found": 0,
            "errors": [],
            "imported_feedbacks": [],
            "not_found_emails": []
        }
        
        # Check required columns
        if 'mentor_email' not in df.columns:
            raise HTTPException(status_code=400, detail="Excel must have 'mentor_email' column")
        if 'feedback' not in df.columns:
            raise HTTPException(status_code=400, detail="Excel must have 'feedback' column")
        
        for idx, row in df.iterrows():
            try:
                mentor_email = str(row.get('mentor_email', '')).strip().lower()
                feedback_text = str(row.get('feedback', '')).strip()
                
                if not mentor_email or mentor_email == 'nan':
                    results["errors"].append(f"Row {idx + 2}: Missing mentor_email")
                    continue
                    
                if not feedback_text or feedback_text == 'nan':
                    results["errors"].append(f"Row {idx + 2}: Missing feedback text")
                    continue
                
                # Find mentor
                mentor = await db.mentors.find_one({"email": {"$regex": f"^{mentor_email}$", "$options": "i"}})
                if not mentor:
                    results["mentor_not_found"] += 1
                    results["not_found_emails"].append(mentor_email)
                    continue
                
                # Parse rating (optional)
                rating = 5  # Default rating
                if 'rating' in df.columns and pd.notna(row.get('rating')):
                    try:
                        rating = float(row.get('rating'))
                        if not (1 <= rating <= 5):
                            rating = 5
                    except:
                        rating = 5
                
                # Parse candidate name (optional)
                candidate_name = "Previous Client"
                if 'candidate_name' in df.columns and pd.notna(row.get('candidate_name')):
                    name = str(row.get('candidate_name')).strip()
                    if name and name != 'nan':
                        candidate_name = name
                
                # Parse date (optional)
                feedback_date = datetime.utcnow()
                if 'date' in df.columns and pd.notna(row.get('date')):
                    try:
                        date_val = row.get('date')
                        if isinstance(date_val, datetime):
                            feedback_date = date_val
                        elif isinstance(date_val, str):
                            # Try common date formats
                            for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y']:
                                try:
                                    feedback_date = datetime.strptime(date_val, fmt)
                                    break
                                except:
                                    continue
                    except:
                        pass
                
                # Create feedback document
                feedback_doc = {
                    "id": f"hist-fb-{uuid.uuid4().hex[:12]}",
                    "mentor_id": mentor["id"],
                    "candidate_id": None,  # Historical, no linked candidate
                    "candidate_name_override": candidate_name,  # Store name directly
                    "booking_id": None,  # Historical, no linked booking
                    "rating_overall": round(rating, 1),
                    "other_feedback": feedback_text,
                    "is_historical": True,  # Mark as historical import
                    "created_at": feedback_date,
                    "imported_at": datetime.utcnow(),
                    "imported_by": "admin"
                }
                
                await db.candidate_feedbacks.insert_one(feedback_doc)
                results["imported"] += 1
                results["imported_feedbacks"].append({
                    "mentor": mentor.get("name", mentor_email),
                    "candidate": candidate_name,
                    "rating": round(rating, 1)
                })
                
            except Exception as e:
                results["errors"].append(f"Row {idx + 2}: {str(e)}")
        
        return results
        
    finally:
        import os
        if os.path.exists(temp_path):
            os.remove(temp_path)


@router.get("/mentors/feedback-template")
async def get_feedback_import_template(request: Request):
    """Get the expected format for feedback Excel import"""
    await verify_admin(request)
    
    return {
        "description": "Excel template for importing historical mentor feedback/testimonials",
        "required_columns": ["mentor_email", "feedback"],
        "optional_columns": ["rating", "candidate_name", "date"],
        "example_rows": [
            {
                "mentor_email": "coach@example.com",
                "feedback": "Amazing session! Really helped me prepare for my McKinsey interview.",
                "rating": 5,
                "candidate_name": "Rahul S.",
                "date": "2024-06-15"
            },
            {
                "mentor_email": "mentor@example.com",
                "feedback": "Very structured approach to case solving. Highly recommended!",
                "rating": 4.5,
                "candidate_name": "Priya M.",
                "date": "2024-08-20"
            }
        ],
        "notes": [
            "mentor_email: Required - email to identify the mentor",
            "feedback: Required - the qualitative feedback text",
            "rating: Optional - 1-5 scale, defaults to 5",
            "candidate_name: Optional - name to display, defaults to 'Previous Client'",
            "date: Optional - when feedback was given (YYYY-MM-DD format)"
        ]
    }


@router.get("/users/{user_id}")
async def get_user(user_id: str, request: Request):
    """Get a specific user"""
    await verify_admin(request)
    db = get_db(request)
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    return user


@router.get("/users/{user_id}/details")
async def get_user_details(user_id: str, request: Request):
    """Get detailed user information including activity and progress"""
    await verify_admin(request)
    db = get_db(request)
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Get coaching bookings
    bookings = await db.bookings.find({"user_id": user_id}, {"_id": 0}).to_list(100)
    
    # Get peer practice sessions (fix: use requester_id and partner_id, not user_id and peer_id)
    peer_sessions = await db.peer_sessions.find(
        {"$or": [{"requester_id": user_id}, {"partner_id": user_id}]},
        {"_id": 0}
    ).to_list(100)
    
    # Get feedbacks received
    feedbacks = await db.session_feedbacks.find({"user_id": user_id}, {"_id": 0}).to_list(50)
    
    # Get video progress (stored in progress collection, not video_progress)
    user_progress = await db.progress.find_one({"user_id": user_id}, {"_id": 0})
    videos_watched = len(user_progress.get("videos_completed", [])) if user_progress else 0
    
    # Get drill attempts
    drill_attempts = await db.drill_attempts.find({"user_id": user_id}, {"_id": 0}).to_list(100)
    
    # Get invoices/payments
    invoices = await db.invoices.find({"user_id": user_id}, {"_id": 0}).to_list(50)
    
    # Calculate stats
    total_coaching_sessions = len(bookings)
    completed_coaching_sessions = len([b for b in bookings if b.get("status") == "completed"])
    total_peer_sessions = len(peer_sessions)
    completed_peer_sessions = len([p for p in peer_sessions if p.get("status") == "completed"])
    # videos_watched already calculated above from progress collection
    drills_completed = len([d for d in drill_attempts if d.get("completed")])
    total_spent = sum(inv.get("amount", 0) for inv in invoices if inv.get("status") == "paid")
    
    # Calculate average rating from feedbacks
    ratings = [f.get("rating", 0) for f in feedbacks if f.get("rating")]
    avg_rating = sum(ratings) / len(ratings) if ratings else 0
    
    return {
        "user": user,
        "stats": {
            "total_coaching_sessions": total_coaching_sessions,
            "completed_coaching_sessions": completed_coaching_sessions,
            "total_peer_sessions": total_peer_sessions,
            "completed_peer_sessions": completed_peer_sessions,
            "videos_watched": videos_watched,
            "drills_completed": drills_completed,
            "total_spent": total_spent,
            "average_rating": round(avg_rating, 2)
        },
        "recent_bookings": bookings[:10],
        "recent_peer_sessions": peer_sessions[:10],
        "recent_feedbacks": feedbacks[:10],
        "video_ids_completed": user_progress.get("videos_completed", []) if user_progress else [],
        "drill_attempts": drill_attempts[:20],
        "invoices": invoices
    }


class UserCreate(BaseModel):
    name: str
    email: str
    plan: str = "free_trial"
    coaching_sessions_total: int = 0
    is_mentor: bool = False
    is_admin: bool = False


@router.post("/users")
async def create_user(user_data: UserCreate, request: Request):
    """Create a new user"""
    await verify_admin(request)
    db = get_db(request)
    
    # Check if email already exists
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="User with this email already exists")
    
    user_id = f"user-{str(uuid.uuid4())[:8]}"
    mentor_id = None
    
    # If user is being created as a mentor, auto-create mentor card
    if user_data.is_mentor:
        mentor_id = f"mentor-{str(uuid.uuid4())[:8]}"
        mentor_card = {
            "id": mentor_id,
            "name": user_data.name,
            "email": user_data.email,
            "title": "Consultant",
            "company": "",
            "bio": "",
            "expertise": [],
            "picture": f"https://ui-avatars.com/api/?name={user_data.name.replace(' ', '+')}&background=0D8ABC&color=fff",
            "linkedin": "",
            "specialization": "General",
            "hourly_rate": 12000,
            "rating": 5.0,
            "sessions_conducted": 0,
            "sessions_done": 0,
            "years_experience": 0,
            "is_active": True,
            "is_hidden": True,  # Hidden by default until admin makes visible
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow()
        }
        await db.mentors.insert_one(mentor_card)
    
    user = {
        "id": user_id,
        "email": user_data.email,
        "name": user_data.name,
        "picture": f"https://ui-avatars.com/api/?name={user_data.name.replace(' ', '+')}&background=random",
        "google_id": f"manual-{user_id}",
        "plan": user_data.plan,
        "plan_start_date": datetime.utcnow(),
        "plan_end_date": datetime.utcnow() + timedelta(days=90),
        "coaching_sessions_total": user_data.coaching_sessions_total,
        "coaching_sessions_used": 0,
        "cohort_batch": None,
        "is_mentor": user_data.is_mentor,
        "is_admin": user_data.is_admin,
        "mentor_id": mentor_id,
        "peer_rating": 5.0,
        "peer_sessions_done": 0,
        "peer_availability": [],
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow()
    }
    
    await db.users.insert_one(user)
    
    # Send mentor invite email if user is being created as a mentor
    email_sent = False
    if user_data.is_mentor:
        email_sent = await send_mentor_invite_email(db, user_data.email, user_data.name)
    
    return {
        "message": "User created successfully", 
        "user_id": user_id, 
        "mentor_id": mentor_id,
        "mentor_invite_sent": email_sent if user_data.is_mentor else None
    }


@router.put("/users/{user_id}")
async def update_user(user_id: str, user_data: UserUpdate, request: Request):
    """Update user details"""
    await verify_admin(request)
    db = get_db(request)
    
    # Get current user to check if mentor status is changing
    current_user = await db.users.find_one({"id": user_id})
    if not current_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Filter out None values but KEEP zero values (0 is a valid value for credits)
    update_data = {k: v for k, v in user_data.dict().items() if v is not None}
    
    # Log what we're updating for debugging
    logger.info(f"Updating user {user_id}: {list(update_data.keys())}")
    if "strategy_calls_total" in update_data:
        logger.info(f"Admin explicitly set strategy_calls_total to: {update_data['strategy_calls_total']}")
    
    # Check what was explicitly provided by admin (exclude unset fields)
    explicitly_set_fields = user_data.dict(exclude_unset=True)
    logger.info(f"Explicitly set fields by admin: {list(explicitly_set_fields.keys())}")
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No update data provided")
    
    # If plan is provided, always sync features from the plan
    if "plan" in update_data:
        new_plan_key = update_data["plan"]
        # Find the plan in the database
        plan_doc = await db.plans.find_one({"plan_key": new_plan_key})
        if not plan_doc:
            # Try finding by name
            plan_doc = await db.plans.find_one({"name": {"$regex": f"^{new_plan_key}$", "$options": "i"}})
        
        if plan_doc:
            plan_features = plan_doc.get("features", {})
            # Update user's feature access based on the plan
            update_data["features"] = plan_features
            # Update coaching sessions if plan provides them
            if plan_features.get("coaching_sessions", 0) > 0:
                update_data["coaching_sessions_total"] = plan_features.get("coaching_sessions", 0)
            
            # IMPORTANT: Only set strategy_calls_total from plan if admin didn't explicitly set it
            # Check if strategy_calls_total was explicitly provided in the request
            if "strategy_calls_total" not in user_data.dict(exclude_unset=True):
                # Admin didn't set it explicitly, use plan default
                if plan_features.get("strategy_calls", 0) > 0:
                    update_data["strategy_calls_total"] = plan_features.get("strategy_calls", 0)
            # else: Admin explicitly set strategy_calls_total, keep their value (don't overwrite)
            
            # Only reset plan dates if plan is actually changing
            if update_data["plan"] != current_user.get("plan"):
                update_data["plan_start_date"] = datetime.utcnow().isoformat()
                
                # Check if admin explicitly changed the plan_end_date
                # If the provided date is the same as the current date, admin did NOT override it
                # (they just sent the existing value along with the plan change)
                from datetime import timedelta
                current_end_date = current_user.get("plan_end_date")
                provided_end_date = user_data.plan_end_date
                
                # Normalize dates for comparison (handle timezone differences)
                admin_explicitly_changed_date = False
                if provided_end_date and current_end_date:
                    # Strip timezone info for comparison
                    provided_str = str(provided_end_date).split('+')[0].split('Z')[0].split('.')[0]
                    current_str = str(current_end_date).split('+')[0].split('Z')[0].split('.')[0]
                    admin_explicitly_changed_date = provided_str != current_str
                elif provided_end_date and not current_end_date:
                    # Admin provided a date where none existed
                    admin_explicitly_changed_date = True
                
                # Recalculate plan_end_date based on new plan (unless admin explicitly changed it)
                if not admin_explicitly_changed_date:
                    if update_data["plan"] == "free_trial":
                        # Free trial is always 7 days
                        update_data["plan_end_date"] = (datetime.utcnow() + timedelta(days=7)).isoformat()
                    elif plan_doc.get("category") in ["subscription", "coaching", "cohort"]:
                        # Get duration from plan config, default to 30 days
                        duration_days = plan_doc.get("duration_days") or 30
                        if plan_doc.get("duration_months"):
                            duration_days = plan_doc.get("duration_months") * 30
                        update_data["plan_end_date"] = (datetime.utcnow() + timedelta(days=duration_days)).isoformat()
                
                # Set category-specific end date fields
                end_date_to_use = update_data.get("plan_end_date") or user_data.plan_end_date
                if end_date_to_use:
                    if plan_doc.get("category") == "coaching":
                        update_data["coaching_program_end_date"] = end_date_to_use
                    elif plan_doc.get("category") == "subscription":
                        update_data["subscription_end_date"] = end_date_to_use
    
    # Check if is_mentor is being set to True and user wasn't a mentor before
    becoming_mentor = update_data.get("is_mentor") == True and not current_user.get("is_mentor")
    
    if becoming_mentor:
        # Check if mentor card already exists for this email
        existing_mentor = await db.mentors.find_one({"email": current_user.get("email")})
        
        if not existing_mentor:
            # Create new mentor card (hidden by default)
            mentor_id = f"mentor-{str(uuid.uuid4())[:8]}"
            mentor_card = {
                "id": mentor_id,
                "name": current_user.get("name", ""),
                "email": current_user.get("email", ""),
                "title": "Consultant",
                "company": "",
                "bio": "",
                "expertise": [],
                "picture": current_user.get("picture") or f"https://ui-avatars.com/api/?name={current_user.get('name', 'Mentor').replace(' ', '+')}&background=0D8ABC&color=fff",
                "linkedin": "",
                "specialization": "General",
                "hourly_rate": 12000,
                "rating": 5.0,
                "sessions_conducted": 0,
                "sessions_done": 0,
                "years_experience": 0,
                "is_active": True,
                "is_hidden": True,  # Hidden by default until admin makes visible
                "created_at": datetime.utcnow(),
                "updated_at": datetime.utcnow()
            }
            await db.mentors.insert_one(mentor_card)
            update_data["mentor_id"] = mentor_id
        else:
            # Link to existing mentor card
            update_data["mentor_id"] = existing_mentor.get("id")
    
    # Sync ALL date fields when plan_end_date is updated (single source of truth)
    if "plan_end_date" in update_data:
        end_date = update_data["plan_end_date"]
        # Always sync to all category-specific fields for consistency
        update_data["subscription_end"] = end_date
        update_data["subscription_end_date"] = end_date
        update_data["coaching_program_end_date"] = end_date
    
    update_data["updated_at"] = datetime.utcnow()
    
    result = await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    # Send mentor invite email if user is being made a mentor
    email_sent = False
    if becoming_mentor:
        email_sent = await send_mentor_invite_email(
            db, 
            current_user.get("email"), 
            current_user.get("name", "Mentor")
        )
    
    return {
        "message": "User updated successfully",
        "features_synced": "plan" in user_data.dict() and user_data.plan is not None,
        "mentor_invite_sent": email_sent if becoming_mentor else None
    }


@router.delete("/users/{user_id}")
async def delete_user(user_id: str, request: Request):
    """Delete a user and all associated data.
    
    Cascade cleanup covers every collection that holds a `user_id` /
    `requester_id` / `accepter_id` / `mentor_id` reference for this user
    so we don't leave orphaned rows that can break unique indexes on
    a future re-signup with the same email. Each cleanup is best-effort
    (wrapped in try/except) — even if one collection has a transient
    issue we still attempt to delete the user document itself, which
    is the user-facing outcome of this action.
    """
    await verify_admin(request)
    db = get_db(request)
    
    # Verify the user exists first so we can return a friendlier 404
    existing = await db.users.find_one({"id": user_id}, {"_id": 0, "email": 1, "role": 1, "is_admin": 1})
    if not existing:
        raise HTTPException(status_code=404, detail="User not found")
    
    user_email = (existing.get("email") or "").lower()
    
    # Cascade across all collections that reference the user.
    cleanup_errors: list = []
    cascade_targets = [
        ("peer_profiles",      {"user_id": user_id}),
        ("peer_sessions",      {"$or": [{"requester_id": user_id}, {"accepter_id": user_id}]}),
        ("bookings",           {"user_id": user_id}),
        ("session_feedbacks",  {"user_id": user_id}),
        ("user_sessions",      {"user_id": user_id}),
        ("slot_reservations",  {"user_id": user_id}),
        ("payments",           {"user_id": user_id}),
        ("payment_orders",     {"user_id": user_id}),
        ("discount_usage",     {"user_id": user_id}),
        ("notifications",      {"user_id": user_id}),
        ("user_plan_assignments", {"user_id": user_id}),
        ("subscriptions",      {"user_id": user_id}),
        ("workshop_registrations", {"user_id": user_id}),
        ("competition_entries", {"user_id": user_id}),
        ("user_workshop_attendance", {"user_id": user_id}),
    ]
    for coll, query in cascade_targets:
        try:
            await db[coll].delete_many(query)
        except Exception as e:
            cleanup_errors.append(f"{coll}: {e}")
    
    # Also nuke OTP records / pending password reset state by email so
    # the deleted user can be re-invited without conflicts.
    if user_email:
        try:
            await db.otp_codes.delete_many({"email": user_email})
        except Exception as e:
            cleanup_errors.append(f"otp_codes: {e}")
    
    # Finally remove the user document itself
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    if cleanup_errors:
        logger.warning(f"Cascade cleanup partial for user {user_id}: {cleanup_errors}")
    
    return {
        "message": "User and associated data deleted successfully",
        "user_id": user_id,
        "cleanup_warnings": cleanup_errors or None,
    }


class BulkDeleteUsersRequest(BaseModel):
    user_ids: List[str]


@router.post("/users/bulk-delete")
async def bulk_delete_users(body: BulkDeleteUsersRequest, request: Request):
    """Delete multiple users (any role) in one request, with cascade
    cleanup. Returns per-user success/failure so the admin UI can show
    a clear summary instead of a generic "failed" toast when only some
    deletions failed.
    """
    await verify_admin(request)
    db = get_db(request)
    
    if not body.user_ids:
        raise HTTPException(status_code=400, detail="No user_ids provided")
    
    deleted: list = []
    failed: list = []
    for uid in body.user_ids:
        try:
            existing = await db.users.find_one({"id": uid}, {"_id": 0, "email": 1})
            if not existing:
                failed.append({"user_id": uid, "error": "Not found"})
                continue
            email = (existing.get("email") or "").lower()
            for coll, query in [
                ("peer_profiles",      {"user_id": uid}),
                ("peer_sessions",      {"$or": [{"requester_id": uid}, {"accepter_id": uid}]}),
                ("bookings",           {"user_id": uid}),
                ("session_feedbacks",  {"user_id": uid}),
                ("user_sessions",      {"user_id": uid}),
                ("slot_reservations",  {"user_id": uid}),
                ("payments",           {"user_id": uid}),
                ("payment_orders",     {"user_id": uid}),
                ("discount_usage",     {"user_id": uid}),
                ("notifications",      {"user_id": uid}),
                ("user_plan_assignments", {"user_id": uid}),
                ("subscriptions",      {"user_id": uid}),
                ("workshop_registrations", {"user_id": uid}),
                ("competition_entries", {"user_id": uid}),
                ("user_workshop_attendance", {"user_id": uid}),
            ]:
                try:
                    await db[coll].delete_many(query)
                except Exception:
                    pass
            if email:
                try:
                    await db.otp_codes.delete_many({"email": email})
                except Exception:
                    pass
            res = await db.users.delete_one({"id": uid})
            if res.deleted_count == 1:
                deleted.append(uid)
            else:
                failed.append({"user_id": uid, "error": "Delete returned 0"})
        except Exception as e:
            logger.exception(f"Bulk delete failed for user {uid}: {e}")
            failed.append({"user_id": uid, "error": str(e)})
    
    return {
        "deleted_count": len(deleted),
        "failed_count": len(failed),
        "deleted": deleted,
        "failed": failed,
    }


@router.post("/users/{user_id}/access")
async def update_user_access(user_id: str, access_data: AccessUpdate, request: Request):
    """Grant or revoke specific access for a user"""
    await verify_admin(request)
    db = get_db(request)
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Get or create custom_access field
    custom_access = user.get("custom_access", {})
    custom_access[access_data.access_type] = access_data.granted
    
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"custom_access": custom_access, "updated_at": datetime.utcnow()}}
    )
    
    return {"message": f"Access {'granted' if access_data.granted else 'revoked'} successfully"}


# ============ Mentor Management ============

@router.get("/mentors")
async def get_all_mentors(request: Request):
    """Get all mentors with dynamic ratings (including deleted ones for admin view)"""
    await verify_admin(request)
    db = get_db(request)
    
    # Get all mentors including deleted ones (for admin to see), sorted by display_order
    mentors = await db.mentors.find({}, {"_id": 0}).sort("display_order", 1).to_list(100)
    
    # Calculate dynamic ratings for each mentor
    result = []
    for m in mentors:
        mentor_id = m.get("id")
        
        # Get completed sessions count
        completed_sessions = await db.bookings.count_documents({
            "mentor_id": mentor_id,
            "status": "completed"
        })
        
        # Calculate dynamic rating from candidate feedbacks
        candidate_feedbacks = await db.candidate_feedbacks.find(
            {"mentor_id": mentor_id},
            {"rating": 1}
        ).to_list(500)
        
        if candidate_feedbacks:
            total_rating = sum(f.get("rating", 5) for f in candidate_feedbacks)
            m["rating"] = round(total_rating / len(candidate_feedbacks), 1)
            m["total_reviews"] = len(candidate_feedbacks)
        elif completed_sessions > 0:
            # Has sessions but no feedback yet
            m["rating"] = None  # Will show as "NA" in frontend
            m["total_reviews"] = 0
        else:
            # No sessions done
            m["rating"] = None
            m["total_reviews"] = 0
        
        m["sessions_done"] = completed_sessions
        result.append(m)
    
    return {"mentors": result}


@router.get("/mentors/{mentor_id}/details")
async def get_mentor_details(mentor_id: str, request: Request):
    """Get detailed mentor information including journey, sessions, and feedback"""
    await verify_admin(request)
    db = get_db(request)
    
    mentor = await db.mentors.find_one({"id": mentor_id}, {"_id": 0})
    if not mentor:
        raise HTTPException(status_code=404, detail="Mentor not found")
    
    # Get all sessions conducted by this mentor
    bookings = await db.bookings.find({"mentor_id": mentor_id}, {"_id": 0}).to_list(500)
    
    # Get feedbacks given by this mentor
    feedbacks_given = await db.session_feedbacks.find({"mentor_id": mentor_id}, {"_id": 0}).to_list(200)
    
    # Get feedbacks received from candidates
    candidate_feedbacks = await db.mentor_feedbacks.find({"mentor_id": mentor_id}, {"_id": 0}).to_list(200)
    
    # Get weekly availability
    availability = await db.mentor_weekly_availability.find(
        {"mentor_id": mentor_id}, 
        {"_id": 0, "mentor_id": 0, "updated_at": 0, "admin_override": 0}
    ).to_list(10)
    
    # Get earnings/payments
    payments = await db.mentor_payments.find({"mentor_id": mentor_id}, {"_id": 0}).to_list(100)
    
    # Calculate stats
    total_sessions = len(bookings)
    completed_sessions = len([b for b in bookings if b.get("status") == "completed"])
    cancelled_sessions = len([b for b in bookings if b.get("status") == "cancelled"])
    upcoming_sessions = len([b for b in bookings if b.get("status") == "scheduled"])
    
    # Calculate earnings
    total_earnings = sum(p.get("amount", 0) for p in payments if p.get("status") == "paid")
    pending_earnings = sum(p.get("amount", 0) for p in payments if p.get("status") == "pending")
    
    # Calculate average rating
    ratings = [f.get("rating", 0) for f in candidate_feedbacks if f.get("rating")]
    # Only show rating if there are actual ratings, don't default to 5.0
    avg_rating = sum(ratings) / len(ratings) if ratings else None
    
    # Session breakdown by month
    sessions_by_month = {}
    for booking in bookings:
        if booking.get("date"):
            month = booking.get("date")[:7]  # YYYY-MM
            sessions_by_month[month] = sessions_by_month.get(month, 0) + 1
    
    return {
        "mentor": mentor,
        "stats": {
            "total_sessions": total_sessions,
            "completed_sessions": completed_sessions,
            "cancelled_sessions": cancelled_sessions,
            "upcoming_sessions": upcoming_sessions,
            "total_earnings": total_earnings,
            "pending_earnings": pending_earnings,
            "average_rating": round(avg_rating, 2),
            "total_feedbacks_given": len(feedbacks_given),
            "total_feedbacks_received": len(candidate_feedbacks)
        },
        "recent_sessions": sorted(bookings, key=lambda x: x.get("date", ""), reverse=True)[:10],
        "feedbacks_given": feedbacks_given[:10],
        "candidate_feedbacks": candidate_feedbacks[:10],
        "availability": availability,
        "payments": payments[:20],
        "sessions_by_month": sessions_by_month
    }


@router.post("/mentors/invite")
async def invite_mentor(request: Request):
    """Invite a new mentor (creates mentor record with comprehensive profile)"""
    await verify_admin(request)
    db = get_db(request)
    
    data = await request.json()
    mentor_id = f"mentor-{str(uuid.uuid4())[:8]}"
    mentor_name = data.get("name", "Mentor")
    mentor_email = data.get("email")
    
    # Get picture URL or generate default
    picture = data.get("picture")
    if not picture:
        picture = f"https://ui-avatars.com/api/?name={mentor_name.replace(' ', '+')}&background=0D8ABC&color=fff"
    
    mentor = {
        "id": mentor_id,
        "name": mentor_name,
        "email": mentor_email,
        "phone": data.get("phone", ""),
        "linkedin": data.get("linkedin", ""),
        "location": data.get("location", ""),
        
        # Consulting experience
        "consulting_position": data.get("consulting_position", ""),
        "consulting_firm": data.get("consulting_firm", ""),
        "consulting_firm_logo": data.get("consulting_firm_logo"),
        "consulting_is_current": data.get("consulting_is_current", False),
        "college": data.get("college", ""),
        
        # Current & previous companies
        "current_company": data.get("current_company", ""),
        "current_company_logo": data.get("current_company_logo"),
        "previous_company_1": data.get("previous_company_1", ""),
        "previous_company_2": data.get("previous_company_2", ""),
        "years_experience": data.get("years_experience", ""),
        
        # For backward compatibility
        "specialization": data.get("specialization") or data.get("consulting_firm", "General"),
        "title": data.get("title") or data.get("consulting_position", "Consultant"),
        "company": data.get("company") or data.get("current_company", ""),
        
        # Pricing
        "hourly_rate": data.get("hourly_rate", 12000),
        "price_per_session": data.get("price_per_session", 1500),
        
        # Optional fields
        "headline": data.get("headline", ""),
        "is_top_coach": data.get("is_top_coach", False),
        "bio": data.get("bio", ""),
        
        # Profile
        "picture": picture,
        "rating": None,  # No rating yet
        "sessions_done": 0,
        "availability": [],
        "is_hidden": True,  # Hidden by default until admin makes visible
        "status": "invited",
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow()
    }
    
    await db.mentors.insert_one(mentor)
    
    # Also create a user account for the mentor
    user_id = f"user-mentor-{str(uuid.uuid4())[:8]}"
    mentor_user = {
        "id": user_id,
        "email": mentor_email,
        "name": mentor_name,
        "phone": data.get("phone", ""),
        "picture": mentor["picture"],
        "plan": "mentor",
        "is_mentor": True,
        "mentor_id": mentor_id,
        "is_admin": False,
        "created_at": datetime.utcnow()
    }
    await db.users.insert_one(mentor_user)
    
    # Send mentor invite email
    email_sent = await send_mentor_invite_email(db, mentor_email, mentor_name)
    
    return {
        "message": "Mentor invited successfully", 
        "mentor_id": mentor_id,
        "email_sent": email_sent
    }


@router.post("/migrate-session-statuses")
async def migrate_session_statuses(request: Request):
    """
    Migrate old session statuses to new status system:
    - pending → confirmed
    - cancelled, cancelled_by_candidate, cancelled_by_mentor, cancelled_by_admin → candidate_cancelled or mentor_cancelled based on cancelled_by field
    - no_show → both_no_show (default, since we can't determine who no-showed)
    - rescheduled → mentor_rescheduled or candidate_rescheduled based on rescheduled_by field
    """
    await verify_admin(request)
    db = get_db(request)
    
    stats = {
        "pending_to_confirmed": 0,
        "cancelled_migrated": 0,
        "no_show_migrated": 0,
        "rescheduled_migrated": 0,
        "errors": 0
    }
    
    # Migrate pending → confirmed
    result = await db.bookings.update_many(
        {"status": "pending"},
        {"$set": {"status": "confirmed", "migrated_from": "pending"}}
    )
    stats["pending_to_confirmed"] = result.modified_count
    
    # Migrate cancelled statuses
    cancelled_sessions = await db.bookings.find({
        "status": {"$in": ["cancelled", "cancelled_by_candidate", "cancelled_by_mentor", "cancelled_by_admin"]}
    }).to_list(10000)
    
    for session in cancelled_sessions:
        try:
            old_status = session.get("status")
            cancelled_by = session.get("cancelled_by", "")
            
            if old_status == "cancelled_by_candidate" or cancelled_by == "candidate":
                new_status = "candidate_cancelled"
            elif old_status == "cancelled_by_mentor" or cancelled_by == "mentor":
                new_status = "mentor_cancelled"
            else:
                # Default to candidate_cancelled if we can't determine
                new_status = "candidate_cancelled"
            
            await db.bookings.update_one(
                {"id": session.get("id")},
                {"$set": {"status": new_status, "migrated_from": old_status}}
            )
            stats["cancelled_migrated"] += 1
        except Exception as e:
            logger.error(f"Error migrating cancelled session {session.get('id')}: {e}")
            stats["errors"] += 1
    
    # Migrate no_show → both_no_show
    result = await db.bookings.update_many(
        {"status": "no_show"},
        {"$set": {"status": "both_no_show", "migrated_from": "no_show"}}
    )
    stats["no_show_migrated"] = result.modified_count
    
    # Migrate rescheduled based on rescheduled_by field
    rescheduled_sessions = await db.bookings.find({
        "status": "rescheduled"
    }).to_list(10000)
    
    for session in rescheduled_sessions:
        try:
            rescheduled_by = session.get("rescheduled_by", "")
            
            if rescheduled_by == "mentor":
                new_status = "mentor_rescheduled"
            elif rescheduled_by == "candidate":
                new_status = "candidate_rescheduled"
            else:
                # Default to candidate_rescheduled if we can't determine
                new_status = "candidate_rescheduled"
            
            await db.bookings.update_one(
                {"id": session.get("id")},
                {"$set": {"status": new_status, "migrated_from": "rescheduled"}}
            )
            stats["rescheduled_migrated"] += 1
        except Exception as e:
            logger.error(f"Error migrating rescheduled session {session.get('id')}: {e}")
            stats["errors"] += 1
    
    return {
        "message": "Migration completed",
        "stats": stats
    }


@router.get("/mentors/template")
async def get_mentor_template(request: Request):
    """Generate and return Excel template for bulk mentor upload"""
    await verify_admin(request)
    
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Please install it first.")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mentors"
    
    # Define headers
    headers = [
        "Full Name*", "Email*", "Phone*", "LinkedIn*", "Location*",
        "Consulting Position*", "Consulting Firm*", "Current Company*",
        "Consulting Is Current (Y/N)", "Previous Company 1", "Previous Company 2",
        "Years Experience*", "Hourly Rate*", "Session Price*", "Headline", "Top Coach (Y/N)"
    ]
    
    # Style for header row
    header_fill = PatternFill(start_color="0D8ABC", end_color="0D8ABC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # Add sample row
    sample_data = [
        "John Smith", "john.smith@example.com", "+91 98765 43210", "linkedin.com/in/johnsmith",
        "Mumbai, India", "Senior Consultant", "McKinsey & Company", "Google",
        "N", "Amazon", "Microsoft", "8", "12000", "1500", "Ex-McKinsey | 100+ Cases", "N"
    ]
    
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Add instructions sheet
    instructions = wb.create_sheet("Instructions")
    instructions_text = [
        ["Mentor Bulk Upload Instructions"],
        [""],
        ["Required Fields (marked with *):"],
        ["- Full Name: Mentor's complete name"],
        ["- Email: Valid email address (will be used for login)"],
        ["- Phone: Phone number with country code"],
        ["- LinkedIn: LinkedIn profile URL or username"],
        ["- Location: City, Country format"],
        ["- Consulting Position: Last position at consulting firm"],
        ["- Consulting Firm: Name of the consulting firm (e.g., McKinsey, BCG, Bain)"],
        ["- Current Company: Current employer (or consulting firm if still there)"],
        ["- Years Experience: Total years of work experience (number)"],
        ["- Hourly Rate: Hourly coaching rate in INR (number)"],
        ["- Session Price: Single session price in INR (number)"],
        [""],
        ["Optional Fields:"],
        ["- Consulting Is Current: Y if consulting firm is current company, N otherwise"],
        ["- Previous Company 1 & 2: Previous employers"],
        ["- Headline: Short bio/tagline for the mentor card"],
        ["- Top Coach: Y to mark as featured top coach, N otherwise"],
        [""],
        ["Notes:"],
        ["- Profile photos will use default avatars initially"],
        ["- You can update photos individually after upload"],
        ["- Mentors will be created as 'Hidden' - make them visible from admin panel"],
        ["- Delete the sample row before uploading"],
    ]
    
    for row, text in enumerate(instructions_text, 1):
        cell = instructions.cell(row=row, column=1, value=text[0] if text else "")
        if row == 1:
            cell.font = Font(bold=True, size=14)
        instructions.column_dimensions['A'].width = 80
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=mentor_upload_template.xlsx"}
    )


@router.post("/mentors/bulk-upload")
async def bulk_upload_mentors(request: Request, file: UploadFile = File(...)):
    """Bulk upload mentors from Excel file"""
    await verify_admin(request)
    db = get_db(request)
    
    # Check file type
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Please upload an Excel file (.xlsx or .xls)")
    
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    # Read file
    import io
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents))
    ws = wb.active
    
    # Get headers from first row
    headers = [cell.value for cell in ws[1]]
    
    # Expected column mapping
    column_map = {
        "Full Name*": "name",
        "Email*": "email",
        "Phone*": "phone",
        "LinkedIn*": "linkedin",
        "Location*": "location",
        "Consulting Position*": "consulting_position",
        "Consulting Firm*": "consulting_firm",
        "Current Company*": "current_company",
        "Consulting Is Current (Y/N)": "consulting_is_current",
        "Previous Company 1": "previous_company_1",
        "Previous Company 2": "previous_company_2",
        "Years Experience*": "years_experience",
        "Hourly Rate*": "hourly_rate",
        "Session Price*": "session_price",
        "Headline": "headline",
        "Top Coach (Y/N)": "is_top_coach"
    }
    
    # Map header indices
    header_indices = {}
    for idx, header in enumerate(headers):
        if header in column_map:
            header_indices[column_map[header]] = idx
    
    # Process rows
    created_count = 0
    errors = []
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip empty rows
        if not row or not row[0]:
            continue
        
        try:
            # Extract data
            def get_value(field, default=""):
                idx = header_indices.get(field)
                if idx is not None and idx < len(row):
                    return row[idx] if row[idx] is not None else default
                return default
            
            name = str(get_value("name", "")).strip()
            email = str(get_value("email", "")).strip()
            
            # Skip if required fields missing
            if not name or not email:
                errors.append(f"Row {row_num}: Missing name or email")
                continue
            
            # Check if email already exists
            existing = await db.mentors.find_one({"email": email})
            if existing:
                errors.append(f"Row {row_num}: Email {email} already exists")
                continue
            
            # Parse boolean fields
            consulting_is_current = str(get_value("consulting_is_current", "N")).upper() in ["Y", "YES", "TRUE", "1"]
            is_top_coach = str(get_value("is_top_coach", "N")).upper() in ["Y", "YES", "TRUE", "1"]
            
            # Parse numeric fields
            try:
                years_exp = int(float(get_value("years_experience", 0) or 0))
            except:
                years_exp = 0
            
            try:
                hourly_rate = int(float(get_value("hourly_rate", 12000) or 12000))
            except:
                hourly_rate = 12000
            
            try:
                session_price = int(float(get_value("session_price", 1500) or 1500))
            except:
                session_price = 1500
            
            consulting_firm = str(get_value("consulting_firm", "")).strip()
            current_company = str(get_value("current_company", "")).strip()
            
            if consulting_is_current:
                current_company = consulting_firm
            
            # Generate default avatar
            picture = f"https://ui-avatars.com/api/?name={name.replace(' ', '+')}&background=0D8ABC&color=fff&size=200"
            
            # Create mentor record
            mentor_id = f"mentor-{str(uuid.uuid4())[:8]}"
            mentor = {
                "id": mentor_id,
                "name": name,
                "email": email,
                "phone": str(get_value("phone", "")).strip(),
                "linkedin": str(get_value("linkedin", "")).strip(),
                "location": str(get_value("location", "")).strip(),
                "consulting_position": str(get_value("consulting_position", "")).strip(),
                "consulting_firm": consulting_firm,
                "consulting_is_current": consulting_is_current,
                "current_company": current_company,
                "previous_company_1": str(get_value("previous_company_1", "")).strip(),
                "previous_company_2": str(get_value("previous_company_2", "")).strip(),
                "years_experience": years_exp,
                "hourly_rate": hourly_rate,
                "price_per_session": session_price,
                "headline": str(get_value("headline", "")).strip(),
                "is_top_coach": is_top_coach,
                "picture": picture,
                # For backward compatibility
                "specialization": consulting_firm,
                "title": str(get_value("consulting_position", "Consultant")).strip(),
                "company": current_company,
                "bio": "",
                "rating": None,
                "sessions_done": 0,
                "availability": [],
                "is_hidden": True,  # Hidden by default
                "status": "active",
                "created_at": datetime.utcnow(),
                "updated_at": datetime.utcnow()
            }
            
            await db.mentors.insert_one(mentor)
            
            # Create user account
            user_id = f"user-mentor-{str(uuid.uuid4())[:8]}"
            mentor_user = {
                "id": user_id,
                "email": email,
                "name": name,
                "phone": mentor["phone"],
                "picture": picture,
                "plan": "mentor",
                "is_mentor": True,
                "mentor_id": mentor_id,
                "is_admin": False,
                "created_at": datetime.utcnow()
            }
            await db.users.insert_one(mentor_user)
            
            created_count += 1
            
        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")
    
    return {
        "message": "Bulk upload completed",
        "created": created_count,
        "errors": errors,
        "total_rows_processed": row_num - 1 if 'row_num' in locals() else 0
    }


@router.delete("/mentors/clear-all")
async def clear_all_mentors(request: Request):
    """Delete all mentors from the system"""
    await verify_admin(request)
    db = get_db(request)
    
    # Count before deletion
    mentor_count = await db.mentors.count_documents({})
    
    # Delete all mentors
    await db.mentors.delete_many({})
    
    # Delete mentor user accounts
    await db.users.delete_many({"is_mentor": True})
    
    # Delete mentor availability
    await db.mentor_weekly_availability.delete_many({})
    
    return {
        "message": "All mentors cleared",
        "deleted_count": mentor_count
    }


@router.post("/mentors")
async def create_mentor(mentor_data: MentorCreate, request: Request):
    """Create a new mentor"""
    await verify_admin(request)
    db = get_db(request)
    
    mentor_id = f"mentor-{str(uuid.uuid4())[:8]}"
    
    mentor = {
        "id": mentor_id,
        **mentor_data.dict(),
        "rating": 5.0,
        "sessions_conducted": 0,
        "years_experience": 0,
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow()
    }
    
    await db.mentors.insert_one(mentor)
    
    # Also create a user account for the mentor
    user_id = f"user-mentor-{str(uuid.uuid4())[:8]}"
    mentor_user = {
        "id": user_id,
        "email": mentor_data.email,
        "name": mentor_data.name,
        "picture": mentor_data.picture or "https://images.unsplash.com/photo-1472099645785-5658abf4ff4e?w=100&h=100&fit=crop&crop=face",
        "plan": "mentor",
        "is_mentor": True,
        "mentor_id": mentor_id,
        "is_admin": False,
        "created_at": datetime.utcnow()
    }
    await db.users.insert_one(mentor_user)
    
    return {"message": "Mentor created successfully", "mentor_id": mentor_id}


@router.put("/mentors/{mentor_id}")
async def update_mentor(mentor_id: str, request: Request):
    """Update mentor details with comprehensive profile"""
    await verify_admin(request)
    db = get_db(request)
    
    data = await request.json()
    
    # Build update data from all provided fields
    update_fields = [
        "name", "email", "phone", "linkedin", "location",
        "consulting_position", "consulting_firm", "consulting_firm_logo",
        "consulting_is_current", "current_company", "current_company_logo",
        "previous_company_1", "previous_company_2", "years_experience",
        "hourly_rate", "price_per_session", "headline", "is_top_coach",
        "is_landing_featured",  # Show on public landing-page mentor carousel
        "bio", "picture", "specialization", "title", "company",
        "can_take_strategy_calls", "college"  # Strategy call eligibility + college
    ]
    
    update_data = {}
    for field in update_fields:
        if field in data:
            update_data[field] = data[field]
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No update data provided")
    
    update_data["updated_at"] = datetime.utcnow()
    
    result = await db.mentors.update_one({"id": mentor_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Mentor not found")
    
    # Invalidate mentor list cache so the next request picks up the change
    try:
        from routes.mentors import invalidate_mentor_cache
        invalidate_mentor_cache()
    except Exception:
        pass

    # Also update the user account if email or name changed
    if "email" in update_data or "name" in update_data or "picture" in update_data or "phone" in update_data:
        user_update = {}
        if "email" in update_data:
            user_update["email"] = update_data["email"]
        if "name" in update_data:
            user_update["name"] = update_data["name"]
        if "picture" in update_data:
            user_update["picture"] = update_data["picture"]
        if "phone" in update_data:
            user_update["phone"] = update_data["phone"]
        
        if user_update:
            await db.users.update_many(
                {"mentor_id": mentor_id},
                {"$set": user_update}
            )
    
    return {"message": "Mentor updated successfully"}


@router.delete("/mentors/{mentor_id}")
async def delete_mentor(mentor_id: str, request: Request):
    """
    Hard-delete a mentor - completely removes the mentor profile from the database.
    This will:
    1. Delete the mentor record from the database
    2. Remove mentor access from the associated user account (user becomes regular candidate)
    3. Cancel any upcoming bookings
    4. Delete mentor's weekly availability settings
    5. If the same email tries to login again, they'll be treated as a fresh user
    """
    await verify_admin(request)
    db = get_db(request)
    
    # Find the mentor
    mentor = await db.mentors.find_one({"id": mentor_id})
    if not mentor:
        raise HTTPException(status_code=404, detail="Mentor not found")
    
    mentor_email = mentor.get("email")
    mentor_name = mentor.get("name", "Mentor")
    user_id = mentor.get("user_id")
    
    # Cancel any upcoming coaching bookings for this mentor
    from datetime import date
    today = date.today().isoformat()
    
    # First, get all bookings that will be cancelled so we can delete their calendar events
    upcoming_bookings = await db.bookings.find({
        "mentor_id": mentor_id,
        "date": {"$gte": today},
        "status": {"$in": ["scheduled", "confirmed", "pending"]}
    }).to_list(1000)
    
    # Delete calendar events for all these bookings
    from services.calendar_service import get_calendar_service
    calendar_service = get_calendar_service()
    if calendar_service.is_available():
        for booking in upcoming_bookings:
            calendar_event_id = booking.get("calendar_event_id")
            if calendar_event_id:
                try:
                    calendar_service.cancel_event(calendar_event_id, notify_attendees=True)
                    logger.info(f"Cancelled calendar event {calendar_event_id} due to mentor deletion")
                except Exception as cal_err:
                    logger.warning(f"Failed to cancel calendar event {calendar_event_id}: {cal_err}")
            
            hidden_event_id = booking.get("hidden_event_id")
            if hidden_event_id:
                try:
                    calendar_service.cancel_event(hidden_event_id, notify_attendees=False)
                except Exception as cal_err:
                    logger.warning(f"Failed to cancel hidden event {hidden_event_id}: {cal_err}")
    
    cancelled_bookings = await db.bookings.update_many(
        {
            "mentor_id": mentor_id,
            "date": {"$gte": today},
            "status": {"$in": ["scheduled", "confirmed", "pending"]}
        },
        {
            "$set": {
                "status": "cancelled",
                "cancelled_by": "admin",
                "cancellation_reason": "Mentor profile deleted",
                "updated_at": datetime.utcnow()
            }
        }
    )
    
    # Remove mentor access from the associated user account
    # This allows them to use the platform as a regular candidate
    # If they login again, they won't have mentor privileges
    user_update_result = await db.users.update_many(
        {"$or": [
            {"mentor_id": mentor_id},
            {"email": mentor_email},
            {"id": user_id} if user_id else {"_id": None}
        ]},
        {
            "$set": {
                "is_mentor": False,
                "updated_at": datetime.utcnow()
            },
            "$unset": {
                "mentor_id": "",
                "mentor_deleted": "",
                "mentor_deleted_at": ""
            }
        }
    )
    
    # Delete mentor's weekly availability
    await db.mentor_weekly_availability.delete_many({"mentor_id": mentor_id})
    
    # Delete mentor's daily availability overrides
    await db.mentor_availability.delete_many({"mentor_id": mentor_id})
    
    # Hard delete the mentor record
    delete_result = await db.mentors.delete_one({"id": mentor_id})
    
    return {
        "message": f"Mentor '{mentor_name}' permanently deleted",
        "mentor_id": mentor_id,
        "mentor_email": mentor_email,
        "deleted": delete_result.deleted_count > 0,
        "user_accounts_updated": user_update_result.modified_count,
        "bookings_cancelled": cancelled_bookings.modified_count,
        "note": "If this email registers again, they will be treated as a fresh user"
    }


@router.post("/mentors/{mentor_id}/restore")
async def restore_mentor(mentor_id: str, request: Request):
    """Restore a soft-deleted mentor - Note: This only works for soft-deleted mentors, not hard-deleted ones"""
    await verify_admin(request)
    db = get_db(request)
    
    # Find the mentor
    mentor = await db.mentors.find_one({"id": mentor_id})
    if not mentor:
        raise HTTPException(status_code=404, detail="Mentor not found")
    
    if not mentor.get("is_deleted"):
        raise HTTPException(status_code=400, detail="Mentor is not deleted")
    
    mentor_email = mentor.get("email")
    
    # Restore the mentor
    await db.mentors.update_one(
        {"id": mentor_id},
        {
            "$set": {
                "is_deleted": False,
                "is_active": True,
                "updated_at": datetime.utcnow()
            },
            "$unset": {
                "deleted_at": ""
            }
        }
    )
    
    # Restore mentor access for the user
    await db.users.update_many(
        {"email": mentor_email},
        {
            "$set": {
                "is_mentor": True,
                "mentor_id": mentor_id,
                "updated_at": datetime.utcnow()
            },
            "$unset": {
                "mentor_deleted": "",
                "mentor_deleted_at": ""
            }
        }
    )
    
    return {"message": "Mentor restored successfully", "mentor_id": mentor_id}


@router.post("/mentors/reorder")
async def reorder_mentors(request: Request):
    """Reorder mentors display order"""
    await verify_admin(request)
    db = get_db(request)
    
    data = await request.json()
    orders = data.get("orders", [])
    
    if not orders:
        raise HTTPException(status_code=400, detail="No order data provided")
    
    # Update display_order for each mentor
    for item in orders:
        mentor_id = item.get("id")
        display_order = item.get("display_order")
        
        if mentor_id is not None and display_order is not None:
            await db.mentors.update_one(
                {"id": mentor_id},
                {"$set": {"display_order": display_order, "updated_at": datetime.utcnow()}}
            )
    
    return {"message": "Mentors reordered successfully", "count": len(orders)}


@router.put("/mentors/{mentor_id}/availability")
async def override_mentor_availability(mentor_id: str, data: AvailabilityOverride, request: Request):
    """Override mentor weekly availability and blocked days (admin override)"""
    await verify_admin(request)
    db = get_db(request)
    
    # Clear existing weekly availability template
    await db.mentor_weekly_availability.delete_many({"mentor_id": mentor_id})
    
    # Clear cached per-day availability (stale data from previous template)
    await db.mentor_availability.delete_many({"mentor_id": mentor_id})
    
    # Insert new weekly availability
    for day_data in data.availability:
        day_name = day_data.get("day")
        slots = day_data.get("slots", [])
        
        if day_name:
            await db.mentor_weekly_availability.insert_one({
                "mentor_id": mentor_id,
                "day": day_name,
                "slots": slots,
                "admin_override": True,
                "updated_at": datetime.utcnow()
            })
    
    # Update blocked days in mentors collection
    blocked_days = data.blocked_days if hasattr(data, 'blocked_days') and data.blocked_days else []
    await db.mentors.update_one(
        {"id": mentor_id},
        {"$set": {
            "blocked_days": blocked_days,
            "updated_at": datetime.utcnow()
        }}
    )
    
    return {"message": "Mentor availability and blocked days updated successfully"}


@router.get("/mentors/{mentor_id}/availability")
async def get_mentor_availability(mentor_id: str, request: Request):
    """Get mentor's current weekly availability and blocked days for admin view"""
    await verify_admin(request)
    db = get_db(request)
    
    # Check if mentor exists and get blocked_days
    mentor = await db.mentors.find_one({"id": mentor_id}, {"_id": 0, "name": 1, "email": 1, "blocked_days": 1})
    if not mentor:
        raise HTTPException(status_code=404, detail="Mentor not found")
    
    # Get weekly availability from mentor_weekly_availability collection
    availability_cursor = db.mentor_weekly_availability.find(
        {"mentor_id": mentor_id},
        {"_id": 0, "day": 1, "slots": 1, "admin_override": 1, "updated_at": 1}
    )
    availability_docs = await availability_cursor.to_list(10)
    
    # Convert to the format expected by frontend
    availability = []
    has_admin_override = False
    last_updated = None
    
    for doc in availability_docs:
        day = doc.get("day")
        slots = doc.get("slots", [])
        if doc.get("admin_override"):
            has_admin_override = True
        if doc.get("updated_at"):
            last_updated = doc.get("updated_at")
        
        if day and slots:
            availability.append({
                "day": day,
                "slots": slots
            })
    
    return {
        "mentor_id": mentor_id,
        "mentor_name": mentor.get("name"),
        "availability": availability,
        "blocked_days": mentor.get("blocked_days", []),
        "has_admin_override": has_admin_override,
        "last_updated": last_updated.isoformat() if last_updated else None,
        "is_empty": len(availability) == 0
    }


@router.put("/mentors/{mentor_id}/visibility")
async def toggle_mentor_visibility(mentor_id: str, request: Request):
    """Toggle mentor visibility on candidate dashboard (regular mentor list)"""
    await verify_admin(request)
    db = get_db(request)
    
    # Get current mentor
    mentor = await db.mentors.find_one({"id": mentor_id})
    if not mentor:
        raise HTTPException(status_code=404, detail="Mentor not found")
    
    # Toggle is_hidden field
    current_hidden = mentor.get("is_hidden", False)
    new_hidden = not current_hidden
    
    await db.mentors.update_one(
        {"id": mentor_id},
        {"$set": {"is_hidden": new_hidden, "updated_at": datetime.utcnow()}}
    )
    
    return {
        "message": f"Mentor {'hidden from' if new_hidden else 'visible on'} candidate dashboard",
        "is_hidden": new_hidden
    }


@router.put("/mentors/{mentor_id}/strategy-call-visibility")
async def toggle_mentor_strategy_call_visibility(mentor_id: str, request: Request):
    """Toggle mentor visibility in strategy call selection (separate from general visibility)"""
    await verify_admin(request)
    db = get_db(request)
    
    # Get current mentor
    mentor = await db.mentors.find_one({"id": mentor_id})
    if not mentor:
        raise HTTPException(status_code=404, detail="Mentor not found")
    
    # Toggle is_hidden_from_strategy_calls field
    current_hidden = mentor.get("is_hidden_from_strategy_calls", False)
    new_hidden = not current_hidden
    
    await db.mentors.update_one(
        {"id": mentor_id},
        {"$set": {"is_hidden_from_strategy_calls": new_hidden, "updated_at": datetime.utcnow()}}
    )
    
    return {
        "message": f"Mentor {'hidden from' if new_hidden else 'visible in'} strategy call selection",
        "is_hidden_from_strategy_calls": new_hidden
    }


# ============ Admin Calendar View ============

@router.get("/calendar/day/{date}")
async def get_calendar_day_view(date: str, request: Request):
    """
    Get all mentors' availability and booked sessions for a specific day.
    Returns data formatted for the admin calendar day view.
    """
    await verify_admin(request)
    db = get_db(request)
    
    try:
        target_date = datetime.strptime(date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    day_name = target_date.strftime("%A")  # e.g., "Monday"
    
    # Get all active mentors (not deleted, not hidden)
    mentors = await db.mentors.find(
        {"is_deleted": {"$ne": True}},
        {"_id": 0, "id": 1, "name": 1, "consulting_firm": 1, "picture": 1, "blocked_days": 1}
    ).to_list(100)
    
    # Get all bookings for this date
    bookings = await db.bookings.find(
        {"date": date},
        {"_id": 0}
    ).to_list(500)
    
    # Create a map of mentor_id -> bookings
    bookings_by_mentor = {}
    for booking in bookings:
        mentor_id = booking.get("mentor_id")
        if mentor_id not in bookings_by_mentor:
            bookings_by_mentor[mentor_id] = []
        bookings_by_mentor[mentor_id].append(booking)
    
    # Build calendar data for each mentor
    calendar_data = []
    
    for mentor in mentors:
        mentor_id = mentor.get("id")
        
        # Check if this day is blocked
        blocked_days = mentor.get("blocked_days", [])
        is_blocked = date in blocked_days
        
        # Get mentor's weekly availability for this day
        availability_doc = await db.mentor_weekly_availability.find_one(
            {"mentor_id": mentor_id, "day": day_name},
            {"_id": 0, "slots": 1}
        )
        
        available_slots = []
        if availability_doc and not is_blocked:
            available_slots = availability_doc.get("slots", [])
        
        # Get bookings for this mentor on this date
        mentor_bookings = bookings_by_mentor.get(mentor_id, [])
        
        # Build time slots (Full 24 hours in 30-min increments)
        time_slots = []
        for hour in range(0, 24):  # 12 AM to 11:30 PM
            for minute in [0, 30]:
                slot_time = f"{hour:02d}:{minute:02d}"
                slot_end = f"{hour:02d}:{minute + 30:02d}" if minute == 0 else f"{(hour + 1) % 24:02d}:00"
                
                # Check if this slot is available
                is_available = False
                for avail in available_slots:
                    avail_from = avail.get("from", "")
                    avail_to = avail.get("to", "")
                    if avail_from <= slot_time < avail_to:
                        is_available = True
                        break
                
                # Check if this slot is booked
                booked_session = None
                for booking in mentor_bookings:
                    booking_time = booking.get("time", "")
                    if booking_time == slot_time:
                        booked_session = {
                            "id": booking.get("id"),
                            "candidate_name": booking.get("user_name", "Unknown"),
                            "candidate_email": booking.get("user_email", ""),
                            "session_type": booking.get("session_type", "coaching"),
                            "status": booking.get("status", "scheduled")
                        }
                        break
                
                time_slots.append({
                    "time": slot_time,
                    "time_display": datetime.strptime(slot_time, "%H:%M").strftime("%I:%M %p"),
                    "is_available": is_available and not is_blocked,
                    "is_booked": booked_session is not None,
                    "is_blocked": is_blocked,
                    "booking": booked_session
                })
        
        calendar_data.append({
            "mentor_id": mentor_id,
            "mentor_name": mentor.get("name"),
            "firm": mentor.get("consulting_firm", ""),
            "picture": mentor.get("picture", ""),
            "is_blocked": is_blocked,
            "total_available": len([s for s in time_slots if s["is_available"] and not s["is_booked"]]),
            "total_booked": len([s for s in time_slots if s["is_booked"]]),
            "time_slots": time_slots
        })
    
    return {
        "date": date,
        "day_name": day_name,
        "mentors": calendar_data
    }


@router.get("/calendar/week/{start_date}")
async def get_calendar_week_view(start_date: str, request: Request):
    """
    Get all mentors' availability summary for a week starting from the given date.
    Returns aggregated data for the admin calendar week view.
    """
    await verify_admin(request)
    db = get_db(request)
    
    try:
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    # Generate dates for the week
    week_dates = []
    for i in range(7):
        d = start + timedelta(days=i)
        week_dates.append({
            "date": d.strftime("%Y-%m-%d"),
            "day_name": d.strftime("%A"),
            "day_short": d.strftime("%a"),
            "day_num": d.day
        })
    
    # Get all active mentors
    mentors = await db.mentors.find(
        {"is_deleted": {"$ne": True}},
        {"_id": 0, "id": 1, "name": 1, "consulting_firm": 1, "picture": 1, "blocked_days": 1}
    ).to_list(100)
    
    # Get all bookings for this week
    date_strings = [d["date"] for d in week_dates]
    bookings = await db.bookings.find(
        {"date": {"$in": date_strings}},
        {"_id": 0, "mentor_id": 1, "date": 1, "time": 1, "status": 1}
    ).to_list(1000)
    
    # Create booking counts by mentor and date
    booking_counts = {}
    for booking in bookings:
        key = f"{booking.get('mentor_id')}_{booking.get('date')}"
        if key not in booking_counts:
            booking_counts[key] = 0
        if booking.get("status") != "cancelled":
            booking_counts[key] += 1
    
    # Get all weekly availability
    all_availability = await db.mentor_weekly_availability.find(
        {},
        {"_id": 0, "mentor_id": 1, "day": 1, "slots": 1}
    ).to_list(1000)
    
    # Create availability map
    availability_map = {}
    for avail in all_availability:
        key = f"{avail.get('mentor_id')}_{avail.get('day')}"
        slots = avail.get("slots", [])
        # Count 30-min slots
        total_slots = 0
        for slot in slots:
            try:
                from_time = datetime.strptime(slot.get("from", "09:00"), "%H:%M")
                to_time = datetime.strptime(slot.get("to", "17:00"), "%H:%M")
                duration = (to_time - from_time).seconds // 60
                total_slots += duration // 30
            except:
                pass
        availability_map[key] = total_slots
    
    # Build week data for each mentor
    calendar_data = []
    
    for mentor in mentors:
        mentor_id = mentor.get("id")
        blocked_days = mentor.get("blocked_days", [])
        
        week_summary = []
        for day_info in week_dates:
            date_str = day_info["date"]
            day_name = day_info["day_name"]
            
            is_blocked = date_str in blocked_days
            avail_key = f"{mentor_id}_{day_name}"
            booking_key = f"{mentor_id}_{date_str}"
            
            total_slots = availability_map.get(avail_key, 0) if not is_blocked else 0
            booked_count = booking_counts.get(booking_key, 0)
            
            week_summary.append({
                "date": date_str,
                "day_short": day_info["day_short"],
                "day_num": day_info["day_num"],
                "is_blocked": is_blocked,
                "total_slots": total_slots,
                "booked": booked_count,
                "available": max(0, total_slots - booked_count)
            })
        
        calendar_data.append({
            "mentor_id": mentor_id,
            "mentor_name": mentor.get("name"),
            "firm": mentor.get("consulting_firm", ""),
            "picture": mentor.get("picture", ""),
            "week": week_summary
        })
    
    return {
        "start_date": start_date,
        "week_dates": week_dates,
        "mentors": calendar_data
    }


# ============ Mentor Profile Approval Workflow ============

class MentorProfileChangeRequest(BaseModel):
    name: Optional[str] = None
    title: Optional[str] = None
    company: Optional[str] = None
    bio: Optional[str] = None
    expertise: Optional[List[str]] = None
    picture: Optional[str] = None
    linkedin: Optional[str] = None
    specialization: Optional[str] = None
    hourly_rate: Optional[int] = None


@router.get("/mentors/pending-changes")
async def get_pending_mentor_changes(request: Request):
    """Get all mentors with pending profile changes awaiting approval"""
    await verify_admin(request)
    db = get_db(request)
    
    # Find mentors with pending_changes field
    mentors_with_changes = await db.mentors.find(
        {"pending_changes": {"$exists": True, "$ne": None}},
        {"_id": 0}
    ).to_list(100)
    
    return {"pending_approvals": mentors_with_changes}


@router.get("/mentors/pending-changes/count")
async def get_pending_mentor_changes_count(request: Request):
    """Get count of mentors with pending profile changes"""
    await verify_admin(request)
    db = get_db(request)
    
    count = await db.mentors.count_documents(
        {"pending_changes": {"$exists": True, "$ne": None}}
    )
    
    return {"count": count}


@router.post("/mentors/{mentor_id}/approve-changes")
async def approve_mentor_changes(mentor_id: str, request: Request):
    """Approve pending profile changes for a mentor"""
    await verify_admin(request)
    db = get_db(request)
    
    mentor = await db.mentors.find_one({"id": mentor_id})
    if not mentor:
        raise HTTPException(status_code=404, detail="Mentor not found")
    
    pending_changes = mentor.get("pending_changes")
    if not pending_changes:
        raise HTTPException(status_code=400, detail="No pending changes to approve")
    
    # Apply pending changes to the mentor profile
    update_data = {k: v for k, v in pending_changes.items() if v is not None and k != "submitted_at"}
    update_data["updated_at"] = datetime.utcnow()
    
    # Remove pending_changes field after approval
    await db.mentors.update_one(
        {"id": mentor_id},
        {
            "$set": update_data,
            "$unset": {"pending_changes": ""}
        }
    )
    
    # Record approval in history
    await db.mentor_change_history.insert_one({
        "id": f"change-{str(uuid.uuid4())[:8]}",
        "mentor_id": mentor_id,
        "changes": pending_changes,
        "status": "approved",
        "approved_at": datetime.utcnow(),
        "approved_by": (await get_current_user(request)).get("id")
    })
    
    return {"message": "Mentor profile changes approved and applied"}


@router.post("/mentors/{mentor_id}/reject-changes")
async def reject_mentor_changes(mentor_id: str, request: Request):
    """Reject pending profile changes for a mentor"""
    await verify_admin(request)
    db = get_db(request)
    
    body = await request.json()
    rejection_reason = body.get("reason", "")
    
    mentor = await db.mentors.find_one({"id": mentor_id})
    if not mentor:
        raise HTTPException(status_code=404, detail="Mentor not found")
    
    pending_changes = mentor.get("pending_changes")
    if not pending_changes:
        raise HTTPException(status_code=400, detail="No pending changes to reject")
    
    # Remove pending_changes field
    await db.mentors.update_one(
        {"id": mentor_id},
        {"$unset": {"pending_changes": ""}}
    )
    
    # Record rejection in history
    await db.mentor_change_history.insert_one({
        "id": f"change-{str(uuid.uuid4())[:8]}",
        "mentor_id": mentor_id,
        "changes": pending_changes,
        "status": "rejected",
        "rejection_reason": rejection_reason,
        "rejected_at": datetime.utcnow(),
        "rejected_by": (await get_current_user(request)).get("id")
    })
    
    return {"message": "Mentor profile changes rejected"}



# ============ Videos Management ============

@router.get("/videos")
async def get_all_videos(request: Request):
    """Get all videos"""
    await verify_admin(request)
    db = get_db(request)
    
    videos = await db.videos.find({}, {"_id": 0}).to_list(200)
    return {"videos": videos}


@router.post("/videos")
async def create_video(video_data: VideoCreate, request: Request):
    """Create a new video"""
    await verify_admin(request)
    db = get_db(request)
    
    video_id = f"video-{str(uuid.uuid4())[:8]}"
    
    video = {
        "id": video_id,
        **video_data.dict(),
        "locked": not video_data.is_free,
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow()
    }
    
    await db.videos.insert_one(video)
    return {"message": "Video created successfully", "video_id": video_id}


@router.put("/videos/{video_id}")
async def update_video(video_id: str, video_data: VideoUpdate, request: Request):
    """Update video details"""
    await verify_admin(request)
    db = get_db(request)
    
    update_data = {k: v for k, v in video_data.dict().items() if v is not None}
    if "is_free" in update_data:
        update_data["locked"] = not update_data["is_free"]
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No update data provided")
    
    update_data["updated_at"] = datetime.utcnow()
    
    result = await db.videos.update_one({"id": video_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Video not found")
    
    return {"message": "Video updated successfully"}


@router.delete("/videos/{video_id}")
async def delete_video(video_id: str, request: Request):
    """Delete a video"""
    await verify_admin(request)
    db = get_db(request)
    
    result = await db.videos.delete_one({"id": video_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Video not found")
    
    return {"message": "Video deleted successfully"}


# ============ Courses Management ============

@router.get("/courses")
async def get_all_courses(request: Request):
    """Get all courses with full hierarchy (3-level: Course -> Module -> Session)"""
    await verify_admin(request)
    db = get_db(request)
    
    # Get all courses - sort by order, with null/missing order at the end
    courses_cursor = db.courses.find({}, {"_id": 0})
    all_courses = await courses_cursor.to_list(100)
    
    # Sort courses: those with order first (by order), then those without order
    courses_with_order = [c for c in all_courses if c.get("order") is not None]
    courses_without_order = [c for c in all_courses if c.get("order") is None]
    
    courses_with_order.sort(key=lambda x: x.get("order", 0))
    courses_without_order.sort(key=lambda x: x.get("created_at", "") or x.get("title", ""))
    
    courses = courses_with_order + courses_without_order
    
    result = []
    for course in courses:
        course_id = course.get("id")
        
        # Get modules - sort by order, null values at end
        all_modules = await db.course_modules.find(
            {"course_id": course_id}, {"_id": 0}
        ).to_list(100)
        
        modules_with_order = [m for m in all_modules if m.get("order") is not None]
        modules_without_order = [m for m in all_modules if m.get("order") is None]
        modules_with_order.sort(key=lambda x: x.get("order", 0))
        modules = modules_with_order + modules_without_order
        
        course_modules = []
        for module in modules:
            module_id = module.get("id")
            
            # Get sessions - sort by order, null values at end
            all_sessions = await db.course_sessions.find(
                {"module_id": module_id}, {"_id": 0}
            ).to_list(100)
            
            sessions_with_order = [s for s in all_sessions if s.get("order") is not None]
            sessions_without_order = [s for s in all_sessions if s.get("order") is None]
            sessions_with_order.sort(key=lambda x: x.get("order", 0))
            sessions = sessions_with_order + sessions_without_order
            
            course_modules.append({
                **module,
                "sessions": sessions
            })
        
        result.append({
            **course,
            "modules": course_modules
        })
    
    return {"courses": result}


@router.post("/courses")
async def create_course(course_data: CourseCreate, request: Request):
    """Create a new course"""
    await verify_admin(request)
    db = get_db(request)
    
    course_id = f"course-{str(uuid.uuid4())[:8]}"
    
    course = {
        "id": course_id,
        **course_data.dict(),
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow()
    }
    
    await db.courses.insert_one(course)
    return {"message": "Course created successfully", "course_id": course_id}


@router.put("/courses/{course_id}")
async def update_course(course_id: str, course_data: CourseUpdate, request: Request):
    """Update course details"""
    await verify_admin(request)
    db = get_db(request)
    
    update_data = {k: v for k, v in course_data.dict().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No update data provided")
    
    update_data["updated_at"] = datetime.utcnow()
    
    result = await db.courses.update_one({"id": course_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Course not found")
    
    return {"message": "Course updated successfully"}


@router.delete("/courses/{course_id}")
async def delete_course(course_id: str, request: Request):
    """Delete a course and all its content (3-level structure)"""
    await verify_admin(request)
    db = get_db(request)
    
    # Get all modules in this course
    modules = await db.course_modules.find({"course_id": course_id}).to_list(100)
    module_ids = [m["id"] for m in modules]
    
    # Delete all sessions under these modules
    await db.course_sessions.delete_many({"module_id": {"$in": module_ids}})
    
    # Delete all modules
    await db.course_modules.delete_many({"course_id": course_id})
    
    # Delete course
    result = await db.courses.delete_one({"id": course_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Course not found")
    
    return {"message": "Course and all content deleted successfully"}


# Modules
@router.post("/courses/modules")
async def create_module(module_data: ModuleCreate, request: Request):
    """Create a new module"""
    await verify_admin(request)
    db = get_db(request)
    
    module_id = f"module-{str(uuid.uuid4())[:8]}"
    
    module = {
        "id": module_id,
        **module_data.dict(),
        "created_at": datetime.utcnow()
    }
    
    await db.course_modules.insert_one(module)
    return {"message": "Module created successfully", "module_id": module_id}


@router.put("/courses/modules/{module_id}")
async def update_module(module_id: str, module_data: ModuleUpdate, request: Request):
    """Update module details"""
    await verify_admin(request)
    db = get_db(request)
    
    update_data = {k: v for k, v in module_data.dict().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No update data provided")
    
    result = await db.course_modules.update_one({"id": module_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Module not found")
    
    return {"message": "Module updated successfully"}


@router.delete("/courses/modules/{module_id}")
async def delete_module(module_id: str, request: Request):
    """Delete a module and all its sessions (3-level structure)"""
    await verify_admin(request)
    db = get_db(request)
    
    # Delete sessions directly under this module
    await db.course_sessions.delete_many({"module_id": module_id})
    
    # Delete module
    result = await db.course_modules.delete_one({"id": module_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Module not found")
    
    return {"message": "Module deleted successfully"}


# Note: Submodule endpoints are deprecated - the hierarchy is now Course -> Module -> Session
# Keeping endpoints for backwards compatibility but they will be removed in future
@router.post("/courses/submodules")
async def create_submodule(submodule_data: SubmoduleCreate, request: Request):
    """Create a new submodule"""
    await verify_admin(request)
    db = get_db(request)
    
    submodule_id = f"submodule-{str(uuid.uuid4())[:8]}"
    
    submodule = {
        "id": submodule_id,
        **submodule_data.dict(),
        "created_at": datetime.utcnow()
    }
    
    await db.course_submodules.insert_one(submodule)
    return {"message": "Submodule created successfully", "submodule_id": submodule_id}


@router.put("/courses/submodules/{submodule_id}")
async def update_submodule(submodule_id: str, submodule_data: SubmoduleUpdate, request: Request):
    """Update submodule details"""
    await verify_admin(request)
    db = get_db(request)
    
    update_data = {k: v for k, v in submodule_data.dict().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No update data provided")
    
    result = await db.course_submodules.update_one({"id": submodule_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Submodule not found")
    
    return {"message": "Submodule updated successfully"}


@router.delete("/courses/submodules/{submodule_id}")
async def delete_submodule(submodule_id: str, request: Request):
    """Delete a submodule and all its sessions"""
    await verify_admin(request)
    db = get_db(request)
    
    # Delete sessions
    await db.course_sessions.delete_many({"submodule_id": submodule_id})
    
    # Delete submodule
    result = await db.course_submodules.delete_one({"id": submodule_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Submodule not found")
    
    return {"message": "Submodule deleted successfully"}


# Sessions
@router.post("/courses/sessions")
async def create_session(session_data: SessionCreate, request: Request):
    """Create a new session (video)"""
    await verify_admin(request)
    db = get_db(request)
    
    session_id = f"session-{str(uuid.uuid4())[:8]}"
    
    session = {
        "id": session_id,
        **session_data.dict(),
        "created_at": datetime.utcnow()
    }
    
    # Map content_url to video_url/pdf_url based on content_type
    if session_data.content_url:
        if session_data.content_type == 'pdf':
            session["pdf_url"] = session_data.content_url
        else:
            session["video_url"] = session_data.content_url
    
    await db.course_sessions.insert_one(session)
    return {"message": "Session created successfully", "session_id": session_id}


@router.put("/courses/sessions/{session_id}")
async def update_session(session_id: str, session_data: SessionUpdate, request: Request):
    """Update session details"""
    await verify_admin(request)
    db = get_db(request)
    
    update_data = {k: v for k, v in session_data.dict().items() if v is not None}
    
    # Map content_url to video_url/pdf_url based on content_type
    if "content_url" in update_data:
        content_type = update_data.get("content_type")
        if not content_type:
            existing = await db.course_sessions.find_one({"id": session_id})
            content_type = existing.get("content_type", "video") if existing else "video"
        if content_type == "pdf":
            update_data["pdf_url"] = update_data["content_url"]
        else:
            update_data["video_url"] = update_data["content_url"]
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No update data provided")
    
    result = await db.course_sessions.update_one({"id": session_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Session not found")
    
    return {"message": "Session updated successfully"}


@router.delete("/courses/sessions/{session_id}")
async def delete_session(session_id: str, request: Request):
    """Delete a session"""
    await verify_admin(request)
    db = get_db(request)
    
    result = await db.course_sessions.delete_one({"id": session_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Session not found")
    
    return {"message": "Session deleted successfully"}


# ============ Workshops Management ============

@router.get("/workshops")
async def get_all_workshops(request: Request):
    """Get all workshops with registration counts"""
    await verify_admin(request)
    db = get_db(request)
    
    workshops = await db.workshops.find({}, {"_id": 0}).to_list(100)
    
    # OPTIMIZATION: Get all registration counts in a single aggregation query
    workshop_ids = [w.get("id") for w in workshops if w.get("id")]
    registration_counts = {}
    if workshop_ids:
        pipeline = [
            {"$match": {"workshop_id": {"$in": workshop_ids}}},
            {"$group": {"_id": "$workshop_id", "count": {"$sum": 1}}}
        ]
        async for doc in db.workshop_registrations.aggregate(pipeline):
            registration_counts[doc["_id"]] = doc["count"]
    
    # Assign counts to workshops
    for workshop in workshops:
        workshop_id = workshop.get("id")
        workshop["registration_count"] = registration_counts.get(workshop_id, 0)
    
    return {"workshops": workshops}


@router.post("/workshops")
async def create_workshop(workshop_data: WorkshopCreate, request: Request):
    """Create a new workshop"""
    await verify_admin(request)
    db = get_db(request)
    
    workshop_id = f"workshop-{str(uuid.uuid4())[:8]}"
    
    workshop = {
        "id": workshop_id,
        **workshop_data.dict(),
        "locked": not workshop_data.is_free,
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow()
    }
    
    await db.workshops.insert_one(workshop)
    return {"message": "Workshop created successfully", "workshop_id": workshop_id}


@router.put("/workshops/{workshop_id}")
async def update_workshop(workshop_id: str, workshop_data: WorkshopUpdate, request: Request):
    """Update workshop details"""
    await verify_admin(request)
    db = get_db(request)
    
    # Debug logging for thumbnails
    logger.info(f"=== WORKSHOP UPDATE DEBUG ===")
    logger.info(f"Workshop ID: {workshop_id}")
    logger.info(f"Received thumbnail: {workshop_data.thumbnail}")
    logger.info(f"Received thumbnail_hero: {workshop_data.thumbnail_hero}")
    logger.info(f"Received thumbnail_card: {workshop_data.thumbnail_card}")
    logger.info(f"Received thumbnail_recording: {workshop_data.thumbnail_recording}")
    
    update_data = {k: v for k, v in workshop_data.dict().items() if v is not None}
    
    logger.info(f"After None filter - update_data keys: {list(update_data.keys())}")
    logger.info(f"Thumbnails in update_data: thumbnail={update_data.get('thumbnail')}, thumbnail_hero={update_data.get('thumbnail_hero')}, thumbnail_card={update_data.get('thumbnail_card')}, thumbnail_recording={update_data.get('thumbnail_recording')}")
    
    if "is_free" in update_data:
        update_data["locked"] = not update_data["is_free"]
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No update data provided")
    
    update_data["updated_at"] = datetime.utcnow()
    
    result = await db.workshops.update_one({"id": workshop_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Workshop not found")
    
    logger.info(f"Update result: matched={result.matched_count}, modified={result.modified_count}")
    
    return {"message": "Workshop updated successfully"}


@router.delete("/workshops/{workshop_id}")
async def delete_workshop(workshop_id: str, request: Request):
    """Delete a workshop"""
    await verify_admin(request)
    db = get_db(request)
    
    result = await db.workshops.delete_one({"id": workshop_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Workshop not found")
    
    return {"message": "Workshop deleted successfully"}


# ─── WORKSHOP WHATSAPP BROADCAST ENDPOINTS ────────────────────────────

@router.post("/workshops/{workshop_id}/whatsapp-broadcast")
async def send_workshop_broadcast(workshop_id: str, request: Request):
    """Send WhatsApp broadcast about a workshop to ALL users with phone numbers"""
    await verify_admin(request)
    db = get_db(request)
    
    # Get workshop
    workshop = await db.workshops.find_one({"id": workshop_id}, {"_id": 0})
    if not workshop:
        raise HTTPException(status_code=404, detail="Workshop not found")
    
    # Get all users with phone numbers
    users = await db.users.find(
        {"phone_number": {"$exists": True, "$nin": ["", None]}},
        {"_id": 0, "id": 1, "name": 1, "first_name": 1, "phone_number": 1}
    ).to_list(10000)
    
    if not users:
        return {"message": "No users with phone numbers found", "sent": 0, "failed": 0}
    
    from services.wati_service import wati_service
    
    sent = 0
    failed = 0
    errors = []
    
    for user in users:
        phone = user.get("phone_number", "")
        if not phone:
            continue
        
        name = user.get("first_name") or user.get("name", "").split()[0] if user.get("name") else "there"
        
        parameters = [
            {"name": "1", "value": name},
            {"name": "2", "value": workshop.get("title", "")},
            {"name": "3", "value": workshop.get("instructor", workshop.get("host", ""))},
            {"name": "4", "value": workshop.get("date", "")},
            {"name": "5", "value": workshop.get("time", "")},
            {"name": "6", "value": workshop.get("duration", "")}
        ]
        
        try:
            await wati_service.send_template_message(
                recipient_number=phone,
                template_name="workshop_announcement",
                parameters=parameters
            )
            sent += 1
        except Exception as e:
            failed += 1
            errors.append(f"{phone}: {str(e)[:100]}")
            logger.error(f"WhatsApp broadcast failed for {phone}: {e}")
    
    logger.info(f"Workshop broadcast for '{workshop.get('title')}': sent={sent}, failed={failed}")
    
    return {
        "message": f"Broadcast sent to {sent} users, {failed} failed",
        "sent": sent,
        "failed": failed,
        "total_users": len(users),
        "errors": errors[:10]  # Return first 10 errors for debugging
    }


@router.post("/workshops/{workshop_id}/whatsapp-register-reminder")
async def send_workshop_register_reminder(workshop_id: str, request: Request):
    """Send WhatsApp reminder to register to users who haven't registered for the workshop"""
    await verify_admin(request)
    db = get_db(request)
    
    # Get workshop
    workshop = await db.workshops.find_one({"id": workshop_id}, {"_id": 0})
    if not workshop:
        raise HTTPException(status_code=404, detail="Workshop not found")
    
    # Get registered user IDs for this workshop
    registrations = await db.workshop_registrations.find(
        {"workshop_id": workshop_id},
        {"_id": 0, "user_id": 1}
    ).to_list(10000)
    registered_user_ids = {r.get("user_id") for r in registrations}
    
    # Get all users with phone numbers who are NOT registered
    users = await db.users.find(
        {"phone_number": {"$exists": True, "$nin": ["", None]}},
        {"_id": 0, "id": 1, "name": 1, "first_name": 1, "phone_number": 1}
    ).to_list(10000)
    
    unregistered_users = [u for u in users if u.get("id") not in registered_user_ids]
    
    if not unregistered_users:
        return {"message": "All users with phone numbers are already registered!", "sent": 0, "failed": 0}
    
    from services.wati_service import wati_service
    
    sent = 0
    failed = 0
    errors = []
    
    for user in unregistered_users:
        phone = user.get("phone_number", "")
        if not phone:
            continue
        
        name = user.get("first_name") or user.get("name", "").split()[0] if user.get("name") else "there"
        
        parameters = [
            {"name": "1", "value": name},
            {"name": "2", "value": workshop.get("title", "")},
            {"name": "3", "value": workshop.get("instructor", workshop.get("host", ""))},
            {"name": "4", "value": workshop.get("date", "")},
            {"name": "5", "value": workshop.get("time", "")}
        ]
        
        try:
            await wati_service.send_template_message(
                recipient_number=phone,
                template_name="workshop_register_reminder",
                parameters=parameters
            )
            sent += 1
        except Exception as e:
            failed += 1
            errors.append(f"{phone}: {str(e)[:100]}")
            logger.error(f"WhatsApp register reminder failed for {phone}: {e}")
    
    logger.info(f"Workshop register reminder for '{workshop.get('title')}': sent={sent}, failed={failed}")
    
    return {
        "message": f"Reminder sent to {sent} unregistered users, {failed} failed",
        "sent": sent,
        "failed": failed,
        "total_unregistered": len(unregistered_users),
        "errors": errors[:10]
    }


@router.post("/workshops/{workshop_id}/whatsapp-post-workshop")
async def send_post_workshop_message(workshop_id: str, request: Request):
    """Send WhatsApp follow-up message to all registered participants after workshop completion"""
    await verify_admin(request)
    db = get_db(request)
    
    # Get workshop
    workshop = await db.workshops.find_one({"id": workshop_id}, {"_id": 0})
    if not workshop:
        raise HTTPException(status_code=404, detail="Workshop not found")
    
    # Get all registered participants for this workshop
    registrations = await db.workshop_registrations.find(
        {"workshop_id": workshop_id},
        {"_id": 0, "user_id": 1}
    ).to_list(10000)
    
    if not registrations:
        return {"message": "No registered participants found", "sent": 0, "failed": 0}
    
    registered_user_ids = [r.get("user_id") for r in registrations]
    
    # Get user details with phone numbers
    users = await db.users.find(
        {
            "id": {"$in": registered_user_ids},
            "phone_number": {"$exists": True, "$nin": ["", None]}
        },
        {"_id": 0, "id": 1, "name": 1, "first_name": 1, "phone_number": 1}
    ).to_list(10000)
    
    if not users:
        return {"message": "No registered participants with phone numbers found", "sent": 0, "failed": 0}
    
    from services.wati_service import wati_service
    
    sent = 0
    failed = 0
    errors = []
    
    for user in users:
        phone = user.get("phone_number", "")
        if not phone:
            continue
        
        # Only update workshop_name attribute (no message sending)
        try:
            await wati_service.update_contact_attribute(
                recipient_number=phone,
                attribute_name="workshop_name",
                attribute_value=workshop.get("title", "")
            )
            sent += 1
            logger.info(f"Updated workshop_name attribute for {phone} to '{workshop.get('title')}'")
        except Exception as e:
            failed += 1
            errors.append(f"{phone}: {str(e)[:100]}")
            logger.error(f"Failed to update attribute for {phone}: {e}")
    
    logger.info(f"Post-workshop follow-up for '{workshop.get('title')}': sent={sent}, failed={failed} to {len(users)} registered participants")
    
    return {
        "message": f"Post-workshop message sent to {sent} participants, {failed} failed",
        "sent": sent,
        "failed": failed,
        "total_registered": len(registrations),
        "total_with_phone": len(users),
        "errors": errors[:10]
    }


# ─── PUBLIC WORKSHOPS ENDPOINT (No Auth Required) ────────────────────────────

@router.get("/workshops/public")
async def get_public_workshops(request: Request):
    """Get all published workshops for public landing page - no auth required"""
    db = get_db(request)
    
    now = datetime.utcnow()
    
    # Get all workshops
    all_workshops = await db.workshops.find({}, {"_id": 0}).to_list(100)
    
    # OPTIMIZATION: Get all registration counts in a single aggregation query
    workshop_ids = [w.get("id") for w in all_workshops if w.get("id")]
    registration_counts = {}
    if workshop_ids:
        pipeline = [
            {"$match": {"workshop_id": {"$in": workshop_ids}}},
            {"$group": {"_id": "$workshop_id", "count": {"$sum": 1}}}
        ]
        async for doc in db.workshop_registrations.aggregate(pipeline):
            registration_counts[doc["_id"]] = doc["count"]
    
    upcoming = []
    past = []
    
    for workshop in all_workshops:
        # Get registration count from pre-fetched data
        workshop_id = workshop.get("id")
        workshop["registration_count"] = registration_counts.get(workshop_id, 0)
        
        # Determine if workshop is upcoming or past based on date and status
        is_past = workshop.get("is_past", False)
        status = workshop.get("status", "upcoming")
        
        # Try to parse the date for more accurate categorization
        try:
            workshop_date_str = workshop.get("date", "")
            # Parse date like "March 20, 2026" or "2026-03-20"
            if workshop_date_str:
                from dateutil import parser
                workshop_date = parser.parse(workshop_date_str)
                if workshop_date.date() < now.date():
                    is_past = True
        except:
            pass
        
        if is_past or status == "completed":
            past.append(workshop)
        else:
            upcoming.append(workshop)
    
    # Sort upcoming by date (soonest first), past by date (most recent first)
    def parse_date_for_sort(w):
        try:
            from dateutil import parser
            return parser.parse(w.get("date", "2099-12-31"))
        except:
            return datetime(2099, 12, 31)
    
    upcoming.sort(key=parse_date_for_sort)
    past.sort(key=parse_date_for_sort, reverse=True)
    
    return {
        "upcoming": upcoming,
        "past": past
    }


@router.get("/workshops/{workshop_id}/registrations")
async def get_workshop_registrations(workshop_id: str, request: Request):
    """Get all registrations for a workshop with user details"""
    await verify_admin(request)
    db = get_db(request)
    
    # Verify workshop exists
    workshop = await db.workshops.find_one({"id": workshop_id}, {"_id": 0, "title": 1})
    if not workshop:
        raise HTTPException(status_code=404, detail="Workshop not found")
    
    registrations = await db.workshop_registrations.find(
        {"workshop_id": workshop_id},
        {"_id": 0}
    ).sort("registered_at", -1).to_list(500)
    
    # Enrich registrations with user details (phone number and current plan)
    enriched_registrations = []
    for reg in registrations:
        user_id = reg.get("user_id")
        if user_id:
            user = await db.users.find_one({"id": user_id}, {"_id": 0, "phone_number": 1, "plan": 1, "plan_name": 1})
            if user:
                reg["user_phone"] = user.get("phone_number", "")
                reg["current_plan"] = user.get("plan_name") or user.get("plan") or "Free"
            else:
                reg["user_phone"] = ""
                reg["current_plan"] = "Unknown"
        else:
            reg["user_phone"] = ""
            reg["current_plan"] = "Unknown"
        enriched_registrations.append(reg)
    
    return {
        "workshop_id": workshop_id,
        "workshop_title": workshop.get("title"),
        "registrations": enriched_registrations,
        "total": len(enriched_registrations)
    }


@router.delete("/workshops/{workshop_id}/registrations/{registration_id}")
async def admin_remove_registration(workshop_id: str, registration_id: str, request: Request):
    """Admin remove a user's registration from a workshop"""
    await verify_admin(request)
    db = get_db(request)
    
    registration = await db.workshop_registrations.find_one_and_delete({
        "workshop_id": workshop_id,
        "id": registration_id
    })
    
    if not registration:
        raise HTTPException(status_code=404, detail="Registration not found")
    
    # Try to delete calendar event
    if registration.get("calendar_event_id"):
        try:
            from services.calendar_service import GoogleCalendarService
            calendar_service = GoogleCalendarService()
            if calendar_service.is_available():
                calendar_service.delete_event(registration["calendar_event_id"])
        except Exception as e:
            import logging
            logging.error(f"Failed to delete calendar event: {e}")
    
    # Notify user about removal (optional)
    try:
        from services.email_service import send_email
        
        workshop = await db.workshops.find_one({"id": workshop_id}, {"_id": 0, "title": 1})
        
        email_html = f"""
        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <h2 style="color: #1e293b;">Workshop Registration Cancelled</h2>
            <p>Hi {registration.get("user_name", "")},</p>
            <p>Your registration for the workshop <strong>"{workshop.get("title", "")}"</strong> has been cancelled by the admin.</p>
            <p>If you have any questions, please contact us at hi@gradnext.co</p>
            <p>Team gradnext</p>
        </div>
        """
        
        await send_email(
            to_email=registration.get("user_email"),
            subject=f"Workshop Registration Cancelled: {workshop.get('title', '')}",
            html_content=email_html,
            from_email="hi@mail.gradnext.co"
        )
    except Exception as e:
        import logging
        logging.error(f"Failed to send cancellation email: {e}")
    
    return {"success": True, "message": "Registration removed successfully"}


@router.post("/workshops/{workshop_id}/send-updated-invites")
async def send_updated_workshop_invites(workshop_id: str, request: Request):
    """
    Send updated calendar invites with the new meeting link to all registered users.
    This is useful when the admin updates the meeting link for a workshop.
    """
    await verify_admin(request)
    db = get_db(request)
    
    # Get the workshop
    workshop = await db.workshops.find_one({"id": workshop_id}, {"_id": 0})
    if not workshop:
        raise HTTPException(status_code=404, detail="Workshop not found")
    
    meeting_link = workshop.get("meeting_link")
    if not meeting_link:
        raise HTTPException(status_code=400, detail="Workshop has no meeting link set. Please add a meeting link first.")
    
    # Get all registrations for this workshop
    registrations = await db.workshop_registrations.find({"workshop_id": workshop_id}).to_list(500)
    
    if not registrations:
        return {"success": True, "sent": 0, "failed": 0, "message": "No registrations found"}
    
    sent_count = 0
    failed_count = 0
    errors = []
    
    # Import required services
    try:
        from services.calendar_service import GoogleCalendarService
        from services.email_service import send_email
    except ImportError as e:
        raise HTTPException(status_code=500, detail=f"Required service not available: {str(e)}")
    
    calendar_service = GoogleCalendarService()
    
    for registration in registrations:
        user_email = registration.get("user_email")
        user_name = registration.get("user_name", "")
        old_calendar_event_id = registration.get("calendar_event_id")
        
        try:
            # Delete old calendar event if exists
            if old_calendar_event_id and calendar_service.is_available():
                try:
                    calendar_service.delete_event(old_calendar_event_id)
                except Exception as e:
                    import logging
                    logging.warning(f"Could not delete old calendar event {old_calendar_event_id}: {e}")
            
            # Create new calendar event with updated meeting link
            new_calendar_event_id = None
            if calendar_service.is_available():
                workshop_date = workshop.get("date")
                workshop_time = workshop.get("time", "10:00")
                duration_str = workshop.get("duration", "1 hour")
                
                # Parse duration
                duration_minutes = 60
                if "hour" in duration_str.lower():
                    try:
                        hours = float(duration_str.lower().replace("hours", "").replace("hour", "").strip())
                        duration_minutes = int(hours * 60)
                    except:
                        pass
                
                event_result = calendar_service.create_workshop_event_with_link(
                    workshop_title=workshop.get("title"),
                    workshop_description=workshop.get("description", ""),
                    instructor_name=workshop.get("instructor", ""),
                    workshop_date=workshop_date,
                    workshop_time=workshop_time,
                    duration_minutes=duration_minutes,
                    attendee_email=user_email,
                    attendee_name=user_name,
                    meeting_link=meeting_link
                )
                
                if event_result:
                    new_calendar_event_id = event_result.get("event_id")
            
            # Update registration with new calendar event ID
            if new_calendar_event_id:
                await db.workshop_registrations.update_one(
                    {"id": registration.get("id")},
                    {"$set": {
                        "calendar_event_id": new_calendar_event_id,
                        "meet_link": meeting_link,
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }}
                )
            
            # Also send an email notification about the updated link
            email_html = f"""
            <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
                <h2 style="color: #1e293b;">Workshop Meeting Link Updated</h2>
                <p>Hi {user_name},</p>
                <p>The meeting link for the workshop <strong>"{workshop.get("title", "")}"</strong> has been updated.</p>
                
                <div style="background-color: #f1f5f9; padding: 15px; border-radius: 8px; margin: 20px 0;">
                    <p style="margin: 0 0 10px 0;"><strong>Workshop:</strong> {workshop.get("title", "")}</p>
                    <p style="margin: 0 0 10px 0;"><strong>Date:</strong> {workshop.get("date", "")}</p>
                    <p style="margin: 0 0 10px 0;"><strong>Time:</strong> {workshop.get("time", "")} IST</p>
                    <p style="margin: 0;"><strong>New Meeting Link:</strong><br>
                    <a href="{meeting_link}" style="color: #2563eb; word-break: break-all;">{meeting_link}</a></p>
                </div>
                
                <p>📅 An updated calendar invite has also been sent to your email.</p>
                <p>Please use the new link to join the workshop.</p>
                
                <p style="margin-top: 30px;">Team gradnext</p>
            </div>
            """
            
            await send_email(
                to_email=user_email,
                subject=f"Updated Meeting Link: {workshop.get('title', '')}",
                html_content=email_html,
                from_email="hi@mail.gradnext.co"
            )
            
            sent_count += 1
            
        except Exception as e:
            import logging
            logging.error(f"Failed to send updated invite to {user_email}: {str(e)}")
            failed_count += 1
            errors.append(f"{user_email}: {str(e)}")
    
    return {
        "success": True,
        "sent": sent_count,
        "failed": failed_count,
        "total": len(registrations),
        "errors": errors[:10] if errors else [],  # Return first 10 errors max
        "message": f"Sent updated invites to {sent_count}/{len(registrations)} users"
    }


# ============ Drills Management ============

@router.get("/drills")
async def get_all_drills(request: Request):
    """Get all drills"""
    await verify_admin(request)
    db = get_db(request)
    
    drills = await db.drills.find({}, {"_id": 0}).to_list(200)
    return {"drills": drills}


@router.post("/drills")
async def create_drill(drill_data: DrillCreate, request: Request):
    """Create a new drill"""
    await verify_admin(request)
    db = get_db(request)
    
    drill_id = f"drill-{str(uuid.uuid4())[:8]}"
    
    drill = {
        "id": drill_id,
        **drill_data.dict(),
        "locked": not drill_data.is_free,
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow()
    }
    
    await db.drills.insert_one(drill)
    return {"message": "Drill created successfully", "drill_id": drill_id}


@router.put("/drills/{drill_id}")
async def update_drill(drill_id: str, drill_data: DrillUpdate, request: Request):
    """Update drill details"""
    await verify_admin(request)
    db = get_db(request)
    
    update_data = {k: v for k, v in drill_data.dict().items() if v is not None}
    if "is_free" in update_data:
        update_data["locked"] = not update_data["is_free"]
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No update data provided")
    
    update_data["updated_at"] = datetime.utcnow()
    
    result = await db.drills.update_one({"id": drill_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Drill not found")
    
    return {"message": "Drill updated successfully"}


@router.delete("/drills/{drill_id}")
async def delete_drill(drill_id: str, request: Request):
    """Delete a drill"""
    await verify_admin(request)
    db = get_db(request)
    
    result = await db.drills.delete_one({"id": drill_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Drill not found")
    
    return {"message": "Drill deleted successfully"}


@router.get("/drills/tags")
async def get_all_tags(request: Request):
    """Get all unique drill tags"""
    await verify_admin(request)
    db = get_db(request)
    
    drills = await db.drills.find({}, {"tags": 1}).to_list(500)
    all_tags = set()
    for drill in drills:
        all_tags.update(drill.get("tags", []))
    
    return {"tags": list(all_tags)}


@router.get("/drills/categories")
async def get_all_categories(request: Request):
    """Get all unique drill categories"""
    await verify_admin(request)
    db = get_db(request)
    
    categories = await db.drills.distinct("category")
    return {"categories": categories}


@router.get("/drills/difficulties")
async def get_all_difficulties(request: Request):
    """Get all difficulty levels"""
    return {"difficulties": ["beginner", "intermediate", "advanced"]}


# ============ Materials Management ============

@router.get("/materials")
async def get_all_materials(request: Request):
    """Get all materials"""
    await verify_admin(request)
    db = get_db(request)
    
    materials = await db.materials.find({}, {"_id": 0}).to_list(200)
    return {"materials": materials}


@router.post("/materials")
async def create_material(material_data: MaterialCreate, request: Request):
    """Create a new material"""
    await verify_admin(request)
    db = get_db(request)
    
    material_id = f"material-{str(uuid.uuid4())[:8]}"
    
    material = {
        "id": material_id,
        **material_data.dict(),
        "locked": not material_data.is_free,
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow()
    }
    
    await db.materials.insert_one(material)
    return {"message": "Material created successfully", "material_id": material_id}


@router.put("/materials/{material_id}")
async def update_material(material_id: str, material_data: MaterialUpdate, request: Request):
    """Update material details"""
    await verify_admin(request)
    db = get_db(request)
    
    update_data = {k: v for k, v in material_data.dict().items() if v is not None}
    if "is_free" in update_data:
        update_data["locked"] = not update_data["is_free"]
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No update data provided")
    
    update_data["updated_at"] = datetime.utcnow()
    
    result = await db.materials.update_one({"id": material_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Material not found")
    
    return {"message": "Material updated successfully"}


@router.delete("/materials/{material_id}")
async def delete_material(material_id: str, request: Request):
    """Delete a material"""
    await verify_admin(request)
    db = get_db(request)
    
    result = await db.materials.delete_one({"id": material_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Material not found")
    
    return {"message": "Material deleted successfully"}


# ============ Peer Practice Management ============

@router.get("/peer-practice/users")
async def get_peer_practice_users(request: Request):
    """Get all users with peer practice data - excludes mentors and admins"""
    await verify_admin(request)
    db = get_db(request)
    
    # Exclude mentors and admins from peer practice - they should not appear here
    users = await db.users.find(
        {
            "$or": [
                {"is_mentor": {"$ne": True}},
                {"is_mentor": {"$exists": False}}
            ],
            "$and": [
                {
                    "$or": [
                        {"is_admin": {"$ne": True}},
                        {"is_admin": {"$exists": False}}
                    ]
                }
            ]
        },
        {"_id": 0, "id": 1, "name": 1, "email": 1, "picture": 1, "peer_availability": 1, "peer_rating": 1, "peer_sessions_done": 1, "peer_practice_status": 1}
    ).to_list(500)
    
    return {"users": users}


@router.put("/peer-practice/users/{user_id}/availability")
async def update_peer_availability(user_id: str, request: Request):
    """Update a user's peer practice availability"""
    await verify_admin(request)
    db = get_db(request)
    
    body = await request.json()
    availability = body.get("availability", [])
    
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {"peer_availability": availability, "updated_at": datetime.utcnow()}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "Peer availability updated successfully"}


@router.put("/peer-practice/users/{user_id}/status")
async def update_peer_practice_status(user_id: str, request: Request):
    """Pause, resume, or remove user from peer practice"""
    await verify_admin(request)
    db = get_db(request)
    
    body = await request.json()
    status = body.get("status")  # 'active', 'paused', 'removed'
    
    if status not in ['active', 'paused', 'removed']:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    update_data = {"peer_practice_status": status, "updated_at": datetime.utcnow()}
    
    if status == 'removed':
        update_data["peer_availability"] = []
    
    result = await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": f"User peer practice status updated to {status}"}


# ============ Peer Sessions Tracking ============

@router.get("/peer-sessions")
async def get_all_peer_sessions(
    request: Request,
    page: int = 1,
    limit: int = 20,
    status: str = None,
    date_from: str = None,
    date_to: str = None,
    user_id: str = None,
    search: str = None
):
    """Get all peer practice sessions with detailed tracking info - excludes mentor sessions"""
    await verify_admin(request)
    db = get_db(request)
    
    # Get list of mentor user IDs to exclude their sessions
    mentor_ids = set()
    mentors_from_users = await db.users.find(
        {"$or": [{"is_mentor": True}, {"role": "mentor"}]},
        {"_id": 0, "id": 1}
    ).to_list(500)
    for m in mentors_from_users:
        if m.get("id"):
            mentor_ids.add(m["id"])
    
    # Also check the mentors collection
    mentors_collection = await db.mentors.find({}, {"_id": 0, "id": 1, "user_id": 1}).to_list(500)
    for m in mentors_collection:
        if m.get("id"):
            mentor_ids.add(m["id"])
        if m.get("user_id"):
            mentor_ids.add(m["user_id"])
    
    # Build query using $and for all conditions
    conditions = []
    
    # Exclude mentor sessions
    if mentor_ids:
        conditions.append({"requester_id": {"$nin": list(mentor_ids)}})
        conditions.append({"partner_id": {"$nin": list(mentor_ids)}})
    
    # Status filter
    if status and status != 'all':
        conditions.append({"status": status})
    
    # Date range filter
    if date_from and date_to:
        conditions.append({"date": {"$gte": date_from, "$lte": date_to}})
    elif date_from:
        conditions.append({"date": {"$gte": date_from}})
    elif date_to:
        conditions.append({"date": {"$lte": date_to}})
    
    # User ID filter
    if user_id:
        conditions.append({"$or": [{"requester_id": user_id}, {"partner_id": user_id}]})
    
    # Search by name or email
    if search and search.strip():
        search_term = search.strip()
        conditions.append({
            "$or": [
                {"requester_name": {"$regex": search_term, "$options": "i"}},
                {"requester_email": {"$regex": search_term, "$options": "i"}},
                {"partner_name": {"$regex": search_term, "$options": "i"}},
                {"partner_email": {"$regex": search_term, "$options": "i"}}
            ]
        })
    
    # Build final query
    query = {"$and": conditions} if conditions else {}
    
    # Debug log the query
    import json
    print(f"[PEER-SESSIONS] Search term: {search}, Query conditions count: {len(conditions)}")
    
    # Get total count for pagination
    total = await db.peer_sessions.count_documents(query)
    print(f"[PEER-SESSIONS] Total matching: {total}")
    
    # Get sessions with pagination, sorted by date descending (most recent first)
    skip = (page - 1) * limit
    sessions = await db.peer_sessions.find(query).sort([("date", -1), ("created_at", -1)]).skip(skip).limit(limit).to_list(limit)
    
    # Process sessions to add computed fields
    processed_sessions = []
    for session in sessions:
        # Determine feedback status
        requester_feedback = session.get("requester_feedback")
        partner_feedback = session.get("partner_feedback")
        
        # Get rating from feedback if exists
        requester_rating = None
        partner_rating = None
        
        if requester_feedback:
            requester_rating = requester_feedback.get("average_rating") or requester_feedback.get("rating_overall")
        
        if partner_feedback:
            partner_rating = partner_feedback.get("average_rating") or partner_feedback.get("rating_overall")
        
        processed_session = {
            "id": session.get("id") or str(session.get("_id")),
            "date": session.get("date"),
            "time_slot": session.get("time_slot"),
            "status": session.get("status"),
            "session_type": session.get("session_type"),
            "case_type": session.get("case_type"),
            "duration_minutes": session.get("duration_minutes", 90),
            # Requester info
            "requester_id": session.get("requester_id"),
            "requester_name": session.get("requester_name"),
            "requester_email": session.get("requester_email"),
            "requester_picture": session.get("requester_picture"),
            "requester_feedback_given": requester_feedback is not None,
            "requester_rating": requester_rating,
            "requester_checked_in": session.get("requester_checked_in", False),
            "requester_checked_in_at": session.get("requester_checked_in_at"),
            # Partner info
            "partner_id": session.get("partner_id"),
            "partner_name": session.get("partner_name"),
            "partner_email": session.get("partner_email"),
            "partner_picture": session.get("partner_picture"),
            "partner_feedback_given": partner_feedback is not None,
            "partner_rating": partner_rating,
            "partner_checked_in": session.get("partner_checked_in", False),
            "partner_checked_in_at": session.get("partner_checked_in_at"),
            # Timestamps
            "created_at": session.get("created_at").isoformat() if session.get("created_at") else None,
            "meet_link": session.get("meet_link"),
            # Reschedule info
            "reschedule_requested": session.get("reschedule_requested", False),
            "previous_date": session.get("previous_date"),
            "previous_time_slot": session.get("previous_time_slot"),
        }
        processed_sessions.append(processed_session)
    
    return {
        "sessions": processed_sessions,
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": (total + limit - 1) // limit
    }


@router.delete("/peer-sessions/cleanup-mentor-sessions")
async def cleanup_mentor_sessions(request: Request):
    """Delete all peer sessions involving mentors - one-time cleanup"""
    await verify_admin(request)
    db = get_db(request)
    
    # Get all mentor IDs from users and mentors collections
    mentor_ids = set()
    
    mentors_from_users = await db.users.find(
        {"$or": [{"is_mentor": True}, {"role": "mentor"}]},
        {"_id": 0, "id": 1}
    ).to_list(500)
    for m in mentors_from_users:
        if m.get("id"):
            mentor_ids.add(m["id"])
    
    mentors_collection = await db.mentors.find({}, {"_id": 0, "id": 1, "user_id": 1}).to_list(500)
    for m in mentors_collection:
        if m.get("id"):
            mentor_ids.add(m["id"])
        if m.get("user_id"):
            mentor_ids.add(m["user_id"])
    
    if not mentor_ids:
        return {"message": "No mentors found", "deleted_count": 0}
    
    # Delete sessions involving mentors
    result = await db.peer_sessions.delete_many({
        "$or": [
            {"requester_id": {"$in": list(mentor_ids)}},
            {"partner_id": {"$in": list(mentor_ids)}}
        ]
    })
    
    return {
        "message": f"Deleted {result.deleted_count} peer sessions involving mentors",
        "deleted_count": result.deleted_count,
        "mentor_ids_checked": list(mentor_ids)
    }


@router.get("/peer-sessions/stats")
async def get_peer_sessions_stats(request: Request):
    """Get statistics for peer practice sessions - excludes mentor sessions"""
    await verify_admin(request)
    db = get_db(request)
    
    # Get mentor IDs to exclude
    mentor_ids = set()
    mentors_from_users = await db.users.find(
        {"$or": [{"is_mentor": True}, {"role": "mentor"}]},
        {"_id": 0, "id": 1}
    ).to_list(500)
    for m in mentors_from_users:
        if m.get("id"):
            mentor_ids.add(m["id"])
    
    mentors_collection = await db.mentors.find({}, {"_id": 0, "id": 1, "user_id": 1}).to_list(500)
    for m in mentors_collection:
        if m.get("id"):
            mentor_ids.add(m["id"])
        if m.get("user_id"):
            mentor_ids.add(m["user_id"])
    
    # Base filter to exclude mentor sessions
    base_filter = {
        "requester_id": {"$nin": list(mentor_ids)},
        "partner_id": {"$nin": list(mentor_ids)}
    }
    
    # Get total counts by status (excluding mentor sessions)
    total = await db.peer_sessions.count_documents(base_filter)
    pending = await db.peer_sessions.count_documents({**base_filter, "status": "pending"})
    confirmed = await db.peer_sessions.count_documents({**base_filter, "status": "confirmed"})
    completed = await db.peer_sessions.count_documents({**base_filter, "status": "completed"})
    cancelled = await db.peer_sessions.count_documents({**base_filter, "status": "cancelled"})
    declined = await db.peer_sessions.count_documents({**base_filter, "status": "declined"})
    reschedule_pending = await db.peer_sessions.count_documents({**base_filter, "status": "reschedule_pending"})
    
    # Get today's sessions
    today = datetime.utcnow().strftime("%Y-%m-%d")
    sessions_today = await db.peer_sessions.count_documents({**base_filter, "date": today})
    
    # Get sessions this week
    from datetime import timedelta
    week_start = (datetime.utcnow() - timedelta(days=datetime.utcnow().weekday())).strftime("%Y-%m-%d")
    sessions_this_week = await db.peer_sessions.count_documents({**base_filter, "date": {"$gte": week_start}})
    
    # Count sessions with feedback
    with_both_feedback = await db.peer_sessions.count_documents({
        **base_filter,
        "requester_feedback": {"$ne": None},
        "partner_feedback": {"$ne": None}
    })
    
    with_requester_feedback = await db.peer_sessions.count_documents({
        **base_filter,
        "requester_feedback": {"$ne": None}
    })
    
    with_partner_feedback = await db.peer_sessions.count_documents({
        **base_filter,
        "partner_feedback": {"$ne": None}
    })
    
    return {
        "total": total,
        "by_status": {
            "pending": pending,
            "confirmed": confirmed,
            "completed": completed,
            "cancelled": cancelled,
            "declined": declined,
            "reschedule_pending": reschedule_pending
        },
        "sessions_today": sessions_today,
        "sessions_this_week": sessions_this_week,
        "feedback_stats": {
            "both_feedback": with_both_feedback,
            "requester_feedback_only": with_requester_feedback - with_both_feedback,
            "partner_feedback_only": with_partner_feedback - with_both_feedback,
            "no_feedback": max(0, completed - with_both_feedback - (with_requester_feedback - with_both_feedback) - (with_partner_feedback - with_both_feedback))
        }
    }


@router.get("/peer-sessions/{session_id}")
async def get_peer_session_details(session_id: str, request: Request):
    """Get detailed information about a specific peer session"""
    await verify_admin(request)
    db = get_db(request)
    
    from bson import ObjectId
    
    # Find session
    session = await db.peer_sessions.find_one({
        "$or": [
            {"id": session_id},
            {"_id": ObjectId(session_id) if len(session_id) == 24 else None}
        ]
    })
    
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Get requester user details
    requester = await db.users.find_one({"id": session.get("requester_id")}, {"_id": 0, "id": 1, "name": 1, "email": 1, "picture": 1, "plan": 1})
    
    # Get partner user details
    partner = await db.users.find_one({"id": session.get("partner_id")}, {"_id": 0, "id": 1, "name": 1, "email": 1, "picture": 1, "plan": 1})
    
    return {
        "session": {
            "id": session.get("id") or str(session.get("_id")),
            "date": session.get("date"),
            "time_slot": session.get("time_slot"),
            "status": session.get("status"),
            "session_type": session.get("session_type"),
            "case_type": session.get("case_type"),
            "duration_minutes": session.get("duration_minutes", 90),
            "meet_link": session.get("meet_link"),
            "requester_notes": session.get("requester_notes"),
            "created_at": session.get("created_at").isoformat() if session.get("created_at") else None,
            "reschedule_requested": session.get("reschedule_requested", False),
            "reschedule_requester": session.get("reschedule_requester"),
            "reschedule_reason": session.get("reschedule_reason"),
            "reschedule_new_date": session.get("reschedule_new_date"),
            "reschedule_new_time": session.get("reschedule_new_time"),
            "previous_date": session.get("previous_date"),
            "previous_time_slot": session.get("previous_time_slot"),
        },
        "requester": requester,
        "partner": partner,
        "requester_feedback": session.get("requester_feedback"),
        "partner_feedback": session.get("partner_feedback")
    }


@router.post("/peer-sessions/{session_id}/update-status")
async def admin_update_peer_session_status(session_id: str, request: Request):
    """Admin can update peer session status - includes admin-specific statuses"""
    await verify_admin(request)
    db = get_db(request)
    
    from bson import ObjectId
    
    body = await request.json()
    new_status = body.get("status")
    notes = body.get("notes", "")
    
    # Extended status list to include admin-specific statuses
    allowed_statuses = [
        "pending", "confirmed", "completed", "cancelled", "declined",
        "admin_cancelled", "admin_rescheduled", "reschedule_pending"
    ]
    
    if new_status not in allowed_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Allowed: {', '.join(allowed_statuses)}")
    
    # Get admin user info
    admin_user = await get_current_user(request)
    admin_email = admin_user.get("email") if isinstance(admin_user, dict) else admin_user.email
    admin_name = admin_user.get("name", admin_email) if isinstance(admin_user, dict) else getattr(admin_user, "name", admin_email)
    
    # Build admin action log entry
    action_log_entry = {
        "timestamp": datetime.utcnow(),
        "admin_email": admin_email,
        "admin_name": admin_name,
        "action": f"Status changed to {new_status}",
        "notes": notes
    }
    
    # Find and update session
    result = await db.peer_sessions.update_one(
        {"$or": [
            {"id": session_id},
            {"_id": ObjectId(session_id) if len(session_id) == 24 else None}
        ]},
        {
            "$set": {
                "status": new_status,
                "admin_notes": notes,
                "admin_updated_at": datetime.utcnow(),
                "last_admin_email": admin_email
            },
            "$push": {
                "admin_action_history": action_log_entry
            }
        }
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Session not found")
    
    return {
        "success": True,
        "message": f"Session status updated to {new_status}",
        "new_status": new_status
    }


@router.put("/peer-sessions/{session_id}/participants")
@router.post("/peer-sessions/{session_id}/participants")
async def admin_manage_peer_session_participants(session_id: str, request: Request):
    """
    Admin endpoint to manage peer session participants.
    Can remove participants or swap them.
    
    Body:
    {
        "action": "remove_requester" | "remove_partner" | "swap_requester" | "swap_partner",
        "new_user_id": "user123" (required for swap actions),
        "notes": "Optional reason for change"
    }
    """
    await verify_admin(request)
    db = get_db(request)
    
    from bson import ObjectId
    
    body = await request.json()
    action = body.get("action")
    new_user_id = body.get("new_user_id")
    notes = body.get("notes", "")
    
    allowed_actions = ["remove_requester", "remove_partner", "swap_requester", "swap_partner"]
    
    if action not in allowed_actions:
        raise HTTPException(status_code=400, detail=f"Invalid action. Allowed: {', '.join(allowed_actions)}")
    
    # Get admin user info
    admin_user = await get_current_user(request)
    admin_email = admin_user.get("email") if isinstance(admin_user, dict) else admin_user.email
    admin_name = admin_user.get("name", admin_email) if isinstance(admin_user, dict) else getattr(admin_user, "name", admin_email)
    
    # Find session
    session = await db.peer_sessions.find_one({
        "$or": [
            {"id": session_id},
            {"_id": ObjectId(session_id) if len(session_id) == 24 else None}
        ]
    })
    
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    update_data = {}
    action_description = ""
    
    if action == "remove_requester":
        # Mark requester as removed
        update_data["requester_removed_by_admin"] = True
        update_data["requester_removed_at"] = datetime.utcnow()
        update_data["status"] = "admin_cancelled"
        action_description = f"Requester ({session.get('requester_name')}) removed from session"
        
    elif action == "remove_partner":
        # Mark partner as removed
        update_data["partner_removed_by_admin"] = True
        update_data["partner_removed_at"] = datetime.utcnow()
        update_data["status"] = "admin_cancelled"
        action_description = f"Partner ({session.get('partner_name')}) removed from session"
        
    elif action == "swap_requester":
        if not new_user_id:
            raise HTTPException(status_code=400, detail="new_user_id required for swap action")
        
        # Get new user details
        new_user = await db.users.find_one({"id": new_user_id}, {"_id": 0})
        if not new_user:
            raise HTTPException(status_code=404, detail="New user not found")
        
        old_requester_name = session.get("requester_name")
        
        update_data["requester_id"] = new_user_id
        update_data["requester_name"] = new_user.get("name", "Unknown")
        update_data["requester_email"] = new_user.get("email", "")
        update_data["requester_picture"] = new_user.get("picture", "")
        update_data["admin_swapped_requester"] = True
        action_description = f"Requester swapped: {old_requester_name} → {new_user.get('name')}"
        
    elif action == "swap_partner":
        if not new_user_id:
            raise HTTPException(status_code=400, detail="new_user_id required for swap action")
        
        # Get new user details
        new_user = await db.users.find_one({"id": new_user_id}, {"_id": 0})
        if not new_user:
            raise HTTPException(status_code=404, detail="New user not found")
        
        old_partner_name = session.get("partner_name")
        
        update_data["partner_id"] = new_user_id
        update_data["partner_name"] = new_user.get("name", "Unknown")
        update_data["partner_email"] = new_user.get("email", "")
        update_data["partner_picture"] = new_user.get("picture", "")
        update_data["admin_swapped_partner"] = True
        action_description = f"Partner swapped: {old_partner_name} → {new_user.get('name')}"
    
    # Log admin action
    action_log_entry = {
        "timestamp": datetime.utcnow(),
        "admin_email": admin_email,
        "admin_name": admin_name,
        "action": action_description,
        "notes": notes
    }
    
    update_data["admin_updated_at"] = datetime.utcnow()
    update_data["last_admin_email"] = admin_email
    
    # Update session
    result = await db.peer_sessions.update_one(
        {"$or": [
            {"id": session_id},
            {"_id": ObjectId(session_id) if len(session_id) == 24 else None}
        ]},
        {
            "$set": update_data,
            "$push": {
                "admin_action_history": action_log_entry
            }
        }
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Session not found")
    
    return {
        "success": True,
        "message": action_description,
        "action": action
    }


@router.get("/peer-sessions/{session_id}/action-history")
async def get_peer_session_action_history(session_id: str, request: Request):
    """Get admin action history for a peer session"""
    await verify_admin(request)
    db = get_db(request)
    
    from bson import ObjectId
    
    session = await db.peer_sessions.find_one(
        {"$or": [
            {"id": session_id},
            {"_id": ObjectId(session_id) if len(session_id) == 24 else None}
        ]},
        {"_id": 0, "admin_action_history": 1}
    )
    
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    history = session.get("admin_action_history", [])
    
    # Convert datetime to ISO string
    for entry in history:
        if isinstance(entry.get("timestamp"), datetime):
            entry["timestamp"] = entry["timestamp"].isoformat()
    
    return {
        "session_id": session_id,
        "action_history": history,
        "total_actions": len(history)
    }



# ============ Coaching Sessions Admin ============


# ============ Peer Practice Profiles Admin ============

@router.get("/peer-profiles")
async def get_all_peer_profiles(
    request: Request,
    page: int = 1,
    limit: int = 20,
    search: str = None,
    is_listed: str = None,  # "true", "false", or None for all
    sort_by: str = "created_at",  # created_at, name, sessions_done, rating
    sort_order: str = "desc"
):
    """Get all peer practice profiles with admin controls"""
    await verify_admin(request)
    db = get_db(request)
    
    # Build query
    query = {}
    
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"university": {"$regex": search, "$options": "i"}}
        ]
    
    if is_listed is not None:
        if is_listed.lower() == "true":
            query["is_listed"] = True
        elif is_listed.lower() == "false":
            query["is_listed"] = False
    
    # Count total
    total = await db.peer_profiles.count_documents(query)
    
    # Determine sort
    sort_field = sort_by
    sort_direction = -1 if sort_order == "desc" else 1
    
    # Get profiles
    skip = (page - 1) * limit
    profiles_cursor = db.peer_profiles.find(query, {"_id": 0}).sort(sort_field, sort_direction).skip(skip).limit(limit)
    profiles = await profiles_cursor.to_list(length=limit)
    
    # Enrich with session data
    enriched_profiles = []
    for profile in profiles:
        user_id = profile.get("user_id")
        
        # Get session stats
        total_sessions = await db.peer_sessions.count_documents({
            "$or": [
                {"requester_id": user_id},
                {"partner_id": user_id}
            ]
        })
        
        completed_sessions = await db.peer_sessions.count_documents({
            "$or": [
                {"requester_id": user_id},
                {"partner_id": user_id}
            ],
            "status": "completed"
        })
        
        pending_sessions = await db.peer_sessions.count_documents({
            "$or": [
                {"requester_id": user_id},
                {"partner_id": user_id}
            ],
            "status": {"$in": ["pending", "confirmed"]}
        })
        
        cancelled_sessions = await db.peer_sessions.count_documents({
            "$or": [
                {"requester_id": user_id},
                {"partner_id": user_id}
            ],
            "status": {"$in": ["cancelled", "admin_cancelled", "declined"]}
        })
        
        # Get user data for plan info
        user_data = await db.users.find_one({"id": user_id}, {"_id": 0, "plan": 1, "email": 1})
        
        enriched_profiles.append({
            "user_id": user_id,
            "name": profile.get("name"),
            "email": profile.get("email") or (user_data.get("email") if user_data else ""),
            "university": profile.get("university"),
            "ug_college": profile.get("ug_college"),
            "pg_college": profile.get("pg_college"),
            "firms_targeting": profile.get("firms_targeting", []),
            "cases_done": profile.get("cases_done", 0),
            "profile_picture": profile.get("profile_picture"),
            "linkedin_url": profile.get("linkedin_url"),
            "location": profile.get("location"),
            "years_of_experience": profile.get("years_of_experience"),
            "preparation_level": profile.get("preparation_level"),
            "is_listed": profile.get("is_listed", False),
            "peer_rating": profile.get("peer_rating"),
            "peer_sessions_done": profile.get("peer_sessions_done", 0),
            "created_at": profile.get("created_at").isoformat() if isinstance(profile.get("created_at"), datetime) else profile.get("created_at"),
            "updated_at": profile.get("updated_at").isoformat() if isinstance(profile.get("updated_at"), datetime) else profile.get("updated_at"),
            "google_calendar_connected": profile.get("google_calendar_connected", False),
            "plan": user_data.get("plan") if user_data else None,
            "session_stats": {
                "total": total_sessions,
                "completed": completed_sessions,
                "pending": pending_sessions,
                "cancelled": cancelled_sessions
            }
        })
    
    total_pages = (total + limit - 1) // limit
    
    return {
        "profiles": enriched_profiles,
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": total_pages
    }


@router.post("/peer-profiles/{user_id}/toggle-visibility")
async def admin_toggle_peer_profile_visibility(user_id: str, request: Request):
    """Admin can toggle whether a peer profile is visible on the website"""
    await verify_admin(request)
    db = get_db(request)
    
    body = await request.json()
    is_listed = body.get("is_listed")
    admin_notes = body.get("notes", "")
    
    if is_listed is None:
        raise HTTPException(status_code=400, detail="is_listed field is required")
    
    # Get admin user info
    admin_user = await get_current_user(request)
    admin_email = admin_user.get("email") if isinstance(admin_user, dict) else admin_user.email
    
    # Update profile
    result = await db.peer_profiles.update_one(
        {"user_id": user_id},
        {
            "$set": {
                "is_listed": is_listed,
                "admin_visibility_updated_at": datetime.utcnow(),
                "admin_visibility_updated_by": admin_email,
                "admin_visibility_notes": admin_notes
            }
        }
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Peer profile not found")
    
    visibility_status = "visible" if is_listed else "hidden"
    
    return {
        "success": True,
        "message": f"Profile visibility set to {visibility_status}",
        "user_id": user_id,
        "is_listed": is_listed
    }


@router.get("/peer-profiles/stats")
async def get_peer_profiles_stats(request: Request):
    """Get aggregate stats for peer profiles"""
    await verify_admin(request)
    db = get_db(request)
    
    total_profiles = await db.peer_profiles.count_documents({})
    listed_profiles = await db.peer_profiles.count_documents({"is_listed": True})
    unlisted_profiles = await db.peer_profiles.count_documents({"is_listed": False})
    
    # Profiles with calendar connected
    calendar_connected = await db.peer_profiles.count_documents({"google_calendar_connected": True})
    
    # Average rating
    pipeline = [
        {"$match": {"peer_rating": {"$exists": True, "$ne": None}}},
        {"$group": {"_id": None, "avg_rating": {"$avg": "$peer_rating"}}}
    ]
    rating_result = await db.peer_profiles.aggregate(pipeline).to_list(1)
    avg_rating = rating_result[0]["avg_rating"] if rating_result else 0
    
    return {
        "total_profiles": total_profiles,
        "listed_profiles": listed_profiles,
        "unlisted_profiles": unlisted_profiles,
        "calendar_connected": calendar_connected,
        "average_rating": round(avg_rating, 2)
    }


@router.get("/coaching-sessions")
async def get_all_coaching_sessions(
    request: Request,
    page: int = 1,
    limit: int = 20,
    status: str = None,
    mentor_id: str = None,
    date_from: str = None,
    date_to: str = None,
    search: str = None,
    booking_type: str = None  # NEW: Filter by 'coaching', 'strategy_call', or None for all
):
    """Get all coaching sessions AND strategy calls with filtering and pagination for admin view"""
    await verify_admin(request)
    db = get_db(request)
    
    # Build query for coaching bookings
    coaching_query = {}
    
    if status:
        coaching_query["status"] = status
    
    if mentor_id:
        coaching_query["mentor_id"] = mentor_id
    
    if date_from:
        coaching_query["date"] = {"$gte": date_from}
    
    if date_to:
        if "date" in coaching_query:
            coaching_query["date"]["$lte"] = date_to
        else:
            coaching_query["date"] = {"$lte": date_to}
    
    if search:
        coaching_query["$or"] = [
            {"candidate_name": {"$regex": search, "$options": "i"}},
            {"candidate_email": {"$regex": search, "$options": "i"}},
            {"mentor_name": {"$regex": search, "$options": "i"}},
        ]
    
    # Build query for strategy calls (similar filters)
    strategy_query = {}
    
    if status:
        # Map status for strategy calls
        status_map = {"pending": "scheduled", "confirmed": "confirmed", "completed": "completed", "cancelled": "cancelled"}
        strategy_query["status"] = status_map.get(status, status)
    
    if mentor_id:
        strategy_query["mentor_id"] = mentor_id
    
    if date_from:
        strategy_query["date"] = {"$gte": date_from}
    
    if date_to:
        if "date" in strategy_query:
            strategy_query["date"]["$lte"] = date_to
        else:
            strategy_query["date"] = {"$lte": date_to}
    
    if search:
        strategy_query["$or"] = [
            {"user_name": {"$regex": search, "$options": "i"}},
            {"user_email": {"$regex": search, "$options": "i"}},
            {"mentor_name": {"$regex": search, "$options": "i"}},
        ]
    
    # Get counts based on booking_type filter
    coaching_sessions = []
    strategy_sessions = []
    
    if booking_type != "strategy_call":
        coaching_sessions = await db.bookings.find(coaching_query, {"_id": 0}).sort([("date", -1), ("time_slot", -1)]).to_list(1000)
    
    if booking_type != "coaching":
        strategy_sessions = await db.strategy_call_sessions.find(strategy_query, {"_id": 0}).sort([("date", -1), ("time", -1)]).to_list(1000)
    
    # Process coaching sessions
    processed_coaching = []
    for session in coaching_sessions:
        mentor_feedback = await db.mentor_feedbacks.find_one({"booking_id": session.get("id")}, {"_id": 0})
        candidate_feedback = await db.candidate_feedbacks.find_one({"booking_id": session.get("id")}, {"_id": 0})
        mentor = await db.mentors.find_one({"id": session.get("mentor_id")}, {"_id": 0, "name": 1, "email": 1, "picture": 1, "timezone": 1})
        candidate = await db.users.find_one({"id": session.get("user_id")}, {"_id": 0, "name": 1, "email": 1, "picture": 1, "plan": 1, "timezone": 1})
        
        processed_coaching.append({
            "id": session.get("id"),
            "booking_type": "coaching",  # NEW: Identify type
            "date": session.get("date"),
            "time_slot": session.get("time_slot"),
            "status": session.get("status"),
            "session_type": session.get("session_type"),  # FIT Interview, Case Interview, etc.
            "case_type": session.get("case_type"),  # Profitability, Market Entry, etc.
            "duration_minutes": session.get("duration_minutes", 45),
            "mentor_id": session.get("mentor_id"),
            "mentor_name": mentor.get("name") if mentor else session.get("mentor_name"),
            "mentor_email": mentor.get("email") if mentor else None,
            "mentor_picture": mentor.get("picture") if mentor else None,
            "mentor_timezone": (mentor.get("timezone") if mentor else None) or "Asia/Kolkata",
            "candidate_timezone": (candidate.get("timezone") if candidate else None) or "Asia/Kolkata",
            "mentor_checked_in": session.get("mentor_checked_in", False),
            "mentor_checked_in_at": session.get("mentor_checked_in_at"),
            "mentor_feedback_given": mentor_feedback is not None,
            "mentor_feedback_rating": mentor_feedback.get("rating_overall") if mentor_feedback else None,
            "candidate_id": session.get("user_id"),
            "candidate_name": candidate.get("name") if candidate else session.get("candidate_name"),
            "candidate_email": candidate.get("email") if candidate else session.get("candidate_email"),
            "candidate_picture": candidate.get("picture") if candidate else None,
            "candidate_plan": candidate.get("plan") if candidate else None,
            "candidate_checked_in": session.get("candidate_checked_in", False),
            "candidate_checked_in_at": session.get("candidate_checked_in_at"),
            "candidate_feedback_given": candidate_feedback is not None,
            "candidate_feedback_rating": candidate_feedback.get("rating_overall") if candidate_feedback else None,
            "completion_status": session.get("completion_status"),
            "completion_notes": session.get("completion_notes"),
            "created_at": session.get("created_at"),
            "meet_link": session.get("meet_link"),
            "meet_space_name": session.get("meet_space_name"),
            "recording_url": session.get("recording_url"),
            "transcript_url": session.get("transcript_url"),
            "meet_artifacts_checked_at": session.get("meet_artifacts_checked_at"),
            "reschedule_requested": session.get("reschedule_requested", False),
            # Reschedule info - for original sessions that were rescheduled
            "rescheduled_by": session.get("rescheduled_by"),
            "rescheduled_by_name": session.get("rescheduled_by_name"),
            "rescheduled_at": session.get("rescheduled_at"),
            "rescheduled_to_id": session.get("rescheduled_to_id"),
            "rescheduled_to_date": session.get("rescheduled_to_date"),
            "rescheduled_to_time": session.get("rescheduled_to_time"),
            # Reschedule info - for new sessions created from reschedule
            "rescheduled_from_id": session.get("rescheduled_from_id"),
            "rescheduled_from_date": session.get("rescheduled_from_date"),
            "rescheduled_from_time": session.get("rescheduled_from_time"),
        })
    
    # Process strategy call sessions
    processed_strategy = []
    for session in strategy_sessions:
        mentor = await db.mentors.find_one({"id": session.get("mentor_id")}, {"_id": 0, "name": 1, "email": 1, "picture": 1, "timezone": 1})
        candidate = await db.users.find_one({"id": session.get("user_id")}, {"_id": 0, "name": 1, "email": 1, "picture": 1, "plan": 1, "timezone": 1})
        
        processed_strategy.append({
            "id": session.get("id"),
            "booking_type": "strategy_call",  # NEW: Identify type
            "date": session.get("date"),
            "time_slot": session.get("time"),  # Strategy calls use 'time' not 'time_slot'
            "status": session.get("status"),
            "session_type": "Strategy Call",  # Fixed type
            "case_type": None,  # Strategy calls don't have case types
            "duration_minutes": session.get("duration_minutes", 30),
            "mentor_id": session.get("mentor_id"),
            "mentor_name": mentor.get("name") if mentor else session.get("mentor_name"),
            "mentor_email": mentor.get("email") if mentor else session.get("mentor_email"),
            "mentor_picture": mentor.get("picture") if mentor else None,
            "mentor_timezone": (mentor.get("timezone") if mentor else None) or "Asia/Kolkata",
            "candidate_timezone": (candidate.get("timezone") if candidate else None) or "Asia/Kolkata",
            "mentor_checked_in": session.get("mentor_checked_in", False),
            "mentor_checked_in_at": session.get("mentor_checked_in_at"),
            "mentor_feedback_given": False,  # Strategy calls don't have feedback yet
            "mentor_feedback_rating": None,
            "candidate_id": session.get("user_id"),
            "candidate_name": candidate.get("name") if candidate else session.get("user_name"),
            "candidate_email": candidate.get("email") if candidate else session.get("user_email"),
            "candidate_picture": candidate.get("picture") if candidate else None,
            "candidate_plan": candidate.get("plan") if candidate else None,
            "candidate_checked_in": session.get("candidate_checked_in", False),
            "candidate_checked_in_at": session.get("candidate_checked_in_at"),
            "candidate_feedback_given": False,
            "candidate_feedback_rating": None,
            "completion_status": session.get("completion_status"),
            "completion_notes": session.get("notes"),
            "created_at": session.get("created_at"),
            "meet_link": session.get("meet_link"),
            # Recording fields — must match coaching sessions so admin UI renders them
            "meet_space_name": session.get("meet_space_name"),
            "recording_url": session.get("recording_url"),
            "transcript_url": session.get("transcript_url"),
            "meet_artifacts_checked_at": session.get("meet_artifacts_checked_at"),
            "reschedule_requested": False,
            # Reschedule info - for original sessions that were rescheduled
            "rescheduled_by": session.get("rescheduled_by"),
            "rescheduled_by_name": session.get("rescheduled_by_name"),
            "rescheduled_at": session.get("rescheduled_at"),
            "rescheduled_to_id": session.get("rescheduled_to_id"),
            "rescheduled_to_date": session.get("rescheduled_to_date"),
            "rescheduled_to_time": session.get("rescheduled_to_time"),
            # Reschedule info - for new sessions created from reschedule
            "rescheduled_from_id": session.get("rescheduled_from_id"),
            "rescheduled_from_date": session.get("rescheduled_from_date"),
            "rescheduled_from_time": session.get("rescheduled_from_time") or session.get("rescheduled_from_time"),
        })
    
    # Combine and sort all sessions by date (most recent first)
    all_sessions = processed_coaching + processed_strategy
    all_sessions.sort(key=lambda x: (x.get("date", ""), x.get("time_slot", "")), reverse=True)
    
    # Apply pagination
    total = len(all_sessions)
    skip = (page - 1) * limit
    paginated_sessions = all_sessions[skip:skip + limit]
    
    return {
        "sessions": paginated_sessions,
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": (total + limit - 1) // limit,
        "coaching_count": len(processed_coaching),
        "strategy_call_count": len(processed_strategy)
    }


@router.get("/coaching-sessions/stats")
async def get_coaching_sessions_stats(request: Request):
    """Get statistics for coaching sessions"""
    await verify_admin(request)
    db = get_db(request)
    
    from datetime import datetime, timedelta
    
    # Total sessions
    total = await db.bookings.count_documents({})
    
    # Today's sessions
    today = datetime.now().strftime("%Y-%m-%d")
    sessions_today = await db.bookings.count_documents({"date": today})
    
    # This week's sessions
    week_start = (datetime.now() - timedelta(days=datetime.now().weekday())).strftime("%Y-%m-%d")
    sessions_this_week = await db.bookings.count_documents({"date": {"$gte": week_start}})
    
    # By status - new status system
    status_counts = {}
    for status in ["confirmed", "completed", "mentor_no_show", "candidate_no_show", "both_no_show", "mentor_cancelled", "candidate_cancelled", "mentor_rescheduled", "candidate_rescheduled"]:
        count = await db.bookings.count_documents({"status": status})
        status_counts[status] = count
    
    # Feedback stats
    total_completed = await db.bookings.count_documents({"status": "completed"})
    mentor_feedback_count = await db.mentor_feedbacks.count_documents({})
    candidate_feedback_count = await db.candidate_feedbacks.count_documents({})
    
    # Check-in stats
    mentor_checked_in = await db.bookings.count_documents({"mentor_checked_in": True})
    candidate_checked_in = await db.bookings.count_documents({"candidate_checked_in": True})
    both_checked_in = await db.bookings.count_documents({
        "mentor_checked_in": True,
        "candidate_checked_in": True
    })
    
    # Get mentors list for filter dropdown
    mentors = await db.mentors.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
    
    return {
        "total": total,
        "sessions_today": sessions_today,
        "sessions_this_week": sessions_this_week,
        "by_status": status_counts,
        "feedback_stats": {
            "total_completed": total_completed,
            "mentor_feedback_given": mentor_feedback_count,
            "candidate_feedback_given": candidate_feedback_count,
        },
        "check_in_stats": {
            "mentor_checked_in": mentor_checked_in,
            "candidate_checked_in": candidate_checked_in,
            "both_checked_in": both_checked_in,
        },
        "mentors": mentors
    }


@router.get("/coaching-sessions/mentors-list")
async def get_mentors_for_session(request: Request):
    """Get list of active mentors for manual session creation"""
    await verify_admin(request)
    db = get_db(request)
    
    # Get all mentors (don't filter by is_active since many might not have this field)
    mentors = await db.mentors.find(
        {},
        {"_id": 0, "id": 1, "name": 1, "email": 1, "firm": 1, "picture": 1}
    ).sort("name", 1).to_list(100)
    
    # Filter out inactive mentors manually
    active_mentors = [m for m in mentors if m.get("is_active") != False]
    
    return {"mentors": active_mentors}


@router.get("/coaching-sessions/candidates-list")
async def get_candidates_for_session(request: Request, search: str = ""):
    """Get list of candidates for manual session creation with search"""
    await verify_admin(request)
    db = get_db(request)
    
    query = {"is_mentor": {"$ne": True}, "is_admin": {"$ne": True}}
    
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}}
        ]
    
    candidates = await db.users.find(
        query,
        {"_id": 0, "id": 1, "name": 1, "email": 1, "picture": 1, "plan": 1}
    ).sort("name", 1).limit(50).to_list(50)
    
    return {"candidates": candidates}


@router.get("/coaching-sessions/export-excel")
async def export_coaching_sessions_to_excel(
    request: Request,
    status: str = None,
    mentor_id: str = None,
    date_from: str = None,
    date_to: str = None,
    search: str = None,
    booking_type: str = None
):
    """Export coaching sessions to Excel file with optional filters"""
    await verify_admin(request)
    db = get_db(request)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        from fastapi.responses import StreamingResponse
        
        # Build query based on filters
        query = {}
        
        if status:
            query["status"] = status
        
        if mentor_id:
            query["mentor_id"] = mentor_id
        
        if date_from:
            query["date"] = {"$gte": date_from}
        
        if date_to:
            if "date" in query:
                query["date"]["$lte"] = date_to
            else:
                query["date"] = {"$lte": date_to}
        
        if search:
            query["$or"] = [
                {"candidate_name": {"$regex": search, "$options": "i"}},
                {"candidate_email": {"$regex": search, "$options": "i"}},
                {"mentor_name": {"$regex": search, "$options": "i"}},
            ]
        
        # Fetch sessions based on filters
        sessions = await db.bookings.find(query, {"_id": 0}).to_list(10000)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Coaching Sessions"
        
        # Define headers
        headers = [
            "Session ID", "Candidate Name", "Candidate Email", "Mentor Name", 
            "Mentor Email", "Date", "Time", "Status", "Session Type", "Case Type",
            "Booking Type", "Meet Link", "Created At", "Updated At", "Admin Remarks"
        ]
        
        # Style header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data rows
        for row_idx, session in enumerate(sessions, 2):
            ws.cell(row=row_idx, column=1, value=session.get("id", ""))
            ws.cell(row=row_idx, column=2, value=session.get("user_name", ""))
            ws.cell(row=row_idx, column=3, value=session.get("user_email", ""))
            ws.cell(row=row_idx, column=4, value=session.get("mentor_name", ""))
            ws.cell(row=row_idx, column=5, value=session.get("mentor_email", ""))
            ws.cell(row=row_idx, column=6, value=session.get("date", ""))
            ws.cell(row=row_idx, column=7, value=session.get("time_slot", ""))
            ws.cell(row=row_idx, column=8, value=session.get("status", ""))
            ws.cell(row=row_idx, column=9, value=session.get("session_type", ""))
            ws.cell(row=row_idx, column=10, value=session.get("case_type", ""))
            ws.cell(row=row_idx, column=11, value=session.get("booking_type", "coaching"))
            ws.cell(row=row_idx, column=12, value=session.get("meet_link", ""))
            ws.cell(row=row_idx, column=13, value=session.get("created_at", ""))
            ws.cell(row=row_idx, column=14, value=session.get("updated_at", ""))
            ws.cell(row=row_idx, column=15, value=session.get("admin_remarks", ""))
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Return as streaming response
        filename = f"coaching_sessions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error exporting coaching sessions: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to export sessions: {str(e)}")


@router.get("/coaching-sessions/{session_id}")
async def get_coaching_session_details(session_id: str, request: Request):
    """Get detailed info about a specific coaching session"""
    await verify_admin(request)
    db = get_db(request)
    
    # Find session
    session = await db.bookings.find_one({"id": session_id}, {"_id": 0})
    
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Get mentor details
    mentor = await db.mentors.find_one(
        {"id": session.get("mentor_id")},
        {"_id": 0, "name": 1, "email": 1, "picture": 1, "title": 1, "company": 1}
    )
    
    # Get candidate details
    candidate = await db.users.find_one(
        {"id": session.get("user_id")},
        {"_id": 0, "id": 1, "name": 1, "email": 1, "picture": 1, "plan": 1}
    )
    
    # Get mentor feedback
    mentor_feedback = await db.mentor_feedbacks.find_one(
        {"booking_id": session_id},
        {"_id": 0}
    )
    
    # Get candidate feedback
    candidate_feedback = await db.candidate_feedbacks.find_one(
        {"booking_id": session_id},
        {"_id": 0}
    )
    
    return {
        "session": session,
        "mentor": mentor,
        "candidate": candidate,
        "mentor_feedback": mentor_feedback,
        "candidate_feedback": candidate_feedback
    }


@router.post("/coaching-sessions/{session_id}/sync-recording")
async def admin_sync_session_recording(session_id: str, request: Request):
    """Manually pull the Meet recording + transcript URLs from Google
    onto this booking. Used when an admin wants the artifacts before
    the next scheduler cycle. Idempotent. Searches both `bookings` and
    `strategy_call_sessions` collections."""
    await verify_admin(request)
    db = get_db(request)

    target_coll = None
    booking = await db.bookings.find_one({"id": session_id}, {"_id": 0})
    if booking:
        target_coll = db.bookings
    else:
        booking = await db.strategy_call_sessions.find_one({"id": session_id}, {"_id": 0})
        if booking:
            target_coll = db.strategy_call_sessions
    if not booking or target_coll is None:
        raise HTTPException(status_code=404, detail="Session not found")
    if not booking.get("meet_space_name"):
        raise HTTPException(
            status_code=400,
            detail="No Meet space is associated with this session — recording sync is only available for sessions booked after the Meet REST API integration was deployed.",
        )
    from services.meet_artifacts_service import sync_artifacts_for_record
    res = await sync_artifacts_for_record(target_coll, booking)
    fresh = await target_coll.find_one(
        {"id": session_id},
        {"_id": 0, "recording_url": 1, "transcript_url": 1, "meet_artifacts_checked_at": 1, "recording_drive_moved": 1},
    ) or {}
    return {
        "success": True,
        "result": res,
        "recording_url": fresh.get("recording_url"),
        "transcript_url": fresh.get("transcript_url"),
        "checked_at": fresh.get("meet_artifacts_checked_at"),
        "recording_drive_moved": fresh.get("recording_drive_moved", False),
    }


@router.patch("/coaching-sessions/{session_id}/set-recording")
async def admin_set_session_recording_url(session_id: str, request: Request):
    """Allow admin to manually set a recording URL (and optional transcript URL)
    for any coaching session or strategy call. Useful when:
     - The session was recorded by a participant (not auto-recorded via API)
     - The auto-recording was not attached to the Meet space
     - The recording exists in Drive but wasn't synced automatically

    Body: {"recording_url": "https://...", "transcript_url": "https://..." (optional)}
    """
    await verify_admin(request)
    db = get_db(request)
    body = await request.json()
    recording_url = (body.get("recording_url") or "").strip()
    transcript_url = (body.get("transcript_url") or "").strip()
    meet_space_name = (body.get("meet_space_name") or "").strip()

    if not recording_url:
        raise HTTPException(status_code=400, detail="recording_url is required")

    target_coll = None
    booking = await db.bookings.find_one({"id": session_id}, {"_id": 0})
    if booking:
        target_coll = db.bookings
    else:
        booking = await db.strategy_call_sessions.find_one({"id": session_id}, {"_id": 0})
        if booking:
            target_coll = db.strategy_call_sessions
    if not booking or target_coll is None:
        raise HTTPException(status_code=404, detail="Session not found")

    now = datetime.now(timezone.utc).isoformat()
    update_fields: dict = {
        "recording_url": recording_url,
        "meet_artifacts_checked_at": now,
        "updated_at": now,
        "recording_set_manually": True,
    }
    if transcript_url:
        update_fields["transcript_url"] = transcript_url
    if meet_space_name:
        update_fields["meet_space_name"] = meet_space_name

    await target_coll.update_one({"id": session_id}, {"$set": update_fields})
    logger.info(
        "admin set-recording: session %s → %s (transcript: %s)",
        session_id, recording_url[:60], transcript_url[:60] if transcript_url else "none",
    )
    return {"success": True, "session_id": session_id, "recording_url": recording_url}


@router.post("/recordings/bulk-sync")
async def admin_bulk_sync_recordings(request: Request):
    """Trigger the artifact scheduler immediately for all pending sessions
    (both coaching bookings and strategy call sessions). Returns a summary
    of how many sessions were found, synced, and skipped.
    This is the same job the background scheduler runs every 5 minutes."""
    await verify_admin(request)
    db = get_db(request)
    from services.meet_artifacts_service import sync_pending_recordings
    result = await sync_pending_recordings(db, max_per_run=200, hours_lookback=168)  # 7-day lookback
    return {"success": True, **result}


# ============================================================================
# Recording Health Check + Diagnostics
# ============================================================================

@router.get("/recordings/config")
async def admin_recordings_config(request: Request):
    """Surface the current recording configuration so admin UI can show
    "Drive folder: ✓" / "Auto-record: ✓" status at a glance. No secrets."""
    await verify_admin(request)
    from services.calendar_service import (
        RECORDINGS_DRIVE_FOLDER_ID,
        MEET_AUTORECORD_ENABLED,
        IMPERSONATE_EMAIL,
        get_calendar_service,
    )
    cal = get_calendar_service()
    return {
        "auto_record_enabled": MEET_AUTORECORD_ENABLED,
        "calendar_service_available": cal.is_available(),
        "impersonate_email": IMPERSONATE_EMAIL,
        "recordings_drive_folder_id": RECORDINGS_DRIVE_FOLDER_ID or None,
        "recordings_drive_folder_url": (
            f"https://drive.google.com/drive/folders/{RECORDINGS_DRIVE_FOLDER_ID}"
            if RECORDINGS_DRIVE_FOLDER_ID else None
        ),
    }


@router.post("/recordings/self-test")
async def admin_recordings_self_test(request: Request):
    """End-to-end recording infrastructure health check.

    Steps performed:
      1. Verifies calendar service is initialized.
      2. Calls `_create_meet_space_with_recording(return_diagnostics=True)`
         — returns which tier succeeded/failed + raw HTTP error bodies.
      3. If a destination Drive folder is configured, verifies the
         impersonated user can access it (HEAD / GET on the folder).
      4. Returns a structured report the admin UI can render.

    Note: this DOES create a real Meet space. The space costs nothing
    and is harmless — it stays unused and Google garbages it after 60
    days. We don't run it on a schedule; it's manual-only.
    """
    await verify_admin(request)
    from services.calendar_service import (
        get_calendar_service,
        RECORDINGS_DRIVE_FOLDER_ID,
        IMPERSONATE_EMAIL,
        MEET_AUTORECORD_ENABLED,
    )

    report: Dict[str, Any] = {
        "auto_record_enabled": MEET_AUTORECORD_ENABLED,
        "impersonate_email": IMPERSONATE_EMAIL,
        "recordings_drive_folder_id": RECORDINGS_DRIVE_FOLDER_ID or None,
        "steps": [],
    }

    # Step 1 — calendar service available
    cal = get_calendar_service()
    if not cal.is_available():
        report["steps"].append({
            "name": "calendar_service_init",
            "ok": False,
            "detail": "GoogleCalendarService failed to initialize. Check GOOGLE_SERVICE_ACCOUNT_JSON env var.",
        })
        return report
    report["steps"].append({"name": "calendar_service_init", "ok": True, "detail": "Service initialized"})

    # Step 2 — try to create a Meet space with recording
    if not MEET_AUTORECORD_ENABLED:
        report["steps"].append({
            "name": "meet_space_create",
            "ok": False,
            "detail": "MEET_AUTORECORD_ENABLED is false — auto-recording is disabled.",
        })
        return report

    space_diag = cal._create_meet_space_with_recording(return_diagnostics=True)
    if not space_diag or not space_diag.get("meeting_uri"):
        report["steps"].append({
            "name": "meet_space_create",
            "ok": False,
            "detail": (space_diag or {}).get("error", "No diagnostics returned"),
            "attempts": (space_diag or {}).get("attempts", []),
            "remediation": _meet_remediation_hint((space_diag or {}).get("attempts", [])),
        })
        return report

    report["test_meeting"] = {
        "meeting_uri": space_diag.get("meeting_uri"),
        "meeting_code": space_diag.get("meeting_code"),
        "space_name": space_diag.get("space_name"),
        "tier": space_diag.get("tier"),
    }
    report["steps"].append({
        "name": "meet_space_create",
        "ok": True,
        "detail": (
            f"Created test Meet space at tier '{space_diag.get('tier')}'. "
            f"Space: {space_diag.get('space_name')}"
        ),
        "attempts": space_diag.get("attempts", []),
    })

    # Step 3 — verify the destination Drive folder is reachable
    if RECORDINGS_DRIVE_FOLDER_ID:
        drive_check = cal.check_drive_folder_access(RECORDINGS_DRIVE_FOLDER_ID)
        report["steps"].append({
            "name": "drive_folder_access",
            "ok": drive_check.get("ok"),
            "detail": drive_check.get("detail"),
            "folder_name": drive_check.get("name"),
            "remediation": (
                None if drive_check.get("ok")
                else f"Two checks needed: (1) Add 'https://www.googleapis.com/auth/drive' to the service account's allowed scopes in admin.google.com → Security → API Controls → Domain-wide Delegation. (2) Add {IMPERSONATE_EMAIL} as Manager (or Content Manager) of the Shared Drive containing folder {RECORDINGS_DRIVE_FOLDER_ID}."
            ),
        })
    else:
        report["steps"].append({
            "name": "drive_folder_access",
            "ok": False,
            "detail": "RECORDINGS_DRIVE_FOLDER_ID not configured. Recordings will stay in the host's My Drive.",
        })

    report["overall_ok"] = all(s.get("ok") for s in report["steps"])
    return report


def _meet_remediation_hint(attempts: List[Dict[str, Any]]) -> Optional[str]:
    """Suggest a fix based on the HTTP status codes returned by the Meet API."""
    if not attempts:
        return "No attempts logged — check backend logs for 'Meet API' lines."
    last = attempts[-1]
    code = last.get("status_code")
    if code == 401:
        return "Meet API returned 401 Unauthorized — service account credentials are invalid or expired."
    if code == 403:
        return (
            "Meet API returned 403 Forbidden — the service account is not authorized for the "
            "Meet REST API scopes. Go to admin.google.com → Security → API Controls → Domain-wide "
            "Delegation, and add 'meetings.space.created' and 'meetings.space.settings' to the "
            "service account's allowed scopes."
        )
    if code == 400:
        return (
            "All artifact-generation tiers were rejected with 400. The Workspace plan on this "
            "host may not include recording — verify the plan is Business Standard / Plus / "
            "Enterprise (Business Starter does NOT support recording)."
        )
    if code and code >= 500:
        return f"Meet API returned {code} server error — likely a transient Google outage; retry in a few minutes."
    return f"Meet API returned status {code} — check backend logs for full error body."


@router.get("/recordings/diagnose/{session_id}")
async def admin_recordings_diagnose(session_id: str, request: Request):
    """Per-session recording diagnostic. Returns:
      - the booking's meet_space_name + recording_url status
      - raw conferenceRecords API response for that space
      - per-record recordings/transcripts subresource responses
    so we can pinpoint where in the pipeline a recording is missing.
    """
    await verify_admin(request)
    db = get_db(request)
    booking = None
    coll_name = None
    for cn in ("bookings", "strategy_call_sessions"):
        b = await db[cn].find_one({"id": session_id}, {"_id": 0})
        if b:
            booking, coll_name = b, cn
            break
    if not booking:
        raise HTTPException(status_code=404, detail="Session not found")

    space_name = booking.get("meet_space_name")
    info: Dict[str, Any] = {
        "session_id": session_id,
        "collection": coll_name,
        "date": booking.get("date"),
        "time_slot": booking.get("time_slot"),
        "meet_link": booking.get("meet_link"),
        "meet_space_name": space_name,
        "recording_url": booking.get("recording_url"),
        "transcript_url": booking.get("transcript_url"),
        "meet_artifacts_checked_at": booking.get("meet_artifacts_checked_at"),
        "recording_drive_moved": booking.get("recording_drive_moved", False),
        "recording_drive_move_results": booking.get("recording_drive_move_results"),
        "stored_meet_artifacts": booking.get("meet_artifacts"),
    }

    if not space_name:
        info["diagnosis"] = (
            "This session has no meet_space_name. It was either booked before "
            "the Meet REST API integration was deployed, OR the Meet API call "
            "failed at create time and the legacy non-recording flow was used. "
            "Try POST /api/admin/meet/backfill-access?session_id=" + session_id
        )
        return info

    # Live fetch from Meet API
    try:
        from services.meet_artifacts_service import fetch_artifacts_for_space
        live = fetch_artifacts_for_space(space_name)
        info["live_artifacts"] = live
        if live is None:
            info["diagnosis"] = (
                "Meet REST API returned no conferenceRecords for this space. "
                "Either the meeting was never joined (no recording produced), "
                "or Google is still finalizing artifacts (typically 5–30 min after meeting end)."
            )
        elif not live.get("recordings"):
            info["diagnosis"] = (
                "conferenceRecords exist but no recording artifacts were produced. "
                "Likely cause: the meeting was held but recording was not started. "
                "Verify the host's Workspace plan supports recording AND that "
                "auto-record was actually applied to this space (check the create-time tier)."
            )
        else:
            info["diagnosis"] = (
                f"Meet API returned {len(live.get('recordings', []))} recording(s). "
                "If recording_url is empty above, try POST "
                f"/api/admin/coaching-sessions/{session_id}/sync-recording "
                "to force a sync."
            )
    except Exception as e:  # noqa: BLE001
        info["live_artifacts_error"] = str(e)

    return info


@router.post("/recordings/sync-all-pending")
async def admin_recordings_sync_all_pending(request: Request):
    """Force an immediate full pass of the artifact sync. Replicates what
    the background scheduler does, but on-demand. Useful when an admin
    just held a session and wants the recording link NOW."""
    await verify_admin(request)
    db = get_db(request)
    from services.meet_artifacts_service import sync_pending_recordings
    stats = await sync_pending_recordings(db, max_per_run=100, hours_lookback=168)
    return {"success": True, "stats": stats}


@router.get("/recordings/global-diagnose")
async def admin_recordings_global_diagnose(request: Request):
    """One-shot diagnostic — answers "why is no recording showing on the
    dashboard?" without needing a session_id. Returns:
      * scheduler heartbeat (last run time, last stats, last error)
      * counts: bookings/strategy_calls in last 7 days WITH meet_space_name,
        of those how many already have recording_url, how many are stuck
      * sample (up to 10) of stuck sessions with id + date + mentor +
        candidate so admin can pick one to force-sync
      * Drive folder access status (impersonated user can read the
        Shared Drive folder)
    """
    await verify_admin(request)
    db = get_db(request)
    from datetime import datetime, timezone, timedelta
    from services.calendar_service import (
        RECORDINGS_DRIVE_FOLDER_ID,
        IMPERSONATE_EMAIL,
        MEET_AUTORECORD_ENABLED,
        get_calendar_service,
    )

    now = datetime.now(timezone.utc)
    since_iso = (now - timedelta(days=7)).isoformat()

    out: Dict[str, Any] = {
        "checked_at": now.isoformat(),
        "auto_record_enabled": MEET_AUTORECORD_ENABLED,
        "impersonate_email": IMPERSONATE_EMAIL,
        "recordings_drive_folder_id": RECORDINGS_DRIVE_FOLDER_ID or None,
        "recordings_drive_folder_url": (
            f"https://drive.google.com/drive/folders/{RECORDINGS_DRIVE_FOLDER_ID}"
            if RECORDINGS_DRIVE_FOLDER_ID else None
        ),
    }

    # 1. Scheduler heartbeat
    hb = await db.system_status.find_one({"_id": "recording_scheduler"}, {"_id": 0})
    out["scheduler_heartbeat"] = hb or {"warning": "No heartbeat written yet — scheduler may not be running. Restart backend or wait 5 min."}
    if hb and hb.get("last_completed_at"):
        try:
            last_dt = datetime.fromisoformat(hb["last_completed_at"].replace("Z", "+00:00"))
            mins_ago = round((now - last_dt).total_seconds() / 60, 1)
            out["scheduler_heartbeat"]["minutes_since_last_run"] = mins_ago
            out["scheduler_alive"] = mins_ago < 15
        except Exception:  # noqa: BLE001
            out["scheduler_alive"] = None

    # 2. Counts across both collections
    counts: Dict[str, Any] = {}
    stuck_samples: List[Dict[str, Any]] = []
    for cname in ("bookings", "strategy_call_sessions"):
        coll = db[cname]
        with_space = await coll.count_documents({
            "meet_space_name": {"$nin": [None, ""], "$exists": True},
            "$or": [{"created_at": {"$gte": since_iso}}, {"date": {"$gte": (now - timedelta(days=7)).strftime("%Y-%m-%d")}}],
        })
        with_recording = await coll.count_documents({
            "meet_space_name": {"$nin": [None, ""], "$exists": True},
            "recording_url": {"$nin": [None, ""], "$exists": True},
            "$or": [{"created_at": {"$gte": since_iso}}, {"date": {"$gte": (now - timedelta(days=7)).strftime("%Y-%m-%d")}}],
        })
        moved = await coll.count_documents({
            "meet_space_name": {"$nin": [None, ""], "$exists": True},
            "recording_drive_moved": True,
            "$or": [{"created_at": {"$gte": since_iso}}, {"date": {"$gte": (now - timedelta(days=7)).strftime("%Y-%m-%d")}}],
        })
        counts[cname] = {
            "with_meet_space_name_last_7d": with_space,
            "with_recording_url": with_recording,
            "stuck_no_recording": with_space - with_recording,
            "moved_to_shared_drive": moved,
        }
        # Sample stuck sessions
        async for b in coll.find(
            {
                "meet_space_name": {"$nin": [None, ""], "$exists": True},
                "$or": [
                    {"recording_url": {"$exists": False}},
                    {"recording_url": {"$in": [None, ""]}},
                ],
                "$and": [
                    {"$or": [{"created_at": {"$gte": since_iso}}, {"date": {"$gte": (now - timedelta(days=7)).strftime("%Y-%m-%d")}}]},
                ],
            },
            {"_id": 0, "id": 1, "date": 1, "time_slot": 1, "meet_space_name": 1,
             "user_id": 1, "user_email": 1, "mentor_id": 1, "mentor_email": 1,
             "meet_artifacts_checked_at": 1, "status": 1},
        ).sort([("date", -1)]).limit(10):
            stuck_samples.append({**b, "collection": cname})
    out["counts_by_collection"] = counts
    out["stuck_session_samples"] = stuck_samples

    # 3. Drive folder access
    cal = get_calendar_service()
    out["calendar_service_available"] = cal.is_available()
    if RECORDINGS_DRIVE_FOLDER_ID and cal.is_available():
        try:
            out["drive_folder_check"] = cal.check_drive_folder_access(RECORDINGS_DRIVE_FOLDER_ID)
        except Exception as e:  # noqa: BLE001
            out["drive_folder_check"] = {"ok": False, "error": repr(e)}

    # 4. One-line diagnosis
    diagnosis_lines: List[str] = []
    if not out.get("scheduler_alive"):
        diagnosis_lines.append(
            "⚠️ Scheduler heartbeat is missing or stale — the artifact poller may not be running. "
            "Try restarting the backend (`sudo supervisorctl restart backend`)."
        )
    total_stuck = sum(c["stuck_no_recording"] for c in counts.values())
    if total_stuck > 0:
        diagnosis_lines.append(
            f"📌 {total_stuck} session(s) in the last 7 days have a Meet space but no recording_url — "
            f"hit POST /api/admin/recordings/sync-all-pending to force a sync now, OR use "
            f"POST /api/admin/recordings/find-and-force-sync to target a specific session."
        )
    drive = out.get("drive_folder_check") or {}
    if RECORDINGS_DRIVE_FOLDER_ID and not drive.get("ok"):
        diagnosis_lines.append(
            f"⚠️ Drive folder check FAILED: {drive.get('detail') or drive.get('error')}. "
            f"Recordings will land in {IMPERSONATE_EMAIL}'s My Drive but will NOT auto-move "
            f"to the Shared Drive folder."
        )
    if not diagnosis_lines:
        diagnosis_lines.append("✅ All recording-pipeline systems look healthy.")
    out["diagnosis"] = diagnosis_lines
    return out


class FindAndForceSyncRequest(BaseModel):
    """Locate a session by date + mentor email and immediately force a sync."""
    date: Optional[str] = None  # YYYY-MM-DD
    time_slot: Optional[str] = None  # e.g. "6:20 PM" or "18:20"
    mentor_email: Optional[str] = None
    candidate_email: Optional[str] = None
    session_id: Optional[str] = None  # bypass lookup if known


@router.post("/recordings/find-and-force-sync")
async def admin_recordings_find_and_force_sync(body: FindAndForceSyncRequest, request: Request):
    """Find a session by (date + mentor_email) or (date + candidate_email)
    or session_id, then immediately pull artifacts from the Meet REST API
    and persist on the booking. Skips the 8-min wait used by the
    on-completion hook. Returns the live Meet API response + the booking
    after the sync.
    """
    await verify_admin(request)
    db = get_db(request)

    # Try to find the session
    candidates: List[Dict[str, Any]] = []
    target_coll = None
    target_record = None

    if body.session_id:
        for cname in ("bookings", "strategy_call_sessions"):
            rec = await db[cname].find_one({"id": body.session_id}, {"_id": 0})
            if rec:
                target_coll = db[cname]
                target_record = rec
                break
    else:
        # Build query
        q: Dict[str, Any] = {}
        if body.date:
            q["date"] = body.date
        if body.mentor_email:
            q["$or"] = [
                {"mentor_email": body.mentor_email},
                {"mentor_email": body.mentor_email.lower()},
            ]
        if body.candidate_email:
            q["user_email"] = body.candidate_email
        for cname in ("bookings", "strategy_call_sessions"):
            async for rec in db[cname].find(q, {"_id": 0}):
                if body.time_slot and body.time_slot not in (rec.get("time_slot") or ""):
                    continue
                candidates.append({**rec, "_collection": cname})
        if not candidates:
            raise HTTPException(
                status_code=404,
                detail=f"No session found for date={body.date}, mentor={body.mentor_email}, candidate={body.candidate_email}",
            )
        # Take the most recent if multiple
        target_record = max(candidates, key=lambda r: r.get("created_at") or "")
        target_coll = db[target_record.pop("_collection")]

    if not target_record:
        raise HTTPException(status_code=404, detail="Session not found")

    if not body.session_id:
        # (lookup happened above; we're past the lookup now)
        pass
    from services.calendar_service import IMPERSONATE_EMAIL
    space_name = target_record.get("meet_space_name")
    response: Dict[str, Any] = {
        "session": {
            "id": target_record.get("id"),
            "collection": target_coll.name,
            "date": target_record.get("date"),
            "time_slot": target_record.get("time_slot"),
            "mentor_email": target_record.get("mentor_email"),
            "user_email": target_record.get("user_email"),
            "meet_link": target_record.get("meet_link"),
            "meet_space_name": space_name,
            "recording_url_before": target_record.get("recording_url"),
            "recording_drive_moved_before": target_record.get("recording_drive_moved", False),
        },
        "candidates_found": len(candidates) if candidates else 1,
    }

    if not space_name:
        response["error"] = (
            "This session has no meet_space_name. Either it was booked before the Meet REST "
            "API integration was deployed, OR Meet API space-create failed at booking time "
            "and the legacy non-recording flow was used. Recording cannot be retrieved — "
            "no Meet space exists to pull artifacts from."
        )
        return response

    # Live fetch from Meet REST API
    from services.meet_artifacts_service import fetch_artifacts_for_space, sync_artifacts_for_record
    live = fetch_artifacts_for_space(space_name)
    response["live_meet_api"] = live

    # Sync onto the booking
    sync_result = await sync_artifacts_for_record(target_coll, target_record)
    response["sync_result"] = sync_result

    # Refetch booking after sync
    fresh = await target_coll.find_one({"id": target_record.get("id")}, {"_id": 0}) or {}
    response["session"]["recording_url_after"] = fresh.get("recording_url")
    response["session"]["transcript_url"] = fresh.get("transcript_url")
    response["session"]["recording_drive_moved_after"] = fresh.get("recording_drive_moved", False)
    response["session"]["recording_drive_move_results"] = fresh.get("recording_drive_move_results")

    # Diagnosis
    diag: List[str] = []
    if not live:
        diag.append(
            "❌ Meet REST API returned NO conferenceRecords for this space. Either the meeting "
            "was never joined (no recording produced), Google is still finalizing artifacts "
            "(typically 5–30 min after meeting end), OR the impersonated service account lost "
            "access to this space."
        )
    elif not (live.get("recordings") or []):
        diag.append(
            "❌ Meet API returned conferenceRecords but no recording artifacts. The host's "
            "Workspace plan may not include recording, OR auto-record was not actually applied "
            "to this space at create time. Check the Meet space's create-time tier with "
            "POST /api/admin/recordings/self-test."
        )
    elif not fresh.get("recording_url"):
        diag.append(
            "⚠️ Meet API has the recording but sync didn't persist a recording_url. "
            "Check sync_result above for an error."
        )
    elif not fresh.get("recording_drive_moved"):
        diag.append(
            "⚠️ Recording exists and was synced, but failed to move to the Shared Drive folder. "
            "Likely cause: missing 'https://www.googleapis.com/auth/drive' DWD scope, OR "
            f"{IMPERSONATE_EMAIL} is not a Manager of the Shared Drive containing the "
            f"target folder. The recording is still accessible at the URL above."
        )
    else:
        diag.append("✅ Recording fetched and moved to Shared Drive successfully.")
    response["diagnosis"] = diag
    return response


@router.post("/meet/backfill-access")
async def admin_backfill_meet_access(request: Request):
    """Manually re-run the Meet access backfill — patches existing
    Meet spaces or regenerates links for upcoming sessions stuck on
    the legacy "Wait for the host" gate.

    Optional query params:
      * `reset=true` — clears the `meet_access_backfilled_at` stamps
        first so already-processed sessions are retried (e.g. when
        we ship an improved patch path).
      * `session_id=<id>` — limits the run to a single session
        (across bookings / strategy_call_sessions / case_competition_sessions).
    """
    await verify_admin(request)
    db = get_db(request)

    reset = (request.query_params.get("reset") or "").lower() in ("1", "true", "yes")
    session_id = request.query_params.get("session_id")

    collections = ["bookings", "strategy_call_sessions", "case_competition_sessions"]

    if reset:
        for coll_name in collections:
            q = {"meet_access_backfilled_at": {"$exists": True}}
            if session_id:
                q["id"] = session_id
            await db[coll_name].update_many(
                q,
                {"$unset": {
                    "meet_access_backfilled_at": "",
                    "meet_access_backfill_method": "",
                }},
            )

    if session_id:
        # Fast path: backfill a single session
        from services.calendar_service import get_calendar_service
        cal = get_calendar_service()
        if not cal.is_available():
            raise HTTPException(status_code=503, detail="Calendar service unavailable")

        target = None
        target_coll = None
        for coll_name in collections:
            doc = await db[coll_name].find_one({"id": session_id}, {"_id": 0})
            if doc:
                target, target_coll = doc, coll_name
                break
        if not target:
            raise HTTPException(status_code=404, detail="Session not found")

        space_name = target.get("meet_space_name")
        method = "failed"
        update_fields = {
            "meet_access_backfilled_at": datetime.now(timezone.utc).isoformat()
        }
        if space_name:
            ok = cal.update_meet_space_access_open(space_name)
            if ok:
                method = "patched"
        else:
            meet_space = cal._create_meet_space_with_recording()
            if meet_space and meet_space.get("meeting_uri"):
                update_fields["meet_link"] = meet_space["meeting_uri"]
                update_fields["meet_space_name"] = meet_space.get("space_name")
                method = "regenerated"

        update_fields["meet_access_backfill_method"] = method
        await db[target_coll].update_one({"id": session_id}, {"$set": update_fields})
        return {
            "success": method != "failed",
            "session_id": session_id,
            "collection": target_coll,
            "method": method,
            "meet_link": update_fields.get("meet_link") or target.get("meet_link"),
        }

    # Full-sweep path
    from migrations.startup_migrations import backfill_old_meet_access_to_open
    await backfill_old_meet_access_to_open(db)
    return {"success": True, "message": "Backfill triggered — see backend logs for counts."}




@router.post("/coaching-sessions/{session_id}/update-status")
async def admin_update_coaching_session_status(session_id: str, request: Request):
    """Admin can update status of coaching sessions OR strategy calls."""
    await verify_admin(request)
    db = get_db(request)
    
    body = await request.json()
    new_status = body.get("status")
    notes = body.get("notes", "")
    
    if new_status not in ["confirmed", "completed", "mentor_no_show", "candidate_no_show", "both_no_show", "mentor_cancelled", "candidate_cancelled", "admin_cancelled", "mentor_rescheduled", "candidate_rescheduled", "admin_rescheduled"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    # Locate the session in either coaching (db.bookings) or strategy
    # (db.strategy_call_sessions) — the admin list view shows both types.
    booking = await db.bookings.find_one({"id": session_id})
    is_strategy_call = False
    if not booking:
        booking = await db.strategy_call_sessions.find_one({"id": session_id})
        is_strategy_call = booking is not None

    if not booking:
        raise HTTPException(status_code=404, detail="Session not found")
    
    cancel_statuses = {"mentor_cancelled", "candidate_cancelled", "admin_cancelled"}
    reschedule_statuses = {"mentor_rescheduled", "candidate_rescheduled", "admin_rescheduled"}

    # For both cancellations and admin-marked reschedules, remove the original
    # calendar event so the coach's Google Calendar reflects reality.
    if new_status in cancel_statuses or new_status in reschedule_statuses:
        from services.calendar_service import get_calendar_service
        calendar_service = get_calendar_service()
        if calendar_service.is_available():
            calendar_event_id = booking.get("calendar_event_id")
            if calendar_event_id:
                try:
                    calendar_service.cancel_event(calendar_event_id, notify_attendees=(new_status in cancel_statuses))
                    logger.info(f"Admin cleared calendar event {calendar_event_id} for {new_status} on session {session_id}")
                except Exception as cal_err:
                    logger.warning(f"Failed to cancel calendar event {calendar_event_id}: {cal_err}")
            
            hidden_event_id = booking.get("hidden_event_id")
            if hidden_event_id:
                try:
                    calendar_service.cancel_event(hidden_event_id, notify_attendees=False)
                    logger.info(f"Admin cleared hidden calendar event {hidden_event_id} for {new_status} on session {session_id}")
                except Exception as cal_err:
                    logger.warning(f"Failed to cancel hidden calendar event {hidden_event_id}: {cal_err}")

    # Free up the slot only for cancellations (reschedules will be re-booked
    # to a new slot which will manage its own availability).
    if new_status in cancel_statuses and not is_strategy_call:
        slot_value = booking.get("time_slot") or booking.get("time")
        if slot_value:
            await db.mentor_availability.update_one(
                {"mentor_id": booking.get("mentor_id"), "date": booking.get("date")},
                {"$pull": {"booked_slots": slot_value}}
            )
            try:
                hour, minute = map(int, slot_value.split(':'))
                next_slot_minutes = hour * 60 + minute + 30
                next_hour = next_slot_minutes // 60
                next_minute = next_slot_minutes % 60
                next_slot = f"{next_hour:02d}:{next_minute:02d}"
                await db.mentor_availability.update_one(
                    {"mentor_id": booking.get("mentor_id"), "date": booking.get("date")},
                    {"$pull": {"booked_slots": next_slot}}
                )
            except (ValueError, AttributeError):
                pass
    
    update_data = {
        "status": new_status,
        "admin_notes": notes,
        "admin_updated_at": datetime.utcnow(),
    }
    
    if new_status in cancel_statuses:
        update_data["cancelled_at"] = datetime.utcnow()
        update_data["cancelled_by"] = "admin"
        
        # CREDIT SESSION BACK to candidate when admin cancels
        candidate = await db.users.find_one({"id": booking.get("user_id")})
        if candidate:
            candidate_plan = (candidate.get("plan") or "").lower()
            is_candidate_unlimited = (
                candidate_plan == "pinnacle" or
                candidate.get("is_unlimited_coaching", False) or
                candidate.get("coaching_sessions_total") == -1
            )
            
            if is_candidate_unlimited:
                # Unlimited users: just decrement usage counter
                if candidate.get("coaching_sessions_used", 0) > 0:
                    await db.users.update_one(
                        {"id": booking.get("user_id")},
                        {"$inc": {"coaching_sessions_used": -1}}
                    )
                    logger.info(f"Admin cancel: Decremented coaching_sessions_used for unlimited user {candidate.get('email')}")
            else:
                # For users with limited sessions: restore the credit
                update_ops = {}
                
                if candidate.get("coaching_sessions_used", 0) > 0:
                    update_ops["$inc"] = {"coaching_sessions_used": -1}
                
                # Also restore coaching_sessions_remaining for users not on coaching plans
                # (they purchased sessions directly)
                coaching_plans = ["last_mile", "mid_mile", "full_prep", "cohort_premium", "cohort_elite"]
                if candidate_plan not in coaching_plans:
                    if "$inc" not in update_ops:
                        update_ops["$inc"] = {}
                    update_ops["$inc"]["coaching_sessions_remaining"] = 1
                
                if update_ops:
                    await db.users.update_one(
                        {"id": booking.get("user_id")},
                        update_ops
                    )
                    logger.info(f"Admin cancel: Credited session back to user {candidate.get('email')}, ops={update_ops}")
        else:
            logger.warning(f"Admin cancel: Could not find candidate {booking.get('user_id')} to credit session back")
    
    if new_status in reschedule_statuses:
        update_data["rescheduled_by"] = "admin"
        update_data["rescheduled_at"] = datetime.utcnow()
    
    target_collection = db.strategy_call_sessions if is_strategy_call else db.bookings
    result = await target_collection.update_one(
        {"id": session_id},
        {"$set": update_data}
    )
    
    if new_status in cancel_statuses:
        try:
            candidate = await db.users.find_one({"id": booking.get("user_id")})
            mentor = await db.mentors.find_one({"id": booking.get("mentor_id")})
            
            if candidate and mentor:
                session_type_label = "Strategy Call" if is_strategy_call else booking.get("session_type", "Coaching session")
                await send_admin_cancellation_whatsapp_notifications(
                    candidate_name=candidate.get("name", "Candidate"),
                    candidate_phone=candidate.get("phone_number"),
                    candidate_country_code=candidate.get("phone_country_code", "+91"),
                    mentor_name=mentor.get("name", "Mentor"),
                    mentor_phone=mentor.get("phone_number") or mentor.get("phone"),
                    mentor_country_code=mentor.get("phone_country_code", "+91"),
                    session_date=booking.get("date"),
                    session_time=booking.get("time_slot") or booking.get("time"),
                    session_type=session_type_label
                )
        except Exception as wa_error:
            logger.warning(f"WhatsApp cancellation notification failed (non-critical): {wa_error}")
    
    return {"message": f"Session status updated to {new_status}"}


class ManualSessionCreate(BaseModel):
    mentor_id: str
    candidate_id: str
    date: str  # YYYY-MM-DD format
    time_slot: str  # HH:MM format
    session_type: Optional[str] = None  # Only for coaching: Case session, Fit Interview, PEI session, CV review session, General discussion
    case_type: Optional[str] = None  # Only for Case sessions
    admin_remarks: Optional[str] = None
    booking_type: str = "coaching"  # coaching or strategy_call
    deduct_credit: bool = False  # Whether to deduct a session credit from candidate


@router.post("/coaching-sessions/manual")
async def admin_create_manual_session(session_data: ManualSessionCreate, request: Request):
    """Admin can manually create a coaching session, bypassing availability checks"""
    await verify_admin(request)
    db = get_db(request)
    
    # Validate mentor exists
    mentor = await db.mentors.find_one({"id": session_data.mentor_id}, {"_id": 0})
    if not mentor:
        raise HTTPException(status_code=404, detail="Mentor not found")
    
    # Validate candidate exists
    candidate = await db.users.find_one({"id": session_data.candidate_id}, {"_id": 0})
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")
    
    # Determine duration based on booking type
    # Coaching sessions = 45 minutes, Strategy calls = 30 minutes
    duration_minutes = 45 if session_data.booking_type == "coaching" else 30
    
    # Build session notes for calendar
    if session_data.booking_type == "coaching":
        calendar_notes = f"Session Type: {session_data.session_type}"
        if session_data.session_type == "Case session" and session_data.case_type:
            calendar_notes += f"\nCase Type: {session_data.case_type}"
    else:
        calendar_notes = "Strategy Call"
    
    if session_data.admin_remarks:
        calendar_notes += f"\n\nAdmin Remarks:\n{session_data.admin_remarks}"
    calendar_notes += "\n\n[This session was manually created by admin]"
    
    # Create calendar event with Google Meet link
    calendar_result = None
    try:
        from services.calendar_service import create_coaching_session_event
        calendar_result = create_coaching_session_event(
            mentor_name=mentor.get("name", "Mentor"),
            mentor_email=mentor.get("email", ""),
            candidate_name=candidate.get("name", "Candidate"),
            candidate_email=candidate.get("email", ""),
            session_date=session_data.date,
            session_time=session_data.time_slot,
            duration_minutes=duration_minutes,
            session_notes=calendar_notes
        )
    except Exception as e:
        logger.warning(f"Failed to create calendar event: {e}")
    
    # Create the booking record
    booking_id = f"manual-{uuid.uuid4().hex[:12]}"
    booking_record = {
        "id": booking_id,
        "user_id": session_data.candidate_id,
        "mentor_id": session_data.mentor_id,
        "date": session_data.date,
        "time_slot": session_data.time_slot,
        "status": "confirmed",
        "session_type": session_data.session_type if session_data.booking_type == "coaching" else "Strategy Call",
        "booking_type": session_data.booking_type,
        "created_at": datetime.utcnow(),
        "created_by_admin": True,
        "admin_remarks": session_data.admin_remarks,
        "candidate_name": candidate.get("name", ""),
        "candidate_email": candidate.get("email", ""),
        "mentor_name": mentor.get("name", ""),
        "mentor_email": mentor.get("email", ""),
        "duration_minutes": duration_minutes,
    }
    
    # Add case type if applicable (only for coaching sessions with Case session type)
    if session_data.booking_type == "coaching" and session_data.session_type == "Case session" and session_data.case_type:
        booking_record["case_type"] = session_data.case_type
    
    # Add calendar event details if created
    if calendar_result:
        booking_record["calendar_event_id"] = calendar_result.get("event_id")
        booking_record["hidden_event_id"] = calendar_result.get("hidden_event_id")
        booking_record["meet_link"] = calendar_result.get("meet_link")
        # Persist meet_space_name so the artifacts scheduler can pull
        # the recording for admin-created sessions too.
        booking_record["meet_space_name"] = calendar_result.get("meet_space_name")
        booking_record["calendar_html_link"] = calendar_result.get("html_link")
    
    # Insert booking
    await db.bookings.insert_one(booking_record)
    
    # Update mentor availability to mark slot as booked (even if bypassing availability)
    await db.mentor_availability.update_one(
        {"mentor_id": session_data.mentor_id, "date": session_data.date},
        {"$addToSet": {"booked_slots": session_data.time_slot}},
        upsert=True
    )
    
    # Deduct credit if admin chose to do so
    credit_deducted = False
    if session_data.deduct_credit and session_data.booking_type == "coaching":
        # Check candidate's coaching access type
        has_purchased_sessions = candidate.get("coaching_sessions_remaining", 0) > 0
        is_unlimited_coaching = candidate.get("plan_key") in ["pinnacle", "ultimate"] or candidate.get("unlimited_coaching", False)
        
        if is_unlimited_coaching:
            # Unlimited users: just track usage, don't deduct
            await db.users.update_one(
                {"id": session_data.candidate_id},
                {"$inc": {"coaching_sessions_used": 1}}
            )
            credit_deducted = True
        elif has_purchased_sessions:
            # Users who purchased sessions: deduct from coaching_sessions_remaining
            await db.users.update_one(
                {"id": session_data.candidate_id},
                {
                    "$inc": {
                        "coaching_sessions_remaining": -1,
                        "coaching_sessions_used": 1
                    }
                }
            )
            credit_deducted = True
        else:
            # Users with coaching plans or cohort: track usage against their total
            await db.users.update_one(
                {"id": session_data.candidate_id},
                {"$inc": {"coaching_sessions_used": 1}}
            )
            credit_deducted = True
        
        # Mark booking as credit deducted
        await db.bookings.update_one(
            {"id": booking_id},
            {"$set": {"credit_deducted": True}}
        )
    
    # Remove _id for response
    booking_record.pop("_id", None)
    
    # Send WhatsApp notifications to candidate and mentor (fire and forget)
    try:
        await send_admin_booking_whatsapp_notifications(
            candidate_name=candidate.get("name", "Candidate"),
            candidate_phone=candidate.get("phone_number"),
            candidate_country_code=candidate.get("phone_country_code", "+91"),
            mentor_name=mentor.get("name", "Mentor"),
            mentor_phone=mentor.get("phone_number") or mentor.get("phone"),  # Try both field names
            mentor_country_code=mentor.get("phone_country_code", "+91"),
            session_date=session_data.date,
            session_time=session_data.time_slot,
            session_type=session_data.session_type if session_data.booking_type == "coaching" else "Strategy Call"
        )
    except Exception as wa_error:
        logger.warning(f"WhatsApp notification failed (non-critical): {wa_error}")
    
    return {
        "message": "Manual session created successfully",
        "booking": booking_record,
        "calendar_invite_sent": calendar_result is not None,
        "credit_deducted": credit_deducted
    }


async def send_admin_booking_whatsapp_notifications(
    candidate_name: str,
    candidate_phone: str,
    candidate_country_code: str,
    mentor_name: str,
    mentor_phone: str,
    mentor_country_code: str,
    session_date: str,
    session_time: str,
    session_type: str
):
    """Send WhatsApp notifications when admin creates a session"""
    
    logger.info(f"[WhatsApp Admin Booking] Starting notifications...")
    logger.info(f"[WhatsApp Admin Booking] Candidate: {candidate_name}, Phone: {candidate_phone}, Country: {candidate_country_code}")
    logger.info(f"[WhatsApp Admin Booking] Mentor: {mentor_name}, Phone: {mentor_phone}, Country: {mentor_country_code}")
    
    # Format phone numbers
    def format_phone(phone: str, country_code: str) -> str:
        if not phone:
            logger.warning(f"[WhatsApp Admin Booking] Phone number is None or empty")
            return None
        phone = str(phone).replace(" ", "").replace("-", "")
        if not phone.startswith("+"):
            country_code = country_code if country_code else "+91"
            phone = f"{country_code}{phone}"
        return phone
    
    candidate_full_phone = format_phone(candidate_phone, candidate_country_code)
    mentor_full_phone = format_phone(mentor_phone, mentor_country_code)
    
    logger.info(f"[WhatsApp Admin Booking] Formatted phones - Candidate: {candidate_full_phone}, Mentor: {mentor_full_phone}")
    
    # Send to candidate
    if candidate_full_phone:
        try:
            await wati_service.send_template_message(
                recipient_number=candidate_full_phone,
                template_name="candidate_coaching_session_booking",
                parameters=[
                    {"name": "1", "value": candidate_name},
                    {"name": "2", "value": session_type},
                    {"name": "3", "value": mentor_name},
                    {"name": "4", "value": session_date},
                    {"name": "5", "value": session_time}
                ]
            )
            logger.info(f"[WhatsApp Admin Booking] ✅ Sent to candidate: {candidate_full_phone}")
        except Exception as e:
            logger.warning(f"[WhatsApp Admin Booking] ❌ Failed to send to candidate: {e}")
    else:
        logger.warning(f"[WhatsApp Admin Booking] ⚠️ Skipping candidate - no phone number")
    
    # Send to mentor
    if mentor_full_phone:
        try:
            await wati_service.send_template_message(
                recipient_number=mentor_full_phone,
                template_name="mentor_coaching_session_booking",
                parameters=[
                    {"name": "1", "value": mentor_name},
                    {"name": "2", "value": session_type},
                    {"name": "3", "value": candidate_name},
                    {"name": "4", "value": session_date},
                    {"name": "5", "value": session_time}
                ]
            )
            logger.info(f"[WhatsApp Admin Booking] ✅ Sent to mentor: {mentor_full_phone}")
        except Exception as e:
            logger.warning(f"[WhatsApp Admin Booking] ❌ Failed to send to mentor: {e}")
    else:
        logger.warning(f"[WhatsApp Admin Booking] ⚠️ Skipping mentor - no phone number")


async def send_admin_cancellation_whatsapp_notifications(
    candidate_name: str,
    candidate_phone: str,
    candidate_country_code: str,
    mentor_name: str,
    mentor_phone: str,
    mentor_country_code: str,
    session_date: str,
    session_time: str,
    session_type: str
):
    """Send WhatsApp notifications when admin cancels a session"""
    
    logger.info(f"[WhatsApp Admin Cancel] Starting notifications...")
    
    # Format phone numbers
    def format_phone(phone: str, country_code: str) -> str:
        if not phone:
            return None
        phone = str(phone).replace(" ", "").replace("-", "")
        if not phone.startswith("+"):
            country_code = country_code if country_code else "+91"
            phone = f"{country_code}{phone}"
        return phone
    
    candidate_full_phone = format_phone(candidate_phone, candidate_country_code)
    mentor_full_phone = format_phone(mentor_phone, mentor_country_code)
    
    # Send to candidate
    if candidate_full_phone:
        try:
            await wati_service.send_template_message(
                recipient_number=candidate_full_phone,
                template_name="candidate_coaching_session_cancellation",
                parameters=[
                    {"name": "1", "value": candidate_name},
                    {"name": "2", "value": session_type},
                    {"name": "3", "value": mentor_name},
                    {"name": "4", "value": session_date},
                    {"name": "5", "value": session_time}
                ]
            )
            logger.info(f"[WhatsApp Admin Cancel] ✅ Sent to candidate: {candidate_full_phone}")
        except Exception as e:
            logger.warning(f"[WhatsApp Admin Cancel] ❌ Failed to send to candidate: {e}")
    
    # Send to mentor
    if mentor_full_phone:
        try:
            await wati_service.send_template_message(
                recipient_number=mentor_full_phone,
                template_name="mentor_session_cancellation",
                parameters=[
                    {"name": "1", "value": mentor_name},
                    {"name": "2", "value": session_type},
                    {"name": "3", "value": candidate_name},
                    {"name": "4", "value": session_date},
                    {"name": "5", "value": session_time}
                ]
            )
            logger.info(f"[WhatsApp Admin Cancel] ✅ Sent to mentor: {mentor_full_phone}")
        except Exception as e:
            logger.warning(f"[WhatsApp Admin Cancel] ❌ Failed to send to mentor: {e}")


async def send_admin_reschedule_whatsapp_notifications(
    candidate_name: str,
    candidate_phone: str,
    candidate_country_code: str,
    mentor_name: str,
    mentor_phone: str,
    mentor_country_code: str,
    new_date: str,
    new_time: str,
    session_type: str
):
    """Send WhatsApp notifications when admin reschedules a session"""
    
    logger.info(f"[WhatsApp Admin Reschedule] Starting notifications...")
    
    # Format phone numbers
    def format_phone(phone: str, country_code: str) -> str:
        if not phone:
            return None
        phone = str(phone).replace(" ", "").replace("-", "")
        if not phone.startswith("+"):
            country_code = country_code if country_code else "+91"
            phone = f"{country_code}{phone}"
        return phone
    
    candidate_full_phone = format_phone(candidate_phone, candidate_country_code)
    mentor_full_phone = format_phone(mentor_phone, mentor_country_code)
    
    # Send to candidate
    if candidate_full_phone:
        try:
            await wati_service.send_template_message(
                recipient_number=candidate_full_phone,
                template_name="candidate_coaching_session_reschedule",
                parameters=[
                    {"name": "1", "value": candidate_name},
                    {"name": "2", "value": session_type},
                    {"name": "3", "value": mentor_name},
                    {"name": "4", "value": new_date},
                    {"name": "5", "value": new_time}
                ]
            )
            logger.info(f"[WhatsApp Admin Reschedule] ✅ Sent to candidate: {candidate_full_phone}")
        except Exception as e:
            logger.warning(f"[WhatsApp Admin Reschedule] ❌ Failed to send to candidate: {e}")
    
    # Send to mentor
    if mentor_full_phone:
        try:
            await wati_service.send_template_message(
                recipient_number=mentor_full_phone,
                template_name="mentor_coaching_session_reschedule_v2",
                parameters=[
                    {"name": "1", "value": mentor_name},
                    {"name": "2", "value": session_type},
                    {"name": "3", "value": candidate_name},
                    {"name": "4", "value": new_date},
                    {"name": "5", "value": new_time}
                ]
            )
            logger.info(f"[WhatsApp Admin Reschedule] ✅ Sent to mentor: {mentor_full_phone}")
        except Exception as e:
            logger.warning(f"[WhatsApp Admin Reschedule] ❌ Failed to send to mentor: {e}")


# ============ App Settings (Single Session Price) ============

@router.get("/settings/topup")
async def get_topup_settings(request: Request):
    """Get the top-up session pricing settings including base price and discount tiers"""
    await verify_admin(request)
    db = get_db(request)
    
    settings = await db.app_settings.find_one({"key": "topup_settings"})
    
    if settings and settings.get("value"):
        return {
            "base_price": settings["value"].get("base_price", 2999),
            "discount_tiers": settings["value"].get("discount_tiers", [
                {"min_sessions": 5, "discount": 5},
                {"min_sessions": 10, "discount": 10},
                {"min_sessions": 15, "discount": 15},
                {"min_sessions": 20, "discount": 20}
            ])
        }
    
    # Return defaults
    return {
        "base_price": 2999,
        "discount_tiers": [
            {"min_sessions": 5, "discount": 5},
            {"min_sessions": 10, "discount": 10},
            {"min_sessions": 15, "discount": 15},
            {"min_sessions": 20, "discount": 20}
        ]
    }


@router.put("/settings/topup")
async def update_topup_settings(request: Request):
    """Update the top-up session pricing settings including base price and discount tiers"""
    await verify_admin(request)
    db = get_db(request)
    
    body = await request.json()
    base_price = body.get("base_price", 2999)
    discount_tiers = body.get("discount_tiers", [])
    
    if not isinstance(base_price, (int, float)) or base_price <= 0:
        raise HTTPException(status_code=400, detail="Base price must be a positive number")
    
    # Validate discount tiers
    if not isinstance(discount_tiers, list):
        raise HTTPException(status_code=400, detail="Discount tiers must be a list")
    
    for tier in discount_tiers:
        if not isinstance(tier.get("min_sessions"), int) or tier.get("min_sessions") < 1:
            raise HTTPException(status_code=400, detail="Each tier must have a valid min_sessions (>= 1)")
        if not isinstance(tier.get("discount"), (int, float)) or tier.get("discount") < 0 or tier.get("discount") > 100:
            raise HTTPException(status_code=400, detail="Each tier must have a valid discount (0-100)")
    
    # Sort tiers by min_sessions
    discount_tiers = sorted(discount_tiers, key=lambda x: x.get("min_sessions", 0))
    
    await db.app_settings.update_one(
        {"key": "topup_settings"},
        {"$set": {
            "key": "topup_settings",
            "value": {
                "base_price": int(base_price),
                "discount_tiers": discount_tiers
            },
            "updated_at": datetime.utcnow().isoformat()
        }},
        upsert=True
    )
    
    return {
        "message": "Top-up settings updated",
        "base_price": int(base_price),
        "discount_tiers": discount_tiers
    }


# Keep old endpoint for backwards compatibility but redirect to new settings
@router.get("/settings/single-session-price")
async def get_single_session_price(request: Request):
    """Deprecated: Use /settings/topup instead. Returns base price for backwards compatibility."""
    await verify_admin(request)
    db = get_db(request)
    
    settings = await db.app_settings.find_one({"key": "topup_settings"})
    if settings and settings.get("value"):
        return {"price": settings["value"].get("base_price", 2999)}
    
    return {"price": 2999}


@router.put("/settings/single-session-price")
async def update_single_session_price(request: Request):
    """Deprecated: Use /settings/topup instead. Updates base price for backwards compatibility."""
    await verify_admin(request)
    db = get_db(request)
    
    body = await request.json()
    price = body.get("price", 2999)
    
    if not isinstance(price, (int, float)) or price <= 0:
        raise HTTPException(status_code=400, detail="Price must be a positive number")
    
    # Get existing discount tiers or use defaults
    existing = await db.app_settings.find_one({"key": "topup_settings"})
    existing_tiers = []
    if existing and existing.get("value"):
        existing_tiers = existing["value"].get("discount_tiers", [])
    
    if not existing_tiers:
        existing_tiers = [
            {"min_sessions": 5, "discount": 5},
            {"min_sessions": 10, "discount": 10},
            {"min_sessions": 15, "discount": 15},
            {"min_sessions": 20, "discount": 20}
        ]
    
    await db.app_settings.update_one(
        {"key": "topup_settings"},
        {"$set": {
            "key": "topup_settings",
            "value": {
                "base_price": int(price),
                "discount_tiers": existing_tiers
            },
            "updated_at": datetime.utcnow().isoformat()
        }},
        upsert=True
    )
    
    return {"message": "Single session price updated", "price": int(price)}


# ============ Cohort Management ============

@router.get("/cohorts")
async def get_all_cohorts(request: Request):
    """Get all cohorts with their sections and resources"""
    await verify_admin(request)
    db = get_db(request)
    
    cohorts = await db.cohorts.find({}, {"_id": 0}).to_list(50)
    
    # Get sections and resources for each cohort
    for cohort in cohorts:
        # Get sections
        sections = await db.cohort_sections.find(
            {"cohort_id": cohort["id"]}, 
            {"_id": 0}
        ).sort("order", 1).to_list(50)
        cohort["sections"] = sections
        
        # Get resources
        resources = await db.cohort_resources.find(
            {"cohort_id": cohort["id"]}, 
            {"_id": 0}
        ).to_list(100)
        cohort["resources"] = resources
        
        # Count members
        members_count = await db.users.count_documents({"cohort_batch": cohort["name"]})
        cohort["members_count"] = members_count
    
    return {"cohorts": cohorts}


@router.post("/cohorts")
async def create_cohort(request: Request):
    """Create a new cohort with lifecycle validation"""
    await verify_admin(request)
    db = get_db(request)
    
    data = await request.json()
    status = data.get("status", "registering")
    
    # Validate lifecycle: only 1 active and 1 registering allowed
    if status == "active":
        existing_active = await db.cohorts.find_one({"status": "active"})
        if existing_active:
            raise HTTPException(
                status_code=400, 
                detail=f"An active cohort already exists: '{existing_active['name']}'. Please complete or archive it first."
            )
    elif status == "registering":
        existing_registering = await db.cohorts.find_one({"status": "registering"})
        if existing_registering:
            raise HTTPException(
                status_code=400, 
                detail=f"A registering cohort already exists: '{existing_registering['name']}'. Please activate or archive it first."
            )
    
    cohort_id = f"cohort-{str(uuid.uuid4())[:8]}"
    
    cohort = {
        "id": cohort_id,
        "name": data.get("name"),
        "description": data.get("description", ""),
        "start_date": data.get("start_date"),
        "end_date": data.get("end_date"),
        "status": status,
        "max_participants": data.get("max_participants", 50),
        "price": data.get("price", 0),
        "features": data.get("features", []),
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow()
    }
    
    await db.cohorts.insert_one(cohort)
    try:
        from routes.cohorts import invalidate_public_cohorts_cache
        invalidate_public_cohorts_cache()
    except Exception:
        pass
    return {"message": "Cohort created successfully", "cohort_id": cohort_id}


@router.put("/cohorts/{cohort_id}/status")
async def update_cohort_status(cohort_id: str, request: Request):
    """Update cohort status with lifecycle validation"""
    await verify_admin(request)
    db = get_db(request)
    
    data = await request.json()
    new_status = data.get("status")
    
    if new_status not in ["registering", "active", "completed", "archived"]:
        raise HTTPException(status_code=400, detail="Invalid status. Must be: registering, active, completed, or archived")
    
    # Get current cohort
    cohort = await db.cohorts.find_one({"id": cohort_id})
    if not cohort:
        raise HTTPException(status_code=404, detail="Cohort not found")
    
    current_status = cohort.get("status")
    
    # Validate lifecycle transitions
    if new_status == "active":
        # Check if another active cohort exists
        existing_active = await db.cohorts.find_one({"status": "active", "id": {"$ne": cohort_id}})
        if existing_active:
            raise HTTPException(
                status_code=400, 
                detail=f"Cannot activate: '{existing_active['name']}' is already active. Complete or archive it first."
            )
    elif new_status == "registering":
        # Check if another registering cohort exists
        existing_registering = await db.cohorts.find_one({"status": "registering", "id": {"$ne": cohort_id}})
        if existing_registering:
            raise HTTPException(
                status_code=400, 
                detail=f"Cannot set to registering: '{existing_registering['name']}' is already accepting registrations."
            )
    
    # Update status
    await db.cohorts.update_one(
        {"id": cohort_id},
        {"$set": {"status": new_status, "updated_at": datetime.utcnow()}}
    )
    try:
        from routes.cohorts import invalidate_public_cohorts_cache
        invalidate_public_cohorts_cache()
    except Exception:
        pass
    
    return {"message": f"Cohort status updated to '{new_status}'", "previous_status": current_status}


@router.put("/cohorts/{cohort_id}")
async def update_cohort(cohort_id: str, request: Request):
    """Update cohort details"""
    await verify_admin(request)
    db = get_db(request)
    
    data = await request.json()
    
    # Don't allow status update through this endpoint
    if "status" in data:
        del data["status"]
    
    data["updated_at"] = datetime.utcnow()
    
    result = await db.cohorts.update_one(
        {"id": cohort_id},
        {"$set": data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Cohort not found")
    
    try:
        from routes.cohorts import invalidate_public_cohorts_cache
        invalidate_public_cohorts_cache()
    except Exception:
        pass
    return {"message": "Cohort updated successfully"}


@router.post("/cohorts/{cohort_id}/enroll")
async def enroll_user_in_cohort(cohort_id: str, request: Request):
    """Enroll a user in a cohort (admin action)"""
    await verify_admin(request)
    db = get_db(request)
    
    data = await request.json()
    user_id = data.get("user_id")
    
    if not user_id:
        raise HTTPException(status_code=400, detail="user_id is required")
    
    # Get cohort
    cohort = await db.cohorts.find_one({"id": cohort_id})
    if not cohort:
        raise HTTPException(status_code=404, detail="Cohort not found")
    
    # Check if user exists
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Update user's cohort enrollment
    await db.users.update_one(
        {"id": user_id},
        {"$set": {
            "cohort_id": cohort_id,
            "cohort_batch": cohort["name"],
            "cohort_enrolled_at": datetime.utcnow()
        }}
    )
    
    return {"message": f"User enrolled in cohort '{cohort['name']}'"}


@router.delete("/cohorts/{cohort_id}/enroll/{user_id}")
async def unenroll_user_from_cohort(cohort_id: str, user_id: str, request: Request):
    """Remove a user from a cohort"""
    await verify_admin(request)
    db = get_db(request)
    
    await db.users.update_one(
        {"id": user_id, "cohort_id": cohort_id},
        {"$unset": {"cohort_id": "", "cohort_batch": "", "cohort_enrolled_at": ""}}
    )
    
    return {"message": "User removed from cohort"}


@router.get("/cohorts/{cohort_id}/members")
async def get_cohort_members(cohort_id: str, request: Request):
    """Get all members enrolled in a cohort"""
    await verify_admin(request)
    db = get_db(request)
    
    members = await db.users.find(
        {"cohort_id": cohort_id},
        {"_id": 0, "id": 1, "name": 1, "email": 1, "picture": 1, "cohort_enrolled_at": 1}
    ).to_list(200)
    
    return {"members": members, "count": len(members)}


@router.post("/cohorts/{cohort_id}/sections")
async def add_cohort_section(cohort_id: str, request: Request):
    """Add a section to a cohort"""
    await verify_admin(request)
    db = get_db(request)
    
    data = await request.json()
    section_id = f"section-{str(uuid.uuid4())[:8]}"
    
    section = {
        "id": section_id,
        "cohort_id": cohort_id,
        "title": data.get("title"),
        "description": data.get("description", ""),
        "order": data.get("order", 0),
        "created_at": datetime.utcnow()
    }
    
    await db.cohort_sections.insert_one(section)
    return {"message": "Section added successfully", "section_id": section_id}


@router.post("/cohorts/{cohort_id}/resources")
async def add_cohort_resource(cohort_id: str, request: Request):
    """Add a resource to a cohort (optionally within a section)"""
    await verify_admin(request)
    db = get_db(request)
    
    data = await request.json()
    resource_id = f"resource-{str(uuid.uuid4())[:8]}"
    
    resource = {
        "id": resource_id,
        "cohort_id": cohort_id,
        "section_id": data.get("section_id"),  # Optional - if None, it's a general resource
        "title": data.get("title"),
        "type": data.get("type", "document"),
        "file_url": data.get("file_url"),
        "created_at": datetime.utcnow()
    }
    
    await db.cohort_resources.insert_one(resource)
    return {"message": "Resource added successfully", "resource_id": resource_id}


@router.get("/cohort/sections")
async def get_cohort_sections(request: Request):
    """Get all cohort sections"""
    await verify_admin(request)
    db = get_db(request)
    
    sections = await db.cohort_sections.find({}, {"_id": 0}).to_list(50)
    return {"sections": sections}


@router.post("/cohort/sections")
async def create_cohort_section(section_data: CohortSectionCreate, request: Request):
    """Create a new cohort section"""
    await verify_admin(request)
    db = get_db(request)
    
    section_id = f"section-{str(uuid.uuid4())[:8]}"
    
    section = {
        "id": section_id,
        **section_data.dict(),
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow()
    }
    
    await db.cohort_sections.insert_one(section)
    return {"message": "Section created successfully", "section_id": section_id}


@router.put("/cohort/sections/{section_id}")
async def update_cohort_section(section_id: str, request: Request):
    """Update a cohort section"""
    await verify_admin(request)
    db = get_db(request)
    
    body = await request.json()
    update_data = {k: v for k, v in body.items() if v is not None and k != "id"}
    update_data["updated_at"] = datetime.utcnow()
    
    result = await db.cohort_sections.update_one({"id": section_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Section not found")
    
    return {"message": "Section updated successfully"}


@router.delete("/cohort/sections/{section_id}")
async def delete_cohort_section(section_id: str, request: Request):
    """Delete a cohort section and its resources"""
    await verify_admin(request)
    db = get_db(request)
    
    result = await db.cohort_sections.delete_one({"id": section_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Section not found")
    
    # Also delete associated resources
    await db.cohort_resources.delete_many({"section_id": section_id})
    
    return {"message": "Section and resources deleted successfully"}


@router.get("/cohort/resources")
async def get_cohort_resources(request: Request, section_id: str = None):
    """Get cohort resources, optionally filtered by section"""
    await verify_admin(request)
    db = get_db(request)
    
    query = {"section_id": section_id} if section_id else {}
    resources = await db.cohort_resources.find(query, {"_id": 0}).to_list(200)
    
    return {"resources": resources}


@router.post("/cohort/resources")
async def create_cohort_resource(resource_data: CohortResourceCreate, request: Request):
    """Create a new cohort resource"""
    await verify_admin(request)
    db = get_db(request)
    
    resource_id = f"resource-{str(uuid.uuid4())[:8]}"
    
    resource = {
        "id": resource_id,
        **resource_data.dict(),
        "created_at": datetime.utcnow(),
        "updated_at": datetime.utcnow()
    }
    
    await db.cohort_resources.insert_one(resource)
    return {"message": "Resource created successfully", "resource_id": resource_id}


@router.put("/cohort/resources/{resource_id}")
async def update_cohort_resource(resource_id: str, request: Request):
    """Update a cohort resource"""
    await verify_admin(request)
    db = get_db(request)
    
    body = await request.json()
    update_data = {k: v for k, v in body.items() if v is not None and k != "id"}
    update_data["updated_at"] = datetime.utcnow()
    
    result = await db.cohort_resources.update_one({"id": resource_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Resource not found")
    
    return {"message": "Resource updated successfully"}


@router.delete("/cohort/resources/{resource_id}")
async def delete_cohort_resource(resource_id: str, request: Request):
    """Delete a cohort resource"""
    await verify_admin(request)
    db = get_db(request)
    
    result = await db.cohort_resources.delete_one({"id": resource_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Resource not found")
    
    return {"message": "Resource deleted successfully"}


# ============ File Upload ============

UPLOAD_DIR = "/app/uploads"
CHUNK_DIR = "/app/uploads/chunks"
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(CHUNK_DIR, exist_ok=True)

# Store active chunked uploads
active_uploads = {}


@router.post("/upload")
async def upload_file(
    request: Request,
    file: UploadFile = File(...),
    category: str = Form("general"),
    persist_to_db: str = Form("false"),  # For small images that need to persist across deployments
    use_cloud: str = Form("true")  # Use cloud storage (recommended for all files)
):
    """Upload a file (thumbnail, video, document) to cloud storage or local
    
    For videos and large files: Always use cloud storage (use_cloud=true)
    For small images: Can persist to MongoDB (persist_to_db=true) or cloud
    
    Returns URL that can be used directly in video_url or thumbnail fields.
    """
    import asyncio
    
    await verify_admin(request)
    db = get_db(request)
    
    # Read file content with error handling
    try:
        file_content = await file.read()
    except Exception as e:
        logging.error(f"Failed to read uploaded file: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to read file: {str(e)}")
    
    file_size = len(file_content)
    file_ext = os.path.splitext(file.filename)[1].lower()
    
    logging.info(f"Upload received: {file.filename}, size={file_size}, ext={file_ext}, category={category}")
    
    # Determine if this is a video file
    video_extensions = ['.mp4', '.webm', '.mov', '.avi', '.mkv', '.m4v', '.wmv', '.flv']
    is_video = file_ext in video_extensions
    
    # For videos or when use_cloud=true, use cloud storage
    from services import cloud_storage_service
    
    if (use_cloud.lower() == "true" or is_video) and cloud_storage_service.is_enabled():
        try:
            # Run synchronous cloud upload in thread pool to avoid blocking event loop
            result = await asyncio.to_thread(
                cloud_storage_service.upload_file,
                data=file_content,
                filename=file.filename,
                folder=category,
                content_type=file.content_type
            )
            
            # Return cloud storage URL
            file_url = f"/api/files/{result['storage_path']}"
            
            return {
                "message": "File uploaded to cloud storage successfully",
                "url": file_url,
                "filename": file.filename,
                "storage_path": result['storage_path'],
                "size": result['size'],
                "cloud_stored": True,
                "is_video": is_video
            }
        except Exception as e:
            logging.exception(f"Cloud upload failed for {file.filename} ({file_size} bytes): {e}")
            # Fall through to local storage as backup
    
    # For small images (< 5MB) with persist_to_db=true, store in MongoDB
    image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.webp', '.svg', '.ico', '.avif']
    should_persist = persist_to_db.lower() == "true" and file_ext in image_extensions and file_size < 5 * 1024 * 1024
    
    if should_persist:
        import base64
        
        # Determine content type
        content_types = {
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.png': 'image/png',
            '.gif': 'image/gif',
            '.webp': 'image/webp',
            '.svg': 'image/svg+xml',
            '.ico': 'image/x-icon',
            '.avif': 'image/avif'
        }
        content_type = content_types.get(file_ext, 'image/png')
        
        # Generate unique ID
        file_id = f"img_{uuid.uuid4().hex[:12]}"
        
        # Store in MongoDB
        image_doc = {
            "id": file_id,
            "filename": file.filename,
            "category": category,
            "content_type": content_type,
            "data": base64.b64encode(file_content).decode('utf-8'),
            "size": file_size,
            "created_at": datetime.utcnow().isoformat()
        }
        
        await db.persistent_images.insert_one(image_doc)
        
        # Return URL that serves from MongoDB
        file_url = f"/api/images/{file_id}"
        
        return {"message": "File uploaded successfully", "url": file_url, "filename": file.filename, "persisted": True, "cloud_stored": False}
    
    # For large files or non-persisted files, save to disk (existing behavior)
    # Create category directory
    category_dir = os.path.join(UPLOAD_DIR, category)
    os.makedirs(category_dir, exist_ok=True)
    
    # Generate unique filename
    unique_filename = f"{str(uuid.uuid4())[:8]}{file_ext}"
    file_path = os.path.join(category_dir, unique_filename)
    
    # Save file
    with open(file_path, "wb") as buffer:
        buffer.write(file_content)
    
    # Return URL path (use /api/uploads for proper routing through ingress)
    file_url = f"/api/uploads/{category}/{unique_filename}"
    
    return {"message": "File uploaded successfully", "url": file_url, "filename": unique_filename, "persisted": False, "cloud_stored": False}


class ChunkedUploadInit(BaseModel):
    filename: str
    filesize: int
    filetype: str
    total_chunks: int
    upload_id: str
    category: str = "general"


@router.post("/upload/init")
async def init_chunked_upload(data: ChunkedUploadInit, request: Request):
    """Initialize a chunked upload session - creates temp directory for chunks on disk."""
    await verify_admin(request)
    
    logger = logging.getLogger(__name__)
    logger.info(f"Upload init: {data.filename}, size={data.filesize}, chunks={data.total_chunks}, id={data.upload_id}")
    
    # Create temp directory for this upload's chunks (disk-based, not MongoDB)
    upload_temp_dir = "/tmp/gradnext_uploads"
    os.makedirs(upload_temp_dir, exist_ok=True)
    upload_dir = os.path.join(upload_temp_dir, data.upload_id)
    os.makedirs(upload_dir, exist_ok=True)
    
    # Write metadata file
    import json
    meta = {
        "upload_id": data.upload_id,
        "filename": data.filename,
        "filesize": data.filesize,
        "filetype": data.filetype,
        "total_chunks": data.total_chunks,
        "category": data.category,
    }
    with open(os.path.join(upload_dir, "_meta.json"), "w") as f:
        json.dump(meta, f)
    
    return {"success": True, "upload_id": data.upload_id, "message": "Upload initialized"}


@router.post("/upload/chunk")
async def upload_chunk(
    request: Request,
    chunk: UploadFile = File(...),
    upload_id: str = Form(...),
    chunk_index: int = Form(...),
    total_chunks: int = Form(...)
):
    """Upload a single chunk - writes directly to disk (no base64, no MongoDB overhead)."""
    import logging
    logger = logging.getLogger(__name__)
    
    await verify_admin(request)
    
    upload_temp_dir = "/tmp/gradnext_uploads"
    upload_dir = os.path.join(upload_temp_dir, upload_id)
    
    # Auto-create directory if missing (handles server restart edge case)
    if not os.path.exists(upload_dir):
        os.makedirs(upload_dir, exist_ok=True)
        logger.warning(f"Upload dir recreated for {upload_id} (chunk {chunk_index})")
    
    # Read chunk data
    try:
        chunk_data = await chunk.read()
    except Exception as e:
        logger.error(f"Failed to read chunk {chunk_index} for {upload_id}: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to read chunk: {str(e)}")
    
    if not chunk_data:
        raise HTTPException(status_code=400, detail="Empty chunk received")
    
    # Write chunk directly to disk as binary (NO base64, NO MongoDB)
    chunk_path = os.path.join(upload_dir, f"chunk_{chunk_index:05d}")
    try:
        with open(chunk_path, "wb") as f:
            f.write(chunk_data)
    except Exception as e:
        logger.error(f"Failed to write chunk {chunk_index} to disk: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to save chunk: {str(e)}")
    
    # Count received chunks
    received = len([f for f in os.listdir(upload_dir) if f.startswith("chunk_")])
    
    logger.info(f"Chunk {chunk_index}/{total_chunks} for {upload_id} ({len(chunk_data)} bytes) - {received} received")
    
    return {
        "success": True,
        "chunk_index": chunk_index,
        "received": received,
        "total": total_chunks
    }


class ChunkedUploadFinalize(BaseModel):
    upload_id: str
    filename: str
    total_chunks: int
    category: str = "general"


@router.post("/upload/finalize")
async def finalize_chunked_upload(data: ChunkedUploadFinalize, request: Request):
    """Kick off finalization in the background and return immediately.
    
    Why background?
        Cloudflare (the production proxy on app.gradnext.co) enforces a 100s
        upper bound on the response time from the origin. Combining 1.5GB
        of chunks on disk and PUT-ing the result to Emergent Object Storage
        easily exceeds that and causes a 520 even though the upload is
        actually still in progress on our backend. Returning immediately
        and letting the client poll dodges the proxy timeout entirely.
    
    Client must poll GET /api/admin/upload/status/{upload_id} until the
    response contains state == "done" (success) or state == "failed".
    """
    import asyncio
    import logging
    logger = logging.getLogger(__name__)
    
    await verify_admin(request)
    
    upload_id = data.upload_id
    upload_temp_dir = "/tmp/gradnext_uploads"
    upload_dir = os.path.join(upload_temp_dir, upload_id)
    
    logger.info(f"Finalize received: {upload_id}, filename={data.filename}, total_chunks={data.total_chunks}")
    
    # Validate up front so the client gets a synchronous error if the
    # upload session is broken — no point queueing background work then.
    if not os.path.exists(upload_dir):
        raise HTTPException(status_code=400, detail="Upload session not found or expired. Please try uploading again.")
    
    chunk_files = sorted([f for f in os.listdir(upload_dir) if f.startswith("chunk_")])
    if len(chunk_files) < data.total_chunks:
        raise HTTPException(
            status_code=400,
            detail=f"Missing chunks: found {len(chunk_files)} of {data.total_chunks}. Please retry upload.",
        )
    
    # Initialize status, then schedule the heavy work
    _write_upload_status(upload_id, {
        "state": "processing",
        "phase": "queued",
        "upload_id": upload_id,
        "filename": data.filename,
        "started_at": datetime.now(timezone.utc).isoformat(),
    })
    
    asyncio.create_task(
        _run_finalize_in_background(
            upload_id=upload_id,
            filename=data.filename,
            total_chunks=data.total_chunks,
            category=data.category,
        )
    )
    
    return {
        "success": True,
        "status": "processing",
        "upload_id": upload_id,
        "message": "Upload received. Processing on the server — poll /api/admin/upload/status/{upload_id} for completion.",
    }


# ---- Status tracking for chunked-upload finalization ----
# In-memory map (fast path) backed by on-disk JSON (so a worker recycle or
# multi-worker setup still serves status correctly across requests).
_upload_status_mem: Dict[str, Dict[str, Any]] = {}
_UPLOAD_STATUS_DIR = "/tmp/gradnext_uploads/_status"
os.makedirs(_UPLOAD_STATUS_DIR, exist_ok=True)


def _write_upload_status(upload_id: str, status: Dict[str, Any]) -> None:
    import json
    _upload_status_mem[upload_id] = status
    try:
        path = os.path.join(_UPLOAD_STATUS_DIR, f"{upload_id}.json")
        with open(path, "w") as f:
            json.dump(status, f)
    except Exception as e:
        logger.warning(f"Failed to persist upload status for {upload_id}: {e}")


def _read_upload_status(upload_id: str) -> Optional[Dict[str, Any]]:
    import json
    if upload_id in _upload_status_mem:
        return _upload_status_mem[upload_id]
    try:
        path = os.path.join(_UPLOAD_STATUS_DIR, f"{upload_id}.json")
        if os.path.exists(path):
            with open(path) as f:
                return json.load(f)
    except Exception:
        pass
    return None


async def _run_finalize_in_background(upload_id: str, filename: str, total_chunks: int, category: str) -> None:
    """Background worker: combine chunks → push to cloud → record final URL."""
    import asyncio
    
    upload_temp_dir = "/tmp/gradnext_uploads"
    upload_dir = os.path.join(upload_temp_dir, upload_id)
    combined_path: Optional[str] = None
    
    def _combine_chunks_sync() -> tuple:
        """Concat chunk files into a single file. Sync; runs in thread pool."""
        chunk_files = sorted([f for f in os.listdir(upload_dir) if f.startswith("chunk_")])
        file_ext = os.path.splitext(filename)[1].lower() or ".bin"
        out_path = os.path.join(upload_temp_dir, f"{upload_id}_final{file_ext}")
        final_size = 0
        with open(out_path, "wb") as outfile:
            for cf in chunk_files:
                cp = os.path.join(upload_dir, cf)
                with open(cp, "rb") as infile:
                    while True:
                        block = infile.read(1024 * 1024)  # 1MB blocks for fast local I/O
                        if not block:
                            break
                        outfile.write(block)
                        final_size += len(block)
        return out_path, final_size, file_ext, len(chunk_files)
    
    try:
        # ----- Phase 1: Combine chunks on disk -----
        _write_upload_status(upload_id, {
            "state": "processing",
            "phase": "combining",
            "upload_id": upload_id,
            "filename": filename,
        })
        combined_path, final_size, file_ext, num_chunks = await asyncio.to_thread(_combine_chunks_sync)
        logger.info(f"[bg] Combined {num_chunks} chunks: {final_size} bytes -> {combined_path}")
        
        video_extensions = ['.mp4', '.webm', '.mov', '.avi', '.mkv', '.m4v', '.wmv', '.flv']
        is_video = file_ext in video_extensions
        
        # ----- Phase 2: Push to cloud storage -----
        _write_upload_status(upload_id, {
            "state": "processing",
            "phase": "uploading_to_cloud",
            "upload_id": upload_id,
            "filename": filename,
            "size": final_size,
        })
        
        from services import cloud_storage_service
        
        if cloud_storage_service.is_enabled():
            try:
                result = await asyncio.to_thread(
                    cloud_storage_service.upload_file_stream,
                    file_path=combined_path,
                    filename=filename,
                    folder=category,
                )
                file_url = f"/api/files/{result['storage_path']}"
                logger.info(f"[bg] Cloud upload done: {upload_id} -> {file_url} ({final_size} bytes)")
                _write_upload_status(upload_id, {
                    "state": "done",
                    "phase": "complete",
                    "upload_id": upload_id,
                    "success": True,
                    "url": file_url,
                    "filename": filename,
                    "storage_path": result['storage_path'],
                    "size": result.get('size', final_size),
                    "cloud_stored": True,
                    "is_video": is_video,
                    "message": "File uploaded to cloud storage successfully",
                    "completed_at": datetime.now(timezone.utc).isoformat(),
                })
                return
            except Exception as e:
                logger.exception(f"[bg] Cloud upload failed for {upload_id}: {e}; falling back to local")
        
        # ----- Fallback: local disk -----
        category_dir = os.path.join(UPLOAD_DIR, category)
        os.makedirs(category_dir, exist_ok=True)
        unique_filename = f"{uuid.uuid4().hex[:8]}{file_ext}"
        final_path = os.path.join(category_dir, unique_filename)
        shutil.move(combined_path, final_path)
        combined_path = None  # already moved
        file_url = f"/api/uploads/{category}/{unique_filename}"
        _write_upload_status(upload_id, {
            "state": "done",
            "phase": "complete",
            "upload_id": upload_id,
            "success": True,
            "url": file_url,
            "filename": filename,
            "size": final_size,
            "cloud_stored": False,
            "is_video": is_video,
            "message": "File uploaded locally",
            "completed_at": datetime.now(timezone.utc).isoformat(),
        })
    except Exception as e:
        logger.exception(f"[bg] Finalize failed for {upload_id}: {e}")
        _write_upload_status(upload_id, {
            "state": "failed",
            "upload_id": upload_id,
            "error": str(e),
            "completed_at": datetime.now(timezone.utc).isoformat(),
        })
    finally:
        # Cleanup temp chunks + combined file (always, even on failure)
        try:
            if os.path.exists(upload_dir):
                shutil.rmtree(upload_dir, ignore_errors=True)
        except Exception:
            pass
        try:
            if combined_path and os.path.exists(combined_path):
                os.unlink(combined_path)
        except Exception:
            pass


@router.get("/upload/status/{upload_id}")
async def get_upload_status(upload_id: str, request: Request):
    """Poll the current state of a chunked-upload finalization.
    
    Returns: { state: "processing" | "done" | "failed", phase, url?, error? }
    """
    await verify_admin(request)
    status = _read_upload_status(upload_id)
    if not status:
        raise HTTPException(status_code=404, detail="Upload status not found or expired")
    return status


# ============ Plan Management Endpoints ============

# Comprehensive Default Plans based on new structure
DEFAULT_PLANS = [
    # ========== FREE TRIAL ==========
    {
        "id": "plan-free-trial",
        "plan_key": "free_trial",
        "name": "Free Trial",
        "category": "subscription",
        "description": "7-day trial with limited Pro+ access",
        "pricing": {"one_month": 0, "six_month": 0, "one_time": 0},
        "currency": "INR",
        "duration_months": None,
        "duration_days": 7,
        "is_auto_renew": False,
        "features": {
            "course_recordings": True,
            "course_recordings_limited": True,
            "drills_exercises": True,
            "drills_limited": True,
            "case_materials": True,
            "case_materials_limited": True,
            "workshops": "only_recorded",
            "workshops_limited": True,
            "peer_to_peer": "1_only",
            "peer_sessions_per_month": 1,
            "coaching_sessions": 0,
            "strategy_calls": 0,
            "dedicated_coach": False
        },
        "display_features": ["Limited course access", "Practice drills", "Case materials"],
        "is_active": True,
        "is_hidden": False,
        "order": 0,
        "highlight": False,
        "badge": "Free",
        "application_only": False,
        "show_on_pages": ["home", "pricing"]
    },
    # ========== SUBSCRIPTION PLANS ==========
    {
        "id": "plan-basic",
        "plan_key": "basic_plan",
        "name": "Basic Plan",
        "category": "subscription",
        "description": "Essential access to course recordings and materials",
        "pricing": {"one_month": 499, "six_month": 399, "one_time": None},
        "currency": "INR",
        "duration_months": 1,
        "is_auto_renew": True,
        "features": {
            "course_recordings": True,
            "course_recordings_limited": False,
            "drills_exercises": True,
            "drills_limited": False,
            "case_materials": True,
            "case_materials_limited": False,
            "workshops": "only_recorded",
            "workshops_limited": False,
            "peer_to_peer": "none",
            "peer_sessions_per_month": 0,
            "coaching_sessions": 0,
            "strategy_calls": 0,
            "dedicated_coach": False,
            "industry_primers": False,
            "knowledge_sessions": False
        },
        "display_features": ["Full course access", "Drills & exercises", "Case materials", "Recorded workshops"],
        "is_active": True,
        "is_hidden": False,
        "order": 1,
        "highlight": False,
        "badge": None,
        "application_only": False,
        "show_on_pages": ["home", "pricing"]
    },
    {
        "id": "plan-pro",
        "plan_key": "pro_plan",
        "name": "Pro Plan",
        "category": "subscription",
        "description": "Full access with live workshops and peer practice",
        "pricing": {"one_month": 699, "six_month": 599, "one_time": None},
        "currency": "INR",
        "duration_months": 1,
        "is_auto_renew": True,
        "features": {
            "course_recordings": True,
            "course_recordings_limited": False,
            "drills_exercises": True,
            "drills_limited": False,
            "case_materials": True,
            "case_materials_limited": False,
            "workshops": "recorded_and_live",
            "workshops_limited": False,
            "peer_to_peer": "1_per_week",
            "peer_sessions_per_month": 4,
            "coaching_sessions": 0,
            "strategy_calls": 0,
            "dedicated_coach": False,
            "industry_primers": True,
            "knowledge_sessions": True
        },
        "display_features": ["Full course access", "Drills & exercises", "Case materials", "Live + recorded workshops", "4 peer practice sessions/month"],
        "is_active": True,
        "is_hidden": False,
        "order": 2,
        "highlight": True,
        "badge": "Most Popular",
        "application_only": False,
        "show_on_pages": ["home", "pricing"]
    },
    {
        "id": "plan-pro-plus",
        "plan_key": "pro_plus",
        "name": "Pro+",
        "category": "subscription",
        "description": "Premium subscription with unlimited peer practice",
        "pricing": {"one_month": 1299, "six_month": 999, "one_time": None},
        "currency": "INR",
        "duration_months": 1,
        "is_auto_renew": True,
        "features": {
            "course_recordings": True,
            "course_recordings_limited": False,
            "drills_exercises": True,
            "drills_limited": False,
            "case_materials": True,
            "case_materials_limited": False,
            "workshops": "recorded_and_live",
            "workshops_limited": False,
            "peer_to_peer": "unlimited",
            "peer_sessions_per_month": 8,
            "coaching_sessions": 0,
            "strategy_calls": 0,
            "dedicated_coach": False,
            "industry_primers": True,
            "knowledge_sessions": True
        },
        "display_features": ["Full course access", "Drills & exercises", "Case materials", "Live + recorded workshops", "Unlimited peer practice"],
        "is_active": True,
        "is_hidden": False,
        "order": 3,
        "highlight": False,
        "badge": "Best Value",
        "application_only": False,
        "show_on_pages": ["home", "pricing"]
    },
    # ========== COACHING PLANS ==========
    {
        "id": "plan-single-session",
        "plan_key": "single_session",
        "name": "Single Session",
        "category": "coaching",
        "description": "One coaching session with an MBB consultant",
        "pricing": {"one_month": None, "six_month": None, "one_time": 2999},
        "currency": "INR",
        "duration_months": None,
        "is_auto_renew": False,
        "features": {
            "course_recordings": False,
            "course_recordings_limited": False,
            "drills_exercises": False,
            "drills_limited": False,
            "case_materials": False,
            "case_materials_limited": False,
            "workshops": "none",
            "workshops_limited": False,
            "peer_to_peer": "none",
            "coaching_sessions": 1,
            "strategy_calls": 0,
            "dedicated_coach": False
        },
        "display_features": ["1 coaching session", "MBB consultant", "45-min session", "Personalized feedback"],
        "is_active": True,
        "is_hidden": False,
        "order": 10,
        "highlight": False,
        "badge": None,
        "application_only": False,
        "show_on_pages": ["coaching"]
    },
    {
        "id": "plan-last-mile",
        "plan_key": "last_mile",
        "name": "Last Mile",
        "category": "coaching",
        "description": "2-month intensive with 5 coaching sessions",
        "pricing": {"one_month": None, "six_month": None, "one_time": 16999},
        "currency": "INR",
        "duration_months": 2,
        "is_auto_renew": False,
        "features": {
            "course_recordings": True,
            "course_recordings_limited": False,
            "drills_exercises": True,
            "drills_limited": False,
            "case_materials": True,
            "case_materials_limited": False,
            "workshops": "only_recorded",
            "workshops_limited": False,
            "peer_to_peer": "none",
            "coaching_sessions": 5,
            "strategy_calls": 1,
            "dedicated_coach": False
        },
        "is_active": True,
        "is_hidden": False,
        "order": 11,
        "highlight": False,
        "badge": None,
        "application_only": False,
        "show_on_pages": ["coaching"]
    },
    {
        "id": "plan-mid-mile",
        "plan_key": "mid_mile",
        "name": "Mid Mile",
        "category": "coaching",
        "description": "3-month program with 10 coaching sessions",
        "pricing": {"one_month": None, "six_month": None, "one_time": 31999},
        "currency": "INR",
        "duration_months": 3,
        "is_auto_renew": False,
        "features": {
            "course_recordings": True,
            "course_recordings_limited": False,
            "drills_exercises": True,
            "drills_limited": False,
            "case_materials": True,
            "case_materials_limited": False,
            "workshops": "recorded_and_live",
            "workshops_limited": False,
            "peer_to_peer": "1_per_week",
            "coaching_sessions": 10,
            "strategy_calls": 2,
            "dedicated_coach": False
        },
        "is_active": True,
        "is_hidden": False,
        "order": 12,
        "highlight": True,
        "badge": "Most Popular",
        "application_only": False,
        "show_on_pages": ["coaching"]
    },
    {
        "id": "plan-full-prep",
        "plan_key": "full_prep",
        "name": "Full Prep",
        "category": "coaching",
        "description": "6-month comprehensive program with dedicated coach",
        "pricing": {"one_month": None, "six_month": None, "one_time": 44999},
        "currency": "INR",
        "duration_months": 6,
        "is_auto_renew": False,
        "features": {
            "course_recordings": True,
            "course_recordings_limited": False,
            "drills_exercises": True,
            "drills_limited": False,
            "case_materials": True,
            "case_materials_limited": False,
            "workshops": "recorded_and_live",
            "workshops_limited": False,
            "peer_to_peer": "unlimited",
            "coaching_sessions": 15,
            "strategy_calls": 3,
            "dedicated_coach": True
        },
        "is_active": True,
        "is_hidden": False,
        "order": 13,
        "highlight": False,
        "badge": "Best Value",
        "application_only": False,
        "show_on_pages": ["coaching"]
    },
    {
        "id": "plan-pinnacle",
        "plan_key": "pinnacle",
        "name": "Pinnacle",
        "category": "coaching",
        "description": "6-month elite program with unlimited coaching",
        "pricing": {"one_month": None, "six_month": None, "one_time": None},
        "currency": "INR",
        "duration_months": 6,
        "is_auto_renew": False,
        "features": {
            "course_recordings": True,
            "course_recordings_limited": False,
            "drills_exercises": True,
            "drills_limited": False,
            "case_materials": True,
            "case_materials_limited": False,
            "workshops": "recorded_and_live",
            "workshops_limited": False,
            "peer_to_peer": "unlimited",
            "coaching_sessions": -1,
            "strategy_calls": -1,
            "dedicated_coach": True
        },
        "is_active": True,
        "is_hidden": False,
        "order": 14,
        "highlight": False,
        "badge": "Elite",
        "application_only": True,
        "show_on_pages": ["coaching"]
    },
    # ========== COHORT PLANS ==========
    {
        "id": "plan-cohort-premium",
        "plan_key": "cohort_premium",
        "name": "Cohort Premium",
        "category": "cohort",
        "description": "2-month cohort program with group sessions",
        "pricing": {"one_month": None, "six_month": None, "one_time": 12999},
        "currency": "INR",
        "duration_months": 2,
        "is_auto_renew": False,
        "features": {
            "course_recordings": False,
            "course_recordings_limited": False,
            "drills_exercises": True,
            "drills_limited": False,
            "case_materials": True,
            "case_materials_limited": False,
            "workshops": "only_recorded",
            "workshops_limited": False,
            "peer_to_peer": "2_per_week",
            "coaching_sessions": 1,
            "strategy_calls": 0,
            "dedicated_coach": False
        },
        "is_active": True,
        "is_hidden": False,
        "order": 20,
        "highlight": False,
        "badge": None,
        "application_only": False,
        "show_on_pages": ["cohort"]
    },
    {
        "id": "plan-cohort-elite",
        "plan_key": "cohort_elite",
        "name": "Cohort Elite",
        "category": "cohort",
        "description": "2-month elite cohort with live workshops",
        "pricing": {"one_month": None, "six_month": None, "one_time": 19999},
        "currency": "INR",
        "duration_months": 2,
        "is_auto_renew": False,
        "features": {
            "course_recordings": False,
            "course_recordings_limited": False,
            "drills_exercises": True,
            "drills_limited": False,
            "case_materials": True,
            "case_materials_limited": False,
            "workshops": "recorded_and_live",
            "workshops_limited": False,
            "peer_to_peer": "2_per_week",
            "coaching_sessions": 3,
            "strategy_calls": 0,
            "dedicated_coach": False
        },
        "is_active": True,
        "is_hidden": False,
        "order": 21,
        "highlight": True,
        "badge": "Popular",
        "application_only": False,
        "show_on_pages": ["cohort"]
    },
    # ========== ADD-ONS ==========
    {
        "id": "addon-peer-session",
        "plan_key": "addon_peer_session",
        "name": "Peer-to-Peer Sessions",
        "category": "addon",
        "description": "Unlimited peer practice sessions - auto-included with subscriptions",
        "pricing": {"one_month": 199, "six_month": 149, "one_time": None},
        "currency": "INR",
        "duration_months": 1,
        "is_auto_renew": True,
        "features": {
            "course_recordings": False,
            "course_recordings_limited": False,
            "drills_exercises": False,
            "drills_limited": False,
            "case_materials": False,
            "case_materials_limited": False,
            "workshops": "none",
            "workshops_limited": False,
            "peer_to_peer": "unlimited",
            "coaching_sessions": 0,
            "strategy_calls": 0,
            "dedicated_coach": False
        },
        "is_active": True,
        "is_hidden": False,
        "order": 30,
        "highlight": False,
        "badge": "Add-on",
        "application_only": False,
        "requires_base_plan": True,
        "auto_add_to_subscription": True,
        "show_on_pages": ["home", "pricing"]
    },
    {
        "id": "addon-strategy-call",
        "plan_key": "addon_strategy_call",
        "name": "Strategy Call",
        "category": "addon",
        "description": "30-minute strategy session with mentor",
        "pricing": {"one_month": None, "six_month": None, "one_time": 1199},
        "currency": "INR",
        "duration_months": None,
        "is_auto_renew": False,
        "features": {
            "course_recordings": False,
            "course_recordings_limited": False,
            "drills_exercises": False,
            "drills_limited": False,
            "case_materials": False,
            "case_materials_limited": False,
            "workshops": "none",
            "workshops_limited": False,
            "peer_to_peer": "none",
            "coaching_sessions": 0,
            "strategy_calls": 1,
            "dedicated_coach": False
        },
        "is_active": True,
        "is_hidden": False,
        "order": 31,
        "highlight": False,
        "badge": "Add-on",
        "application_only": False,
        "requires_base_plan": True,
        "show_on_pages": ["pricing"]
    },
    {
        "id": "addon-live-workshop",
        "plan_key": "addon_live_workshop",
        "name": "Live Workshop",
        "category": "addon",
        "description": "Access to one live workshop",
        "pricing": {"one_month": None, "six_month": None, "one_time": 199},
        "currency": "INR",
        "duration_months": None,
        "is_auto_renew": False,
        "features": {
            "course_recordings": False,
            "course_recordings_limited": False,
            "drills_exercises": False,
            "drills_limited": False,
            "case_materials": False,
            "case_materials_limited": False,
            "workshops": "recorded_and_live",
            "workshops_limited": False,
            "peer_to_peer": "none",
            "coaching_sessions": 0,
            "strategy_calls": 0,
            "dedicated_coach": False
        },
        "is_active": True,
        "is_hidden": False,
        "order": 32,
        "highlight": False,
        "badge": "Add-on",
        "application_only": False,
        "requires_base_plan": True,
        "show_on_pages": ["pricing"]
    }
]


def normalize_plan_schema(plan: dict) -> dict:
    """Normalize plan to new comprehensive schema"""
    # If plan already has new schema, return as-is
    if "pricing" in plan and isinstance(plan.get("pricing"), dict) and "one_month" in plan.get("pricing", {}):
        # Still add new fields if missing
        if "show_on_pages" not in plan:
            plan["show_on_pages"] = ["home"]
        if "auto_add_to_subscription" not in plan:
            plan["auto_add_to_subscription"] = False
        if "requires_base_plan" not in plan:
            plan["requires_base_plan"] = False
        return plan
    
    # Convert old schema to new
    old_price = plan.get("price", 0)
    old_features = plan.get("features", {})
    
    normalized = {
        **plan,
        "category": plan.get("category", "subscription"),
        "pricing": plan.get("pricing") if isinstance(plan.get("pricing"), dict) else {
            "one_month": old_price if old_price else None,
            "six_month": None,
            "one_time": None
        },
        "duration_months": plan.get("duration_months"),
        "duration_days": plan.get("duration_days"),
        "is_auto_renew": plan.get("is_auto_renew", plan.get("is_subscription", False)),
        "features": {
            "course_recordings": old_features.get("courses", old_features.get("course_recordings", True)),
            "course_recordings_limited": old_features.get("course_recordings_limited", False),
            "drills_exercises": old_features.get("drills", old_features.get("drills_exercises", True)),
            "drills_limited": old_features.get("drills_limited", False),
            "case_materials": old_features.get("materials", old_features.get("case_materials", True)),
            "case_materials_limited": old_features.get("case_materials_limited", False),
            "workshops": old_features.get("workshops") if isinstance(old_features.get("workshops"), str) else ("only_recorded" if old_features.get("workshops") else "none"),
            "workshops_limited": old_features.get("workshops_limited", False),
            "peer_to_peer": old_features.get("peer_to_peer", "1_per_week" if old_features.get("peer_practice") else "none"),
            "coaching_sessions": plan.get("coaching_sessions", old_features.get("coaching_sessions", 0)),
            "strategy_calls": old_features.get("strategy_calls", 0),
            "dedicated_coach": old_features.get("dedicated_coach", False)
        },
        "display_features": plan.get("display_features", []),
        "is_active": plan.get("is_active", True),
        "is_hidden": plan.get("is_hidden", False),
        "order": plan.get("order", 0),
        "highlight": plan.get("highlight", False),
        "badge": plan.get("badge"),
        "application_only": plan.get("application_only", False),
        "show_on_pages": plan.get("show_on_pages", ["home"]),
        "auto_add_to_subscription": plan.get("auto_add_to_subscription", False),
        "requires_base_plan": plan.get("requires_base_plan", False)
    }
    
    # Remove old fields
    normalized.pop("price", None)
    normalized.pop("is_subscription", None)
    
    return normalized


@router.post("/plans/reset-to-defaults")
async def reset_plans_to_defaults(request: Request):
    """Reset all plans to the default comprehensive schema"""
    await verify_admin(request)
    db = get_db(request)
    
    # Delete all existing plans
    await db.plans.delete_many({})
    
    # Insert fresh default plans
    for plan in DEFAULT_PLANS:
        plan_copy = plan.copy()
        plan_copy["created_at"] = datetime.utcnow().isoformat()
        plan_copy["updated_at"] = datetime.utcnow().isoformat()
        await db.plans.insert_one(plan_copy)
    
    return {"message": "Plans reset to defaults", "count": len(DEFAULT_PLANS)}


@router.post("/plans/cleanup-duplicates")
async def cleanup_duplicate_plans(request: Request):
    """Remove duplicate plans, keeping only the most recent one for each plan_key"""
    await verify_admin(request)
    db = get_db(request)
    
    # Find all unique plan_keys
    pipeline = [
        {"$group": {
            "_id": "$plan_key",
            "count": {"$sum": 1},
            "ids": {"$push": "$id"}
        }},
        {"$match": {"count": {"$gt": 1}}}
    ]
    
    duplicates = await db.plans.aggregate(pipeline).to_list(100)
    removed_count = 0
    
    for dup in duplicates:
        plan_key = dup["_id"]
        ids = dup["ids"]
        # Keep the first one (oldest), delete the rest
        ids_to_delete = ids[1:]  # Skip the first ID
        for plan_id in ids_to_delete:
            await db.plans.delete_one({"id": plan_id})
            removed_count += 1
    
    return {"message": f"Removed {removed_count} duplicate plans", "duplicates_found": len(duplicates)}


@router.post("/plans/migrate-schema")
async def migrate_plans_schema(request: Request):
    """Migrate existing plans to new schema without losing data"""
    await verify_admin(request)
    db = get_db(request)
    
    plans = await db.plans.find({}, {"_id": 0}).to_list(100)
    migrated_count = 0
    
    for plan in plans:
        # Check if migration needed
        if "pricing" not in plan or not isinstance(plan.get("pricing"), dict):
            normalized = normalize_plan_schema(plan)
            normalized["updated_at"] = datetime.utcnow().isoformat()
            
            await db.plans.update_one(
                {"id": plan["id"]},
                {"$set": normalized}
            )
            migrated_count += 1
    
    return {"message": f"Migrated {migrated_count} plans to new schema", "total_plans": len(plans)}


@router.get("/plans")
async def get_all_plans(request: Request, category: Optional[str] = None):
    """Get all plans for admin management, optionally filtered by category"""
    await verify_admin(request)
    db = get_db(request)
    
    # Build query
    query = {}
    if category:
        query["category"] = category
    
    # Check if plans collection exists and has data
    plans = await db.plans.find(query, {"_id": 0}).sort("order", 1).to_list(100)
    
    # Seed default plans if none exist (with duplicate prevention)
    if not plans and not category:
        for plan in DEFAULT_PLANS:
            # Check if plan already exists by plan_key
            existing = await db.plans.find_one({"plan_key": plan.get("plan_key")})
            if not existing:
                plan_copy = plan.copy()
                plan_copy["created_at"] = datetime.utcnow().isoformat()
                plan_copy["updated_at"] = datetime.utcnow().isoformat()
                await db.plans.insert_one(plan_copy)
        # Re-fetch all plans after seeding
        plans = await db.plans.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    
    # Normalize all plans to new schema (for backwards compatibility)
    normalized_plans = [normalize_plan_schema(p) for p in plans]
    
    # Group plans by category for easier frontend handling
    grouped = {
        "subscription": [],
        "coaching": [],
        "cohort": [],
        "addon": []
    }
    
    for plan in normalized_plans:
        cat = plan.get("category", "subscription")
        if cat in grouped:
            grouped[cat].append(plan)
    
    return {"plans": normalized_plans, "grouped": grouped}


@router.post("/plans")
async def create_plan(data: PlanCreate, request: Request):
    """Create a new plan"""
    await verify_admin(request)
    db = get_db(request)
    
    # Check if plan_key already exists
    existing = await db.plans.find_one({"plan_key": data.plan_key})
    if existing:
        raise HTTPException(status_code=400, detail="Plan with this key already exists")
    
    plan_id = f"plan-{str(uuid.uuid4())[:8]}"
    
    plan_data = {
        "id": plan_id,
        "plan_key": data.plan_key,
        "name": data.name,
        "category": data.category,
        "description": data.description,
        "pricing": data.pricing.dict() if hasattr(data.pricing, 'dict') else data.pricing,
        "currency": data.currency,
        "duration_months": data.duration_months,
        "is_auto_renew": data.is_auto_renew,
        "features": data.features.dict() if hasattr(data.features, 'dict') else data.features,
        "display_features": data.display_features,
        "is_active": data.is_active,
        "is_hidden": data.is_hidden,
        "order": data.order,
        "highlight": data.highlight,
        "badge": data.badge,
        "application_only": data.application_only,
        "show_on_pages": data.show_on_pages,
        "created_at": datetime.utcnow().isoformat(),
        "updated_at": datetime.utcnow().isoformat()
    }
    
    await db.plans.insert_one(plan_data)
    
    # Remove _id from response (MongoDB adds it during insert)
    plan_data.pop('_id', None)
    
    return {"message": "Plan created successfully", "plan_id": plan_id, "plan": plan_data}


@router.put("/plans/{plan_id}")
async def update_plan(plan_id: str, data: PlanUpdate, request: Request):
    """Update an existing plan"""
    await verify_admin(request)
    db = get_db(request)
    
    # Build update dict
    update_data = {"updated_at": datetime.utcnow().isoformat()}
    
    for field, value in data.dict(exclude_unset=True).items():
        if value is not None:
            update_data[field] = value
    
    # Try to find by id first, then by plan_key
    result = await db.plans.update_one(
        {"$or": [{"id": plan_id}, {"plan_key": plan_id}]},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    # Get updated plan
    updated_plan = await db.plans.find_one(
        {"$or": [{"id": plan_id}, {"plan_key": plan_id}]}, 
        {"_id": 0}
    )
    
    return {"message": "Plan updated successfully", "plan": updated_plan}


@router.delete("/plans/{plan_id}")
async def delete_plan(plan_id: str, request: Request):
    """Delete a plan (soft delete by setting is_active to False)"""
    await verify_admin(request)
    db = get_db(request)
    
    # Try to find by id first, then by plan_key
    plan = await db.plans.find_one(
        {"$or": [{"id": plan_id}, {"plan_key": plan_id}]}, 
        {"_id": 0}
    )
    
    if plan:
        users_on_plan = await db.users.count_documents({"plan": plan.get("plan_key")})
        if users_on_plan > 0:
            # Soft delete - just deactivate
            await db.plans.update_one(
                {"$or": [{"id": plan_id}, {"plan_key": plan_id}]},
                {"$set": {"is_active": False, "updated_at": datetime.utcnow().isoformat()}}
            )
            return {"message": f"Plan deactivated. {users_on_plan} users are still on this plan.", "soft_delete": True}
    
    # Hard delete if no users
    result = await db.plans.delete_one({"$or": [{"id": plan_id}, {"plan_key": plan_id}]})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    return {"message": "Plan deleted successfully", "soft_delete": False}


@router.get("/plans/stats")
async def get_plan_stats(request: Request):
    """Get statistics for all plans"""
    await verify_admin(request)
    db = get_db(request)
    
    plans = await db.plans.find({}, {"_id": 0}).to_list(100)
    
    stats = []
    for plan in plans:
        user_count = await db.users.count_documents({"plan": plan.get("plan_key")})
        assignment_count = await db.users.count_documents({
            "plan_assignments.plan_key": plan.get("plan_key"),
            "plan_assignments.is_active": True
        })
        
        stats.append({
            "plan_id": plan.get("id"),
            "plan_key": plan.get("plan_key"),
            "name": plan.get("name"),
            "category": plan.get("category"),
            "primary_users": user_count,
            "total_assignments": assignment_count,
            "is_active": plan.get("is_active", True),
            "is_hidden": plan.get("is_hidden", False)
        })
    
    return {"stats": stats}


@router.get("/plans/{plan_id}")
async def get_plan(plan_id: str, request: Request):
    """Get a single plan by ID"""
    await verify_admin(request)
    db = get_db(request)
    
    plan = await db.plans.find_one({"id": plan_id}, {"_id": 0})
    
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    return plan


@router.post("/plans/{plan_id}/duplicate")
async def duplicate_plan(plan_id: str, request: Request):
    """Duplicate an existing plan"""
    await verify_admin(request)
    db = get_db(request)
    
    # Get original plan
    original = await db.plans.find_one({"id": plan_id}, {"_id": 0})
    if not original:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    # Create new plan with modified details
    new_plan_id = f"plan-{str(uuid.uuid4())[:8]}"
    new_plan_key = f"{original['plan_key']}_copy"
    
    # Check for duplicate key and increment
    counter = 1
    while await db.plans.find_one({"plan_key": new_plan_key}):
        new_plan_key = f"{original['plan_key']}_copy_{counter}"
        counter += 1
    
    new_plan = {
        **original,
        "id": new_plan_id,
        "plan_key": new_plan_key,
        "name": f"{original['name']} (Copy)",
        "is_active": False,  # Start as inactive
        "created_at": datetime.utcnow().isoformat(),
        "updated_at": datetime.utcnow().isoformat()
    }
    
    await db.plans.insert_one(new_plan)
    
    # Remove _id from response (MongoDB adds it during insert)
    new_plan.pop('_id', None)
    
    return {"message": "Plan duplicated successfully", "plan_id": new_plan_id, "plan": new_plan}


@router.post("/plans/reorder")
async def reorder_plans(request: Request):
    """Reorder plans by updating their order values"""
    await verify_admin(request)
    db = get_db(request)
    
    body = await request.json()
    plan_orders = body.get("plans", [])  # [{id: "plan-xxx", order: 0}, ...]
    
    for item in plan_orders:
        await db.plans.update_one(
            {"id": item["id"]},
            {"$set": {"order": item["order"], "updated_at": datetime.utcnow().isoformat()}}
        )
    
    return {"message": "Plans reordered successfully"}


# ========== User Plan Assignment Endpoints ==========

@router.post("/users/{user_id}/assign-plan")
async def assign_plan_to_user(user_id: str, request: Request):
    """Assign a plan to a user (can have multiple plans)"""
    await verify_admin(request)
    db = get_db(request)
    
    body = await request.json()
    plan_key = body.get("plan_key")
    start_date = body.get("start_date")
    end_date = body.get("end_date")
    is_trial = body.get("is_trial", False)
    coaching_sessions = body.get("coaching_sessions_granted", 0)
    strategy_calls = body.get("strategy_calls_granted", 0)
    
    # Get plan details
    plan = await db.plans.find_one({"plan_key": plan_key}, {"_id": 0})
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    # Get user
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Create plan assignment
    assignment = {
        "id": f"assign-{str(uuid.uuid4())[:8]}",
        "user_id": user_id,
        "plan_key": plan_key,
        "plan_name": plan.get("name"),
        "category": plan.get("category"),
        "start_date": start_date or datetime.utcnow().isoformat(),
        "end_date": end_date,
        "is_trial": is_trial,
        "is_active": True,
        "coaching_sessions_granted": coaching_sessions or plan.get("features", {}).get("coaching_sessions", 0),
        "coaching_sessions_used": 0,
        "strategy_calls_granted": strategy_calls or plan.get("features", {}).get("strategy_calls", 0),
        "strategy_calls_used": 0,
        "created_at": datetime.utcnow().isoformat()
    }
    
    # Add to user's plan_assignments array
    await db.users.update_one(
        {"id": user_id},
        {
            "$push": {"plan_assignments": assignment},
            "$set": {
                "plan": plan_key,  # Set primary plan
                "updated_at": datetime.utcnow().isoformat()
            }
        }
    )
    
    return {"message": "Plan assigned successfully", "assignment": assignment}


@router.get("/users/{user_id}/plans")
async def get_user_plans(user_id: str, request: Request):
    """Get all plans assigned to a user"""
    await verify_admin(request)
    db = get_db(request)
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    assignments = user.get("plan_assignments", [])
    
    return {
        "user_id": user_id,
        "primary_plan": user.get("plan"),
        "plan_assignments": assignments
    }


@router.delete("/users/{user_id}/plans/{assignment_id}")
async def remove_user_plan(user_id: str, assignment_id: str, request: Request):
    """Remove a plan assignment from a user"""
    await verify_admin(request)
    db = get_db(request)
    
    result = await db.users.update_one(
        {"id": user_id},
        {
            "$pull": {"plan_assignments": {"id": assignment_id}},
            "$set": {"updated_at": datetime.utcnow().isoformat()}
        }
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Assignment not found")
    
    return {"message": "Plan assignment removed successfully"}


# ============ Session Tracking ============

@router.get("/sessions")
async def get_all_sessions(
    request: Request,
    mentor_id: Optional[str] = None,
    candidate_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    status: Optional[str] = None,
    completion_status: Optional[str] = None,
    page: int = 1,
    limit: int = 50
):
    """
    Get all sessions with filters for admin tracking.
    Includes mentor feedback status and candidate feedback status.
    """
    await verify_admin(request)
    db = get_db(request)
    
    # Build query
    query = {}
    
    if mentor_id:
        query["mentor_id"] = mentor_id
    
    if candidate_id:
        query["user_id"] = candidate_id
    
    if date_from:
        query["date"] = {"$gte": date_from}
    
    if date_to:
        if "date" in query:
            query["date"]["$lte"] = date_to
        else:
            query["date"] = {"$lte": date_to}
    
    if status:
        query["status"] = status
    
    if completion_status:
        if completion_status == "pending":
            query["completion_status"] = {"$exists": False}
        else:
            query["completion_status"] = completion_status
    
    # Count total
    total = await db.bookings.count_documents(query)
    
    # Get sessions with pagination
    skip = (page - 1) * limit
    sessions_cursor = db.bookings.find(query, {"_id": 0}).sort("date", -1).skip(skip).limit(limit)
    sessions = await sessions_cursor.to_list(limit)
    
    # Enrich with mentor and candidate info
    result = []
    for session in sessions:
        # Get mentor info
        mentor = await db.mentors.find_one({"id": session.get("mentor_id")}, {"_id": 0, "name": 1, "email": 1})
        
        # Get candidate info
        candidate = await db.users.find_one({"id": session.get("user_id")}, {"_id": 0, "name": 1, "email": 1})
        
        # Check for mentor feedback (feedback given BY mentor TO candidate)
        mentor_feedback = await db.mentor_feedbacks.find_one({
            "booking_id": session.get("id")
        }, {"_id": 0})
        
        # Check for candidate feedback (feedback given BY candidate TO mentor)
        candidate_feedback = await db.feedbacks.find_one({
            "booking_id": session.get("id")
        }, {"_id": 0})
        
        result.append({
            "id": session.get("id"),
            "date": session.get("date"),
            "time_slot": session.get("time_slot"),
            "status": session.get("status", "confirmed"),
            "session_type": session.get("session_type", "Coaching Session"),
            # Mentor info
            "mentor_id": session.get("mentor_id"),
            "mentor_name": mentor.get("name") if mentor else session.get("mentor_name", "Unknown"),
            "mentor_email": mentor.get("email") if mentor else None,
            # Candidate info
            "candidate_id": session.get("user_id"),
            "candidate_name": candidate.get("name") if candidate else session.get("candidate_name", "Unknown"),
            "candidate_email": candidate.get("email") if candidate else None,
            # Check-in tracking
            "mentor_checked_in": session.get("mentor_checked_in", False),
            "mentor_checked_in_at": session.get("mentor_checked_in_at"),
            "candidate_checked_in": session.get("candidate_checked_in", False),
            "candidate_checked_in_at": session.get("candidate_checked_in_at"),
            # Completion status
            "completion_status": session.get("completion_status"),
            "completion_notes": session.get("completion_notes"),
            "completion_marked_at": session.get("completion_marked_at"),
            # Feedback status
            "mentor_feedback_given": mentor_feedback is not None,
            "mentor_feedback_rating": mentor_feedback.get("rating") if mentor_feedback else None,
            "candidate_feedback_given": candidate_feedback is not None,
            "candidate_feedback_rating": candidate_feedback.get("rating") if candidate_feedback else None,
            # Reschedule info
            "was_rescheduled": session.get("was_rescheduled", False),
            "previous_date": session.get("previous_date"),
            "previous_time_slot": session.get("previous_time_slot"),
            # Timestamps
            "created_at": session.get("created_at"),
            "updated_at": session.get("updated_at")
        })
    
    return {
        "sessions": result,
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": (total + limit - 1) // limit
    }


@router.get("/sessions/stats")
async def get_session_stats(request: Request):
    """Get session statistics for admin dashboard"""
    await verify_admin(request)
    db = get_db(request)
    
    # Get today's date
    today = datetime.utcnow().strftime("%Y-%m-%d")
    
    # Total sessions
    total_sessions = await db.bookings.count_documents({})
    
    # Sessions by status
    confirmed = await db.bookings.count_documents({"status": "confirmed"})
    completed = await db.bookings.count_documents({"status": "completed"})
    cancelled = await db.bookings.count_documents({"status": "cancelled"})
    no_show = await db.bookings.count_documents({"status": "no_show"})
    
    # Sessions today
    sessions_today = await db.bookings.count_documents({"date": today})
    
    # Sessions pending confirmation (past sessions without completion_status)
    pending_confirmation = await db.bookings.count_documents({
        "date": {"$lt": today},
        "status": {"$in": ["confirmed", "pending"]},
        "completion_status": {"$exists": False}
    })
    
    # Sessions with both feedbacks
    sessions_with_mentor_feedback = await db.mentor_feedbacks.count_documents({})
    sessions_with_candidate_feedback = await db.feedbacks.count_documents({"booking_id": {"$exists": True}})
    
    return {
        "total_sessions": total_sessions,
        "sessions_today": sessions_today,
        "by_status": {
            "confirmed": confirmed,
            "completed": completed,
            "cancelled": cancelled,
            "no_show": no_show
        },
        "pending_confirmation": pending_confirmation,
        "feedback_stats": {
            "mentor_feedbacks_given": sessions_with_mentor_feedback,
            "candidate_feedbacks_given": sessions_with_candidate_feedback
        }
    }


@router.get("/sessions/mentors-list")
async def get_mentors_for_filter(request: Request):
    """Get list of mentors for session filter dropdown"""
    await verify_admin(request)
    db = get_db(request)
    
    mentors = await db.mentors.find({}, {"_id": 0, "id": 1, "name": 1, "email": 1}).to_list(100)
    return mentors


@router.get("/sessions/candidates-list")
async def get_candidates_for_filter(request: Request):
    """Get list of candidates who have booked sessions for filter dropdown"""
    await verify_admin(request)
    db = get_db(request)
    
    # Get unique user IDs from bookings
    pipeline = [
        {"$group": {"_id": "$user_id"}},
        {"$limit": 200}
    ]
    user_ids = await db.bookings.aggregate(pipeline).to_list(200)
    user_ids = [u["_id"] for u in user_ids if u["_id"]]
    
    # Get user details
    candidates = await db.users.find(
        {"id": {"$in": user_ids}},
        {"_id": 0, "id": 1, "name": 1, "email": 1}
    ).to_list(200)
    
    return candidates


@router.get("/sessions/{session_id}")
async def get_session_details(session_id: str, request: Request):
    """Get detailed information about a specific session"""
    await verify_admin(request)
    db = get_db(request)
    
    session = await db.bookings.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Get mentor details
    mentor = await db.mentors.find_one({"id": session.get("mentor_id")}, {"_id": 0})
    
    # Get candidate details
    candidate = await db.users.find_one({"id": session.get("user_id")}, {"_id": 0})
    
    # Get mentor feedback
    mentor_feedback = await db.mentor_feedbacks.find_one({"booking_id": session_id}, {"_id": 0})
    
    # Get candidate feedback
    candidate_feedback = await db.feedbacks.find_one({"booking_id": session_id}, {"_id": 0})
    
    return {
        "session": session,
        "mentor": {
            "id": mentor.get("id") if mentor else None,
            "name": mentor.get("name") if mentor else None,
            "email": mentor.get("email") if mentor else None,
            "company": mentor.get("company") if mentor else None
        } if mentor else None,
        "candidate": {
            "id": candidate.get("id") if candidate else None,
            "name": candidate.get("name") if candidate else None,
            "email": candidate.get("email") if candidate else None,
            "plan": candidate.get("plan") if candidate else None
        } if candidate else None,
        "mentor_feedback": mentor_feedback,
        "candidate_feedback": candidate_feedback
    }


@router.post("/sessions/{session_id}/mark-status")
async def admin_mark_session_status(
    session_id: str,
    request: Request,
    status: str = None,
    notes: str = None
):
    """
    Admin endpoint to mark session completion status.
    Status options: completed, no_show_candidate, no_show_mentor, cancelled
    """
    await verify_admin(request)
    db = get_db(request)
    
    # Validate status
    valid_statuses = [
        "completed", 
        "no_show_candidate", 
        "no_show_mentor", 
        "cancelled_by_candidate",
        "cancelled_by_mentor",
        "cancelled_by_admin",  # NEW: Admin cancelled
        "cancelled"  # Keep for backward compatibility
    ]
    
    # Get request body
    body = await request.json()
    status = body.get("status")
    notes = body.get("notes", "")
    
    if status not in valid_statuses:
        raise HTTPException(
            status_code=400, 
            detail=f"Invalid status. Must be one of: {valid_statuses}"
        )
    
    # Get the session
    session = await db.bookings.find_one({"id": session_id})
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Update session
    update_data = {
        "completion_status": status,
        "completion_marked_at": datetime.utcnow().isoformat(),
        "completion_marked_by": "admin",
        "admin_notes": notes
    }
    
    # Update booking status based on completion and check-in data
    if status == "completed":
        # Get the booking to check actual check-in data
        booking = await db.bookings.find_one({"id": session_id})
        if booking:
            mentor_checked_in = booking.get("mentor_checked_in", False)
            candidate_checked_in = booking.get("candidate_checked_in", False)
            
            if mentor_checked_in and candidate_checked_in:
                update_data["status"] = "completed"
            elif mentor_checked_in and not candidate_checked_in:
                update_data["status"] = "candidate_no_show"
            elif not mentor_checked_in and candidate_checked_in:
                update_data["status"] = "mentor_no_show"
            else:
                update_data["status"] = "both_no_show"
        else:
            update_data["status"] = "completed"  # Fallback if booking not found
    elif status in ["no_show_candidate"]:
        update_data["status"] = "candidate_no_show"
    elif status in ["no_show_mentor"]:
        update_data["status"] = "mentor_no_show"
    elif status in ["cancelled_by_candidate", "cancelled_by_mentor", "cancelled_by_admin", "cancelled", "admin_cancelled"]:
        update_data["status"] = status  # Keep specific cancellation type
        update_data["cancelled_at"] = datetime.utcnow()
        update_data["cancelled_by"] = "admin"
        
        # CREDIT SESSION BACK to candidate when admin cancels
        candidate = await db.users.find_one({"id": session.get("user_id")})
        if candidate:
            candidate_plan = (candidate.get("plan") or "").lower()
            is_candidate_unlimited = (
                candidate_plan == "pinnacle" or
                candidate.get("is_unlimited_coaching", False) or
                candidate.get("coaching_sessions_total") == -1
            )
            
            if is_candidate_unlimited:
                # Unlimited users: just decrement usage counter
                if candidate.get("coaching_sessions_used", 0) > 0:
                    await db.users.update_one(
                        {"id": session.get("user_id")},
                        {"$inc": {"coaching_sessions_used": -1}}
                    )
                    logger.info(f"Admin cancel (legacy): Decremented coaching_sessions_used for unlimited user {candidate.get('email')}")
            else:
                # For users with limited sessions: restore the credit
                update_ops = {}
                
                if candidate.get("coaching_sessions_used", 0) > 0:
                    update_ops["$inc"] = {"coaching_sessions_used": -1}
                
                # Also restore coaching_sessions_remaining for users not on coaching plans
                coaching_plans = ["last_mile", "mid_mile", "full_prep", "cohort_premium", "cohort_elite"]
                if candidate_plan not in coaching_plans:
                    if "$inc" not in update_ops:
                        update_ops["$inc"] = {}
                    update_ops["$inc"]["coaching_sessions_remaining"] = 1
                
                if update_ops:
                    await db.users.update_one(
                        {"id": session.get("user_id")},
                        update_ops
                    )
                    logger.info(f"Admin cancel (legacy): Credited session back to user {candidate.get('email')}, ops={update_ops}")
        else:
            logger.warning(f"Admin cancel (legacy): Could not find candidate {session.get('user_id')} to credit session back")
        
        # Cancel calendar events when admin cancels the session
        from services.calendar_service import get_calendar_service
        calendar_service = get_calendar_service()
        if calendar_service.is_available():
            # Cancel the main calendar event
            calendar_event_id = session.get("calendar_event_id")
            if calendar_event_id:
                try:
                    calendar_service.cancel_event(calendar_event_id, notify_attendees=True)
                    logger.info(f"Admin cancelled calendar event {calendar_event_id} for session {session_id}")
                except Exception as cal_err:
                    logger.warning(f"Failed to cancel calendar event {calendar_event_id}: {cal_err}")
            
            # Also cancel the hidden event that holds the Meet link
            hidden_event_id = session.get("hidden_event_id")
            if hidden_event_id:
                try:
                    calendar_service.cancel_event(hidden_event_id, notify_attendees=False)
                    logger.info(f"Admin cancelled hidden calendar event {hidden_event_id} for session {session_id}")
                except Exception as cal_err:
                    logger.warning(f"Failed to cancel hidden calendar event {hidden_event_id}: {cal_err}")
        
        # Free up the slot
        slot_value = session.get("time_slot") or session.get("time")
        if slot_value:
            await db.mentor_availability.update_one(
                {"mentor_id": session.get("mentor_id"), "date": session.get("date")},
                {"$pull": {"booked_slots": slot_value}}
            )
            # Also pull the next 30-min slot for 60-min sessions
            try:
                hour, minute = map(int, slot_value.split(':'))
                next_slot_minutes = hour * 60 + minute + 30
                next_hour = next_slot_minutes // 60
                next_minute = next_slot_minutes % 60
                next_slot = f"{next_hour:02d}:{next_minute:02d}"
                await db.mentor_availability.update_one(
                    {"mentor_id": session.get("mentor_id"), "date": session.get("date")},
                    {"$pull": {"booked_slots": next_slot}}
                )
            except (ValueError, AttributeError):
                pass
    
    await db.bookings.update_one(
        {"id": session_id},
        {"$set": update_data}
    )
    
    return {
        "success": True,
        "message": f"Session marked as {status}",
        "session_id": session_id
    }


# ============ Testimonials Management ============

class TestimonialCreate(BaseModel):
    name: str
    testimonial: str
    company_joined: Optional[str] = None
    company_joined_logo: Optional[str] = None
    plan_subscribed: Optional[str] = None
    college: Optional[str] = None
    college_logo: Optional[str] = None
    current_company: Optional[str] = None
    current_company_logo: Optional[str] = None
    image_url: Optional[str] = None
    show_on_pages: List[str] = ["home"]
    is_active: bool = True


class TestimonialUpdate(BaseModel):
    name: Optional[str] = None
    testimonial: Optional[str] = None
    company_joined: Optional[str] = None
    company_joined_logo: Optional[str] = None
    plan_subscribed: Optional[str] = None
    college: Optional[str] = None
    college_logo: Optional[str] = None
    current_company: Optional[str] = None
    current_company_logo: Optional[str] = None
    image_url: Optional[str] = None
    show_on_pages: Optional[List[str]] = None
    is_active: Optional[bool] = None


@router.get("/testimonials")
async def get_all_testimonials(request: Request):
    """Get all testimonials for admin management"""
    await verify_admin(request)
    db = get_db(request)
    
    testimonials = await db.testimonials.find({}, {"_id": 0}).to_list(100)
    return {"testimonials": testimonials}


@router.post("/testimonials")
async def create_testimonial(data: TestimonialCreate, request: Request):
    """Create a new testimonial"""
    await verify_admin(request)
    db = get_db(request)
    
    testimonial_id = f"testimonial_{uuid.uuid4().hex[:12]}"
    
    testimonial_doc = {
        "id": testimonial_id,
        "name": data.name,
        "testimonial": data.testimonial,
        "company_joined": data.company_joined,
        "company_joined_logo": data.company_joined_logo,
        "plan_subscribed": data.plan_subscribed,
        "college": data.college,
        "college_logo": data.college_logo,
        "current_company": data.current_company,
        "current_company_logo": data.current_company_logo,
        "image_url": data.image_url,
        "show_on_pages": data.show_on_pages,
        "is_active": data.is_active,
        "created_at": datetime.utcnow().isoformat(),
        "updated_at": datetime.utcnow().isoformat()
    }
    
    await db.testimonials.insert_one(testimonial_doc)
    
    return {"success": True, "testimonial": {k: v for k, v in testimonial_doc.items() if k != "_id"}}


@router.put("/testimonials/{testimonial_id}")
async def update_testimonial(testimonial_id: str, data: TestimonialUpdate, request: Request):
    """Update an existing testimonial"""
    await verify_admin(request)
    db = get_db(request)
    
    existing = await db.testimonials.find_one({"id": testimonial_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Testimonial not found")
    
    update_data = {k: v for k, v in data.dict().items() if v is not None}
    update_data["updated_at"] = datetime.utcnow().isoformat()
    
    await db.testimonials.update_one(
        {"id": testimonial_id},
        {"$set": update_data}
    )
    
    updated = await db.testimonials.find_one({"id": testimonial_id}, {"_id": 0})
    return {"success": True, "testimonial": updated}


@router.delete("/testimonials/{testimonial_id}")
async def delete_testimonial(testimonial_id: str, request: Request):
    """Delete a testimonial"""
    await verify_admin(request)
    db = get_db(request)
    
    result = await db.testimonials.delete_one({"id": testimonial_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Testimonial not found")
    
    return {"success": True, "message": "Testimonial deleted"}


# ============ Logo Repository Management ============

class LogoCreate(BaseModel):
    name: str
    logo_url: str
    category: str  # 'company', 'college', 'consulting_firm'
    show_on_homepage: Optional[bool] = False


class LogoUpdate(BaseModel):
    name: Optional[str] = None
    logo_url: Optional[str] = None
    category: Optional[str] = None
    show_on_homepage: Optional[bool] = None


@router.get("/logos")
async def get_all_logos(request: Request, category: Optional[str] = None):
    """Get all logos, optionally filtered by category"""
    await verify_admin(request)
    db = get_db(request)
    
    query = {}
    if category:
        query["category"] = category
    
    logos = await db.logo_repository.find(query, {"_id": 0}).to_list(200)
    return {"logos": logos}


@router.post("/logos")
async def create_logo(data: LogoCreate, request: Request):
    """Add a new logo to the repository"""
    await verify_admin(request)
    db = get_db(request)
    
    logo_id = f"logo_{uuid.uuid4().hex[:12]}"
    
    logo_doc = {
        "id": logo_id,
        "name": data.name,
        "logo_url": data.logo_url,
        "category": data.category,
        "show_on_homepage": data.show_on_homepage or False,
        "created_at": datetime.utcnow().isoformat()
    }
    
    await db.logo_repository.insert_one(logo_doc)
    
    return {"success": True, "logo": {k: v for k, v in logo_doc.items() if k != "_id"}}


@router.put("/logos/{logo_id}")
async def update_logo(logo_id: str, data: LogoUpdate, request: Request):
    """Update a logo in the repository"""
    await verify_admin(request)
    db = get_db(request)
    
    existing = await db.logo_repository.find_one({"id": logo_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Logo not found")
    
    update_data = {k: v for k, v in data.dict().items() if v is not None}
    
    await db.logo_repository.update_one(
        {"id": logo_id},
        {"$set": update_data}
    )
    
    updated = await db.logo_repository.find_one({"id": logo_id}, {"_id": 0})
    return {"success": True, "logo": updated}


@router.delete("/logos/{logo_id}")
async def delete_logo(logo_id: str, request: Request):
    """Delete a logo from the repository"""
    await verify_admin(request)
    db = get_db(request)
    
    result = await db.logo_repository.delete_one({"id": logo_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Logo not found")
    
    return {"success": True, "message": "Logo deleted"}



# ============ AI-Generated Drills Management (Case Drills Tab) ============

# Import the PRE_GENERATED_DRILLS from ai_drills
from routes.ai_drills import PRE_GENERATED_DRILLS

@router.get("/ai-drills")
async def get_admin_ai_drills(request: Request):
    """
    Get all AI-generated drills with their is_free_trial status.
    Admin can toggle which drills are available for free trial users.
    """
    await verify_admin(request)
    db = get_db(request)
    
    # Get custom drill settings from database (overrides defaults)
    drill_settings = await db.drill_settings.find({}, {"_id": 0}).to_list(1000)
    drill_settings_map = {d["drill_id"]: d for d in drill_settings}
    
    drills = []
    
    # Counter for each drill type
    type_counters = {
        "case_math": 1,
        "case_structuring": 1,
        "charts_exhibits": 1
    }
    
    type_display_names = {
        "case_math": "Case Math Drill",
        "case_structuring": "Case Structuring Drill",
        "charts_exhibits": "Charts & Exhibits"
    }
    
    difficulty_labels = {
        "beginner": "Easy",
        "intermediate": "Medium",
        "advanced": "Hard"
    }
    
    # Default: first 3 drills of each type are free trial
    FREE_TRIAL_DRILLS_PER_TYPE = 3
    
    for drill_type, difficulties in PRE_GENERATED_DRILLS.items():
        for difficulty, drill_list in difficulties.items():
            for drill in drill_list:
                counter = type_counters[drill_type]
                display_name = f"{type_display_names[drill_type]} {counter}"
                
                # Check if admin has customized this drill's settings
                custom_setting = drill_settings_map.get(drill["id"])
                if custom_setting is not None:
                    is_free_trial = custom_setting.get("is_free_trial", False)
                    is_basic_plan = custom_setting.get("is_basic_plan", False)
                else:
                    # Default: first 3 drills of each type are free trial
                    is_free_trial = counter <= FREE_TRIAL_DRILLS_PER_TYPE
                    # Default: first 6 drills of each type are basic plan (double free trial)
                    is_basic_plan = counter <= FREE_TRIAL_DRILLS_PER_TYPE * 2
                
                drills.append({
                    "id": drill["id"],
                    "drill_type": drill_type,
                    "drill_type_label": type_display_names[drill_type].replace(" Drill", ""),
                    "difficulty": difficulty,
                    "difficulty_label": difficulty_labels[difficulty],
                    "name": display_name,
                    "question_count": len(drill.get("questions", [])),
                    "is_free_trial": is_free_trial,
                    "is_basic_plan": is_basic_plan,
                    "is_custom_setting": custom_setting is not None
                })
                
                type_counters[drill_type] += 1
    
    # Calculate stats
    total_drills = len(drills)
    free_trial_count = sum(1 for d in drills if d["is_free_trial"])
    basic_plan_count = sum(1 for d in drills if d["is_basic_plan"])
    
    return {
        "drills": drills,
        "total": total_drills,
        "free_trial_count": free_trial_count,
        "basic_plan_count": basic_plan_count,
        "stats": {
            "case_math": sum(1 for d in drills if d["drill_type"] == "case_math"),
            "case_structuring": sum(1 for d in drills if d["drill_type"] == "case_structuring"),
            "case_math_free": sum(1 for d in drills if d["drill_type"] == "case_math" and d["is_free_trial"]),
            "case_structuring_free": sum(1 for d in drills if d["drill_type"] == "case_structuring" and d["is_free_trial"]),
            "case_math_basic": sum(1 for d in drills if d["drill_type"] == "case_math" and d["is_basic_plan"]),
            "case_structuring_basic": sum(1 for d in drills if d["drill_type"] == "case_structuring" and d["is_basic_plan"])
        }
    }


class DrillUpdateRequest(BaseModel):
    is_free_trial: Optional[bool] = None
    is_basic_plan: Optional[bool] = None


@router.put("/ai-drills/{drill_id}")
async def update_ai_drill_settings(drill_id: str, data: DrillUpdateRequest, request: Request):
    """
    Update a drill's is_free_trial or is_basic_plan setting.
    This allows admins to customize which drills are available for different plan tiers.
    """
    await verify_admin(request)
    db = get_db(request)
    
    # Validate drill exists in PRE_GENERATED_DRILLS
    drill_exists = False
    for drill_type, difficulties in PRE_GENERATED_DRILLS.items():
        for difficulty, drill_list in difficulties.items():
            for drill in drill_list:
                if drill["id"] == drill_id:
                    drill_exists = True
                    break
    
    if not drill_exists:
        raise HTTPException(status_code=404, detail="Drill not found")
    
    # Build update data - only include fields that are set
    update_data = {
        "drill_id": drill_id,
        "updated_at": datetime.utcnow().isoformat()
    }
    
    if data.is_free_trial is not None:
        update_data["is_free_trial"] = data.is_free_trial
    if data.is_basic_plan is not None:
        update_data["is_basic_plan"] = data.is_basic_plan
    
    # Upsert the drill setting
    await db.drill_settings.update_one(
        {"drill_id": drill_id},
        {"$set": update_data},
        upsert=True
    )
    
    return {
        "success": True,
        "message": f"Drill {drill_id} settings updated"
    }


@router.post("/ai-drills/bulk-update")
async def bulk_update_ai_drill_settings(request: Request):
    """
    Bulk update drill settings.
    Body: { "drill_ids": ["id1", "id2"], "is_free_trial": true/false }
    """
    await verify_admin(request)
    db = get_db(request)
    
    body = await request.json()
    drill_ids = body.get("drill_ids", [])
    is_free_trial = body.get("is_free_trial", False)
    
    if not drill_ids:
        raise HTTPException(status_code=400, detail="drill_ids is required")
    
    # Update all specified drills
    updated_count = 0
    for drill_id in drill_ids:
        result = await db.drill_settings.update_one(
            {"drill_id": drill_id},
            {
                "$set": {
                    "drill_id": drill_id,
                    "is_free_trial": is_free_trial,
                    "updated_at": datetime.utcnow().isoformat()
                }
            },
            upsert=True
        )
        if result.modified_count > 0 or result.upserted_id:
            updated_count += 1
    
    return {
        "success": True,
        "message": f"Updated {updated_count} drills",
        "updated_count": updated_count
    }


@router.post("/ai-drills/reset-defaults")
async def reset_ai_drill_defaults(request: Request):
    """
    Reset all drill settings to defaults (first 3 of each type are free trial).
    This removes all custom settings from the database.
    """
    await verify_admin(request)
    db = get_db(request)
    
    # Delete all custom drill settings
    result = await db.drill_settings.delete_many({})
    
    return {
        "success": True,
        "message": f"Reset {result.deleted_count} drill settings to defaults",
        "deleted_count": result.deleted_count
    }



# ============ Database Diagnostics & Cleanup ============

@router.get("/plans/diagnostics")
async def diagnose_plans_database(request: Request):
    """
    Diagnose database for duplicate plan records and configuration issues.
    This is safe to run - it only checks, doesn't modify anything.
    """
    await verify_admin(request)
    db = get_db(request)
    
    results = {
        "status": "checked",
        "timestamp": datetime.utcnow().isoformat(),
        "plans_checked": [],
        "issues_found": [],
        "recommendations": []
    }
    
    # Check each subscription plan
    for plan_key in ["basic_plan", "pro_plan", "pro_plus"]:
        # Count records
        count = await db.plans.count_documents({"plan_key": plan_key})
        
        plan_info = {
            "plan_key": plan_key,
            "total_records": count,
            "has_duplicates": count > 1,
            "records": []
        }
        
        # Get all records for this plan
        all_records = await db.plans.find({"plan_key": plan_key}).to_list(length=10)
        
        for rec in all_records:
            features = rec.get('features', {})
            record_data = {
                "_id": str(rec.get('_id')),
                "id": rec.get('id'),
                "category": rec.get('category'),
                "has_id_field": rec.get('id') is not None,
                "peer_to_peer": features.get('peer_to_peer'),
                "peer_sessions_per_month": features.get('peer_sessions_per_month'),
                "is_admin_managed": rec.get('id') is not None and rec.get('category') == 'subscription'
            }
            plan_info["records"].append(record_data)
        
        results["plans_checked"].append(plan_info)
        
        # Identify issues
        if count > 1:
            results["issues_found"].append({
                "plan": plan_key,
                "issue": "duplicate_records",
                "severity": "high",
                "description": f"Found {count} records for {plan_key}. Should be only 1.",
                "impact": "Backend may return inconsistent configurations to users"
            })
            
            # Check if they have different configurations
            peer_values = set([r.get('features', {}).get('peer_to_peer') for r in all_records])
            if len(peer_values) > 1:
                results["issues_found"].append({
                    "plan": plan_key,
                    "issue": "conflicting_configurations",
                    "severity": "critical",
                    "description": f"Records have different peer_to_peer values: {peer_values}",
                    "impact": "Users will get random access - some blocked, some not!"
                })
        
        elif count == 0:
            results["issues_found"].append({
                "plan": plan_key,
                "issue": "missing_plan",
                "severity": "critical",
                "description": f"No records found for {plan_key}",
                "impact": "Users cannot subscribe to this plan"
            })
        
        else:
            # Check configuration
            plan_rec = all_records[0]
            features = plan_rec.get('features', {})
            peer = features.get('peer_to_peer')
            sessions = features.get('peer_sessions_per_month')
            
            # Basic and Pro should have peer access
            if plan_key in ["basic_plan", "pro_plan"]:
                if peer == 'none' or (peer is None and (sessions is None or sessions == 0)):
                    results["issues_found"].append({
                        "plan": plan_key,
                        "issue": "missing_peer_access",
                        "severity": "high",
                        "description": f"{plan_key} has no peer practice access configured",
                        "impact": "Users with this plan will be blocked from peer practice"
                    })
    
    # Generate recommendations
    if len(results["issues_found"]) == 0:
        results["recommendations"].append({
            "action": "none",
            "description": "✅ All plans are correctly configured. No action needed."
        })
    else:
        results["recommendations"].append({
            "action": "cleanup_required",
            "description": "⚠️ Issues found. Use /admin/plans/cleanup endpoint to fix automatically."
        })
    
    return results


@router.post("/plans/cleanup")
async def cleanup_plans_database(request: Request, dry_run: bool = True):
    """
    Clean up duplicate plan records and fix configurations.
    
    Parameters:
    - dry_run: If True (default), only shows what would be changed without making changes.
              Set to False to actually apply the fixes.
    
    This will:
    1. Remove duplicate records (keeps admin-managed ones with 'id' field)
    2. Update remaining records with correct peer practice settings
    """
    await verify_admin(request)
    db = get_db(request)
    
    results = {
        "dry_run": dry_run,
        "timestamp": datetime.utcnow().isoformat(),
        "actions_taken": [],
        "errors": []
    }
    
    # Expected configurations
    expected_configs = {
        "basic_plan": {
            "peer_to_peer": "1_per_week",
            "peer_sessions_per_month": 4
        },
        "pro_plan": {
            "peer_to_peer": "1_per_week",
            "peer_sessions_per_month": 4
        },
        "pro_plus": {
            "peer_to_peer": "unlimited",
            "peer_sessions_per_month": -1
        }
    }
    
    for plan_key, expected in expected_configs.items():
        plan_results = {
            "plan_key": plan_key,
            "duplicates_removed": 0,
            "records_updated": 0,
            "details": []
        }
        
        # Get all records
        all_records = await db.plans.find({"plan_key": plan_key}).to_list(length=10)
        
        if len(all_records) == 0:
            plan_results["details"].append(f"⚠️ No records found for {plan_key}")
        
        elif len(all_records) > 1:
            # Multiple records - need to clean up
            main_record = None
            duplicates = []
            
            # Find the admin-managed record (has 'id' field and category='subscription')
            for rec in all_records:
                if rec.get('id') and rec.get('category') == 'subscription':
                    main_record = rec
                else:
                    duplicates.append(rec)
            
            # If no admin-managed record found, keep the first one with 'id' field
            if not main_record:
                for rec in all_records:
                    if rec.get('id'):
                        main_record = rec
                        break
            
            # If still no main record, keep the first one
            if not main_record:
                main_record = all_records[0]
                duplicates = all_records[1:]
            
            plan_results["details"].append(f"Found {len(all_records)} records")
            plan_results["details"].append(f"Keeping: _id={main_record.get('_id')} (id={main_record.get('id')})")
            
            # Delete duplicates
            for dup in duplicates:
                dup_id = dup.get('_id')
                plan_results["details"].append(f"Deleting duplicate: _id={dup_id}")
                
                if not dry_run:
                    delete_result = await db.plans.delete_one({"_id": dup_id})
                    if delete_result.deleted_count > 0:
                        plan_results["duplicates_removed"] += 1
                else:
                    plan_results["duplicates_removed"] += 1  # Count for dry run
            
            # Update main record with correct configuration
            update_data = {
                f"features.{key}": value 
                for key, value in expected.items()
            }
            
            plan_results["details"].append(f"Updating with: {expected}")
            
            if not dry_run:
                update_result = await db.plans.update_one(
                    {"_id": main_record.get('_id')},
                    {"$set": update_data}
                )
                if update_result.modified_count > 0:
                    plan_results["records_updated"] += 1
            else:
                plan_results["records_updated"] += 1  # Count for dry run
        
        else:
            # Single record - just update configuration if needed
            single_record = all_records[0]
            features = single_record.get('features', {})
            
            needs_update = False
            for key, expected_value in expected.items():
                if features.get(key) != expected_value:
                    needs_update = True
                    break
            
            if needs_update:
                update_data = {
                    f"features.{key}": value 
                    for key, value in expected.items()
                }
                
                plan_results["details"].append("Single record found, updating configuration")
                plan_results["details"].append(f"Setting: {expected}")
                
                if not dry_run:
                    update_result = await db.plans.update_one(
                        {"_id": single_record.get('_id')},
                        {"$set": update_data}
                    )
                    if update_result.modified_count > 0:
                        plan_results["records_updated"] += 1
                else:
                    plan_results["records_updated"] += 1
            else:
                plan_results["details"].append("✅ Single record found with correct configuration")
        
        results["actions_taken"].append(plan_results)
    
    if dry_run:
        results["message"] = "DRY RUN: No changes made. Set dry_run=false to apply fixes."
    else:
        results["message"] = "✅ Cleanup completed successfully!"
    
    return results




# ============ Logs Viewer for Production ============

@router.get("/logs/webhooks")
async def get_webhook_logs(request: Request, limit: int = 20):
    """
    Get recent webhook logs for debugging.
    Shows the last N webhook events received from Razorpay.
    """
    await verify_admin(request)
    db = get_db(request)
    
    # Get recent webhook logs
    logs = await db.webhook_logs.find().sort("received_at", -1).limit(limit).to_list(length=limit)
    
    # Format for easy reading
    formatted_logs = []
    for log in logs:
        formatted_log = {
            "received_at": log.get("received_at"),
            "event": log.get("event"),
            "subscription_id": log.get("subscription_id"),
            "user_id": log.get("user_id"),
            "plan_key": log.get("payload", {}).get("payload", {}).get("subscription", {}).get("entity", {}).get("notes", {}).get("plan_key"),
            "status": log.get("payload", {}).get("payload", {}).get("subscription", {}).get("entity", {}).get("status"),
            "_id": str(log.get("_id"))
        }
        formatted_logs.append(formatted_log)
    
    return {
        "total": len(formatted_logs),
        "logs": formatted_logs
    }


@router.get("/logs/recent-activations")
async def get_recent_activations(request: Request, limit: int = 10):
    """
    Get recently activated subscriptions by checking user documents.
    Shows users whose subscriptions were recently activated.
    """
    await verify_admin(request)
    db = get_db(request)
    
    # Get users with subscriptions sorted by activation date
    users = await db.users.find(
        {"subscription.activated_at": {"$exists": True}},
        {
            "id": 1,
            "email": 1,
            "name": 1,
            "plan": 1,
            "plan_name": 1,
            "subscription.activated_at": 1,
            "subscription.status": 1,
            "subscription.plan_key": 1,
            "subscription.razorpay_subscription_id": 1,
            "plan_assignments": 1,
            "updated_at": 1
        }
    ).sort("subscription.activated_at", -1).limit(limit).to_list(length=limit)
    
    formatted_users = []
    for user in users:
        subscription = user.get("subscription", {})
        plan_assignments = user.get("plan_assignments", [])
        active_assignment = next((a for a in plan_assignments if a.get("is_active")), None)
        
        formatted_users.append({
            "user_id": user.get("id"),
            "email": user.get("email"),
            "name": user.get("name"),
            "plan_field": user.get("plan"),
            "plan_name_field": user.get("plan_name"),
            "subscription_plan_key": subscription.get("plan_key"),
            "subscription_status": subscription.get("status"),
            "subscription_activated_at": subscription.get("activated_at"),
            "razorpay_subscription_id": subscription.get("razorpay_subscription_id"),
            "has_active_assignment": active_assignment is not None,
            "active_assignment_plan": active_assignment.get("plan_key") if active_assignment else None,
            "updated_at": user.get("updated_at")
        })
    
    return {
        "total": len(formatted_users),
        "users": formatted_users
    }


@router.get("/logs/user-subscription/{user_email}")
async def get_user_subscription_details(request: Request, user_email: str):
    """
    Get detailed subscription information for a specific user by email.
    Useful for debugging why a user's plan isn't updating.
    """
    await verify_admin(request)
    db = get_db(request)
    
    user = await db.users.find_one({"email": user_email})
    
    if not user:
        raise HTTPException(status_code=404, detail=f"User not found: {user_email}")
    
    # Get relevant webhook logs for this user
    webhooks = await db.webhook_logs.find(
        {"user_id": user.get("id")}
    ).sort("received_at", -1).limit(5).to_list(length=5)
    
    return {
        "user_id": user.get("id"),
        "email": user.get("email"),
        "name": user.get("name"),
        "plan": user.get("plan"),
        "plan_name": user.get("plan_name"),
        "plan_category": user.get("plan_category"),
        "plan_start_date": user.get("plan_start_date"),
        "plan_end_date": user.get("plan_end_date"),
        "is_subscribed": user.get("is_subscribed"),
        "subscription": user.get("subscription"),
        "plan_assignments": user.get("plan_assignments", []),
        "pending_subscription": user.get("pending_subscription"),
        "recent_webhooks": [
            {
                "event": w.get("event"),
                "received_at": w.get("received_at"),
                "subscription_id": w.get("subscription_id")
            } for w in webhooks
        ],
        "updated_at": user.get("updated_at"),
        "created_at": user.get("created_at")
    }


class FixSubscriptionRequest(BaseModel):
    user_email: str
    plan_key: str
    billing_cycle: str = "6_month"


@router.post("/fix-subscription")
async def fix_user_subscription(fix_data: FixSubscriptionRequest, request: Request):
    """
    Manually fix/activate a subscription for a user.
    Use this when a payment was successful but the user's plan wasn't updated.
    """
    await verify_admin(request)
    db = get_db(request)
    
    from datetime import timezone
    from dateutil.relativedelta import relativedelta
    
    # Find the user
    user = await db.users.find_one({"email": fix_data.user_email})
    if not user:
        raise HTTPException(status_code=404, detail=f"User not found: {fix_data.user_email}")
    
    # Find the plan
    plan = await db.plans.find_one({"plan_key": fix_data.plan_key})
    if not plan:
        raise HTTPException(status_code=404, detail=f"Plan not found: {fix_data.plan_key}")
    
    now = datetime.now(timezone.utc)
    
    # Calculate period end based on billing cycle
    if fix_data.billing_cycle == "monthly":
        period_end = now + relativedelta(months=1)
    else:  # 6_month
        period_end = now + relativedelta(months=6)
    
    # Build subscription data
    subscription_data = {
        "status": "active",
        "plan_key": fix_data.plan_key,
        "billing_cycle": fix_data.billing_cycle,
        "auto_renew": True,
        "current_period_start": now.isoformat(),
        "current_period_end": period_end.isoformat(),
        "activated_at": now.isoformat(),
        "manually_fixed_at": now.isoformat(),
        "fixed_by": "admin"
    }
    
    # Update user with correct plan
    update_result = await db.users.update_one(
        {"id": user.get("id")},
        {
            "$set": {
                "subscription": subscription_data,
                "plan": fix_data.plan_key,
                "plan_name": plan.get("name"),
                "plan_category": plan.get("category", "subscription"),
                "plan_start_date": now.isoformat(),
                "plan_end_date": period_end.isoformat(),
                "subscription_end_date": period_end.isoformat(),
                "is_subscribed": True,
                "plan_features": plan.get("features", {}),
                "features": plan.get("features", {}),  # Also set features for access control
                "updated_at": now.isoformat()
            },
            "$unset": {"pending_subscription": ""}
        }
    )
    
    if update_result.modified_count == 0:
        return {
            "success": False,
            "message": "User found but document not modified - data might already be correct"
        }
    
    logger.info(f"✅ Admin manually fixed subscription for {fix_data.user_email}: {fix_data.plan_key}")
    
    return {
        "success": True,
        "message": f"Subscription fixed for {fix_data.user_email}",
        "user_id": user.get("id"),
        "plan": fix_data.plan_key,
        "plan_name": plan.get("name"),
        "period_end": period_end.isoformat()
    }


@router.get("/debug/mentor-sessions/{mentor_email}")
async def debug_mentor_sessions(mentor_email: str, request: Request):
    """
    Debug why a mentor might not see their sessions.
    Checks mentor record, user record, and bookings for ID mismatches.
    """
    await verify_admin(request)
    db = get_db(request)
    
    # Find mentor by email
    mentor = await db.mentors.find_one({"email": mentor_email}, {"_id": 0})
    
    # Find user by email
    user = await db.users.find_one({"email": mentor_email}, {"_id": 0})
    
    result = {
        "mentor_email": mentor_email,
        "mentor_record": None,
        "user_record": None,
        "id_match": False,
        "bookings_by_mentor_id": [],
        "bookings_by_user_id": [],
        "potential_issues": []
    }
    
    if mentor:
        result["mentor_record"] = {
            "id": mentor.get("id"),
            "name": mentor.get("name"),
            "email": mentor.get("email"),
            "is_active": mentor.get("is_active")
        }
    else:
        result["potential_issues"].append("No mentor record found in 'mentors' collection")
    
    if user:
        result["user_record"] = {
            "id": user.get("id"),
            "name": user.get("name"),
            "email": user.get("email"),
            "is_mentor": user.get("is_mentor"),
            "mentor_id": user.get("mentor_id")
        }
    else:
        result["potential_issues"].append("No user record found in 'users' collection")
    
    # Check ID match
    if mentor and user:
        mentor_id_from_mentor = mentor.get("id")
        mentor_id_from_user = user.get("mentor_id")
        
        if mentor_id_from_user == mentor_id_from_mentor:
            result["id_match"] = True
        else:
            result["potential_issues"].append(
                f"ID MISMATCH: mentor.id='{mentor_id_from_mentor}' vs user.mentor_id='{mentor_id_from_user}'"
            )
    
    # Get bookings by mentor_id (from mentors collection)
    if mentor:
        mentor_id = mentor.get("id")
        bookings_by_mentor = await db.bookings.find(
            {"mentor_id": mentor_id},
            {"_id": 0, "id": 1, "date": 1, "time_slot": 1, "status": 1, "user_id": 1, "mentor_id": 1}
        ).to_list(50)
        result["bookings_by_mentor_id"] = bookings_by_mentor
        result["bookings_count_by_mentor_id"] = len(bookings_by_mentor)
    
    # Get bookings where mentor_id might be user's id instead
    if user:
        user_id = user.get("id")
        bookings_by_user_id = await db.bookings.find(
            {"mentor_id": user_id},
            {"_id": 0, "id": 1, "date": 1, "time_slot": 1, "status": 1, "user_id": 1, "mentor_id": 1}
        ).to_list(50)
        result["bookings_by_user_id"] = bookings_by_user_id
        result["bookings_count_by_user_id"] = len(bookings_by_user_id)
        
        if bookings_by_user_id and mentor:
            result["potential_issues"].append(
                f"Found {len(bookings_by_user_id)} bookings with mentor_id=user.id instead of mentor.id!"
            )
    
    # Check strategy calls too
    if mentor:
        strategy_calls = await db.strategy_call_sessions.find(
            {"mentor_id": mentor.get("id")},
            {"_id": 0, "id": 1, "date": 1, "time": 1, "status": 1}
        ).to_list(50)
        result["strategy_calls_by_mentor_id"] = strategy_calls
        result["strategy_calls_count"] = len(strategy_calls)
    
    return result


@router.post("/fix-mentor-bookings/{mentor_email}")
async def fix_mentor_bookings(mentor_email: str, request: Request):
    """
    Fix bookings where mentor_id was set to user.id instead of mentor.id.
    This corrects the ID mismatch so mentors can see their sessions.
    """
    await verify_admin(request)
    db = get_db(request)
    
    # Find mentor and user
    mentor = await db.mentors.find_one({"email": mentor_email})
    user = await db.users.find_one({"email": mentor_email})
    
    if not mentor:
        raise HTTPException(status_code=404, detail=f"Mentor not found: {mentor_email}")
    if not user:
        raise HTTPException(status_code=404, detail=f"User not found: {mentor_email}")
    
    correct_mentor_id = mentor.get("id")
    wrong_mentor_id = user.get("id")
    
    if correct_mentor_id == wrong_mentor_id:
        return {"message": "No fix needed - mentor.id matches user.id", "fixed_count": 0}
    
    # Fix bookings
    bookings_result = await db.bookings.update_many(
        {"mentor_id": wrong_mentor_id},
        {"$set": {"mentor_id": correct_mentor_id, "mentor_id_fixed": True}}
    )
    
    # Fix strategy calls
    strategy_result = await db.strategy_call_sessions.update_many(
        {"mentor_id": wrong_mentor_id},
        {"$set": {"mentor_id": correct_mentor_id, "mentor_id_fixed": True}}
    )
    
    # Also update user's mentor_id field
    await db.users.update_one(
        {"email": mentor_email},
        {"$set": {"mentor_id": correct_mentor_id}}
    )
    
    return {
        "success": True,
        "message": f"Fixed mentor_id from '{wrong_mentor_id}' to '{correct_mentor_id}'",
        "bookings_fixed": bookings_result.modified_count,
        "strategy_calls_fixed": strategy_result.modified_count
    }


@router.get("/debug/user-plan/{user_email}")
async def debug_user_plan(user_email: str, request: Request):
    """
    Debug user's plan and feature access.
    Shows exactly what the dashboard sees for this user.
    """
    await verify_admin(request)
    db = get_db(request)
    
    from models import PlanType
    
    user = await db.users.find_one({"email": user_email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail=f"User not found: {user_email}")
    
    # Get the plan from database
    plan = await db.plans.find_one({"plan_key": user.get("plan")}, {"_id": 0})
    
    # Check what access functions return
    user_plan = user.get("plan", "")
    user_features = user.get("features")
    user_plan_features = user.get("plan_features")
    
    # Simulate the has_subscription_access check
    subscription_plans = [
        PlanType.BASIC.value, PlanType.PRO.value, PlanType.PRO_PLUS.value,
        PlanType.LAST_MILE.value, PlanType.MID_MILE.value, PlanType.FULL_PREP.value,
        PlanType.PINNACLE.value, PlanType.COHORT_PREMIUM.value, PlanType.COHORT_ELITE.value,
    ]
    
    has_sub_by_plan = user_plan in subscription_plans
    
    # Check features
    has_sub_by_features = False
    if user_features:
        sub_features = ['course_recordings', 'peer_practice', 'peer_to_peer', 'case_drills', 'drills_exercises', 'workshops']
        for feature in sub_features:
            feature_value = user_features.get(feature)
            if feature_value == True or (isinstance(feature_value, str) and feature_value not in ['', 'none', 'false']):
                has_sub_by_features = True
                break
    
    return {
        "user_email": user_email,
        "user_id": user.get("id"),
        "current_plan_data": {
            "plan": user_plan,
            "plan_name": user.get("plan_name"),
            "plan_category": user.get("plan_category"),
            "plan_start_date": user.get("plan_start_date"),
            "plan_end_date": user.get("plan_end_date"),
            "subscription_end_date": user.get("subscription_end_date"),
            "is_subscribed": user.get("is_subscribed")
        },
        "subscription_object": user.get("subscription"),
        "features_field": user_features,
        "plan_features_field": user_plan_features,
        "plan_from_db": {
            "name": plan.get("name") if plan else None,
            "plan_key": plan.get("plan_key") if plan else None,
            "features": plan.get("features") if plan else None
        } if plan else None,
        "access_check_results": {
            "has_subscription_by_plan": has_sub_by_plan,
            "has_subscription_by_features": has_sub_by_features,
            "plan_in_subscription_list": user_plan in subscription_plans,
            "subscription_plans_list": subscription_plans
        },
        "diagnosis": {
            "plan_matches_enum": user_plan in subscription_plans,
            "has_features_field": user_features is not None,
            "has_plan_features_field": user_plan_features is not None,
            "recommendation": (
                "Plan should grant access" if user_plan in subscription_plans 
                else f"Plan '{user_plan}' not in subscription list - need to fix"
            )
        }
    }


class AddSessionsRequest(BaseModel):
    user_email: str
    session_count: int
    razorpay_payment_id: Optional[str] = None
    reason: str = "Manual adjustment"


@router.post("/add-sessions")
async def add_sessions_manually(data: AddSessionsRequest, request: Request):
    """
    Manually add coaching sessions to a user's account.
    Use this when a payment was successful in Razorpay but verification failed.
    """
    await verify_admin(request)
    db = get_db(request)
    
    from datetime import timezone
    
    # Find the user
    user = await db.users.find_one({"email": data.user_email})
    if not user:
        raise HTTPException(status_code=404, detail=f"User not found: {data.user_email}")
    
    current_sessions = user.get("coaching_sessions_remaining", 0) or 0
    new_total = current_sessions + data.session_count
    
    now = datetime.now(timezone.utc)
    
    # Update user's sessions
    await db.users.update_one(
        {"id": user.get("id")},
        {
            "$set": {
                "coaching_sessions_remaining": new_total,
                "updated_at": now.isoformat()
            }
        }
    )
    
    # Log this manual adjustment
    adjustment_log = {
        "id": str(uuid.uuid4()),
        "user_id": user.get("id"),
        "user_email": data.user_email,
        "type": "manual_session_add",
        "sessions_added": data.session_count,
        "previous_balance": current_sessions,
        "new_balance": new_total,
        "razorpay_payment_id": data.razorpay_payment_id,
        "reason": data.reason,
        "added_by": "admin",
        "created_at": now.isoformat()
    }
    await db.admin_adjustments.insert_one(adjustment_log)
    
    logger.info(f"✅ Admin added {data.session_count} sessions to {data.user_email}: {current_sessions} -> {new_total}")
    
    return {
        "success": True,
        "message": f"Added {data.session_count} sessions to {data.user_email}",
        "previous_balance": current_sessions,
        "new_balance": new_total,
        "razorpay_payment_id": data.razorpay_payment_id
    }


# ============ Cancellation Policy Management ============

class CancellationPolicyRequest(BaseModel):
    candidate_hours: int
    mentor_hours: int

@router.get("/cancellation-policy")
async def get_cancellation_policy(request: Request):
    """Get current cancellation policy settings"""
    user = await get_current_user(request)
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    db = get_db(request)
    
    # Get policy from database
    policy = await db.platform_settings.find_one({"type": "cancellation_policy"})
    
    if not policy:
        # Return defaults if not set
        return {
            "candidate_hours": 4,
            "mentor_hours": 4
        }
    
    return {
        "candidate_hours": policy.get("candidate_hours", 4),
        "mentor_hours": policy.get("mentor_hours", 4)
    }

@router.put("/cancellation-policy")
async def update_cancellation_policy(policy_data: CancellationPolicyRequest, request: Request):
    """Update cancellation policy settings"""
    user = await get_current_user(request)
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    db = get_db(request)
    
    # Validate hours (must be positive)
    if policy_data.candidate_hours < 0 or policy_data.mentor_hours < 0:
        raise HTTPException(status_code=400, detail="Hours must be non-negative")
    
    # Update or create policy
    await db.platform_settings.update_one(
        {"type": "cancellation_policy"},
        {"$set": {
            "type": "cancellation_policy",
            "candidate_hours": policy_data.candidate_hours,
            "mentor_hours": policy_data.mentor_hours,
            "updated_at": datetime.utcnow().isoformat()
        }},
        upsert=True
    )
    
    return {
        "message": "Cancellation policy updated successfully",
        "candidate_hours": policy_data.candidate_hours,
        "mentor_hours": policy_data.mentor_hours
    }



# ============ Mentor Payouts Management ============

class MarkPaidRequest(BaseModel):
    amount_override: Optional[int] = None  # Optional manual override of amount


@router.get("/payouts")
async def get_payouts(
    request: Request,
    page: int = 1,
    limit: int = 20,
    mentor_id: str = None,
    candidate_id: str = None,
    status: str = None,  # pending, on_hold, paid
    date_from: str = None,
    date_to: str = None
):
    """
    Get all sessions with payment status for admin payouts management.
    
    Payment Status Logic:
    - pending: Session completed + Feedback given + Not paid yet
    - on_hold: Session completed + Feedback NOT given
    - paid: Marked as paid by admin
    """
    await verify_admin(request)
    db = get_db(request)
    
    # Build query - only completed sessions
    query = {"status": "completed"}
    
    if mentor_id:
        query["mentor_id"] = mentor_id
    
    if candidate_id:
        query["user_id"] = candidate_id
    
    if date_from:
        if "date" not in query:
            query["date"] = {}
        query["date"]["$gte"] = date_from
    
    if date_to:
        if "date" not in query:
            query["date"] = {}
        query["date"]["$lte"] = date_to
    
    # Get total count before filtering by payment status
    all_completed = await db.bookings.find(query, {"_id": 0}).to_list(1000)
    
    # Process sessions to determine payment status
    processed_sessions = []
    for session in all_completed:
        booking_id = session.get("id")
        
        # Check if feedback exists
        mentor_feedback = await db.mentor_feedbacks.find_one({"booking_id": booking_id})
        has_feedback = mentor_feedback is not None
        
        # Check if marked as paid
        is_paid = session.get("payment_status") == "paid"
        
        # Determine payment status
        if is_paid:
            payment_status = "paid"
        elif has_feedback:
            payment_status = "pending"
        else:
            payment_status = "on_hold"
        
        # Filter by status if specified
        if status and payment_status != status:
            continue
        
        # Get mentor details
        mentor = await db.mentors.find_one(
            {"id": session.get("mentor_id")},
            {"_id": 0, "name": 1, "email": 1, "picture": 1, "hourly_rate": 1}
        )
        
        # Get candidate details
        candidate = await db.users.find_one(
            {"id": session.get("user_id")},
            {"_id": 0, "name": 1, "email": 1, "picture": 1}
        )
        
        # Calculate amount (use override if exists, else mentor's hourly rate)
        hourly_rate = mentor.get("hourly_rate", 1500) if mentor else 1500
        amount = session.get("payment_amount_override") or hourly_rate
        
        processed_sessions.append({
            "id": booking_id,
            "date": session.get("date"),
            "time_slot": session.get("time_slot"),
            "session_type": session.get("session_type"),
            "duration_minutes": session.get("duration_minutes", 45),
            # Mentor info
            "mentor_id": session.get("mentor_id"),
            "mentor_name": mentor.get("name") if mentor else "Unknown",
            "mentor_email": mentor.get("email") if mentor else None,
            "mentor_picture": mentor.get("picture") if mentor else None,
            "mentor_hourly_rate": hourly_rate,
            # Candidate info
            "candidate_id": session.get("user_id"),
            "candidate_name": candidate.get("name") if candidate else "Unknown",
            "candidate_email": candidate.get("email") if candidate else None,
            "candidate_picture": candidate.get("picture") if candidate else None,
            # Payment info
            "payment_status": payment_status,
            "amount": amount,
            "amount_override": session.get("payment_amount_override"),
            "has_feedback": has_feedback,
            "paid_at": session.get("paid_at"),
            "paid_by": session.get("paid_by"),
        })
    
    # Sort by date descending
    processed_sessions.sort(key=lambda x: x.get("date", ""), reverse=True)
    
    # Paginate
    total = len(processed_sessions)
    start_idx = (page - 1) * limit
    end_idx = start_idx + limit
    paginated = processed_sessions[start_idx:end_idx]
    
    return {
        "sessions": paginated,
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": (total + limit - 1) // limit if total > 0 else 1
    }


@router.get("/payouts/stats")
async def get_payouts_stats(request: Request):
    """
    Get payout statistics including totals and monthly breakdown.
    """
    await verify_admin(request)
    db = get_db(request)
    
    # Get all completed sessions
    completed_sessions = await db.bookings.find(
        {"status": "completed"},
        {"_id": 0}
    ).to_list(1000)
    
    # Initialize counters
    total_pending = 0
    total_on_hold = 0
    total_paid = 0
    pending_count = 0
    on_hold_count = 0
    paid_count = 0
    
    # Monthly breakdown (last 12 months)
    monthly_data = {}
    
    for session in completed_sessions:
        booking_id = session.get("id")
        
        # Get mentor's hourly rate
        mentor = await db.mentors.find_one(
            {"id": session.get("mentor_id")},
            {"_id": 0, "hourly_rate": 1}
        )
        hourly_rate = mentor.get("hourly_rate", 1500) if mentor else 1500
        amount = session.get("payment_amount_override") or hourly_rate
        
        # Check feedback and payment status
        mentor_feedback = await db.mentor_feedbacks.find_one({"booking_id": booking_id})
        has_feedback = mentor_feedback is not None
        is_paid = session.get("payment_status") == "paid"
        
        # Categorize
        if is_paid:
            total_paid += amount
            paid_count += 1
            
            # Add to monthly data
            paid_date = session.get("paid_at", session.get("date", ""))
            if paid_date:
                month_key = paid_date[:7]  # YYYY-MM
                if month_key not in monthly_data:
                    monthly_data[month_key] = {"paid": 0, "pending": 0, "on_hold": 0}
                monthly_data[month_key]["paid"] += amount
        elif has_feedback:
            total_pending += amount
            pending_count += 1
            
            session_date = session.get("date", "")
            if session_date:
                month_key = session_date[:7]
                if month_key not in monthly_data:
                    monthly_data[month_key] = {"paid": 0, "pending": 0, "on_hold": 0}
                monthly_data[month_key]["pending"] += amount
        else:
            total_on_hold += amount
            on_hold_count += 1
            
            session_date = session.get("date", "")
            if session_date:
                month_key = session_date[:7]
                if month_key not in monthly_data:
                    monthly_data[month_key] = {"paid": 0, "pending": 0, "on_hold": 0}
                monthly_data[month_key]["on_hold"] += amount
    
    # Convert monthly data to sorted list
    monthly_list = [
        {"month": k, **v}
        for k, v in sorted(monthly_data.items(), reverse=True)
    ][:12]  # Last 12 months
    
    # Get mentor list for filter dropdown
    mentors = await db.mentors.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
    
    # Get unique candidates from completed sessions
    candidate_ids = list(set([s.get("user_id") for s in completed_sessions if s.get("user_id")]))
    candidates = await db.users.find(
        {"id": {"$in": candidate_ids}},
        {"_id": 0, "id": 1, "name": 1}
    ).to_list(500)
    
    return {
        "summary": {
            "total_pending": total_pending,
            "total_on_hold": total_on_hold,
            "total_paid": total_paid,
            "pending_count": pending_count,
            "on_hold_count": on_hold_count,
            "paid_count": paid_count,
        },
        "monthly_data": monthly_list,
        "mentors": mentors,
        "candidates": candidates
    }


@router.post("/payouts/{booking_id}/mark-paid")
async def mark_session_paid(booking_id: str, request: Request):
    """
    Mark a session as paid. Optionally override the amount.
    """
    await verify_admin(request)
    db = get_db(request)
    admin = await get_current_user(request)
    
    body = await request.json()
    amount_override = body.get("amount_override")
    
    # Find the booking
    booking = await db.bookings.find_one({"id": booking_id})
    if not booking:
        raise HTTPException(status_code=404, detail="Session not found")
    
    if booking.get("status") != "completed":
        raise HTTPException(status_code=400, detail="Can only mark completed sessions as paid")
    
    # Check if feedback exists
    mentor_feedback = await db.mentor_feedbacks.find_one({"booking_id": booking_id})
    if not mentor_feedback:
        raise HTTPException(status_code=400, detail="Cannot mark as paid - mentor feedback not submitted")
    
    # Update payment status
    update_data = {
        "payment_status": "paid",
        "paid_at": datetime.utcnow().isoformat(),
        "paid_by": admin.get("email") or admin.get("name") or "admin"
    }
    
    if amount_override is not None:
        update_data["payment_amount_override"] = int(amount_override)
    
    await db.bookings.update_one(
        {"id": booking_id},
        {"$set": update_data}
    )
    
    return {
        "success": True,
        "message": "Session marked as paid",
        "booking_id": booking_id,
        "paid_at": update_data["paid_at"]
    }


@router.post("/payouts/{booking_id}/update-amount")
async def update_payment_amount(booking_id: str, request: Request):
    """
    Update the payment amount for a session (manual override).
    """
    await verify_admin(request)
    db = get_db(request)
    
    body = await request.json()
    new_amount = body.get("amount")
    
    if new_amount is None or new_amount < 0:
        raise HTTPException(status_code=400, detail="Invalid amount")
    
    # Find and update the booking
    result = await db.bookings.update_one(
        {"id": booking_id},
        {"$set": {"payment_amount_override": int(new_amount)}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Session not found")
    
    return {
        "success": True,
        "message": "Payment amount updated",
        "booking_id": booking_id,
        "new_amount": int(new_amount)
    }


@router.post("/payouts/bulk-mark-paid")
async def bulk_mark_paid(request: Request):
    """
    Mark multiple sessions as paid at once.
    """
    await verify_admin(request)
    db = get_db(request)
    admin = await get_current_user(request)
    
    body = await request.json()
    booking_ids = body.get("booking_ids", [])
    
    if not booking_ids:
        raise HTTPException(status_code=400, detail="No booking IDs provided")
    
    paid_at = datetime.utcnow().isoformat()
    paid_by = admin.get("email") or admin.get("name") or "admin"
    
    # Update all specified bookings that have feedback
    success_count = 0
    failed_ids = []
    
    for booking_id in booking_ids:
        # Check if feedback exists
        mentor_feedback = await db.mentor_feedbacks.find_one({"booking_id": booking_id})
        if not mentor_feedback:
            failed_ids.append({"id": booking_id, "reason": "No feedback"})
            continue
        
        result = await db.bookings.update_one(
            {"id": booking_id, "status": "completed"},
            {"$set": {
                "payment_status": "paid",
                "paid_at": paid_at,
                "paid_by": paid_by
            }}
        )
        
        if result.modified_count > 0:
            success_count += 1
        else:
            failed_ids.append({"id": booking_id, "reason": "Not found or not completed"})
    
    return {
        "success": True,
        "marked_paid": success_count,
        "failed": failed_ids,
        "total_requested": len(booking_ids)
    }


# ============ Delete Feedback Endpoints ============

@router.delete("/coaching-sessions/{booking_id}/mentor-feedback")
async def delete_mentor_feedback(booking_id: str, request: Request):
    """
    Delete mentor feedback for a coaching session.
    This will:
    1. Delete the feedback record
    2. Reset the mentor_feedback_submitted flag on the booking
    3. Put the payout on hold (by removing feedback, payout status becomes on_hold)
    4. Log the deletion for audit purposes
    """
    await verify_admin(request)
    db = get_db(request)
    admin = await get_current_user(request)
    
    # Find the booking
    booking = await db.bookings.find_one({"id": booking_id})
    if not booking:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Find existing mentor feedback
    mentor_feedback = await db.mentor_feedbacks.find_one({"booking_id": booking_id})
    if not mentor_feedback:
        raise HTTPException(status_code=404, detail="No mentor feedback found for this session")
    
    # Store feedback data for audit log before deleting
    feedback_backup = {
        "feedback_id": str(mentor_feedback.get("_id")),
        "rating_overall": mentor_feedback.get("rating_overall"),
        "qualitative_feedback": mentor_feedback.get("qualitative_feedback"),
        "areas_of_strength": mentor_feedback.get("areas_of_strength"),
        "areas_of_improvement": mentor_feedback.get("areas_of_improvement"),
    }
    
    # Delete the feedback
    await db.mentor_feedbacks.delete_one({"booking_id": booking_id})
    
    # Reset the mentor_feedback_submitted flag on the booking
    await db.bookings.update_one(
        {"id": booking_id},
        {"$set": {
            "mentor_feedback_submitted": False,
            "payment_status": "on_hold"  # Reset payment status
        }}
    )
    
    # Create audit log
    audit_log = {
        "id": str(uuid.uuid4()),
        "action": "delete_mentor_feedback",
        "booking_id": booking_id,
        "deleted_by": admin.get("email") or admin.get("name") or "admin",
        "deleted_at": datetime.utcnow(),
        "feedback_backup": feedback_backup,
        "mentor_id": booking.get("mentor_id"),
        "candidate_id": booking.get("user_id"),
        "session_date": booking.get("date"),
        "reason": "Admin deleted feedback to allow re-submission"
    }
    await db.feedback_deletion_logs.insert_one(audit_log)
    
    logger.info(f"Mentor feedback deleted for booking {booking_id} by {admin.get('email')}")
    
    return {
        "success": True,
        "message": "Mentor feedback deleted successfully. Payout is now on hold. Mentor can submit feedback again.",
        "booking_id": booking_id,
        "audit_log_id": audit_log["id"]
    }


@router.delete("/coaching-sessions/{booking_id}/candidate-feedback")
async def delete_candidate_feedback(booking_id: str, request: Request):
    """
    Delete candidate feedback for a coaching session.
    This will:
    1. Delete the feedback record
    2. Reset the candidate_feedback_submitted flag on the booking
    3. Log the deletion for audit purposes
    The candidate will see the feedback prompt on next login.
    """
    await verify_admin(request)
    db = get_db(request)
    admin = await get_current_user(request)
    
    # Find the booking
    booking = await db.bookings.find_one({"id": booking_id})
    if not booking:
        raise HTTPException(status_code=404, detail="Session not found")
    
    # Find existing candidate feedback
    candidate_feedback = await db.candidate_feedbacks.find_one({"booking_id": booking_id})
    if not candidate_feedback:
        raise HTTPException(status_code=404, detail="No candidate feedback found for this session")
    
    # Store feedback data for audit log before deleting
    feedback_backup = {
        "feedback_id": str(candidate_feedback.get("_id")),
        "rating_overall": candidate_feedback.get("rating_overall"),
        "comments": candidate_feedback.get("comments"),
        "rating_knowledge": candidate_feedback.get("rating_knowledge"),
        "rating_communication": candidate_feedback.get("rating_communication"),
        "rating_professionalism": candidate_feedback.get("rating_professionalism"),
    }
    
    # Delete the feedback
    await db.candidate_feedbacks.delete_one({"booking_id": booking_id})
    
    # Reset the candidate_feedback_submitted flag on the booking
    await db.bookings.update_one(
        {"id": booking_id},
        {"$set": {"candidate_feedback_submitted": False}}
    )
    
    # Create audit log
    audit_log = {
        "id": str(uuid.uuid4()),
        "action": "delete_candidate_feedback",
        "booking_id": booking_id,
        "deleted_by": admin.get("email") or admin.get("name") or "admin",
        "deleted_at": datetime.utcnow(),
        "feedback_backup": feedback_backup,
        "mentor_id": booking.get("mentor_id"),
        "candidate_id": booking.get("user_id"),
        "session_date": booking.get("date"),
        "reason": "Admin deleted feedback to allow re-submission"
    }
    await db.feedback_deletion_logs.insert_one(audit_log)
    
    logger.info(f"Candidate feedback deleted for booking {booking_id} by {admin.get('email')}")
    
    return {
        "success": True,
        "message": "Candidate feedback deleted successfully. Candidate will see feedback prompt on next login.",
        "booking_id": booking_id,
        "audit_log_id": audit_log["id"]
    }


@router.get("/feedback-deletion-logs")
async def get_feedback_deletion_logs(
    request: Request,
    page: int = 1,
    limit: int = 20,
    booking_id: Optional[str] = None
):
    """
    Get audit logs of deleted feedback.
    """
    await verify_admin(request)
    db = get_db(request)
    
    query = {}
    if booking_id:
        query["booking_id"] = booking_id
    
    total = await db.feedback_deletion_logs.count_documents(query)
    
    logs = await db.feedback_deletion_logs.find(query)\
        .sort("deleted_at", -1)\
        .skip((page - 1) * limit)\
        .limit(limit)\
        .to_list(limit)
    
    # Convert ObjectId to string
    for log in logs:
        log["_id"] = str(log["_id"])
    
    return {
        "logs": logs,
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": (total + limit - 1) // limit if total > 0 else 1
    }


# ============ Candidate Upload Template ============

@router.get("/candidates/upload-template")
async def get_candidate_upload_template():
    """Download the candidate upload Excel template"""
    from fastapi.responses import FileResponse
    import os
    
    template_path = "/app/uploads/candidate_upload_template.xlsx"
    
    if not os.path.exists(template_path):
        raise HTTPException(status_code=404, detail="Template file not found")
    
    return FileResponse(
        template_path,
        filename="candidate_upload_template.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



# ============ Historical Feedback Migration ============

@router.post("/feedbacks/fix-historical-flags")
async def fix_historical_feedback_flags(request: Request):
    """
    Migration endpoint to fix feedbacks that are historical but missing is_historical flag.
    Historical feedbacks are identified by:
    - candidate_id is null/None
    - booking_id is null/None  
    - OR id starts with 'hist-fb-'
    - OR candidate_name_override exists
    """
    db = get_db(request)
    user = await get_current_user(request, db)
    
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Find feedbacks that should be historical but aren't marked as such
    query = {
        "$and": [
            {"is_historical": {"$ne": True}},  # Not already marked
            {"$or": [
                {"candidate_id": None},
                {"candidate_id": {"$exists": False}},
                {"booking_id": None},
                {"booking_id": {"$exists": False}},
                {"id": {"$regex": "^hist-fb-"}},
                {"candidate_name_override": {"$exists": True, "$ne": None}}
            ]}
        ]
    }
    
    # First, count how many need fixing
    feedbacks_to_fix = await db.candidate_feedbacks.find(query).to_list(length=1000)
    
    if not feedbacks_to_fix:
        return {
            "message": "No feedbacks need fixing",
            "fixed_count": 0,
            "feedbacks": []
        }
    
    # Update them
    result = await db.candidate_feedbacks.update_many(
        query,
        {"$set": {"is_historical": True}}
    )
    
    # Get details of fixed feedbacks for reporting
    fixed_details = []
    for fb in feedbacks_to_fix:
        mentor = await db.mentors.find_one({"id": fb.get("mentor_id")}, {"_id": 0, "name": 1})
        fixed_details.append({
            "feedback_id": fb.get("id"),
            "mentor": mentor.get("name") if mentor else fb.get("mentor_id"),
            "rating": fb.get("rating_overall"),
            "candidate_name_override": fb.get("candidate_name_override")
        })
    
    return {
        "message": f"Fixed {result.modified_count} feedbacks",
        "fixed_count": result.modified_count,
        "feedbacks": fixed_details
    }


@router.delete("/feedbacks/clear-all-historical")
async def clear_all_historical_data(request: Request):
    """
    Delete all historical ratings and feedbacks to start fresh.
    This will:
    1. Remove all feedbacks marked as is_historical=true
    2. Clear the 'rating' field from all mentors (imported ratings)
    3. Clear 'sessions_conducted' field from all mentors (imported session counts)
    """
    db = get_db(request)
    user = await get_current_user(request, db)
    
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    results = {
        "historical_feedbacks_deleted": 0,
        "mentor_ratings_cleared": 0,
        "mentor_sessions_cleared": 0,
        "mentors_affected": []
    }
    
    # 1. Delete all historical feedbacks
    delete_result = await db.candidate_feedbacks.delete_many({"is_historical": True})
    results["historical_feedbacks_deleted"] = delete_result.deleted_count
    
    # Also delete feedbacks that look historical but weren't marked
    suspicious_delete = await db.candidate_feedbacks.delete_many({
        "$or": [
            {"candidate_id": None, "booking_id": None},
            {"id": {"$regex": "^hist-fb-"}},
            {"candidate_name_override": {"$exists": True, "$ne": None}, "candidate_id": None}
        ]
    })
    results["historical_feedbacks_deleted"] += suspicious_delete.deleted_count
    
    # 2. Get mentors with historical data (ratings or sessions_conducted)
    mentors_with_historical = await db.mentors.find(
        {"$or": [
            {"rating": {"$exists": True, "$ne": None}},
            {"sessions_conducted": {"$exists": True, "$gt": 0}}
        ]},
        {"_id": 0, "id": 1, "name": 1, "rating": 1, "sessions_conducted": 1}
    ).to_list(length=500)
    
    for mentor in mentors_with_historical:
        results["mentors_affected"].append({
            "name": mentor.get("name"),
            "old_rating": mentor.get("rating"),
            "old_sessions_conducted": mentor.get("sessions_conducted")
        })
    
    # 3. Clear rating and sessions_conducted fields from all mentors
    update_result = await db.mentors.update_many(
        {"$or": [
            {"rating": {"$exists": True}},
            {"sessions_conducted": {"$exists": True}}
        ]},
        {"$unset": {"rating": "", "sessions_conducted": ""}}
    )
    results["mentor_ratings_cleared"] = update_result.modified_count
    results["mentor_sessions_cleared"] = update_result.modified_count
    
    return {
        "success": True,
        "message": "All historical ratings, feedbacks, and session counts have been cleared",
        "details": results
    }


@router.get("/feedbacks/check-historical")
async def check_historical_feedbacks(request: Request):
    """
    Check for feedbacks that might be historical but aren't marked as such.
    Useful for diagnosing rating calculation issues.
    """
    db = get_db(request)
    user = await get_current_user(request, db)
    
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get all feedbacks with their historical status
    all_feedbacks = await db.candidate_feedbacks.find(
        {},
        {"_id": 0, "id": 1, "mentor_id": 1, "candidate_id": 1, "booking_id": 1, 
         "is_historical": 1, "rating_overall": 1, "candidate_name_override": 1}
    ).to_list(length=1000)
    
    suspicious = []
    properly_marked = []
    
    for fb in all_feedbacks:
        is_marked_historical = fb.get("is_historical", False)
        looks_historical = (
            fb.get("candidate_id") is None or
            fb.get("booking_id") is None or
            (fb.get("id") or "").startswith("hist-fb-") or
            fb.get("candidate_name_override") is not None
        )
        
        mentor = await db.mentors.find_one({"id": fb.get("mentor_id")}, {"_id": 0, "name": 1})
        mentor_name = mentor.get("name") if mentor else fb.get("mentor_id")
        
        fb_info = {
            "feedback_id": fb.get("id"),
            "mentor": mentor_name,
            "rating": fb.get("rating_overall"),
            "is_historical_flag": is_marked_historical,
            "looks_historical": looks_historical,
            "candidate_id": fb.get("candidate_id"),
            "booking_id": fb.get("booking_id"),
            "candidate_name_override": fb.get("candidate_name_override")
        }
        
        if looks_historical and not is_marked_historical:
            suspicious.append(fb_info)
        elif is_marked_historical:
            properly_marked.append(fb_info)
    
    return {
        "total_feedbacks": len(all_feedbacks),
        "properly_marked_historical": len(properly_marked),
        "suspicious_unmarked": len(suspicious),
        "suspicious_feedbacks": suspicious,
        "note": "Suspicious feedbacks look like historical imports but aren't marked as is_historical=true. Run POST /api/admin/feedbacks/fix-historical-flags to fix them."
    }




# ============ Excel Export Endpoints ============

@router.get("/peer-sessions/export-excel")
async def export_peer_sessions_to_excel(request: Request):
    """Export all peer practice sessions to Excel file"""
    await verify_admin(request)
    db = get_db(request)
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        from fastapi.responses import StreamingResponse
        
        # Fetch ALL peer sessions
        sessions = await db.peer_sessions.find({}, {"_id": 0}).to_list(10000)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Peer Practice Sessions"
        
        # Define headers
        headers = [
            "Session ID", "Requester Name", "Requester Email", "Partner Name", 
            "Partner Email", "Date", "Time", "Status", "Case Type", "Focus Area",
            "Meet Link", "Created At", "Matched At", "Cancelled At", "Remarks"
        ]
        
        # Style header row
        header_fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data rows
        for row_idx, session in enumerate(sessions, 2):
            ws.cell(row=row_idx, column=1, value=session.get("id", ""))
            ws.cell(row=row_idx, column=2, value=session.get("requester_name", ""))
            ws.cell(row=row_idx, column=3, value=session.get("requester_email", ""))
            ws.cell(row=row_idx, column=4, value=session.get("partner_name", ""))
            ws.cell(row=row_idx, column=5, value=session.get("partner_email", ""))
            ws.cell(row=row_idx, column=6, value=session.get("date", ""))
            ws.cell(row=row_idx, column=7, value=session.get("time", ""))
            ws.cell(row=row_idx, column=8, value=session.get("status", ""))
            ws.cell(row=row_idx, column=9, value=session.get("case_type", ""))
            ws.cell(row=row_idx, column=10, value=session.get("focus_area", ""))
            ws.cell(row=row_idx, column=11, value=session.get("meet_link", ""))
            ws.cell(row=row_idx, column=12, value=session.get("created_at", ""))
            ws.cell(row=row_idx, column=13, value=session.get("matched_at", ""))
            ws.cell(row=row_idx, column=14, value=session.get("cancelled_at", ""))
            ws.cell(row=row_idx, column=15, value=session.get("remarks", ""))
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Return as streaming response
        filename = f"peer_sessions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error exporting peer sessions: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to export sessions: {str(e)}")



# ============ Session Reminders ============

@router.post("/reminders/trigger")
async def trigger_session_reminders(request: Request):
    """Manually trigger the session reminder check"""
    await verify_admin(request)
    
    try:
        from services.session_reminder_service import process_reminders
        await process_reminders()
        return {"message": "Reminder check triggered successfully"}
    except Exception as e:
        logger.error(f"Error triggering reminders: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/reminders/status")
async def get_reminder_status(request: Request):
    """Get status of sent reminders"""
    await verify_admin(request)
    db = get_db(request)
    
    try:
        # Get recent reminders
        recent_reminders = await db.session_reminders.find({}).sort("sent_at", -1).limit(50).to_list(50)
        
        # Count by type
        reminder_counts = {
            "24h": 0,
            "4h": 0,
            "15min": 0
        }
        
        for r in recent_reminders:
            reminder_type = r.get("reminder_type")
            if reminder_type in reminder_counts:
                reminder_counts[reminder_type] += 1
        
        return {
            "total_reminders_sent": await db.session_reminders.count_documents({}),
            "recent_reminders": [
                {
                    "booking_id": r.get("booking_id"),
                    "reminder_type": r.get("reminder_type"),
                    "candidate_sent": r.get("candidate_sent"),
                    "mentor_sent": r.get("mentor_sent"),
                    "sent_at": r.get("sent_at").isoformat() if r.get("sent_at") else None
                }
                for r in recent_reminders[:20]
            ],
            "counts_by_type": reminder_counts
        }
    except Exception as e:
        logger.error(f"Error getting reminder status: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/reminders/diagnostics")
async def get_reminder_diagnostics(request: Request):
    """Comprehensive diagnostics for WhatsApp session reminders"""
    await verify_admin(request)
    db = get_db(request)
    
    import pytz
    from datetime import datetime
    import os
    
    IST = pytz.timezone('Asia/Kolkata')
    now_utc = datetime.now(pytz.UTC)
    now_ist = datetime.now(IST)
    
    diagnostics = {
        "current_time": {
            "utc": now_utc.strftime('%Y-%m-%d %H:%M:%S %Z'),
            "ist": now_ist.strftime('%Y-%m-%d %H:%M:%S %Z'),
        },
        "wati_config": {
            "api_token_set": bool(os.environ.get('WATI_API_TOKEN')),
            "api_endpoint_set": bool(os.environ.get('WATI_API_ENDPOINT')),
            "whatsapp_number_set": bool(os.environ.get('WATI_WHATSAPP_NUMBER')),
            "api_endpoint": os.environ.get('WATI_API_ENDPOINT', '')[:50] + "..." if os.environ.get('WATI_API_ENDPOINT') else None,
        },
        "reminder_intervals": {
            "24h": "23.5 - 24.5 hours before session",
            "4h": "3.5 - 4.5 hours before session",
            "15min": "10 - 20 minutes before session"
        },
        "bookings": [],
        "issues_found": [],
        "sent_reminders_count": await db.session_reminders.count_documents({}),
        "phone_number_stats": {
            "bookings_with_candidate_phone": 0,
            "bookings_with_mentor_phone": 0,
            "bookings_missing_candidate_phone": 0,
            "bookings_missing_mentor_phone": 0,
        }
    }
    
    # Check WATI config
    if not os.environ.get('WATI_API_TOKEN'):
        diagnostics["issues_found"].append("WATI_API_TOKEN not configured")
    if not os.environ.get('WATI_API_ENDPOINT'):
        diagnostics["issues_found"].append("WATI_API_ENDPOINT not configured")
    
    # Get active bookings
    bookings = await db.bookings.find({
        "status": {"$in": ["confirmed", "pending"]}
    }).to_list(100)
    
    diagnostics["total_active_bookings"] = len(bookings)
    
    for booking in bookings[:15]:  # Analyze first 15
        # Get user and mentor info for phone number lookup
        user_id = booking.get("user_id")
        mentor_id = booking.get("mentor_id")
        
        # Lookup user phone if not in booking
        candidate_phone = booking.get("candidate_phone")
        candidate_country_code = booking.get("candidate_country_code", "+91")
        if not candidate_phone and user_id:
            user = await db.users.find_one({"id": user_id}, {"phone_number": 1, "phone": 1, "phone_country_code": 1})
            if user:
                candidate_phone = user.get("phone_number") or user.get("phone")
                candidate_country_code = user.get("phone_country_code", "+91")
        
        # Lookup mentor phone if not in booking
        mentor_phone = booking.get("mentor_phone")
        mentor_country_code = booking.get("mentor_country_code", "+91")
        if not mentor_phone and mentor_id:
            mentor = await db.mentors.find_one({"id": mentor_id}, {"phone_number": 1, "phone": 1, "phone_country_code": 1})
            if mentor:
                mentor_phone = mentor.get("phone_number") or mentor.get("phone")
                mentor_country_code = mentor.get("phone_country_code", "+91")
        
        booking_info = {
            "id": booking.get("id"),
            "date": booking.get("date"),
            "time": booking.get("time_slot") or booking.get("time"),
            "status": booking.get("status"),
            "candidate_name": booking.get("candidate_name") or booking.get("user_name"),
            "mentor_name": booking.get("mentor_name"),
            "phones": {
                "candidate_phone_in_booking": booking.get("candidate_phone"),
                "candidate_phone_from_user": candidate_phone if not booking.get("candidate_phone") else None,
                "candidate_phone_final": candidate_phone,
                "candidate_country_code": candidate_country_code,
                "mentor_phone_in_booking": booking.get("mentor_phone"),
                "mentor_phone_from_mentor": mentor_phone if not booking.get("mentor_phone") else None,
                "mentor_phone_final": mentor_phone,
                "mentor_country_code": mentor_country_code,
            },
            "issues": []
        }
        
        # Track phone number stats
        if candidate_phone:
            diagnostics["phone_number_stats"]["bookings_with_candidate_phone"] += 1
        else:
            diagnostics["phone_number_stats"]["bookings_missing_candidate_phone"] += 1
            booking_info["issues"].append("NO CANDIDATE PHONE (not in booking, not in user record)")
            
        if mentor_phone:
            diagnostics["phone_number_stats"]["bookings_with_mentor_phone"] += 1
        else:
            diagnostics["phone_number_stats"]["bookings_missing_mentor_phone"] += 1
            booking_info["issues"].append("NO MENTOR PHONE (not in booking, not in mentor record)")
        
        # Parse and calculate time until session
        date_str = booking.get("date")
        time_str = booking.get("time_slot") or booking.get("time")
        
        if date_str and time_str:
            try:
                time_clean = time_str.upper().replace(".", "").strip()
                try:
                    session_dt = datetime.strptime(f"{date_str} {time_clean}", "%Y-%m-%d %I:%M %p")
                except ValueError:
                    session_dt = datetime.strptime(f"{date_str} {time_clean}", "%Y-%m-%d %H:%M")
                
                session_dt_ist = IST.localize(session_dt)
                hours_until = (session_dt_ist - now_ist).total_seconds() / 3600
                
                booking_info["session_datetime_ist"] = session_dt_ist.strftime('%Y-%m-%d %H:%M:%S %Z')
                booking_info["hours_until_session"] = round(hours_until, 2)
                
                # Check which reminder window it falls into
                if hours_until < 0:
                    booking_info["reminder_status"] = "Session in past"
                elif 23.5 <= hours_until <= 24.5:
                    booking_info["reminder_status"] = "In 24h window NOW"
                elif 3.5 <= hours_until <= 4.5:
                    booking_info["reminder_status"] = "In 4h window NOW"
                elif 10/60 <= hours_until <= 20/60:
                    booking_info["reminder_status"] = "In 15min window NOW"
                elif hours_until > 24.5:
                    booking_info["reminder_status"] = f"Too early - 24h reminder in {round(hours_until - 24, 1)}h"
                elif hours_until > 4.5:
                    booking_info["reminder_status"] = f"Next: 4h reminder in {round(hours_until - 4, 1)}h"
                elif hours_until > 20/60:
                    booking_info["reminder_status"] = f"Next: 15min reminder in {round((hours_until - 0.25) * 60, 0)}min"
                else:
                    booking_info["reminder_status"] = "Between windows"
                    
            except Exception as e:
                booking_info["issues"].append(f"Date/time parse error: {str(e)}")
        else:
            booking_info["issues"].append("Missing date or time")
        
        # Check if reminders already sent
        sent_reminders = await db.session_reminders.find({
            "booking_id": booking.get("id")
        }).to_list(10)
        booking_info["reminders_already_sent"] = [r.get("reminder_type") for r in sent_reminders]
        
        diagnostics["bookings"].append(booking_info)
    
    # Add summary of issues
    if diagnostics["phone_number_stats"]["bookings_missing_candidate_phone"] > 0:
        diagnostics["issues_found"].append(f"{diagnostics['phone_number_stats']['bookings_missing_candidate_phone']} bookings missing candidate phone")
    if diagnostics["phone_number_stats"]["bookings_missing_mentor_phone"] > 0:
        diagnostics["issues_found"].append(f"{diagnostics['phone_number_stats']['bookings_missing_mentor_phone']} bookings missing mentor phone")
    
    return diagnostics


@router.post("/reminders/test-send")
async def test_send_reminder(request: Request):
    """Test sending a WhatsApp reminder to verify WATI integration"""
    await verify_admin(request)
    
    import aiohttp
    import os
    
    body = await request.json()
    phone = body.get("phone", "")
    template_name = body.get("template_name", "candidate_session_reminder_4h")
    
    if not phone:
        raise HTTPException(status_code=400, detail="Phone number required")
    
    # Get WATI config
    api_token = os.environ.get('WATI_API_TOKEN')
    api_endpoint = os.environ.get('WATI_API_ENDPOINT')
    
    if not api_token or not api_endpoint:
        return {"success": False, "error": "WATI not configured"}
    
    # Format phone number - remove + and spaces (same as wati_service)
    phone_clean = phone.replace('+', '').replace(' ', '').replace('-', '')
    if not phone_clean.startswith('91') and len(phone_clean) == 10:
        phone_clean = '91' + phone_clean
    
    # Build request - USE QUERY PARAMETER FORMAT (same as wati_service)
    url = f"{api_endpoint}/api/v1/sendTemplateMessage?whatsappNumber={phone_clean}"
    
    headers = {
        "Authorization": f"Bearer {api_token}",
        "Content-Type": "application/json"
    }
    
    # Template requires 5 parameters:
    # {{1}} - Name, {{2}} - Session type, {{3}} - Other person name, {{4}} - Date, {{5}} - Time
    payload = {
        "template_name": template_name,
        "broadcast_name": f"Session Notification {template_name}",
        "parameters": [
            {"name": "1", "value": "Test Candidate"},
            {"name": "2", "value": "coaching"},
            {"name": "3", "value": "Test Mentor"},
            {"name": "4", "value": "2026-03-12"},
            {"name": "5", "value": "8:00 PM"}
        ]
    }
    
    try:
        async with aiohttp.ClientSession() as session:
            async with session.post(url, json=payload, headers=headers) as response:
                response_text = await response.text()
                
                return {
                    "success": response.status in [200, 201],
                    "request": {
                        "url": url,
                        "phone": phone_clean,
                        "template_name": template_name,
                        "parameters": payload["parameters"]
                    },
                    "response": {
                        "status": response.status,
                        "body": response_text
                    }
                }
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "request": {
                "url": url,
                "phone": phone_clean,
                "template_name": template_name
            }
        }


@router.get("/reminders/whatsapp-diagnostics")
async def whatsapp_diagnostics(request: Request):
    """
    Comprehensive WhatsApp reminder diagnostics for production debugging.
    Checks WATI config, scheduler status, bookings, and template configuration.
    """
    await verify_admin(request)
    db = get_db(request)
    
    import pytz
    from datetime import datetime, timedelta
    import os
    
    IST = pytz.timezone('Asia/Kolkata')
    now_utc = datetime.now(pytz.UTC)
    now_ist = datetime.now(IST)
    
    diagnostics = {
        "timestamp": {
            "utc": now_utc.strftime('%Y-%m-%d %H:%M:%S %Z'),
            "ist": now_ist.strftime('%Y-%m-%d %H:%M:%S %Z'),
        },
        "wati_config": {
            "WATI_API_TOKEN": "SET" if os.environ.get('WATI_API_TOKEN') else "NOT SET",
            "WATI_API_ENDPOINT": os.environ.get('WATI_API_ENDPOINT', 'NOT SET'),
            "WATI_WHATSAPP_NUMBER": os.environ.get('WATI_WHATSAPP_NUMBER', 'NOT SET'),
        },
        "expected_templates": [
            "candidate_session_reminder_24h",
            "mentor_session_reminder_24h",
            "candidate_session_reminder_4h",
            "mentor_session_reminder_4h",
            "candidate_session_reminder_15min",
            "mentor_session_reminder_15min"
        ],
        "reminder_windows": {
            "24h": "23.5 - 24.5 hours before session",
            "4h": "3.5 - 4.5 hours before session",
            "15min": "10 - 20 minutes before session"
        },
        "scheduler_interval": "15 minutes",
        "bookings_analysis": [],
        "sent_whatsapp_reminders": [],
        "issues": []
    }
    
    # Check WATI config issues
    if not os.environ.get('WATI_API_TOKEN'):
        diagnostics["issues"].append("❌ WATI_API_TOKEN is not set")
    if not os.environ.get('WATI_API_ENDPOINT'):
        diagnostics["issues"].append("❌ WATI_API_ENDPOINT is not set")
    if not os.environ.get('WATI_WHATSAPP_NUMBER'):
        diagnostics["issues"].append("❌ WATI_WHATSAPP_NUMBER is not set")
    
    # Get active bookings
    bookings = await db.bookings.find({
        "status": {"$in": ["confirmed", "pending"]}
    }).sort("date", 1).to_list(50)
    
    diagnostics["total_active_bookings"] = len(bookings)
    
    bookings_in_window = 0
    bookings_missing_phones = 0
    
    for booking in bookings[:20]:
        # Get phone numbers
        user_id = booking.get("user_id")
        mentor_id = booking.get("mentor_id")
        
        # Lookup phones from user/mentor if not in booking
        candidate_phone = booking.get("candidate_phone")
        mentor_phone = booking.get("mentor_phone")
        
        if not candidate_phone and user_id:
            user = await db.users.find_one({"id": user_id})
            if user:
                candidate_phone = user.get("phone_number") or user.get("phone")
        
        if not mentor_phone and mentor_id:
            mentor = await db.mentors.find_one({"id": mentor_id})
            if mentor:
                mentor_phone = mentor.get("phone_number") or mentor.get("phone")
        
        # Parse session time
        date_str = booking.get("date")
        time_str = booking.get("time_slot") or booking.get("time")
        hours_until = None
        reminder_window = "N/A"
        
        if date_str and time_str:
            try:
                time_clean = time_str.upper().replace(".", "").strip()
                try:
                    session_dt = datetime.strptime(f"{date_str} {time_clean}", "%Y-%m-%d %I:%M %p")
                except ValueError:
                    session_dt = datetime.strptime(f"{date_str} {time_clean}", "%Y-%m-%d %H:%M")
                
                session_dt_ist = IST.localize(session_dt)
                hours_until = (session_dt_ist - now_ist).total_seconds() / 3600
                
                # Determine window
                if hours_until < 0:
                    reminder_window = "PAST"
                elif 23.5 <= hours_until <= 24.5:
                    reminder_window = "IN 24h WINDOW NOW ✅"
                    bookings_in_window += 1
                elif 3.5 <= hours_until <= 4.5:
                    reminder_window = "IN 4h WINDOW NOW ✅"
                    bookings_in_window += 1
                elif 10/60 <= hours_until <= 20/60:
                    reminder_window = "IN 15min WINDOW NOW ✅"
                    bookings_in_window += 1
                elif hours_until > 24.5:
                    reminder_window = f"24h reminder in {round(hours_until - 24, 1)}h"
                elif hours_until > 4.5:
                    reminder_window = f"4h reminder in {round(hours_until - 4, 1)}h"
                elif hours_until > 20/60:
                    reminder_window = f"15min reminder in {round((hours_until - 0.25) * 60, 0)}min"
                else:
                    reminder_window = "Between windows"
            except Exception as e:
                reminder_window = f"Parse error: {e}"
        
        # Check for missing phones
        phone_status = "OK"
        if not candidate_phone and not mentor_phone:
            phone_status = "❌ BOTH MISSING"
            bookings_missing_phones += 1
        elif not candidate_phone:
            phone_status = "❌ Candidate phone missing"
            bookings_missing_phones += 1
        elif not mentor_phone:
            phone_status = "❌ Mentor phone missing"
            bookings_missing_phones += 1
        
        # Check sent reminders
        sent = await db.session_reminders.find({"booking_id": booking.get("id")}).to_list(10)
        sent_types = [s.get("reminder_type") for s in sent]
        
        diagnostics["bookings_analysis"].append({
            "booking_id": booking.get("id"),
            "session": f"{date_str} {time_str}",
            "hours_until": round(hours_until, 2) if hours_until else None,
            "reminder_window": reminder_window,
            "candidate_phone": candidate_phone[:4] + "****" if candidate_phone else None,
            "mentor_phone": mentor_phone[:4] + "****" if mentor_phone else None,
            "phone_status": phone_status,
            "reminders_sent": sent_types
        })
    
    # Get recent sent WhatsApp reminders
    recent_reminders = await db.session_reminders.find().sort("sent_at", -1).limit(20).to_list(20)
    for r in recent_reminders:
        diagnostics["sent_whatsapp_reminders"].append({
            "booking_id": r.get("booking_id"),
            "type": r.get("reminder_type"),
            "candidate_sent": r.get("candidate_sent"),
            "mentor_sent": r.get("mentor_sent"),
            "sent_at": r.get("sent_at").isoformat() if r.get("sent_at") else None
        })
    
    # Summary stats
    diagnostics["summary"] = {
        "total_active_bookings": len(bookings),
        "bookings_in_reminder_window": bookings_in_window,
        "bookings_missing_phones": bookings_missing_phones,
        "total_whatsapp_reminders_sent": await db.session_reminders.count_documents({})
    }
    
    # Add issues based on analysis
    if bookings_in_window == 0 and len(bookings) > 0:
        diagnostics["issues"].append("⚠️ No bookings currently in any reminder window (24h/4h/15min)")
    if bookings_missing_phones > 0:
        diagnostics["issues"].append(f"⚠️ {bookings_missing_phones} bookings have missing phone numbers")
    if len(diagnostics["sent_whatsapp_reminders"]) == 0:
        diagnostics["issues"].append("⚠️ No WhatsApp reminders have been sent yet")
    
    return diagnostics


@router.post("/reminders/trigger-now")
async def trigger_whatsapp_reminders_now(request: Request):
    """
    Manually trigger the WhatsApp reminder check.
    This runs the same logic as the background scheduler.
    """
    await verify_admin(request)
    
    from services.session_reminder_service import process_reminders
    import traceback
    
    try:
        # Run the reminder process
        await process_reminders()
        
        # Get stats after running
        db = get_db(request)
        total_sent = await db.session_reminders.count_documents({})
        recent = await db.session_reminders.find().sort("sent_at", -1).limit(5).to_list(5)
        
        return {
            "success": True,
            "message": "WhatsApp reminder check completed",
            "total_reminders_in_db": total_sent,
            "recent_reminders": [
                {
                    "booking_id": r.get("booking_id"),
                    "type": r.get("reminder_type"),
                    "candidate_sent": r.get("candidate_sent"),
                    "mentor_sent": r.get("mentor_sent"),
                    "sent_at": r.get("sent_at").isoformat() if r.get("sent_at") else None
                }
                for r in recent
            ]
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "traceback": traceback.format_exc()
        }


@router.get("/reminders/scheduler-status")
async def get_scheduler_status(request: Request):
    """Check if the reminder scheduler is running"""
    await verify_admin(request)
    
    import asyncio
    
    status = {
        "reminder_task_exists": hasattr(request.app.state, 'reminder_task'),
        "reminder_task_done": None,
        "reminder_task_cancelled": None,
        "automations_task_exists": hasattr(request.app.state, 'automations_task'),
        "automations_task_done": None,
        "abandoned_checkout_task_exists": hasattr(request.app.state, 'abandoned_checkout_task'),
    }
    
    if hasattr(request.app.state, 'reminder_task'):
        task = request.app.state.reminder_task
        status["reminder_task_done"] = task.done()
        status["reminder_task_cancelled"] = task.cancelled()
        if task.done() and not task.cancelled():
            try:
                exc = task.exception()
                if exc:
                    status["reminder_task_exception"] = str(exc)
            except:
                pass
    
    if hasattr(request.app.state, 'automations_task'):
        task = request.app.state.automations_task
        status["automations_task_done"] = task.done()
        status["automations_task_cancelled"] = task.cancelled()
        if task.done() and not task.cancelled():
            try:
                exc = task.exception()
                if exc:
                    status["automations_task_exception"] = str(exc)
            except:
                pass
    
    return status


@router.post("/reminders/restart-scheduler")
async def restart_reminder_scheduler(request: Request):
    """Restart the reminder scheduler if it crashed"""
    await verify_admin(request)
    
    import asyncio
    from services.session_reminder_service import start_reminder_scheduler
    
    result = {
        "previous_task_existed": hasattr(request.app.state, 'reminder_task'),
        "previous_task_done": None,
        "new_task_created": False
    }
    
    # Check if old task exists and its status
    if hasattr(request.app.state, 'reminder_task'):
        old_task = request.app.state.reminder_task
        result["previous_task_done"] = old_task.done()
        result["previous_task_cancelled"] = old_task.cancelled()
        
        # If task crashed, get the exception
        if old_task.done() and not old_task.cancelled():
            try:
                exc = old_task.exception()
                if exc:
                    result["previous_task_exception"] = str(exc)
            except:
                pass
        
        # Cancel old task if still running
        if not old_task.done():
            old_task.cancel()
            result["old_task_cancelled"] = True
    
    # Create new task
    try:
        request.app.state.reminder_task = asyncio.create_task(start_reminder_scheduler(interval_minutes=15))
        result["new_task_created"] = True
        result["message"] = "Reminder scheduler restarted successfully"
    except Exception as e:
        result["error"] = str(e)
    
    return result


@router.post("/reminders/trigger-debug")
async def trigger_whatsapp_reminders_debug(request: Request):
    """
    Debug version of the reminder trigger that shows detailed processing info.
    """
    await verify_admin(request)
    db = get_db(request)
    
    import pytz
    from datetime import datetime
    from services.wati_service import wati_service
    import os
    import traceback
    
    IST = pytz.timezone('Asia/Kolkata')
    now_ist = datetime.now(IST)
    
    debug_info = {
        "current_time_ist": now_ist.strftime('%Y-%m-%d %H:%M:%S %Z'),
        "wati_configured": bool(os.environ.get('WATI_API_TOKEN')),
        "bookings_in_window": [],
        "bookings_not_in_window": [],
        "reminders_sent": [],
        "errors": []
    }
    
    REMINDER_INTERVALS = {
        '24h': 24,
        '4h': 4,
        '15min': 0.25
    }
    
    try:
        # Get active bookings
        bookings = await db.bookings.find({
            "status": {"$in": ["confirmed", "pending"]}
        }).to_list(500)
        
        debug_info["total_bookings_found"] = len(bookings)
        
        for booking in bookings:  # Process ALL bookings
            booking_debug = {
                "id": booking.get("id"),
                "date": booking.get("date"),
                "time": booking.get("time_slot") or booking.get("time"),
                "status": booking.get("status"),
                "in_any_window": False,
                "checks": []
            }
            
            # Parse session datetime
            date_str = booking.get("date")
            time_str = booking.get("time_slot") or booking.get("time")
            
            if not date_str or not time_str:
                booking_debug["skip_reason"] = "Missing date or time"
                debug_info["bookings_not_in_window"].append(booking_debug)
                continue
            
            try:
                time_clean = time_str.upper().replace(".", "").strip()
                try:
                    session_dt = datetime.strptime(f"{date_str} {time_clean}", "%Y-%m-%d %I:%M %p")
                except ValueError:
                    session_dt = datetime.strptime(f"{date_str} {time_clean}", "%Y-%m-%d %H:%M")
                
                session_dt_ist = IST.localize(session_dt)
                hours_until = (session_dt_ist - now_ist).total_seconds() / 3600
                
                booking_debug["session_dt_ist"] = session_dt_ist.strftime('%Y-%m-%d %H:%M:%S')
                booking_debug["hours_until"] = round(hours_until, 2)
                
            except Exception as e:
                booking_debug["skip_reason"] = f"Date parse error - {e}"
                debug_info["bookings_not_in_window"].append(booking_debug)
                continue
            
            if hours_until < 0:
                booking_debug["skip_reason"] = "Session in past"
                debug_info["bookings_not_in_window"].append(booking_debug)
                continue
            
            # Get phone numbers - ALWAYS from user/mentor records first
            candidate = await db.users.find_one({"id": booking.get("user_id")})
            mentor = await db.mentors.find_one({"id": booking.get("mentor_id")})
            
            candidate_phone = None
            mentor_phone = None
            candidate_name = booking.get("candidate_name", "Candidate")
            mentor_name = booking.get("mentor_name", "Mentor")
            candidate_country_code = "+91"
            mentor_country_code = "+91"
            
            if candidate:
                candidate_phone = candidate.get("phone_number") or candidate.get("phone")
                candidate_name = candidate.get("name", candidate_name)
                candidate_country_code = candidate.get("phone_country_code", "+91")
            if not candidate_phone:
                candidate_phone = booking.get("candidate_phone")
                candidate_country_code = booking.get("candidate_country_code", "+91")
            
            if mentor:
                mentor_phone = mentor.get("phone_number") or mentor.get("phone")
                mentor_name = mentor.get("name", mentor_name)
                mentor_country_code = mentor.get("phone_country_code", "+91")
            if not mentor_phone:
                mentor_phone = booking.get("mentor_phone")
                mentor_country_code = booking.get("mentor_country_code", "+91")
            
            booking_debug["candidate_phone"] = candidate_phone[:4] + "****" if candidate_phone else None
            booking_debug["mentor_phone"] = mentor_phone[:4] + "****" if mentor_phone else None
            
            if not candidate_phone and not mentor_phone:
                booking_debug["skip_reason"] = "No phone numbers available"
                debug_info["bookings_not_in_window"].append(booking_debug)
                continue
            
            # Check each reminder interval
            for reminder_type, hours_before in REMINDER_INTERVALS.items():
                window_start = hours_before - 0.5
                window_end = hours_before + 0.5
                
                if reminder_type == '15min':
                    window_start = 10/60  # 10 minutes
                    window_end = 20/60    # 20 minutes
                
                in_window = window_start <= hours_until <= window_end
                
                check_result = {
                    "type": reminder_type,
                    "window": f"{window_start:.2f}h - {window_end:.2f}h",
                    "hours_until": round(hours_until, 2),
                    "in_window": in_window
                }
                
                if in_window:
                    booking_debug["in_any_window"] = True
                    
                    # Check if already sent
                    reminder_key = f"{booking.get('id')}_{reminder_type}"
                    existing = await db.session_reminders.find_one({"reminder_key": reminder_key})
                    
                    if existing:
                        check_result["action"] = "SKIP: Already sent"
                    else:
                        check_result["action"] = "SENDING..."
                        
                        # Format phone numbers
                        def format_phone(phone, country_code):
                            if not phone:
                                return None
                            phone = phone.replace(" ", "").replace("-", "")
                            if phone.startswith("+"):
                                return phone
                            if phone.startswith("0"):
                                phone = phone[1:]
                            return f"{country_code}{phone}"
                        
                        candidate_full_phone = format_phone(candidate_phone, candidate_country_code)
                        mentor_full_phone = format_phone(mentor_phone, mentor_country_code)
                        
                        # Try to send
                        try:
                            candidate_sent = False
                            mentor_sent = False
                            
                            # Get session type from booking
                            session_type = booking.get("session_type", "coaching")
                            if "strategy" in session_type.lower():
                                session_type = "strategy call"
                            else:
                                session_type = "coaching"
                            
                            if candidate_full_phone:
                                template_name = f"candidate_session_reminder_{reminder_type}"
                                # Template requires 5 parameters:
                                # {{1}} - Name, {{2}} - Session type, {{3}} - Mentor name, {{4}} - Date, {{5}} - Time
                                result = await wati_service.send_template_message(
                                    recipient_number=candidate_full_phone,
                                    template_name=template_name,
                                    parameters=[
                                        {"name": "1", "value": candidate_name},
                                        {"name": "2", "value": session_type},
                                        {"name": "3", "value": mentor_name},
                                        {"name": "4", "value": date_str},
                                        {"name": "5", "value": time_str}
                                    ]
                                )
                                candidate_sent = True
                                check_result["candidate_result"] = f"SUCCESS to {candidate_full_phone[:6]}****"
                            
                            if mentor_full_phone:
                                template_name = f"mentor_session_reminder_{reminder_type}"
                                # Template requires 5 parameters:
                                # {{1}} - Name, {{2}} - Session type, {{3}} - Candidate name, {{4}} - Date, {{5}} - Time
                                result = await wati_service.send_template_message(
                                    recipient_number=mentor_full_phone,
                                    template_name=template_name,
                                    parameters=[
                                        {"name": "1", "value": mentor_name},
                                        {"name": "2", "value": session_type},
                                        {"name": "3", "value": candidate_name},
                                        {"name": "4", "value": date_str},
                                        {"name": "5", "value": time_str}
                                    ]
                                )
                                mentor_sent = True
                                check_result["mentor_result"] = f"SUCCESS to {mentor_full_phone[:6]}****"
                            
                            # Record the reminder
                            await db.session_reminders.insert_one({
                                "reminder_key": reminder_key,
                                "booking_id": booking.get("id"),
                                "reminder_type": reminder_type,
                                "candidate_sent": candidate_sent,
                                "mentor_sent": mentor_sent,
                                "sent_at": datetime.utcnow()
                            })
                            
                            debug_info["reminders_sent"].append({
                                "booking_id": booking.get("id"),
                                "type": reminder_type,
                                "candidate": candidate_sent,
                                "mentor": mentor_sent
                            })
                            
                        except Exception as e:
                            check_result["action"] = f"ERROR: {str(e)}"
                            debug_info["errors"].append({
                                "booking_id": booking.get("id"),
                                "type": reminder_type,
                                "error": str(e)
                            })
                
                booking_debug["checks"].append(check_result)
            
            # Add to appropriate list
            if booking_debug["in_any_window"]:
                debug_info["bookings_in_window"].append(booking_debug)
            else:
                # Only keep summary for bookings not in window
                debug_info["bookings_not_in_window"].append({
                    "id": booking_debug["id"],
                    "session": f"{date_str} {time_str}",
                    "hours_until": booking_debug["hours_until"]
                })
        
        # Summary
        debug_info["summary"] = {
            "total_bookings": len(bookings),
            "bookings_in_window": len(debug_info["bookings_in_window"]),
            "bookings_not_in_window": len(debug_info["bookings_not_in_window"]),
            "reminders_sent": len(debug_info["reminders_sent"]),
            "errors": len(debug_info["errors"])
        }
        
        return debug_info
        
    except Exception as e:
        debug_info["fatal_error"] = str(e)
        debug_info["traceback"] = traceback.format_exc()
        return debug_info


# ============ Abandoned Checkout Recovery ============

@router.get("/abandoned-checkout/templates")
async def get_abandoned_checkout_templates(request: Request):
    """Get all abandoned checkout email templates"""
    await verify_admin(request)
    db = get_db(request)
    
    from services.abandoned_checkout_service import DEFAULT_TEMPLATES
    
    templates = {}
    for interval in ['1h', '24h', '72h']:
        # Check if custom template exists
        custom = await db.abandoned_checkout_templates.find_one({"interval": interval})
        if custom:
            templates[interval] = {
                "interval": interval,
                "subject": custom.get("subject"),
                "body": custom.get("body"),
                "is_custom": True,
                "updated_at": custom.get("updated_at")
            }
        else:
            templates[interval] = {
                "interval": interval,
                "subject": DEFAULT_TEMPLATES[interval]['subject'],
                "body": DEFAULT_TEMPLATES[interval]['body'],
                "is_custom": False
            }
    
    return {"templates": templates}


@router.put("/abandoned-checkout/templates/{interval}")
async def update_abandoned_checkout_template(interval: str, request: Request):
    """Update an abandoned checkout email template"""
    await verify_admin(request)
    db = get_db(request)
    
    if interval not in ['1h', '24h', '72h']:
        raise HTTPException(status_code=400, detail="Invalid interval. Must be 1h, 24h, or 72h")
    
    body = await request.json()
    subject = body.get("subject")
    email_body = body.get("body")
    
    if not subject or not email_body:
        raise HTTPException(status_code=400, detail="Subject and body are required")
    
    await db.abandoned_checkout_templates.update_one(
        {"interval": interval},
        {
            "$set": {
                "interval": interval,
                "subject": subject,
                "body": email_body,
                "updated_at": datetime.utcnow()
            }
        },
        upsert=True
    )
    
    return {"message": f"Template for {interval} updated successfully"}


@router.delete("/abandoned-checkout/templates/{interval}")
async def reset_abandoned_checkout_template(interval: str, request: Request):
    """Reset an abandoned checkout email template to default"""
    await verify_admin(request)
    db = get_db(request)
    
    if interval not in ['1h', '24h', '72h']:
        raise HTTPException(status_code=400, detail="Invalid interval. Must be 1h, 24h, or 72h")
    
    await db.abandoned_checkout_templates.delete_one({"interval": interval})
    
    return {"message": f"Template for {interval} reset to default"}


@router.get("/abandoned-checkout/stats")
async def get_abandoned_checkout_stats(request: Request):
    """Get abandoned checkout recovery statistics"""
    await verify_admin(request)
    db = get_db(request)
    
    # Count abandoned orders
    abandoned_count = await db.payment_orders.count_documents({"status": "created"})
    
    # Count recovery emails sent
    emails_sent = await db.abandoned_checkout_emails.count_documents({})
    
    # Count by interval
    emails_by_interval = {}
    for interval in ['1h', '24h', '72h']:
        count = await db.abandoned_checkout_emails.count_documents({"interval": interval})
        emails_by_interval[interval] = count
    
    # Recent recovery emails
    recent_emails = await db.abandoned_checkout_emails.find({}).sort("sent_at", -1).limit(20).to_list(20)
    
    return {
        "abandoned_orders": abandoned_count,
        "total_recovery_emails_sent": emails_sent,
        "emails_by_interval": emails_by_interval,
        "recent_emails": [
            {
                "order_id": e.get("order_id"),
                "user_email": e.get("user_email"),
                "interval": e.get("interval"),
                "sent_at": e.get("sent_at").isoformat() if e.get("sent_at") else None
            }
            for e in recent_emails
        ]
    }


@router.post("/abandoned-checkout/trigger")
async def trigger_abandoned_checkout_check(request: Request):
    """Manually trigger abandoned checkout recovery check"""
    await verify_admin(request)
    
    try:
        from services.abandoned_checkout_service import process_abandoned_checkouts
        await process_abandoned_checkouts()
        return {"message": "Abandoned checkout check triggered successfully"}
    except Exception as e:
        logger.error(f"Error triggering abandoned checkout check: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/abandoned-checkout/test-email")
async def send_test_abandoned_checkout_email(request: Request):
    """Send a test abandoned checkout email"""
    await verify_admin(request)
    db = get_db(request)
    
    body = await request.json()
    test_email = body.get("email")
    interval = body.get("interval", "1h")
    
    if not test_email:
        raise HTTPException(status_code=400, detail="Email is required")
    
    if interval not in ['1h', '24h', '72h']:
        raise HTTPException(status_code=400, detail="Invalid interval")
    
    try:
        from services.abandoned_checkout_service import send_recovery_email
        
        sent = await send_recovery_email(
            db=db,
            user_email=test_email,
            user_name="Test User",
            plan_name="Pro Plan",
            amount=7999,
            interval=interval
        )
        
        if sent:
            return {"message": f"Test email ({interval}) sent to {test_email}"}
        else:
            raise HTTPException(status_code=500, detail="Failed to send test email")
    except Exception as e:
        logger.error(f"Error sending test email: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ Data Backfill Endpoints ============

@router.post("/backfill/last-login-at")
async def backfill_last_login_at(request: Request):
    """
    Backfill last_login_at for users who don't have it set.
    Uses the most recent session timestamp for each user.
    """
    await verify_admin(request)
    db = get_db(request)
    
    # Find users without last_login_at
    users_without_login = await db.users.find({
        "$or": [
            {"last_login_at": {"$exists": False}},
            {"last_login_at": None}
        ]
    }, {"_id": 0, "id": 1, "email": 1, "created_at": 1}).to_list(10000)
    
    updated_count = 0
    
    for user in users_without_login:
        user_id = user.get("id")
        
        # Find the most recent session for this user
        session = await db.user_sessions.find_one(
            {"user_id": user_id},
            sort=[("created_at", -1)]
        )
        
        if session and session.get("created_at"):
            # Use session creation time as last_login_at
            last_login = session.get("created_at")
            if isinstance(last_login, str):
                login_time = last_login
            else:
                login_time = last_login.isoformat() if hasattr(last_login, 'isoformat') else str(last_login)
        else:
            # Fall back to user creation time
            created_at = user.get("created_at")
            if created_at:
                if isinstance(created_at, str):
                    login_time = created_at
                else:
                    login_time = created_at.isoformat() if hasattr(created_at, 'isoformat') else str(created_at)
            else:
                continue
        
        # Update the user
        await db.users.update_one(
            {"id": user_id},
            {"$set": {"last_login_at": login_time}}
        )
        updated_count += 1
    
    logger.info(f"Backfilled last_login_at for {updated_count} users")
    
    return {
        "success": True,
        "users_processed": len(users_without_login),
        "users_updated": updated_count,
        "message": f"Backfilled last_login_at for {updated_count} users"
    }


@router.post("/backfill/razorpay-payments")
async def sync_razorpay_payments(
    request: Request,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    count: int = 100
):
    """
    Sync payments from Razorpay API to database.
    
    This fetches all captured payments from Razorpay and stores them in the payments collection.
    Skips payments that already exist (based on razorpay_payment_id).
    
    Query params:
    - from_date: Start date (Unix timestamp or YYYY-MM-DD). Default: Feb 1, 2026
    - to_date: End date (Unix timestamp or YYYY-MM-DD). Default: now
    - count: Number of payments to fetch per request (max 100)
    """
    await verify_admin(request)
    db = get_db(request)
    
    import razorpay
    
    RAZORPAY_KEY_ID = os.environ.get("RAZORPAY_KEY_ID", "")
    RAZORPAY_KEY_SECRET = os.environ.get("RAZORPAY_KEY_SECRET", "")
    
    if not RAZORPAY_KEY_ID or not RAZORPAY_KEY_SECRET:
        raise HTTPException(status_code=400, detail="Razorpay credentials not configured")
    
    razorpay_client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))
    
    # Parse dates
    try:
        if from_date:
            if from_date.isdigit():
                from_timestamp = int(from_date)
            else:
                from_dt = datetime.strptime(from_date, "%Y-%m-%d")
                from_timestamp = int(from_dt.timestamp())
        else:
            # Default: Feb 1, 2026
            from_timestamp = int(datetime(2026, 2, 1).timestamp())
        
        if to_date:
            if to_date.isdigit():
                to_timestamp = int(to_date)
            else:
                to_dt = datetime.strptime(to_date, "%Y-%m-%d")
                to_timestamp = int(to_dt.timestamp())
        else:
            # Default: now
            to_timestamp = int(datetime.now().timestamp())
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Invalid date format: {e}")
    
    # Fetch payments from Razorpay
    all_payments = []
    skip = 0
    
    logger.info(f"Starting Razorpay payment sync from {from_timestamp} to {to_timestamp}")
    
    while True:
        try:
            response = razorpay_client.payment.all({
                "from": from_timestamp,
                "to": to_timestamp,
                "count": count,
                "skip": skip
            })
            
            payments = response.get("items", [])
            if not payments:
                break
            
            all_payments.extend(payments)
            skip += count
            
            # Safety limit
            if skip >= 10000:
                logger.warning("Reached 10000 payment limit, stopping fetch")
                break
                
        except Exception as e:
            logger.error(f"Error fetching payments from Razorpay: {e}")
            raise HTTPException(status_code=500, detail=f"Razorpay API error: {str(e)}")
    
    logger.info(f"Fetched {len(all_payments)} payments from Razorpay")
    
    # Process and sync payments
    synced_count = 0
    skipped_count = 0
    failed_count = 0
    
    for payment in all_payments:
        try:
            razorpay_payment_id = payment.get("id")
            
            # Check if already exists
            existing = await db.payments.find_one({"razorpay_payment_id": razorpay_payment_id})
            if existing:
                skipped_count += 1
                continue
            
            # Only sync captured payments
            if payment.get("status") != "captured":
                skipped_count += 1
                continue
            
            # Extract user email from notes or contact
            user_email = payment.get("email") or payment.get("notes", {}).get("email")
            user_id = payment.get("notes", {}).get("user_id")
            
            # Try to find user by email if user_id not in notes
            if not user_id and user_email:
                user = await db.users.find_one({"email": user_email})
                if user:
                    user_id = user.get("id")
            
            # Extract plan info from notes or description
            notes = payment.get("notes", {})
            plan_key = notes.get("plan_key") or notes.get("plan")
            plan_name = notes.get("plan_name") or notes.get("plan_display_name")
            plan_category = notes.get("plan_category") or notes.get("category")
            
            # Convert amount from paisa to rupees
            amount = payment.get("amount", 0) / 100
            
            # Create payment record
            payment_record = {
                "id": f"pay_{str(uuid.uuid4())[:8]}",
                "order_id": payment.get("order_id") or f"order_sync_{uuid.uuid4().hex[:16]}",  # Unique order_id for index
                "razorpay_payment_id": razorpay_payment_id,
                "razorpay_order_id": payment.get("order_id"),
                "user_id": user_id,
                "user_email": user_email,
                "amount": amount,
                "currency": payment.get("currency", "INR"),
                "status": "captured",
                "plan_key": plan_key,
                "plan_name": plan_name,
                "plan_category": plan_category,
                "method": payment.get("method"),
                "description": payment.get("description"),
                "notes": notes,
                "created_at": datetime.fromtimestamp(payment.get("created_at", 0)).isoformat(),
                "synced_from_razorpay": True,
                "synced_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.payments.insert_one(payment_record)
            synced_count += 1
            
        except Exception as e:
            logger.error(f"Error syncing payment {payment.get('id')}: {e}")
            failed_count += 1
    
    logger.info(f"Razorpay sync complete: synced={synced_count}, skipped={skipped_count}, failed={failed_count}")
    
    return {
        "success": True,
        "total_fetched": len(all_payments),
        "synced": synced_count,
        "skipped": skipped_count,
        "failed": failed_count,
        "date_range": {
            "from": datetime.fromtimestamp(from_timestamp).isoformat(),
            "to": datetime.fromtimestamp(to_timestamp).isoformat()
        },
        "message": f"Synced {synced_count} payments from Razorpay"
    }


# ============ Partner Management ============

class CreatePartnerRequest(BaseModel):
    name: str
    contact_email: EmailStr
    notes: Optional[str] = None


class UpdatePartnerRequest(BaseModel):
    name: Optional[str] = None
    contact_email: Optional[EmailStr] = None
    assigned_mentor_ids: Optional[List[str]] = None
    is_active: Optional[bool] = None
    notes: Optional[str] = None


@router.get("/partners")
async def list_partners(request: Request):
    """List all partners"""
    await verify_admin(request)
    db = get_db(request)
    
    partners = await db.partners.find(
        {},
        {"_id": 0, "api_key_hash": 0}  # Don't expose the hash
    ).sort("created_at", -1).to_list(100)
    
    # Enrich with booking counts and mentor names
    for partner in partners:
        partner_id = partner.get("id")
        
        # Get booking counts
        total_bookings = await db.partner_bookings.count_documents({"partner_id": partner_id})
        scheduled_bookings = await db.partner_bookings.count_documents({
            "partner_id": partner_id,
            "status": "scheduled"
        })
        
        partner["total_bookings"] = total_bookings
        partner["scheduled_bookings"] = scheduled_bookings
        
        # Get assigned mentor names
        mentor_ids = partner.get("assigned_mentor_ids", [])
        if mentor_ids:
            mentors = await db.mentors.find(
                {"id": {"$in": mentor_ids}},
                {"_id": 0, "id": 1, "name": 1}
            ).to_list(100)
            partner["assigned_mentors"] = mentors
        else:
            partner["assigned_mentors"] = []
        
        # Convert datetime fields
        if partner.get("created_at"):
            partner["created_at"] = partner["created_at"].isoformat()
        if partner.get("updated_at"):
            partner["updated_at"] = partner["updated_at"].isoformat()
    
    return {"partners": partners, "count": len(partners)}


@router.post("/partners")
async def create_partner(request: Request, data: CreatePartnerRequest):
    """Create a new partner and generate API key"""
    await verify_admin(request)
    db = get_db(request)
    
    # Check if partner with same email exists
    existing = await db.partners.find_one({"contact_email": data.contact_email})
    if existing:
        raise HTTPException(status_code=400, detail="Partner with this email already exists")
    
    # Generate API key
    from routes.partner_api import generate_api_key
    full_key, key_hash, key_prefix = generate_api_key()
    
    partner_id = str(uuid.uuid4())
    partner_doc = {
        "id": partner_id,
        "name": data.name,
        "contact_email": data.contact_email,
        "api_key_hash": key_hash,
        "api_key_prefix": key_prefix,
        "assigned_mentor_ids": [],
        "is_active": True,
        "notes": data.notes,
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.partners.insert_one(partner_doc)
    
    logger.info(f"Partner created: {data.name} ({partner_id})")
    
    # Return the full API key ONLY on creation (will never be shown again)
    return {
        "success": True,
        "partner": {
            "id": partner_id,
            "name": data.name,
            "contact_email": data.contact_email,
            "api_key_prefix": key_prefix,
            "is_active": True
        },
        "api_key": full_key,  # IMPORTANT: Show this only once!
        "warning": "Save this API key now. It will not be shown again."
    }


@router.get("/partners/{partner_id}")
async def get_partner(partner_id: str, request: Request):
    """Get partner details"""
    await verify_admin(request)
    db = get_db(request)
    
    partner = await db.partners.find_one(
        {"id": partner_id},
        {"_id": 0, "api_key_hash": 0}
    )
    
    if not partner:
        raise HTTPException(status_code=404, detail="Partner not found")
    
    # Get assigned mentor details
    mentor_ids = partner.get("assigned_mentor_ids", [])
    if mentor_ids:
        mentors = await db.mentors.find(
            {"id": {"$in": mentor_ids}},
            {"_id": 0, "id": 1, "name": 1, "picture": 1, "consulting_firm": 1}
        ).to_list(100)
        partner["assigned_mentors"] = mentors
    else:
        partner["assigned_mentors"] = []
    
    # Get booking stats
    partner["total_bookings"] = await db.partner_bookings.count_documents({"partner_id": partner_id})
    partner["scheduled_bookings"] = await db.partner_bookings.count_documents({
        "partner_id": partner_id, "status": "scheduled"
    })
    partner["completed_bookings"] = await db.partner_bookings.count_documents({
        "partner_id": partner_id, "status": "completed"
    })
    partner["cancelled_bookings"] = await db.partner_bookings.count_documents({
        "partner_id": partner_id, "status": "cancelled"
    })
    
    # Convert datetime fields
    if partner.get("created_at"):
        partner["created_at"] = partner["created_at"].isoformat()
    if partner.get("updated_at"):
        partner["updated_at"] = partner["updated_at"].isoformat()
    
    return partner


@router.put("/partners/{partner_id}")
async def update_partner(partner_id: str, request: Request, data: UpdatePartnerRequest):
    """Update partner details"""
    await verify_admin(request)
    db = get_db(request)
    
    partner = await db.partners.find_one({"id": partner_id})
    if not partner:
        raise HTTPException(status_code=404, detail="Partner not found")
    
    update_data = {"updated_at": datetime.now(timezone.utc)}
    
    if data.name is not None:
        update_data["name"] = data.name
    if data.contact_email is not None:
        update_data["contact_email"] = data.contact_email
    if data.assigned_mentor_ids is not None:
        update_data["assigned_mentor_ids"] = data.assigned_mentor_ids
    if data.is_active is not None:
        update_data["is_active"] = data.is_active
    if data.notes is not None:
        update_data["notes"] = data.notes
    
    await db.partners.update_one(
        {"id": partner_id},
        {"$set": update_data}
    )
    
    logger.info(f"Partner updated: {partner_id}")
    
    return {"success": True, "message": "Partner updated successfully"}


@router.delete("/partners/{partner_id}")
async def deactivate_partner(partner_id: str, request: Request):
    """Deactivate a partner (soft delete)"""
    await verify_admin(request)
    db = get_db(request)
    
    partner = await db.partners.find_one({"id": partner_id})
    if not partner:
        raise HTTPException(status_code=404, detail="Partner not found")
    
    await db.partners.update_one(
        {"id": partner_id},
        {"$set": {
            "is_active": False,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    logger.info(f"Partner deactivated: {partner_id}")
    
    return {"success": True, "message": "Partner deactivated"}


@router.post("/partners/{partner_id}/regenerate-key")
async def regenerate_partner_api_key(partner_id: str, request: Request):
    """Regenerate API key for a partner (invalidates old key)"""
    await verify_admin(request)
    db = get_db(request)
    
    partner = await db.partners.find_one({"id": partner_id})
    if not partner:
        raise HTTPException(status_code=404, detail="Partner not found")
    
    # Generate new API key
    from routes.partner_api import generate_api_key
    full_key, key_hash, key_prefix = generate_api_key()
    
    await db.partners.update_one(
        {"id": partner_id},
        {"$set": {
            "api_key_hash": key_hash,
            "api_key_prefix": key_prefix,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    logger.info(f"Partner API key regenerated: {partner_id}")
    
    return {
        "success": True,
        "api_key": full_key,
        "api_key_prefix": key_prefix,
        "warning": "Save this API key now. It will not be shown again. The old key has been invalidated."
    }


@router.get("/partners/{partner_id}/bookings")
async def get_partner_bookings(
    partner_id: str,
    request: Request,
    status: Optional[str] = None,
    limit: int = 50,
    skip: int = 0
):
    """Get bookings for a specific partner"""
    await verify_admin(request)
    db = get_db(request)
    
    partner = await db.partners.find_one({"id": partner_id})
    if not partner:
        raise HTTPException(status_code=404, detail="Partner not found")
    
    query = {"partner_id": partner_id}
    if status:
        query["status"] = status
    
    total = await db.partner_bookings.count_documents(query)
    
    bookings = await db.partner_bookings.find(
        query,
        {"_id": 0}
    ).sort("created_at", -1).skip(skip).limit(min(limit, 100)).to_list(100)
    
    # Enrich with mentor names
    mentor_ids = list(set(b.get("mentor_id") for b in bookings if b.get("mentor_id")))
    mentors = await db.mentors.find(
        {"id": {"$in": mentor_ids}},
        {"_id": 0, "id": 1, "name": 1}
    ).to_list(100)
    mentor_map = {m["id"]: m["name"] for m in mentors}
    
    for booking in bookings:
        booking["mentor_name"] = mentor_map.get(booking.get("mentor_id"), "Unknown")
        if booking.get("created_at"):
            booking["created_at"] = booking["created_at"].isoformat()
        if booking.get("updated_at"):
            booking["updated_at"] = booking["updated_at"].isoformat()
        if booking.get("cancelled_at"):
            booking["cancelled_at"] = booking["cancelled_at"].isoformat()
    
    return {
        "partner_name": partner.get("name"),
        "bookings": bookings,
        "total": total,
        "limit": limit,
        "skip": skip
    }


@router.get("/partners/mentors/available")
async def get_available_mentors_for_partners(
    request: Request, 
    include_hidden: bool = True,
    include_inactive: bool = False
):
    """Get list of mentors that can be assigned to partners.
    
    Query params:
    - include_hidden: If true, includes hidden mentors (default: true)
    - include_inactive: If true, includes inactive mentors (default: false)
    """
    await verify_admin(request)
    db = get_db(request)
    
    # Build query - always exclude deleted
    query = {
        "is_deleted": {"$ne": True}
    }
    
    # Optionally filter by active status
    if not include_inactive:
        query["is_active"] = True
    
    # Optionally filter by hidden status
    if not include_hidden:
        query["is_hidden"] = {"$ne": True}
    
    mentors = await db.mentors.find(
        query,
        {
            "_id": 0,
            "id": 1,
            "name": 1,
            "picture": 1,
            "consulting_firm": 1,
            "title": 1,
            "specialization": 1,
            "is_hidden": 1,
            "is_active": 1
        }
    ).sort("name", 1).to_list(200)
    
    return {"mentors": mentors, "count": len(mentors)}
